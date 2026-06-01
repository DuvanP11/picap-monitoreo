"""
Extractor de facturas PIBOX desde Google Drive (MINTIC v1).
============================================================

Lee PDFs de una carpeta compartida en Google Drive vía Service Account,
extrae los campos clave (numero factura, fechas, total, periodo) usando
regex sobre el texto del PDF y genera un Excel consolidado.

Flujo:
    1. Auth con Service Account (archivo JSON local).
    2. Lista todos los .pdf en FOLDER_ID (recursivo opcional).
    3. Descarga cada PDF a memoria (BytesIO).
    4. Extrae texto con pdfplumber + regex.
    5. Genera Excel local. Opcionalmente lo sube al mismo Drive.

Setup inicial (1 vez):
    Ver README_SETUP.md en esta carpeta.

Uso:
    python procesar_facturas_mintic.py
    python procesar_facturas_mintic.py --folder-id <ID>
    python procesar_facturas_mintic.py --output mi_excel.xlsx
    python procesar_facturas_mintic.py --upload   # sube el Excel al Drive

Requisitos:
    pip install pdfplumber google-api-python-client google-auth openpyxl
"""

from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import pdfplumber
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Windows console por defecto es cp1252 y choca con los emojis del output.
# Forzar UTF-8 en stdout/stderr así corre sin tirar UnicodeEncodeError.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, Exception):
    pass


# ════════════════════════════════════════════════════════════════════════════
# CONFIG — ajustá estas constantes a tu setup
# ════════════════════════════════════════════════════════════════════════════

# Path al archivo JSON del Service Account (descargado de Google Cloud Console).
# Lo más seguro es ponerlo en una carpeta privada y NO subir al git.
SERVICE_ACCOUNT_JSON = os.environ.get(
    "MINTIC_SA_JSON",
    str(Path(__file__).parent / "credentials" / "service_account.json"),
)

# IDs de las carpetas-raíz del Drive (una por año). Cada año fue compartido
# individualmente con el Service Account vía Drive Share.
#
# Convención interna por año (la que descubrimos al explorar con --ls):
#     <YYYY>/ ?? / <n>. <Mes>/...   ← los PDFs viven en "3. Marzo", etc.
#
# Para procesar un mes puntual usá --mes Marzo --anio 2024 (o --subfolder).
# Si querés agregar 2025/2026, compartilos primero con el SA y agregalos acá.
CARPETAS_POR_ANIO = {
    2022: "10Asn1x-vsnKIedrRYpaVcZQv6rgynfZo",
    2023: "1E4Xr3tGwZyRkzQZR-EiQxkHuKlNZAqVL",
    2024: "1OzFBhx87jjYtZB3AR7nZmLYrdGrHezwD",
    2025: "1Z--QA6uptgMBjWy7lKCkuWc7LTlFnXTC",
    2026: "1IZQPtrEJ0v4HBTe3bOlvjJKGaM9BU06V",
}

# Default histórico (para retro-compatibilidad con --folder-id sin --anio).
# Apunta a 2024 (el más reciente disponible al momento de armar el script).
FOLDER_ID = os.environ.get(
    "MINTIC_FOLDER_ID",
    CARPETAS_POR_ANIO[2024],
)

# Scopes mínimos: solo lectura de archivos.
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# Si --upload se pasa, además necesitamos drive.file para subir.
SCOPES_UPLOAD = ["https://www.googleapis.com/auth/drive"]

OUTPUT_DEFAULT = "facturas_pibox_extraidas.xlsx"


# ════════════════════════════════════════════════════════════════════════════
# EXTRACCIÓN — regex sobre el texto del PDF
# ════════════════════════════════════════════════════════════════════════════

# Mapa nombre de mes (en español, normalizado: lowercase + sin tilde) → número.
MES_NOMBRE_A_NUM = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9,  # ambas variantes
    "octubre": 10, "noviembre": 11, "diciembre": 12,
}


def _normalizar_mes(nombre: str) -> str:
    """Lowercase + remueve tildes para hacer match robusto contra MES_NOMBRE_A_NUM."""
    import unicodedata
    s = unicodedata.normalize("NFKD", nombre or "")
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


@dataclass
class FacturaExtraida:
    Archivo: str
    Numero_Factura: str
    NIT_Cliente: str
    Cliente: str
    Ciudad: str
    Fecha_Emision: str
    Fecha_Vencimiento: str
    Total_Pagar: int
    Fecha_Inicio_Periodo: str
    Fecha_Fin_Periodo: str
    Periodo_Completo: str


# Lista de ciudades colombianas comunes, usada para separar nombre del
# cliente y ciudad en la línea "NIT  DV  <NOMBRE>  <CIUDAD>".
# Ordenadas por longitud descendente (más específicas primero) para que
# "Bogotá D.C." matchee antes que "Bogotá".
CIUDADES_CO = sorted([
    "Bogotá D.C.", "Bogota D.C.", "Bogotá D. C.", "Bogota D. C.",
    "Bogotá", "Bogota",
    "Medellín", "Medellin",
    "Santa Marta", "Santa Marta D.T.C.H.",
    "Cartagena de Indias", "Cartagena",
    "Cali", "Barranquilla", "Cúcuta", "Cucuta",
    "Bucaramanga", "Pereira", "Manizales", "Ibagué", "Ibague",
    "Soledad", "Soacha", "Villavicencio", "Pasto",
    "Montería", "Monteria", "Neiva", "Armenia",
    "Popayán", "Popayan", "Valledupar",
    "Tunja", "Sincelejo", "Riohacha", "Quibdó", "Quibdo",
    "Florencia", "Yopal", "Leticia", "Mocoa", "Inírida", "Inirida", "Arauca",
    "Bello", "Itagüí", "Itagui", "Envigado", "Sabaneta", "Rionegro",
    "Chía", "Chia", "Mosquera", "Funza", "Facatativá", "Facatativa",
    "Zipaquirá", "Zipaquira", "Tocancipá", "Tocancipa", "Madrid",
    "Palmira", "Buga", "Buenaventura", "Tuluá", "Tulua",
    "Floridablanca", "Girón", "Giron", "Piedecuesta",
    "Apartadó", "Apartado", "Turbaco", "Magangué", "Magangue",
    "Sogamoso", "Duitama", "Chiquinquirá", "Chiquinquira",
    "Maicao", "Fonseca",
], key=len, reverse=True)


def extraer_campos(pdf_bytes: bytes, archivo: str) -> FacturaExtraida:
    """Extrae los 10 campos del PDF. Si algo no matchea, devuelve string vacío."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        texto = "\n".join((p.extract_text() or "") for p in pdf.pages)

    # 1. Numero_Factura — del nombre del archivo (más confiable que el texto)
    m = re.search(r"EXP-?(\d+)", archivo, re.IGNORECASE)
    num_factura = f"EXP{int(m.group(1))}" if m else _buscar_exp_en_texto(texto)

    # 2-3. Fechas Emisión + Vencimiento — formato 'DD/MM/YYYY HH:MM:SS a.m./p.m.'
    #     Hay 2 fechas con hora en el PDF: emisión (primera) y vencimiento (segunda).
    #     Soporta tanto 'a.m.' como 'p.m.', con o sin espacios/puntos intermedios.
    fechas_con_hora = re.findall(
        r"(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}:\d{2}\s*(?:a|p)\.?\s*m\.?",
        texto,
        re.IGNORECASE,
    )
    fecha_emis = fechas_con_hora[0] if len(fechas_con_hora) >= 1 else ""
    fecha_venc = fechas_con_hora[1] if len(fechas_con_hora) >= 2 else ""

    # 4-6. NIT + Cliente + Ciudad — la línea inmediatamente debajo del header
    #      "NIT / CÉDULA   CLIENTE   CIUDAD". El formato real es:
    #         <NIT> <DV?> <NOMBRE_CLIENTE> <CIUDAD>
    #      Donde NIT son los dígitos del NIT (puede tener DV separado por espacio).
    #      Para separar nombre y ciudad uso una lista de ciudades CO conocidas.
    nit_cliente, cliente, ciudad = "", "", ""
    m = re.search(
        r"NIT\s*/\s*C[ÉE]DULA\s+CLIENTE\s+CIUDAD[\r\n]+([^\r\n]+)",
        texto,
        re.IGNORECASE,
    )
    if m:
        linea = m.group(1).strip()
        # Capturar el NIT al inicio. Formato común: "<nit> <dv> <nombre> <ciudad>"
        # Empresas suelen tener NIT de 9-10 dígitos + DV de 1 dígito separado.
        # Personas naturales suelen tener cédula de 7-10 dígitos sin DV separado.
        m2 = re.match(r"^(\d+)\s+(\d)\s+(.+)$", linea)  # con DV separado
        if m2:
            nit_cliente = m2.group(1)
            cli_cdd = m2.group(3).strip()
        else:
            m2 = re.match(r"^(\d+)\s+(.+)$", linea)  # sin DV (poco común)
            if m2:
                nit_cliente = m2.group(1)
                cli_cdd = m2.group(2).strip()
            else:
                cli_cdd = ""
        if cli_cdd:
            # Buscar ciudad en la lista (longest match first)
            for c in CIUDADES_CO:
                if cli_cdd.lower().endswith(c.lower()):
                    cliente = cli_cdd[: -len(c)].strip()
                    ciudad = c
                    break
            if not cliente:
                # Fallback: asumir que la última palabra es la ciudad
                partes = cli_cdd.rsplit(" ", 1)
                if len(partes) == 2:
                    cliente, ciudad = partes[0].strip(), partes[1].strip()
                else:
                    cliente = cli_cdd

    # 3. Total_Pagar — '629.400,00' → 629400
    m = re.search(r"TOTAL\s+A\s+PAGAR\s+([\d.,]+)", texto, re.IGNORECASE)
    total = 0
    if m:
        raw = m.group(1).replace(".", "").replace(",", ".")
        try:
            total = int(float(raw))
        except ValueError:
            total = 0

    # 4-6. Período — soporta 3 formatos reales observados en facturas Pibox:
    #
    #   FORMATO A: "desde el DD/MM/YYYY hasta el DD/MM/YYYY"  (formato viejo)
    #   FORMATO B: "DEL <D1> AL <D2> de <mes> del año <YYYY>" (quincena)
    #   FORMATO C: "mes de <mes> del año <YYYY>"              (mes entero)
    #
    # Orden: probar primero los más específicos (A y B) y caer a C como fallback.
    fecha_ini = fecha_fin = ""

    # A. "desde el DD/MM/YYYY hasta el DD/MM/YYYY"
    m = re.search(
        r"desde\s+el\s+(\d{2}/\d{2}/\d{4})\s+hasta\s+el\s+(\d{2}/\d{2}/\d{4})",
        texto,
        re.IGNORECASE,
    )
    if m:
        fecha_ini, fecha_fin = m.group(1), m.group(2)

    # B. "DEL 16 AL 30 de noviembre del año 2024"
    if not fecha_ini:
        m = re.search(
            r"DEL\s+(\d{1,2})\s+AL\s+(\d{1,2})\s+de\s+(\w+)\s+del\s+a[ñn]o\s+(\d{4})",
            texto,
            re.IGNORECASE,
        )
        if m:
            d1, d2, mes_nombre, anio = int(m.group(1)), int(m.group(2)), m.group(3).lower(), int(m.group(4))
            mes_num = MES_NOMBRE_A_NUM.get(_normalizar_mes(mes_nombre))
            if mes_num:
                fecha_ini = f"{d1:02d}/{mes_num:02d}/{anio}"
                fecha_fin = f"{d2:02d}/{mes_num:02d}/{anio}"

    # C. "mes de noviembre del año 2024" (mes completo)
    if not fecha_ini:
        m = re.search(
            r"mes\s+de\s+(\w+)\s+del\s+a[ñn]o\s+(\d{4})",
            texto,
            re.IGNORECASE,
        )
        if m:
            mes_nombre, anio = m.group(1).lower(), int(m.group(2))
            mes_num = MES_NOMBRE_A_NUM.get(_normalizar_mes(mes_nombre))
            if mes_num:
                from calendar import monthrange
                ultimo = monthrange(anio, mes_num)[1]
                fecha_ini = f"01/{mes_num:02d}/{anio}"
                fecha_fin = f"{ultimo:02d}/{mes_num:02d}/{anio}"

    periodo = f"{fecha_ini} hasta el {fecha_fin}" if fecha_ini and fecha_fin else ""

    return FacturaExtraida(
        Archivo=archivo,
        Numero_Factura=num_factura,
        NIT_Cliente=nit_cliente,
        Cliente=cliente,
        Ciudad=ciudad,
        Fecha_Emision=fecha_emis,
        Fecha_Vencimiento=fecha_venc,
        Total_Pagar=total,
        Fecha_Inicio_Periodo=fecha_ini,
        Fecha_Fin_Periodo=fecha_fin,
        Periodo_Completo=periodo,
    )


def _buscar_exp_en_texto(texto: str) -> str:
    m = re.search(r"\bEXP\s*0*(\d+)\b", texto, re.IGNORECASE)
    return f"EXP{int(m.group(1))}" if m else ""


# ════════════════════════════════════════════════════════════════════════════
# GOOGLE DRIVE — auth, listar y descargar
# ════════════════════════════════════════════════════════════════════════════

def crear_servicio_drive(scopes: list[str]):
    if not Path(SERVICE_ACCOUNT_JSON).exists():
        sys.exit(
            f"❌ No encuentro el JSON del Service Account en: {SERVICE_ACCOUNT_JSON}\n"
            f"   Configurá MINTIC_SA_JSON o copiá el archivo ahí.\n"
            f"   Ver README_SETUP.md para crear el Service Account."
        )
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON, scopes=scopes
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _drive_list(servicio, query: str, fields: str = "nextPageToken, files(id, name, mimeType, modifiedTime, size)") -> list[dict]:
    """Helper genérico: ejecuta files().list con paginación + soporte shared drives."""
    archivos = []
    page_token = None
    while True:
        resp = (
            servicio.files()
            .list(
                q=query,
                fields=fields,
                pageSize=200,
                pageToken=page_token,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            )
            .execute()
        )
        archivos.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return archivos


def listar_pdfs_en_carpeta(servicio, folder_id: str) -> list[dict]:
    """Lista PDFs directamente en la carpeta (no recursivo)."""
    query = (
        f"'{folder_id}' in parents and "
        f"mimeType='application/pdf' and "
        f"trashed=false"
    )
    return _drive_list(servicio, query)


def listar_subcarpetas(servicio, folder_id: str) -> list[dict]:
    """Lista subcarpetas directas de la carpeta dada."""
    query = (
        f"'{folder_id}' in parents and "
        f"mimeType='application/vnd.google-apps.folder' and "
        f"trashed=false"
    )
    return _drive_list(servicio, query)


def listar_pdfs_recursivo(servicio, folder_id: str, _profundidad: int = 0) -> list[dict]:
    """
    Lista TODOS los PDFs bajo folder_id, recorriendo subcarpetas. Cada archivo
    devuelto trae un campo extra '_path' con la ruta relativa al folder raíz
    (útil para reportar dónde estaba cada PDF).
    """
    if _profundidad > 6:
        # Safety: evita recursión infinita por carpetas auto-referenciadas.
        return []

    pdfs = listar_pdfs_en_carpeta(servicio, folder_id)
    for p in pdfs:
        p.setdefault("_path", p["name"])

    for sub in listar_subcarpetas(servicio, folder_id):
        sub_pdfs = listar_pdfs_recursivo(servicio, sub["id"], _profundidad + 1)
        for p in sub_pdfs:
            p["_path"] = f"{sub['name']}/{p.get('_path', p['name'])}"
        pdfs.extend(sub_pdfs)
    return pdfs


def _normalizar(s: str) -> str:
    """Lowercase + remueve tildes para comparar nombres de carpeta tolerante."""
    import unicodedata
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def resolver_subfolder_path(servicio, root_id: str, path: str) -> tuple[str, list[str]]:
    """
    Navega una ruta tipo 'PDF/Marzo' bajo root_id y devuelve (folder_id, trail).
    Match case-insensitive y tolerante a:
      - tildes (Marzo ↔ MARZO ↔ marzo ↔ Márzo)
      - prefijos numéricos ('03-Marzo' matchea 'Marzo')
      - sufijos de año ('Marzo 2026' matchea 'Marzo')
    Si hay ambigüedad o no encuentra, lanza ValueError con sugerencias.
    """
    actual_id = root_id
    trail = []
    for parte in [p for p in path.split("/") if p.strip()]:
        subs = listar_subcarpetas(servicio, actual_id)
        objetivo = _normalizar(parte)

        # 1. Match exacto (case+tilde-insensitive)
        exactos = [s for s in subs if _normalizar(s["name"]) == objetivo]
        if len(exactos) == 1:
            actual_id = exactos[0]["id"]
            trail.append(exactos[0]["name"])
            continue

        # 2. Match "contains" — captura '03-Marzo', 'Marzo 2026', etc.
        flexibles = [s for s in subs if objetivo in _normalizar(s["name"])]
        if len(flexibles) == 1:
            actual_id = flexibles[0]["id"]
            trail.append(flexibles[0]["name"])
            continue
        if len(flexibles) > 1:
            opciones = ", ".join(s["name"] for s in flexibles)
            raise ValueError(
                f"'{parte}' es ambiguo en este nivel. Coincidencias: {opciones}. "
                f"Usá --subfolder con el nombre completo de la subcarpeta."
            )

        # 3. Nada matcheó: mostrar qué hay para que el usuario corrija.
        disponibles = ", ".join(s["name"] for s in subs) or "(carpeta vacía)"
        raise ValueError(
            f"No encontré '{parte}' bajo '{'/'.join(trail) or '<root>'}'.\n"
            f"   Subcarpetas disponibles aquí: {disponibles}"
        )
    return actual_id, trail


def explorar_drive(servicio, folder_id: str, profundidad_max: int = 2, _nivel: int = 0, _prefijo: str = "") -> None:
    """Imprime el árbol de subcarpetas + count de PDFs por carpeta. Solo para --ls."""
    subs = listar_subcarpetas(servicio, folder_id)
    pdfs_aqui = listar_pdfs_en_carpeta(servicio, folder_id)
    if _nivel == 0 and pdfs_aqui:
        print(f"{_prefijo}📄 {len(pdfs_aqui)} PDF(s) directamente aquí")
    for i, sub in enumerate(subs):
        es_ultimo = (i == len(subs) - 1)
        rama = "└── " if es_ultimo else "├── "
        conteo = listar_pdfs_en_carpeta(servicio, sub["id"])
        sufijo = f"  ({len(conteo)} PDFs)" if conteo else ""
        print(f"{_prefijo}{rama}📁 {sub['name']}{sufijo}")
        if _nivel + 1 < profundidad_max:
            nuevo_prefijo = _prefijo + ("    " if es_ultimo else "│   ")
            explorar_drive(servicio, sub["id"], profundidad_max, _nivel + 1, nuevo_prefijo)


def descargar_pdf(servicio, file_id: str) -> bytes:
    request = servicio.files().get_media(fileId=file_id, supportsAllDrives=True)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request, chunksize=1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def subir_excel(servicio, ruta_local: str, folder_id: str) -> str:
    """Sube un archivo al Drive bajo folder_id. Devuelve file_id."""
    nombre = Path(ruta_local).name
    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    metadata = {"name": nombre, "parents": [folder_id]}
    with open(ruta_local, "rb") as fh:
        media = MediaIoBaseUpload(fh, mimetype=mime, resumable=False)
        resp = (
            servicio.files()
            .create(
                body=metadata,
                media_body=media,
                fields="id, webViewLink",
                supportsAllDrives=True,
            )
            .execute()
        )
    return resp.get("webViewLink") or resp.get("id", "")


# ════════════════════════════════════════════════════════════════════════════
# EXCEL — generación con formato profesional
# ════════════════════════════════════════════════════════════════════════════

def generar_excel(filas: list[FacturaExtraida], ruta_salida: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    encabezados = list(FacturaExtraida.__dataclass_fields__.keys())

    # Estilo header
    header_fill = PatternFill("solid", start_color="0E7490")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Datos
    for row_idx, fac in enumerate(filas, start=2):
        d = asdict(fac)
        for col_idx, h in enumerate(encabezados, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d[h])
            cell.alignment = Alignment(vertical="center")
            if h == "Total_Pagar":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Anchos columnas (cabe el contenido típico)
    anchos = {
        "Archivo": 36,
        "Numero_Factura": 14,
        "NIT_Cliente": 14,
        "Cliente": 38,
        "Ciudad": 16,
        "Fecha_Emision": 16,
        "Fecha_Vencimiento": 16,
        "Total_Pagar": 14,
        "Fecha_Inicio_Periodo": 20,
        "Fecha_Fin_Periodo": 20,
        "Periodo_Completo": 32,
    }
    for col_idx, h in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(h, 18)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    wb.save(ruta_salida)


# ════════════════════════════════════════════════════════════════════════════
# CLI
# ════════════════════════════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extractor facturas PIBOX desde Google Drive",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
    # Listar la estructura del Drive (sin procesar nada)
    python procesar_facturas_mintic.py --ls

    # Procesar PDF/Marzo (atajo cómodo)
    python procesar_facturas_mintic.py --mes Marzo

    # Procesar una subcarpeta arbitraria
    python procesar_facturas_mintic.py --subfolder "PDF/Marzo"

    # Procesar TODOS los PDFs bajo el root (todos los meses)
    python procesar_facturas_mintic.py --recursive

    # Modo debug: solo los primeros 3 PDFs
    python procesar_facturas_mintic.py --mes Marzo --limit 3
""",
    )
    parser.add_argument("--folder-id", help="ID de la carpeta raíz del Drive (override de FOLDER_ID / --anio)")
    parser.add_argument("--anio", type=int, help=f"Año a procesar (usa CARPETAS_POR_ANIO). Disponibles: {sorted(CARPETAS_POR_ANIO.keys())}")
    parser.add_argument("--subfolder", help="Ruta de subcarpeta a procesar (ej. 'PDF/Marzo'). Match case-insensitive.")
    parser.add_argument("--mes", help="Atajo: equivale a --subfolder '<mes>' (si la estructura del año tiene los meses al primer nivel)")
    parser.add_argument("--recursive", action="store_true",
                        help="Procesa TODOS los PDFs bajo el punto de partida (default: solo nivel directo)")
    parser.add_argument("--ls", action="store_true",
                        help="Solo lista la estructura del Drive (subcarpetas + count de PDFs) y sale")
    parser.add_argument("--output", default=OUTPUT_DEFAULT, help="Path del Excel a generar")
    parser.add_argument("--upload", action="store_true", help="Sube el Excel resultante al mismo Drive")
    parser.add_argument("--limit", type=int, default=0, help="Procesar solo los primeros N PDFs (debug)")
    args = parser.parse_args()

    # Prioridad: --folder-id > --anio > FOLDER_ID default
    if args.folder_id:
        folder_id = args.folder_id
    elif args.anio:
        if args.anio not in CARPETAS_POR_ANIO:
            disp = sorted(CARPETAS_POR_ANIO.keys())
            sys.exit(f"❌ Año {args.anio} no configurado. Disponibles: {disp}. "
                     f"Compartilo con el SA y agregalo en CARPETAS_POR_ANIO del script.")
        folder_id = CARPETAS_POR_ANIO[args.anio]
        print(f"🗓  Año {args.anio} → folder {folder_id}")
    else:
        folder_id = FOLDER_ID

    if folder_id == "PEGAR_AQUI_EL_FOLDER_ID":
        sys.exit("❌ Configurá FOLDER_ID en el script o pasalo con --folder-id")

    # --mes es azúcar sobre --subfolder.
    # NOTA: hasta acá no sabemos si la estructura es "PDF/Marzo" o "Marzo" directo.
    # Probamos primero "Marzo" (estructura plana) y, si falla, "PDF/Marzo".
    subfolder = args.subfolder
    if args.mes and not subfolder:
        subfolder = args.mes  # Primero probamos sin prefijo

    scopes = SCOPES_UPLOAD if args.upload else SCOPES
    print(f"🔑 Autenticando con Service Account: {SERVICE_ACCOUNT_JSON}")
    servicio = crear_servicio_drive(scopes)

    # Modo --ls: solo explorar y salir.
    if args.ls:
        print(f"📂 Explorando estructura desde {folder_id}")
        print(f"   (raíz)")
        explorar_drive(servicio, folder_id, profundidad_max=3)
        print("\n✅ Fin del listado.")
        return 0

    # Resolver el punto de partida (root o una subcarpeta navegada).
    # Si --mes no matchea directo, intentamos "PDF/<mes>" como fallback.
    if subfolder:
        try:
            print(f"📂 Resolviendo subcarpeta '{subfolder}' desde root {folder_id} …")
            folder_id, trail = resolver_subfolder_path(servicio, folder_id, subfolder)
            print(f"   ✓ Ruta resuelta: {'/'.join(trail)}  →  {folder_id}")
        except ValueError as e:
            # Fallback: si era --mes solo, probar con "PDF/<mes>"
            if args.mes and not args.subfolder and "/" not in subfolder:
                try:
                    sub2 = f"PDF/{subfolder}"
                    print(f"   ↳ probando fallback '{sub2}' …")
                    folder_id, trail = resolver_subfolder_path(servicio, folder_id, sub2)
                    print(f"   ✓ Ruta resuelta: {'/'.join(trail)}  →  {folder_id}")
                except ValueError as e2:
                    sys.exit(f"❌ {e}\n   Fallback también falló: {e2}")
            else:
                sys.exit(f"❌ {e}")

    # Listar PDFs (recursivo o no).
    if args.recursive:
        print(f"📂 Listando PDFs recursivamente …")
        archivos = listar_pdfs_recursivo(servicio, folder_id)
    else:
        print(f"📂 Listando PDFs directos en {folder_id} …")
        archivos = listar_pdfs_en_carpeta(servicio, folder_id)

    if args.limit:
        archivos = archivos[: args.limit]
    print(f"   Encontrados: {len(archivos)} PDF(s)")
    if not archivos:
        sys.exit(
            "⚠️  No se encontraron PDFs.\n"
            "   Verificá que:\n"
            "   1. Compartiste la carpeta correcta con el Service Account (Viewer).\n"
            "   2. La estructura coincide con --subfolder/--mes (probá --ls para ver el árbol).\n"
            "   3. Los archivos son realmente .pdf (no .docx, .xlsx)."
        )

    filas: list[FacturaExtraida] = []
    for i, arch in enumerate(archivos, start=1):
        nombre = arch["name"]
        etiqueta = arch.get("_path", nombre)
        print(f"   [{i:>3d}/{len(archivos)}] {etiqueta} …", end=" ", flush=True)
        try:
            pdf_bytes = descargar_pdf(servicio, arch["id"])
            fac = extraer_campos(pdf_bytes, nombre)
            filas.append(fac)
            print(f"✓ {fac.Numero_Factura}  ${fac.Total_Pagar:,}".replace(",", "."))
        except Exception as e:
            print(f"❌ {type(e).__name__}: {e}")

    if not filas:
        sys.exit("⚠️  No se pudo extraer nada.")

    # Ordenar por número de factura ascendente
    filas.sort(key=lambda f: int(re.sub(r"\D", "", f.Numero_Factura) or 0))

    print(f"📊 Generando Excel: {args.output}")
    generar_excel(filas, args.output)
    print(f"   ✓ {len(filas)} filas escritas.")

    # JSON paralelo al xlsx — usado por el portal Rails (MinticController) para
    # leer las facturas sin gem 'roo'. Mismas filas, formato consumible directo.
    json_path = Path(args.output).with_suffix(".json")
    with open(json_path, "w", encoding="utf-8") as fh:
        json.dump([asdict(f) for f in filas], fh, ensure_ascii=False, indent=2)
    print(f"   ✓ JSON paralelo: {json_path.name}")

    if args.upload:
        print(f"☁️  Subiendo Excel al Drive …")
        url = subir_excel(servicio, args.output, folder_id)
        print(f"   ✓ {url}")

    print("✅ Listo.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
