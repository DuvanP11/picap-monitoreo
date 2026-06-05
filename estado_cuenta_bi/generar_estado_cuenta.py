"""
Generador del Excel "Estado de cuenta SURTITODO".

Replica la plantilla:
  C:\\Users\\Picap\\Documents\\2026\\Abril\\RECAUDOS BI\\Estado de cuenta SURTITODO Abril 2026.xlsm

Estructura de hojas:
  1. "Resumen"          — Logo Pibox + título + tabla resumen con fórmulas.
  2. "Recaudos"         — Query A (CounterDelivery filtrado por Surtitodo). 4 cols.
  3. "Valor Mensajeria" — Query B (BookingCompanyCharge + Commission). 5 cols.

Uso:
    python generar_estado_cuenta.py --mes 2026-04
    python generar_estado_cuenta.py --mes 2026-05 --output "Mi salida.xlsx"

Credenciales CH se leen de env vars: MINTIC_CH_HOST/USER/PASS.
"""
from __future__ import annotations

import argparse
import calendar
import os
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# UTF-8 en Windows
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ════════════════════════════════════════════════════════════════════════════
# ClickHouse
# ════════════════════════════════════════════════════════════════════════════

def ch_request(query: str, host: str, user: str, password: str, timeout: int = 600) -> str:
    url = host.rstrip("/") + "/"
    params = {"query": query + "\nFORMAT TabSeparatedWithNames"}
    r = requests.get(url, params=params, auth=(user, password), timeout=timeout)
    if r.status_code != 200:
        raise RuntimeError(f"CH HTTP {r.status_code}: {r.text[:500]}")
    return r.text


def parse_tsv(text: str) -> tuple[list[str], list[list[str]]]:
    lines = text.rstrip("\n").split("\n")
    if not lines:
        return [], []
    headers = lines[0].split("\t")
    rows = [ln.split("\t") for ln in lines[1:]] if len(lines) > 1 else []
    return headers, rows


def to_num(v):
    if v is None or v == "" or v == "\\N":
        return None
    try:
        if "." in v or "e" in v.lower():
            return float(v)
        return int(v)
    except (ValueError, TypeError):
        return v


# ════════════════════════════════════════════════════════════════════════════
# Queries
# ════════════════════════════════════════════════════════════════════════════

# Query A: Recaudos (CounterDelivery filtrado por Surtitodo).
# 4 cols: ID TRANSACCION, FECHA, DESCRIPCION, MONTO.
QUERY_RECAUDOS = """
WITH
q_service_types AS (
  SELECT
    _id,
    name_es AS name,
    CASE
      WHEN name_es IN (
        'Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor', 'Moto favor',
        'Carga', 'Carga Carry', 'Carga Moto-Vagón', 'Carga NHR', 'Carga NKR', 'Carga NPR',
        'Mensajería', 'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'
      ) THEN 'Pibox'
      WHEN name_es IN (
        'Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta','Carro Subasta','Taxi', 'Carro',
        'Carro sin conductor', 'Moto VIP', 'Moto Económica', 'Carro Queen','Rapidín','Espero tranqui',
        'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto', 'Grúa Carro', 'Grúa Moto'
      ) THEN 'Picap'
      ELSE 'Other'
    END AS type
  FROM (SELECT * FROM picapmongoprod.service_types FINAL)
),
q_wat_filtered AS (
  SELECT
    wat._id          AS transaction_id,
    wat.booking_id   AS booking_id,
    wat.package_id   AS package_id,
    wat.account_id   AS account_id,
    wat.amount       AS amount,
    wat.created_at   AS created_at,
    b.created_at     AS booking_created_at,
    pck.reference    AS package_reference,
    b.passenger_id   AS passenger_id,
    p.company_id     AS passenger_company_id,
    compp.name       AS company_name
  FROM picapmongoprod.wallet_account_transactions AS wat FINAL
  INNER JOIN picapmongoprod.packages       AS pck   FINAL ON pck._id   = wat.package_id
  INNER JOIN picapmongoprod.bookings       AS b     FINAL ON b._id     = wat.booking_id
  INNER JOIN picapmongoprod.wallet_accounts AS wa   FINAL ON wa._id    = wat.account_id
  INNER JOIN picapmongoprod.passengers     AS p     FINAL ON p._id     = b.passenger_id
  INNER JOIN picapmongoprod.countries      AS c     FINAL ON c._id     = b.country_id
  INNER JOIN picapmongoprod.companies      AS compp FINAL ON compp._id = p.company_id
  WHERE
    JSONExtractString(c.name, 'es') = 'Colombia'
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND pck.counter_delivery = 'true'
    AND wat._type = 'WalletAccountCounterDeliveryPaymentTransaction'
    AND lowerUTF8(compp.name) LIKE '%surtitodo%'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
)
SELECT
  qtf.transaction_id                                            AS ID_TRANSACCION,
  toDate(toTimeZone(qtf.booking_created_at, 'America/Bogota'))  AS FECHA,
  qtf.package_reference                                         AS DESCRIPCION,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100 AS MONTO
FROM q_wat_filtered qtf
ORDER BY FECHA ASC
"""


# Query B: Valor Mensajería (BookingCompanyCharge + CommissionCompanyPayment).
# 5 cols: ID SERVICIO, FECHA, EMPRESA, TIPO VEHICULO, MONTO.
QUERY_VALOR_MENSAJERIA = """
WITH q_service_types AS (
  SELECT
    _id,
    any(name_es) AS name,
    any(
      multiIf(
        name_es IN (
          'Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor',
          'Moto favor', 'Carga', 'Carga Carry', 'Carga Moto-Vagón',
          'Carga NHR', 'Carga NKR', 'Carga NPR', 'Mensajería',
          'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'
        ), 'Pibox',
        name_es IN (
          'Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta', 'Taxi',
          'Carro', 'Rapidín', 'Carro sin conductor', 'Moto VIP',
          'Moto Económica', 'Carro Queen', 'Espero tranqui',
          'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto',
          'Grúa Carro', 'Grúa Moto'
        ), 'Picap',
        'Other'
      )
    ) AS type
  FROM picapmongoprod.service_types
  GROUP BY _id
),
q_transactions_filtered AS (
  SELECT
    wat._id          AS _id,
    wat.booking_id   AS booking_id,
    wat.account_id   AS account_id,
    wat._type        AS txt_type,
    wat.created_at   AS created_at,
    wat.amount       AS amount,
    s.served_vehicle_type_id AS served_vehicle_type_id,
    comp.name        AS company_name,
    st.type          AS service_type
  FROM picapmongoprod.wallet_account_transactions wat FINAL
  ANY LEFT JOIN picapmongoprod.bookings        s    FINAL ON s._id    = wat.booking_id
  ANY LEFT JOIN q_service_types                st         ON st._id   = s.requested_service_type_id
  ANY LEFT JOIN picapmongoprod.wallet_accounts wa   FINAL ON wa._id   = wat.account_id
  ANY LEFT JOIN picapmongoprod.companies       comp FINAL ON comp._id = s.company_id
  WHERE
    wat._type IN (
      'WalletAccountTransactionBookingCompanyCharge',
      'WalletAccountTransactionCommissionCompanyPayment'
    )
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND st.type = 'Pibox'
    AND lowerUTF8(comp.name) LIKE '%surtitodo%'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
)
SELECT
  qtf.booking_id                                                AS ID_SERVICIO,
  toDate(toTimeZone(qtf.created_at, 'America/Bogota'))          AS FECHA,
  qtf.company_name                                              AS EMPRESA,
  JSONExtractString(vt.name, 'es')                              AS TIPO_VEHICULO,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100 AS MONTO
FROM q_transactions_filtered qtf
LEFT JOIN picapmongoprod.vehicle_types AS vt FINAL ON vt._id = qtf.served_vehicle_type_id
ORDER BY FECHA ASC, qtf.booking_id
"""


# ════════════════════════════════════════════════════════════════════════════
# Estilos
# ════════════════════════════════════════════════════════════════════════════

MORADO = "5B21B6"   # Morado Pibox / Surtitodo
WHITE  = "FFFFFF"
HDR_BLUE = "1F4E78"

MORADO_FONT  = Font(bold=True, color=MORADO, size=14)
WHITE_BOLD_ITALIC = Font(bold=True, italic=True, color=WHITE, size=14)
BOLD_FONT = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)

MORADO_FILL = PatternFill("solid", start_color=MORADO)
HDR_FILL    = PatternFill("solid", start_color=HDR_BLUE)
LIGHT_FILL  = PatternFill("solid", start_color="F3E8FF")

THIN = Side(border_style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center")
RIGHT  = Alignment(horizontal="right", vertical="center")

INT_FMT  = "#,##0"
COP_FMT  = '_-"$"* #,##0_-;[Red]-"$"* #,##0_-;_-"$"* "-"_-;_-@_-'


# ════════════════════════════════════════════════════════════════════════════
# Helpers de período
# ════════════════════════════════════════════════════════════════════════════

MESES_ES = {
    1: "enero",    2: "febrero", 3: "marzo",      4: "abril",
    5: "mayo",     6: "junio",   7: "julio",      8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


def mes_a_info(mes_str: str) -> dict:
    """'2026-04' → dict con desde, hasta, mes_nombre, etc."""
    año, mes = mes_str.split("-")
    año_i, mes_i = int(año), int(mes)
    _, last_day = calendar.monthrange(año_i, mes_i)
    desde = f"{año}-{mes}-01"
    hasta = f"{año}-{mes}-{last_day:02d}"
    mes_nombre = MESES_ES[mes_i]                              # 'abril'
    mes_capi   = mes_nombre.capitalize()                      # 'Abril'
    return {
        "desde": desde, "hasta": hasta,
        "last_day": last_day,
        "mes_nombre": mes_nombre, "mes_capi": mes_capi,
        "anio": año_i,
        # "Periodo del 01 al 30 de abril 2026"
        "periodo_txt": f"Periodo del 01 al {last_day:02d} de {mes_nombre} {año_i}",
        # Para el nombre de archivo / título
        "filename_base": f"Estado de cuenta SURTITODO {mes_capi} {año_i}",
    }


# ════════════════════════════════════════════════════════════════════════════
# Hoja "Resumen"
# ════════════════════════════════════════════════════════════════════════════

def build_hoja_resumen(wb, info, logo_path: Path | None):
    """Hoja Resumen: logo Pibox + título + tabla resumen con fórmulas."""
    ws = wb.create_sheet("Resumen")

    # ── Anchos de columnas ──
    # A:B reservadas para el logo (cols anchas para que el logo tenga buen tamaño)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 24
    for col_l in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col_l].width = 14

    # Alturas de filas para que el logo se vea bien
    for r in range(1, 5):
        ws.row_dimensions[r].height = 28
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 22
    for r in range(7, 12):
        ws.row_dimensions[r].height = 22

    # ── Imagen Pibox (A1:B4) ──
    if logo_path and logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            # Tamaño aproximado para encajar en A1:B4 (~ancho 46*2=92 + alto 28*4=112 px)
            img.width = 140
            img.height = 100
            ws.add_image(img, "A1")
        except Exception as e:
            print(f"   ⚠ No se pudo insertar el logo: {e}")

    # ── Título principal (C2:J2 merged, morado bold) ──
    ws.merge_cells("C2:J2")
    cell_titulo = ws.cell(row=2, column=3, value="Informe recaudo (pagos contra entrega) y costo de la operación de mensajería")
    cell_titulo.font = MORADO_FONT
    cell_titulo.alignment = CENTER

    # ── Header SURTITODO (B5:C5 merged, morado fill + white bold italic) ──
    ws.merge_cells("B5:C5")
    cell_surt = ws.cell(row=5, column=2, value="SURTITODO")
    cell_surt.fill = MORADO_FILL
    cell_surt.font = WHITE_BOLD_ITALIC
    cell_surt.alignment = CENTER

    # ── Período (B6:C6 merged, bold) ──
    ws.merge_cells("B6:C6")
    cell_periodo = ws.cell(row=6, column=2, value=info["periodo_txt"])
    cell_periodo.font = BOLD_FONT
    cell_periodo.alignment = CENTER
    cell_periodo.fill = LIGHT_FILL

    # ── Tabla resumen (filas 7-11) ──
    filas = [
        (7,  "Recaudos",                            "=SUM(Recaudos!D:D)"),
        (8,  "Pago Servicios",                      "=SUM('Valor Mensajeria'!E:E)"),
        (9,  "Comisión del 1%",                     "=-(C7*1%)"),
        (10, "ICA",                                 "=(-C8*9.66)/1000"),
        (11, "valor a pagar despues del cruce:",    "=SUM(C7:C10)"),
    ]
    for r, etiqueta, formula in filas:
        c_label = ws.cell(row=r, column=2, value=etiqueta)
        c_label.alignment = LEFT
        c_label.border = BORDER_ALL
        if r == 11:
            c_label.font = BOLD_FONT
            c_label.fill = LIGHT_FILL

        c_value = ws.cell(row=r, column=3, value=formula)
        c_value.alignment = RIGHT
        c_value.border = BORDER_ALL
        c_value.number_format = COP_FMT
        if r == 11:
            c_value.font = BOLD_FONT
            c_value.fill = LIGHT_FILL


# ════════════════════════════════════════════════════════════════════════════
# Hojas de data (Recaudos / Valor Mensajeria)
# ════════════════════════════════════════════════════════════════════════════

def _set_data_sheet(ws, headers_excel: list[str], headers_ch: list[str], rows: list[list[str]],
                    money_col_letter: str = None):
    """Helper: escribe headers azul + data + formato moneda en una col específica."""
    pos_ch = {h: i for i, h in enumerate(headers_ch)}

    # Headers
    for c_idx, h in enumerate(headers_excel, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HDR_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER_ALL

    # Data
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, h_excel in enumerate(headers_excel, start=1):
            # Buscar la col CH correspondiente (mismo orden que headers_excel)
            src_idx = pos_ch.get(headers_ch[c_idx - 1]) if c_idx - 1 < len(headers_ch) else None
            v = r[src_idx] if (src_idx is not None and src_idx < len(r)) else None
            ws.cell(row=r_idx, column=c_idx, value=to_num(v)).border = BORDER_ALL

    last_row = 1 + len(rows)

    # Formato moneda
    if money_col_letter:
        col_idx = ord(money_col_letter) - ord("A") + 1
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).number_format = COP_FMT

    ws.freeze_panes = "A2"


def build_hoja_recaudos(wb, headers_ch, rows):
    """Hoja Recaudos: 4 cols (ID TRANSACCION | FECHA | DESCRIPCION | MONTO)."""
    ws = wb.create_sheet("Recaudos")

    headers_excel = ["ID TRANSACCION", "FECHA", "DESCRIPCION", "MONTO"]
    headers_ch_mapeo = ["ID_TRANSACCION", "FECHA", "DESCRIPCION", "MONTO"]

    _set_data_sheet(ws, headers_excel, headers_ch_mapeo, rows, money_col_letter="D")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


def build_hoja_valor_mensajeria(wb, headers_ch, rows):
    """Hoja Valor Mensajeria: 5 cols (ID SERVICIO | FECHA | EMPRESA | TIPO VEHICULO | MONTO)."""
    ws = wb.create_sheet("Valor Mensajeria")

    headers_excel = ["ID SERVICIO", "FECHA", "EMPRESA", "TIPO VEHICULO", "MONTO"]
    headers_ch_mapeo = ["ID_SERVICIO", "FECHA", "EMPRESA", "TIPO_VEHICULO", "MONTO"]

    _set_data_sheet(ws, headers_excel, headers_ch_mapeo, rows, money_col_letter="E")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

SCRIPT_DIR = Path(__file__).parent
LOGO_PATH  = SCRIPT_DIR / "pibox_logo.png"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera el Excel 'Estado de cuenta SURTITODO' (3 hojas)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--mes", required=True, help="Mes a procesar YYYY-MM (ej. 2026-04)")
    parser.add_argument("--output", help="Nombre del archivo .xlsx")
    parser.add_argument("--ch-host", default=os.environ.get("MINTIC_CH_HOST", ""))
    parser.add_argument("--ch-user", default=os.environ.get("MINTIC_CH_USER", ""))
    parser.add_argument("--ch-pass", default=os.environ.get("MINTIC_CH_PASS", ""))
    args = parser.parse_args()

    if not args.ch_host or not args.ch_user or not args.ch_pass:
        sys.exit(
            "❌ Credenciales ClickHouse no configuradas.\n"
            "   Setealas como env vars (MINTIC_CH_HOST/USER/PASS) o pasalas con\n"
            "   --ch-host --ch-user --ch-pass."
        )

    try:
        info = mes_a_info(args.mes)
    except Exception as e:
        sys.exit(f"❌ --mes inválido: {args.mes} ({e}). Usá formato YYYY-MM (ej. 2026-04).")

    output = args.output or f"{info['filename_base']}.xlsx"
    output_path = Path(output)

    print(f"🚀 Generando Estado de cuenta SURTITODO")
    print(f"   • Período  : {info['desde']} → {info['hasta']} ({info['periodo_txt']})")
    print(f"   • Output   : {output_path}")
    print()

    # ── Query A: Recaudos ──
    print("📊 Query A: Recaudos (CounterDelivery)…")
    t0 = time.time()
    sql_a = QUERY_RECAUDOS.format(desde=info["desde"], hasta=info["hasta"])
    text_a = ch_request(sql_a, args.ch_host, args.ch_user, args.ch_pass)
    headers_a, rows_a = parse_tsv(text_a)
    print(f"   ✓ {len(rows_a)} filas en {time.time() - t0:.1f}s")

    # ── Query B: Valor Mensajería ──
    print("📊 Query B: Valor Mensajería (Charges + Commission)…")
    t0 = time.time()
    sql_b = QUERY_VALOR_MENSAJERIA.format(desde=info["desde"], hasta=info["hasta"])
    text_b = ch_request(sql_b, args.ch_host, args.ch_user, args.ch_pass)
    headers_b, rows_b = parse_tsv(text_b)
    print(f"   ✓ {len(rows_b)} filas en {time.time() - t0:.1f}s")

    print()
    print("📝 Generando Excel…")
    wb = Workbook()
    wb.remove(wb.active)

    # Hoja Resumen (debe ser la primera)
    build_hoja_resumen(wb, info, LOGO_PATH if LOGO_PATH.exists() else None)
    print(f"   ✓ Hoja 1: 'Resumen' (con logo + título + 5 filas resumen)")

    # Hoja Recaudos
    build_hoja_recaudos(wb, headers_a, rows_a)
    print(f"   ✓ Hoja 2: 'Recaudos' ({len(rows_a)} filas)")

    # Hoja Valor Mensajería
    build_hoja_valor_mensajeria(wb, headers_b, rows_b)
    print(f"   ✓ Hoja 3: 'Valor Mensajeria' ({len(rows_b)} filas)")

    wb.save(output_path)
    print()
    print(f"✅ Listo: {output_path}")
    print(f"   Tamaño: {output_path.stat().st_size / 1024:.1f} KB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
