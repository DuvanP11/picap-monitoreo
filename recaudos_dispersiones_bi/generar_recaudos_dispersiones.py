"""
Generador del reporte "Recaudos y Dispersiones" (7 hojas) para BI.

Replica la plantilla:
  C:\\Users\\Picap\\Documents\\2026\\Abril\\RECAUDOS BI\\Recaudos y Dispersiones Abril 2026.xlsx

Estructura de hojas:
  1. "Dispersiones 1 al {N} {mes}"   — Query A (Daviplata CashOut, Dispersión Recaudo). 7 cols.
  2. "Dispersion 1 al {N} {mes}"     — Pivote Company_name + tipo_dispersion + Σ amount_cents.
  3. "Acumulado Dispersion"          — Tabla manual (Company_name + MONTO + CORTE).
  4. "Data Recaudos"                 — Query B (recaudos CounterDelivery). 22 cols.
  5. "TD Recaudos"                   — Pivote User_Company + Σ Transaction_amount + Tipología.
  6. "Recaudo 1 al {N} {MES}"        — Tabla manual Surtitodo (cliente + MONTO).
  7. "Acumulado R"                   — Surtitodo + corte (cliente + MONTO + CORTE).

Uso:
    python generar_recaudos_dispersiones.py --mes 2026-04
    python generar_recaudos_dispersiones.py --mes 2026-04 --output "RyD Abril 2026.xlsx"

Credenciales CH se leen de env vars: MINTIC_CH_HOST/USER/PASS.
"""
from __future__ import annotations

import argparse
import calendar
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# UTF-8 en Windows
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ════════════════════════════════════════════════════════════════════════════
# ClickHouse
# ════════════════════════════════════════════════════════════════════════════

def ch_request(query: str, host: str, user: str, password: str, timeout: int = 600) -> str:
    url = host.rstrip("/") + "/"
    params = {"query": query + "\nFORMAT TabSeparatedWithNames"}
    r = requests.get(url, params=params, auth=(user, password), timeout=timeout)
    if r.status_code != 200:
        raise RuntimeError(f"CH HTTP {r.status_code}: {r.text[:500]}")
    return r.text


def parse_tsv(text: str) -> tuple[list[str], list[list[str]]]:
    lines = text.rstrip("\n").split("\n")
    if not lines:
        return [], []
    headers = lines[0].split("\t")
    rows = [ln.split("\t") for ln in lines[1:]] if len(lines) > 1 else []
    return headers, rows


def to_num(v):
    if v is None or v == "" or v == "\\N":
        return None
    try:
        if "." in v or "e" in v.lower():
            return float(v)
        return int(v)
    except (ValueError, TypeError):
        return v


# ════════════════════════════════════════════════════════════════════════════
# Queries
# ════════════════════════════════════════════════════════════════════════════

# Query A: Dispersiones Daviplata CashOut (solo "Dispersión Recaudo").
# Variables: %{desde}, %{hasta}
QUERY_DISPERSIONES = """
WITH filtered_wat AS (
    SELECT * FROM picapmongoprod.wallet_account_transactions FINAL
    WHERE _type = 'WalletAccountDriverBalanceTransactionDaviplataCashOut'
      AND toDate(toTimeZone(created_at, 'America/Bogota'))
          BETWEEN toDate('{desde}') AND toDate('{hasta}')
)
SELECT DISTINCT
    wat._id AS _id,
    toDate(toTimeZone(wat.created_at, 'America/Bogota')) AS created_at,
    ifNull(JSONExtractFloat(wat.amount, 'cents') / 100, 0) AS amount_cents,
    wat._type AS _type,
    comp._id AS company_id,
    comp.name AS Company_name,
    CASE
        WHEN comp._id IN (
            '5f9b1847dc3d1101c7ece86c',
            '5e908acb4f75ba007912a4fd'
        ) THEN 'Dispersión Recaudo'
        ELSE 'Dispersión Garantía'
    END AS tipo_dispersion
FROM filtered_wat wat
INNER JOIN picapmongoprod.wallet_accounts wa   ON wa._id   = wat.account_id
INNER JOIN picapmongoprod.companies      comp ON comp._id = wa.company_id
WHERE tipo_dispersion = 'Dispersión Recaudo'
ORDER BY created_at ASC
"""

# Query B: Recaudos (CounterDelivery). 22 cols — paridad con la query original
# del usuario (igual a Saldo Recaudos pero con score_rent + score_pibox).
QUERY_RECAUDOS = """
WITH
q_service_types AS (
  SELECT
    _id,
    name_es AS name,
    CASE
      WHEN name_es IN (
        'Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor', 'Moto favor',
        'Carga', 'Carga Carry', 'Carga Moto-Vagón', 'Carga NHR', 'Carga NKR', 'Carga NPR',
        'Mensajería', 'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'
      ) THEN 'Pibox'
      WHEN name_es IN (
        'Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta','Carro Subasta','Taxi', 'Carro',
        'Carro sin conductor', 'Moto VIP', 'Moto Económica', 'Carro Queen','Rapidín','Espero tranqui',
        'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto', 'Grúa Carro', 'Grúa Moto'
      ) THEN 'Picap'
      ELSE 'Other'
    END AS type
  FROM (SELECT * FROM picapmongoprod.service_types FINAL)
),
q_wat_filtered AS (
  SELECT
    wat._id AS transaction_id,
    wat.booking_id AS booking_id,
    wat.package_id AS package_id,
    wat.account_id AS account_id,
    wat.amount AS amount,
    wat.normalized_amount_after_transaction AS normalized_amount_after_transaction,
    wat.transaction_state_cd AS transaction_state_cd,
    wat._type AS tx_type,
    wat.created_at AS created_at,
    b.created_at AS booking_created_at,
    pck.reference AS package_reference,
    pck.declared_value AS package_declared_value,
    pck.counter_delivery AS package_counter_delivery,
    b.passenger_id AS passenger_id,
    b.driver_id AS driver_id,
    b.served_vehicle_type_id AS served_vehicle_type_id,
    b.city_id AS city_id,
    b.requested_service_type_id AS requested_service_type_id,
    b.country_id AS country_id,
    wa._id AS wallet_account_id,
    p.company_id AS passenger_company_id,
    p.name AS passenger_name,
    d.name AS driver_name
  FROM picapmongoprod.wallet_account_transactions AS wat FINAL
  INNER JOIN picapmongoprod.packages AS pck FINAL ON pck._id = wat.package_id
  INNER JOIN picapmongoprod.bookings AS b   FINAL ON b._id   = wat.booking_id
  INNER JOIN picapmongoprod.wallet_accounts AS wa FINAL ON wa._id = wat.account_id
  INNER JOIN picapmongoprod.passengers AS p FINAL ON p._id = b.passenger_id
  INNER JOIN picapmongoprod.passengers AS d FINAL ON d._id = b.driver_id
  INNER JOIN picapmongoprod.countries AS c FINAL ON c._id = b.country_id
  WHERE
    JSONExtractString(c.name, 'es') = 'Colombia'
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND pck.counter_delivery = 'true'
    AND wat._type = 'WalletAccountCounterDeliveryPaymentTransaction'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
)
SELECT
  toDate(toTimeZone(qtf.created_at, 'America/Bogota')) AS Date_transaction,
  JSONExtractString(qtf.amount, 'currency_iso') AS Transaction_currency,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100 AS Transaction_amount,
  qtf.transaction_id AS Transaction_ID,
  toFloat64OrZero(JSONExtractString(qtf.normalized_amount_after_transaction, 'cents')) / 100 AS Normalized_Amount_After_Transaction,
  toDate(toTimeZone(qtf.booking_created_at, 'America/Bogota')) AS Date_booking,
  qtf.booking_id AS ID_Booking,
  qtf.package_id AS ID_Package,
  qtf.package_reference AS Reference,
  qtf.passenger_id AS ID_User,
  compp.name AS User_Company,
  qtf.passenger_name AS User_Name,
  toFloat64OrZero(JSONExtractString(qtf.package_declared_value, 'cents')) / 100 AS Declared_Value,
  qtf.transaction_state_cd AS transaction_state_cd,
  qtf.driver_id AS ID_Driver,
  qtf.driver_name AS Driver_Name,
  st.type AS type,
  st.name AS service_type_name,
  JSONExtractString(cit.name, 'es') AS Ciudad,
  JSONExtractString(vt.name, 'es') AS name_vehicle,
  multiIf(sf.new_final_score_rent < 0, 0,
          sf.new_final_score_rent >= 5, 5,
          sf.new_final_score_rent) AS score_rent_fixed,
  multiIf(sf.new_final_score_pibox < 0, 0,
          sf.new_final_score_pibox >= 5, 5,
          sf.new_final_score_pibox) AS score_pibox_fixed
FROM q_wat_filtered qtf
LEFT JOIN picapmongoprod.companies     AS compp FINAL ON compp._id = qtf.passenger_company_id
LEFT JOIN picapmongoprod.vehicle_types AS vt    FINAL ON vt._id    = qtf.served_vehicle_type_id
LEFT JOIN picapmongoprod.cities        AS cit   FINAL ON cit._id   = qtf.city_id
LEFT JOIN q_service_types              AS st          ON st._id    = qtf.requested_service_type_id
LEFT JOIN picapmongoprod.vw_atr_driver_scoring_with_frauds AS sf FINAL ON sf.driver_id = qtf.driver_id
ORDER BY Date_transaction ASC
"""


# ════════════════════════════════════════════════════════════════════════════
# Estilos
# ════════════════════════════════════════════════════════════════════════════

HDR_FILL   = PatternFill("solid", start_color="1F4E78")  # azul corporativo
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN  = Alignment(horizontal="left", vertical="center")

# Header morado para hojas 3, 6, 7 (tablas manuales). Mismo estilo que la
# plantilla del usuario.
MORADO_FILL = PatternFill("solid", start_color="5B2169")
MORADO_FONT = Font(bold=True, color="FFFFFF", size=11)

PIVOT_HDR_FILL   = PatternFill("solid", start_color="D9E1F2")
PIVOT_TOTAL_FILL = PatternFill("solid", start_color="FFE699")

THIN = Side(border_style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INT_FMT = "#,##0"

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


# ════════════════════════════════════════════════════════════════════════════
# Tipología (Hoja 5 columna "Tipo de cliente")
# ════════════════════════════════════════════════════════════════════════════

PRUEBA_KEYWORDS = ["pibox admin", "testeo", "prueba", "qa", "test"]


def tipo_cliente(empresa: str) -> str:
    """Clasifica el cliente para la columna tipología:
      - Surtitodo y relacionados → 'Reportar'
      - PIBOX ADMIN / Testeo / QA / pruebas → 'prueba'
      - Otros → 'ida y vuelta'
    """
    if not empresa:
        return "ida y vuelta"
    n = str(empresa).lower().strip()
    if "surtitodo" in n:
        return "Reportar"
    if any(kw in n for kw in PRUEBA_KEYWORDS):
        return "prueba"
    return "ida y vuelta"


def es_surtitodo(empresa: str) -> bool:
    return "surtitodo" in str(empresa).lower().strip()


# ════════════════════════════════════════════════════════════════════════════
# Helpers de período
# ════════════════════════════════════════════════════════════════════════════

def mes_a_info(mes_str: str) -> dict:
    """'2026-04' → dict con desde, hasta, last_day, mes_nombre, etc."""
    año, mes = mes_str.split("-")
    año_i, mes_i = int(año), int(mes)
    _, last_day = calendar.monthrange(año_i, mes_i)
    desde = f"{año}-{mes}-01"
    hasta = f"{año}-{mes}-{last_day:02d}"
    mes_nombre = MESES_ES[mes_i]
    return {
        "desde": desde,
        "hasta": hasta,
        "last_day": last_day,
        "mes_nombre": mes_nombre,
        "mes_minus": mes_nombre.lower(),  # "abril"
        "mes_mayus": mes_nombre.upper(),  # "ABRIL"
        "anio": año_i,
        # Strings usados como nombres de hojas y como corte
        "hoja_dispersiones": f"Dispersiones 1 al {last_day} {mes_nombre.lower()}",
        "hoja_dispersion_pivot": f"Dispersion 1 al {last_day} {mes_nombre.lower()}",
        "hoja_recaudo": f"Recaudo 1 al {last_day} {mes_nombre.upper()}",
        "corte": f"1 TO {last_day} {mes_nombre.upper()}",
    }


# ════════════════════════════════════════════════════════════════════════════
# Escritura de hojas
# ════════════════════════════════════════════════════════════════════════════

# Hoja 1: Dispersiones — data cruda 7 cols Query A
HEADERS_DISPERSIONES = [
    "wat._id", "created_at", "amount_cents", "wat._type",
    "company_id", "Company_name", "tipo_dispersion",
]

# Mapeo header CH → header Excel
MAP_DISPERSIONES = {
    "_id": "wat._id",
    "created_at": "created_at",
    "amount_cents": "amount_cents",
    "_type": "wat._type",
    "company_id": "company_id",
    "Company_name": "Company_name",
    "tipo_dispersion": "tipo_dispersion",
}


def write_hoja1_dispersiones(wb, info, headers_ch, rows):
    """Hoja 1: data cruda de Query A con 7 cols."""
    name = info["hoja_dispersiones"]
    ws = wb.create_sheet(name)
    # Headers azul
    for c_idx, h in enumerate(HEADERS_DISPERSIONES, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIGN

    # Map de posiciones CH
    pos_ch = {h: i for i, h in enumerate(headers_ch)}

    for r_idx, r in enumerate(rows, start=2):
        for c_idx, hdr_excel in enumerate(HEADERS_DISPERSIONES, start=1):
            hdr_ch = next((ch for ch, ex in MAP_DISPERSIONES.items() if ex == hdr_excel), hdr_excel)
            src_idx = pos_ch.get(hdr_ch)
            if src_idx is None or src_idx >= len(r):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=to_num(r[src_idx]))
        # Formato moneda en amount_cents (col 3)
        ws.cell(row=r_idx, column=3).number_format = INT_FMT

    ws.freeze_panes = "A2"
    widths = {"A": 28, "B": 14, "C": 16, "D": 50, "E": 26, "F": 26, "G": 24}
    for col_l, w in widths.items():
        ws.column_dimensions[col_l].width = w


def write_hoja2_dispersion_pivot(wb, info, headers_ch, rows):
    """Hoja 2: Pivote por Company_name + tipo_dispersion."""
    name = info["hoja_dispersion_pivot"]
    ws = wb.create_sheet(name)

    # Agrupar por (Company_name, tipo_dispersion)
    pos_ch = {h: i for i, h in enumerate(headers_ch)}
    idx_emp  = pos_ch.get("Company_name")
    idx_tipo = pos_ch.get("tipo_dispersion")
    idx_amt  = pos_ch.get("amount_cents")

    pivot = defaultdict(float)
    for r in rows:
        emp = r[idx_emp] if idx_emp is not None and idx_emp < len(r) else ""
        tip = r[idx_tipo] if idx_tipo is not None and idx_tipo < len(r) else ""
        amt = to_num(r[idx_amt]) if idx_amt is not None and idx_amt < len(r) else 0
        if not emp:
            continue
        try:
            pivot[(emp, tip)] += float(amt or 0)
        except (ValueError, TypeError):
            pass

    sorted_pivot = sorted(pivot.items(), key=lambda kv: kv[0][0].lower())

    # Headers
    hdrs = ["Company_name", "tipo_dispersion", "Suma de amount_cents"]
    for c_idx, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    for i, ((emp, tip), monto) in enumerate(sorted_pivot, start=2):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        ws.cell(row=i, column=2, value=tip).border = BORDER_ALL
        c = ws.cell(row=i, column=3, value=round(monto, 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL

    # Total general
    total = sum(v for _, v in sorted_pivot)
    r_total = 2 + len(sorted_pivot)
    ws.cell(row=r_total, column=1, value="Total general").font = Font(bold=True)
    ws.cell(row=r_total, column=1).fill = PIVOT_TOTAL_FILL
    ws.cell(row=r_total, column=2).fill = PIVOT_TOTAL_FILL
    c = ws.cell(row=r_total, column=3, value=round(total, 2))
    c.number_format = INT_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    # Devolver lookup para hoja 3 (empresa → monto)
    por_empresa = defaultdict(float)
    for (emp, _tip), monto in pivot.items():
        por_empresa[emp] += monto
    return por_empresa


def write_hoja3_acumulado_dispersion(wb, info, por_empresa_dispersion):
    """Hoja 3: Tabla manual con Company_name + MONTO + CORTE (header morado)."""
    ws = wb.create_sheet("Acumulado Dispersion")
    # Headers morado
    hdrs = ["Company_name", "MONTO", "CORTE"]
    for c_idx, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = MORADO_FILL; cell.font = MORADO_FONT; cell.alignment = HDR_ALIGN

    # Filas: una por cada empresa del pivote (idealmente solo Surtitodo en
    # "Dispersión Recaudo" porque el filtro de la query lo asegura).
    for i, (emp, monto) in enumerate(sorted(por_empresa_dispersion.items(),
                                            key=lambda kv: kv[0].lower()),
                                     start=2):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        ws.cell(row=i, column=3, value=info["corte"]).border = BORDER_ALL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22


# ────────────────────────────────────────────────────────────────────────────
# Hoja 4: Data Recaudos (22 cols Query B)
# ────────────────────────────────────────────────────────────────────────────

HEADERS_RECAUDOS = [
    "Date_transaction", "Transaction_currency", "Transaction_amount", "Transaction_ID",
    "Normalized_Amount_After_Transaction", "Date_booking", "ID_Booking", "ID_Package",
    "Reference", "ID_User", "User_Company", "User_Name", "Declared_Value",
    "transaction_state_cd", "ID_Driver", "Driver_Name", "type", "service_type_name",
    "Ciudad", "name_vehicle", "score_rent_fixed", "score_pibox_fixed",
]


def write_hoja4_data_recaudos(wb, headers_ch, rows):
    ws = wb.create_sheet("Data Recaudos")
    # Headers azul
    for c_idx, h in enumerate(HEADERS_RECAUDOS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIGN

    # Mapeo posiciones
    pos_ch = {h: i for i, h in enumerate(headers_ch)}

    for r_idx, r in enumerate(rows, start=2):
        for c_idx, hdr in enumerate(HEADERS_RECAUDOS, start=1):
            src_idx = pos_ch.get(hdr)
            if src_idx is None or src_idx >= len(r):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=to_num(r[src_idx]))

    # Formato moneda en Transaction_amount (col 3), Declared_Value (col 13)
    last_row = 1 + len(rows)
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=3).number_format = INT_FMT
        ws.cell(row=r, column=5).number_format = INT_FMT
        ws.cell(row=r, column=13).number_format = INT_FMT

    ws.freeze_panes = "A2"
    for c in range(1, len(HEADERS_RECAUDOS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


# ────────────────────────────────────────────────────────────────────────────
# Hoja 5: TD Recaudos — pivote User_Company + tipología
# ────────────────────────────────────────────────────────────────────────────

def write_hoja5_td_recaudos(wb, headers_ch, rows):
    ws = wb.create_sheet("TD Recaudos")
    pos_ch = {h: i for i, h in enumerate(headers_ch)}
    idx_emp = pos_ch.get("User_Company")
    idx_amt = pos_ch.get("Transaction_amount")

    pivot = defaultdict(float)
    for r in rows:
        emp = r[idx_emp] if idx_emp is not None and idx_emp < len(r) else ""
        amt = to_num(r[idx_amt]) if idx_amt is not None and idx_amt < len(r) else 0
        if not emp:
            continue
        try:
            pivot[emp] += float(amt or 0)
        except (ValueError, TypeError):
            pass

    # Sort por monto descendente (igual que la plantilla)
    sorted_pivot = sorted(pivot.items(), key=lambda kv: -kv[1])

    # Filas 1-2: vacías (mimic pivot layout). Headers en fila 3.
    ws.cell(row=3, column=1, value="Etiquetas de fila")
    ws.cell(row=3, column=2, value="Suma de Transaction_amount")
    ws.cell(row=3, column=3, value="Tipo de cliente")
    for c_idx in (1, 2, 3):
        cell = ws.cell(row=3, column=c_idx)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    for i, (emp, monto) in enumerate(sorted_pivot, start=4):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        ws.cell(row=i, column=3, value=tipo_cliente(emp)).border = BORDER_ALL

    total = sum(v for _, v in sorted_pivot)
    r_total = 4 + len(sorted_pivot)
    ws.cell(row=r_total, column=1, value="Total general").font = Font(bold=True)
    ws.cell(row=r_total, column=1).fill = PIVOT_TOTAL_FILL
    c = ws.cell(row=r_total, column=2, value=round(total, 2))
    c.number_format = INT_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL
    ws.cell(row=r_total, column=3).fill = PIVOT_TOTAL_FILL

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18

    # Devolver lookup empresa → suma (para hojas 6, 7)
    return dict(pivot)


def write_hoja6_recaudo(wb, info, pivot_recaudos):
    """Hoja 6: Tabla manual con Surtitodo (cliente + MONTO, header morado)."""
    name = info["hoja_recaudo"]
    ws = wb.create_sheet(name)
    hdrs = ["cliente", "MONTO"]
    for c_idx, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = MORADO_FILL; cell.font = MORADO_FONT; cell.alignment = HDR_ALIGN

    # Buscar Surtitodo (puede ser "Surtitodo express", "Surtitodo", etc.)
    surt_monto = 0.0
    for emp, monto in pivot_recaudos.items():
        if es_surtitodo(emp):
            surt_monto += monto

    ws.cell(row=2, column=1, value="Surtitodo").border = BORDER_ALL
    c = ws.cell(row=2, column=2, value=round(surt_monto, 2))
    c.number_format = INT_FMT; c.border = BORDER_ALL

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22


def write_hoja7_acumulado_r(wb, info, pivot_recaudos):
    """Hoja 7: Surtitodo + CORTE (header morado)."""
    ws = wb.create_sheet("Acumulado R")
    hdrs = ["cliente", "MONTO", "CORTE"]
    for c_idx, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = MORADO_FILL; cell.font = MORADO_FONT; cell.alignment = HDR_ALIGN

    surt_monto = 0.0
    for emp, monto in pivot_recaudos.items():
        if es_surtitodo(emp):
            surt_monto += monto

    ws.cell(row=2, column=1, value="Surtitodo").border = BORDER_ALL
    c = ws.cell(row=2, column=2, value=round(surt_monto, 2))
    c.number_format = INT_FMT; c.border = BORDER_ALL
    ws.cell(row=2, column=3, value=info["corte"]).border = BORDER_ALL

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera el Excel 'Recaudos y Dispersiones' (7 hojas)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--mes", required=True, help="Mes a procesar en formato YYYY-MM (ej. 2026-04)")
    parser.add_argument("--output", help="Nombre del archivo .xlsx")
    parser.add_argument("--ch-host", default=os.environ.get("MINTIC_CH_HOST", ""))
    parser.add_argument("--ch-user", default=os.environ.get("MINTIC_CH_USER", ""))
    parser.add_argument("--ch-pass", default=os.environ.get("MINTIC_CH_PASS", ""))
    args = parser.parse_args()

    if not args.ch_host or not args.ch_user or not args.ch_pass:
        sys.exit(
            "❌ Credenciales ClickHouse no configuradas.\n"
            "   Setealas como env vars (MINTIC_CH_HOST/USER/PASS) o pasalas con\n"
            "   --ch-host --ch-user --ch-pass."
        )

    try:
        info = mes_a_info(args.mes)
    except Exception as e:
        sys.exit(f"❌ --mes inválido: {args.mes} ({e}). Usá formato YYYY-MM (ej. 2026-04).")

    output = args.output or f"Recaudos y Dispersiones {info['mes_nombre']} {info['anio']}.xlsx"
    output_path = Path(output)

    print(f"🚀 Generando reporte Recaudos y Dispersiones")
    print(f"   • Período  : {info['desde']} → {info['hasta']} ({info['corte']})")
    print(f"   • Output   : {output_path}")
    print()

    # Query A
    print("📊 Query A: Dispersiones Daviplata CashOut…")
    t0 = time.time()
    sql_a = QUERY_DISPERSIONES.format(desde=info["desde"], hasta=info["hasta"])
    text_a = ch_request(sql_a, args.ch_host, args.ch_user, args.ch_pass)
    headers_a, rows_a = parse_tsv(text_a)
    print(f"   ✓ {len(rows_a)} filas en {time.time() - t0:.1f}s")

    # Query B
    print("📊 Query B: Recaudos (CounterDelivery)…")
    t0 = time.time()
    sql_b = QUERY_RECAUDOS.format(desde=info["desde"], hasta=info["hasta"])
    text_b = ch_request(sql_b, args.ch_host, args.ch_user, args.ch_pass)
    headers_b, rows_b = parse_tsv(text_b)
    print(f"   ✓ {len(rows_b)} filas en {time.time() - t0:.1f}s")

    print()
    print("📝 Generando Excel…")
    wb = Workbook()
    wb.remove(wb.active)

    write_hoja1_dispersiones(wb, info, headers_a, rows_a)
    print(f"   ✓ Hoja 1: '{info['hoja_dispersiones']}' ({len(rows_a)} filas)")

    por_empresa_disp = write_hoja2_dispersion_pivot(wb, info, headers_a, rows_a)
    print(f"   ✓ Hoja 2: '{info['hoja_dispersion_pivot']}' ({len(por_empresa_disp)} empresas)")

    write_hoja3_acumulado_dispersion(wb, info, por_empresa_disp)
    print(f"   ✓ Hoja 3: 'Acumulado Dispersion'")

    write_hoja4_data_recaudos(wb, headers_b, rows_b)
    print(f"   ✓ Hoja 4: 'Data Recaudos' ({len(rows_b)} filas)")

    pivot_recaudos = write_hoja5_td_recaudos(wb, headers_b, rows_b)
    print(f"   ✓ Hoja 5: 'TD Recaudos' ({len(pivot_recaudos)} empresas)")

    write_hoja6_recaudo(wb, info, pivot_recaudos)
    print(f"   ✓ Hoja 6: '{info['hoja_recaudo']}'")

    write_hoja7_acumulado_r(wb, info, pivot_recaudos)
    print(f"   ✓ Hoja 7: 'Acumulado R'")

    wb.save(output_path)
    print()
    print(f"✅ Listo: {output_path}")
    print(f"   Tamaño: {output_path.stat().st_size / 1024:.1f} KB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
