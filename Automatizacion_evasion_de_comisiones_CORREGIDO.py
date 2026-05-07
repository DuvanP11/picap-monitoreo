import clickhouse_connect
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import numpy as np
from flask import Flask, send_from_directory

# ══════════════════════════════════════════════════════════════
# 0. CREAR APP FLASK
# ══════════════════════════════════════════════════════════════
app = Flask(__name__)

# ══════════════════════════════════════════════════════════════
# 1. CONEXIÓN
# ══════════════════════════════════════════════════════════════
def obtener_cliente():
    import os as _os
    return clickhouse_connect.get_client(
        host=_os.environ.get("CLICKHOUSE_HOST", "clickhouse.picap.io"),
        port=int(_os.environ.get("CLICKHOUSE_PORT", "8443")),
        username=_os.environ.get("CLICKHOUSE_USER", "dperilla"),
        password=_os.environ.get("CLICKHOUSE_PASSWORD", ""),
        database=_os.environ.get("CLICKHOUSE_DATABASE", "picapmongoprod"),
        secure=True
    )

# ══════════════════════════════════════════════════════════════
# 2. QUERY
# ══════════════════════════════════════════════════════════════
QUERY = """
WITH base AS (
    SELECT 
        toTimeZone(b.created_at, 'America/Bogota') AS creacion_servicio,
        parseDateTimeBestEffortOrNull(ev_accept)   AS fecha_aceptacion,
        parseDateTimeBestEffortOrNull(ev_cancel)   AS fecha_cancelacion,
        b._id              AS booking_id,
        b.status_cd        AS estado_servicio,
        b.driver_id        AS id_driver,
        pd.name            AS name_driver,
        b.passenger_id     AS id_user,
        pu.name            AS name_user,
        b.company_id       AS id_company,
        JSONExtractString(st.name, 'es') AS type_service,
        JSONExtractString(b.final_cost, 'currency_iso') AS moneda,
        b.g_country        AS pais,
        b.g_adm_area_lv_1  AS ciudad,
        toFloat64OrNull(JSONExtractString(b.final_cost,     'cents')) / 100 AS costo_final,
        toFloat64OrNull(JSONExtractString(b.estimated_cost, 'cents')) / 100 AS costo_estimado,

        extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\\[\\s*([+-]?\\d+\\.\\d+)')      AS ev_cancel_lon_str,
        extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\\[.*?,\\s*([+-]?\\d+\\.\\d+)')  AS ev_cancel_lat_str,
        extract(ifNull(b.events,''), 'event_cd":20.*?created_at":"([^"]+)')                        AS ev_accept,
        extract(ifNull(b.events,''), 'event_cd":26.*?created_at":"([^"]+)')                        AS ev_cancel,

        toFloat64OrNull(ev_cancel_lon_str) AS cancel_lon,
        toFloat64OrNull(ev_cancel_lat_str) AS cancel_lat,
        toFloat64(JSONExtractString(b.end_geojson, 'coordinates', 1)) AS end_lon,
        toFloat64(JSONExtractString(b.end_geojson, 'coordinates', 2)) AS end_lat,

        dateDiff('minute',
            parseDateTimeBestEffortOrNull(ev_accept),
            parseDateTimeBestEffortOrNull(ev_cancel)
        ) AS minutos_entre_eventos,

        if(
            dateDiff('minute',
                parseDateTimeBestEffortOrNull(ev_accept),
                parseDateTimeBestEffortOrNull(ev_cancel)
            ) > 5,
            'Evasion', 'Ok'
        ) AS regla_tiempo,

        ROW_NUMBER() OVER (
            PARTITION BY b._id
            ORDER BY b.created_at DESC
        ) AS rn

    FROM picapmongoprod.bookings b
    LEFT JOIN picapmongoprod.passengers pd ON b.driver_id    = pd._id
    LEFT JOIN picapmongoprod.passengers pu ON b.passenger_id = pu._id
    LEFT JOIN picapmongoprod.service_types st ON st._id = b.requested_service_type_id

    WHERE 
        NOT empty(b.origin_geojson)
        AND NOT empty(b.end_geojson)
        AND b.status_cd IN (100, 102)
        AND b.created_at >= toDateTime('2026-05-06 00:00:00')
        AND b.created_at <= toDateTime('2026-05-07 23:59:59')
)

SELECT
    creacion_servicio, fecha_aceptacion, fecha_cancelacion,
    booking_id, estado_servicio,
    id_driver, name_driver, id_user, name_user, id_company,
    type_service, moneda, pais, ciudad,
    costo_final, costo_estimado,
    cancel_lon, cancel_lat, end_lon, end_lat,
    minutos_entre_eventos, regla_tiempo,
    round(geoDistance(cancel_lon, cancel_lat, end_lon, end_lat), 2) AS distancia_cancel_destino,
    CASE
        WHEN geoDistance(cancel_lon, cancel_lat, end_lon, end_lat) <= 450
        THEN 'CANCELA_CERCA'
        ELSE 'CANCELA_LEJOS'
    END AS regla_cancelacion
FROM base
WHERE rn = 1
"""

# ══════════════════════════════════════════════════════════════
# 3. EJECUTAR QUERY
# ══════════════════════════════════════════════════════════════
def ejecutar_query(client) -> pd.DataFrame:
    print("Conectando a ClickHouse...")
    result = client.query(QUERY)
    df = pd.DataFrame(result.result_rows, columns=result.column_names)
    print(f"Servicios cargados: {len(df):,}")
    return df

# ══════════════════════════════════════════════════════════════
# 4. MOTOR DE DECISIÓN
# ══════════════════════════════════════════════════════════════
def clasificar_vectorizado(df: pd.DataFrame) -> pd.DataFrame:
    flag_tiempo    = df['regla_tiempo']      == 'Evasion'
    flag_distancia = df['regla_cancelacion'] == 'CANCELA_CERCA'
    sin_gps        = df['cancel_lon'].isna() | df['cancel_lat'].isna()

    condiciones = [
        flag_tiempo & flag_distancia,
        flag_tiempo & sin_gps,
        flag_tiempo | flag_distancia,
    ]
    opciones = [
        'EVASION CONFIRMADA',
        'EVASION PROBABLE',
        'EVASION PROBABLE',
    ]
    df['veredicto'] = np.select(condiciones, opciones, default='OK')
    df['nivel']     = np.select(condiciones, [3, 2, 2],  default=0)

    df['flags'] = ''
    df.loc[flag_tiempo,    'flags'] += df.loc[flag_tiempo,    'minutos_entre_eventos'].astype(str) + ' min | '
    df.loc[flag_distancia, 'flags'] += df.loc[flag_distancia, 'distancia_cancel_destino'].astype(str) + 'm del destino | '
    df.loc[sin_gps,        'flags'] += 'sin GPS | '
    df['flags'] = df['flags'].str.rstrip(' | ')

    return df

# ══════════════════════════════════════════════════════════════
# 5. CALCULAR COMISIONES Y PENALIDADES (✅ CORREGIDO)
# ══════════════════════════════════════════════════════════════
def calcular_comisiones(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula comisión evadida y penalización del 5% según lógica de negocio:
    - Comisión evadida = costo_estimado * tasa
    - Penalización = comisión_evadida * 5%
    - Total a cobrar = comisión_evadida + penalización
    """
    # Normalizar nombres de países
    df['pais_normalizado'] = df['pais'].str.upper().str.strip()
    
    # Definir tasa de comisión según país
    def obtener_tasa_comision(pais):
        if pd.isna(pais):
            return 0.12  # Default Colombia
        pais_upper = pais.upper()
        if 'COLOMBIA' in pais_upper or 'CO' == pais_upper:
            return 0.12
        elif 'MEXICO' in pais_upper or 'MX' == pais_upper:
            return 0.10
        elif 'NICARAGUA' in pais_upper or 'NI' == pais_upper:
            return 0.10
        else:
            return 0.12  # Default
    
    # Aplicar tasas de comisión
    df['tasa_comision'] = df['pais_normalizado'].apply(obtener_tasa_comision)
    
    # ✅ CÁLCULO CORRECTO SEGÚN LÓGICA DE NEGOCIO
    # 1. Comisión evadida (base)
    df['comision_evadida'] = df['costo_estimado'] * df['tasa_comision']
    df['comision_evadida'] = df['comision_evadida'].round(0)  # Sin decimales
    
    # 2. Penalización del 5% sobre la comisión evadida
    df['recargo_5pct'] = df['comision_evadida'] * 0.05
    df['recargo_5pct'] = df['recargo_5pct'].round(0)  # Sin decimales
    
    # 3. Total a cobrar (comisión + penalización)
    df['total_a_cobrar'] = df['comision_evadida'] + df['recargo_5pct']
    df['total_a_cobrar'] = df['total_a_cobrar'].round(0)  # Sin decimales
    
    # 4. Valor neto (informativo, no se usa para cobro)
    df['valor_neto'] = df['costo_estimado'] - df['comision_evadida']
    df['valor_neto'] = df['valor_neto'].round(0)  # Sin decimales
    
    # Remover columna temporal
    df = df.drop(columns=['pais_normalizado', 'tasa_comision'])
    
    return df

def limpiar_nan_json(df: pd.DataFrame) -> pd.DataFrame:
    return df.replace({np.nan: None})

def limpiar_timezones(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include=['datetimetz']).columns:
        df[col] = df[col].dt.tz_localize(None)
    return df

# ══════════════════════════════════════════════════════════════
# 6. CREAR HOJA DE RESUMEN
# ══════════════════════════════════════════════════════════════
def crear_hoja_resumen(df: pd.DataFrame, writer):
    """
    Crea una hoja de resumen con estadísticas por país
    """
    resumen_data = []
    
    # Obtener lista de países únicos
    paises = df['pais'].dropna().unique()
    
    for pais in paises:
        df_pais = df[df['pais'] == pais]
        
        # 1. Resumen de veredictos
        total_servicios = len(df_pais)
        veredictos = df_pais['veredicto'].value_counts()
        
        for veredicto, cantidad in veredictos.items():
            porcentaje = (cantidad / total_servicios) * 100
            resumen_data.append({
                'Pais': pais,
                'Categoria': 'Veredicto',
                'Tipo': veredicto,
                'Cantidad': cantidad,
                'Porcentaje': f"{porcentaje:.2f}%",
                'Valor_Total': None
            })
        
        # 2. Resumen de pilotos únicos que cancelaron
        pilotos_unicos = df_pais['id_driver'].nunique()
        total_pilotos_posibles = df['id_driver'].nunique()  # Total global
        porcentaje_pilotos = (pilotos_unicos / total_pilotos_posibles) * 100 if total_pilotos_posibles > 0 else 0
        
        resumen_data.append({
            'Pais': pais,
            'Categoria': 'Pilotos',
            'Tipo': 'Pilotos que cancelaron',
            'Cantidad': pilotos_unicos,
            'Porcentaje': f"{porcentaje_pilotos:.2f}%",
            'Valor_Total': None
        })
        
        # 3. Valores de EVASIÓN CONFIRMADA (✅ CORREGIDO)
        df_evasion = df_pais[df_pais['veredicto'] == 'EVASION CONFIRMADA']
        
        if len(df_evasion) > 0:
            comision_evadida_total = df_evasion['comision_evadida'].sum()
            recargo_total = df_evasion['recargo_5pct'].sum()
            total_cobro = df_evasion['total_a_cobrar'].sum()
            
            resumen_data.append({
                'Pais': pais,
                'Categoria': 'Valores EVASIÓN CONFIRMADA',
                'Tipo': 'Comisión Evadida',
                'Cantidad': len(df_evasion),
                'Porcentaje': f"{(len(df_evasion)/total_servicios)*100:.2f}%",
                'Valor_Total': f"${comision_evadida_total:,.0f}"
            })
            
            resumen_data.append({
                'Pais': pais,
                'Categoria': 'Valores EVASIÓN CONFIRMADA',
                'Tipo': 'Penalización 5%',
                'Cantidad': len(df_evasion),
                'Porcentaje': f"{(len(df_evasion)/total_servicios)*100:.2f}%",
                'Valor_Total': f"${recargo_total:,.0f}"
            })
            
            resumen_data.append({
                'Pais': pais,
                'Categoria': 'Valores EVASIÓN CONFIRMADA',
                'Tipo': 'Total a Cobrar',
                'Cantidad': len(df_evasion),
                'Porcentaje': f"{(len(df_evasion)/total_servicios)*100:.2f}%",
                'Valor_Total': f"${total_cobro:,.0f}"
            })
    
    # Crear DataFrame de resumen
    df_resumen = pd.DataFrame(resumen_data)
    
    # Exportar a Excel
    df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
    
    return df_resumen

# ══════════════════════════════════════════════════════════════
# 7. APLICAR FORMATO AL EXCEL
# ══════════════════════════════════════════════════════════════
def aplicar_formato_excel(writer, sheet_name='Resultados'):
    """
    Aplica formato morado con letras blancas en negrilla,
    congela la primera fila y activa filtros
    """
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Definir estilo de encabezado (morado con blanco)
    header_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Aplicar formato a la primera fila
    for col_num, column in enumerate(worksheet.iter_cols(1, worksheet.max_column, 1, 1), 1):
        cell = column[0]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
        # Ajustar ancho de columna
        column_letter = get_column_letter(col_num)
        adjusted_width = max(len(str(cell.value)) + 2, 12)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    worksheet.freeze_panes = 'A2'
    
    # Activar filtros
    worksheet.auto_filter.ref = worksheet.dimensions

# ══════════════════════════════════════════════════════════════
# 8. EXPORTAR EXCEL CON TODO
# ══════════════════════════════════════════════════════════════
def exportar(df: pd.DataFrame):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Evasion_Comisiones_{timestamp}.xlsx"
    
    df = limpiar_timezones(df)
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Hoja 1: Resultados completos
        df.to_excel(writer, sheet_name='Resultados', index=False)
        aplicar_formato_excel(writer, 'Resultados')
        
        # Hoja 2: Resumen
        crear_hoja_resumen(df, writer)
        aplicar_formato_excel(writer, 'Resumen')
    
    print(f"✅ Reporte generado: {filename}")
    print(f"   📊 Total servicios: {len(df):,}")
    print(f"   🚨 Evasiones confirmadas: {len(df[df['veredicto']=='EVASION CONFIRMADA']):,}")
    print(f"   ⚠️  Evasiones probables: {len(df[df['veredicto']=='EVASION PROBABLE']):,}")
    
    # Mostrar totales de cobro
    total_comision = df[df['veredicto']=='EVASION CONFIRMADA']['comision_evadida'].sum()
    total_penalizacion = df[df['veredicto']=='EVASION CONFIRMADA']['recargo_5pct'].sum()
    total_cobrar = df[df['veredicto']=='EVASION CONFIRMADA']['total_a_cobrar'].sum()
    
    print(f"\n💰 TOTALES A COBRAR (EVASIÓN CONFIRMADA):")
    print(f"   Comisión evadida: ${total_comision:,.0f}")
    print(f"   Penalización 5%: ${total_penalizacion:,.0f}")
    print(f"   TOTAL A COBRAR: ${total_cobrar:,.0f}")

# ══════════════════════════════════════════════════════════════
# 9. PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════
def correr():
    try:
        client = obtener_cliente()
        df = ejecutar_query(client)
        
        print("Clasificando servicios...")
        df = clasificar_vectorizado(df)
        
        print("Calculando comisiones y penalidades...")
        df = calcular_comisiones(df)
        
        df = limpiar_nan_json(df)
        
        print("Generando Excel con formato...")
        exportar(df)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        print("\n⚠️  Verifica que estés conectado a la VPN de Picap")

# ══════════════════════════════════════════════════════════════
# 10. RUTAS FLASK (DASHBOARD)
# ══════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return send_from_directory(".", "dashboard.html")

@app.route("/dashboard.html")
def dashboard():
    return send_from_directory(".", "dashboard.html")

# ══════════════════════════════════════════════════════════════
# 11. MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    correr()
