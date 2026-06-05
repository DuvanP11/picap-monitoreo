"""
Generador del reporte "Comisión Recaudos" (9 hojas) para BI.

Replica la plantilla:
  C:\\Users\\Picap\\Documents\\2026\\Abril\\RECAUDOS BI\\Comisión-recaudos-Abril 2026.xlsx

Estructura de hojas:
  1. "1. Comisión Recaudo"          — Query B (BookingCompanyCollectionFee). 10 cols.
  2. "2. Recaudos"                  — Query A (CounterDeliveryPaymentTransaction). 10 cols + 2 calculadas (%, Comisión Manual).
  3. "3. Cruce company"             — Pivote por Company_name (Hoja 2) + lookups + segunda mini-tabla (Hoja 1).
  4. "4. Cruce Booking"             — Pivote por booking_id + Company_name (Hoja 2) + segunda mini-tabla (Hoja 1).
  5. "Resumen"                      — Resumen final con clientes (% > 0, sin Cruz Verde) + Anticipo/Pendiente/Estado.
  6. "Comisión Recaudo ida y vuelta"— Hoja 1 filtrada (sin Multipaquete/Surtitodo/Pibox Admin/Testeo).
  7. "Recaudos ida y vuelta"        — Hoja 2 filtrada (sin clientes prueba + VAL_AMOUNT > 0).
  8. "TD company ida y vuelta"      — Pivote por Company_name (Hoja 7) + segunda mini-tabla (Hoja 6).
  9. "TD Bookings ida y vuelta"     — Pivote por booking_id (Hoja 7) + segunda mini-tabla (Hoja 6 filtrada).

Uso:
    python generar_comisiones.py --mes 2026-04 --output "Comisiones Abril 2026.xlsx"

Credenciales CH se leen de env vars: MINTIC_CH_HOST/USER/PASS.
"""
from __future__ import annotations

import argparse
import calendar
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# UTF-8 en Windows
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ════════════════════════════════════════════════════════════════════════════
# ClickHouse
# ════════════════════════════════════════════════════════════════════════════

def ch_request(query: str, host: str, user: str, password: str, timeout: int = 600) -> str:
    url = host.rstrip("/") + "/"
    params = {"query": query + "\nFORMAT TabSeparatedWithNames"}
    r = requests.get(url, params=params, auth=(user, password), timeout=timeout)
    if r.status_code != 200:
        raise RuntimeError(f"CH HTTP {r.status_code}: {r.text[:500]}")
    return r.text


def parse_tsv(text: str) -> tuple[list[str], list[list[str]]]:
    lines = text.rstrip("\n").split("\n")
    if not lines:
        return [], []
    headers = lines[0].split("\t")
    rows = [ln.split("\t") for ln in lines[1:]] if len(lines) > 1 else []
    return headers, rows


def to_num(v):
    if v is None or v == "" or v == "\\N":
        return None
    try:
        if "." in v or "e" in v.lower():
            return float(v)
        return int(v)
    except (ValueError, TypeError):
        return v


# ════════════════════════════════════════════════════════════════════════════
# Queries
# ════════════════════════════════════════════════════════════════════════════

# Query A: Recaudos (CounterDelivery) → Hoja 2
QUERY_RECAUDOS = """
WITH q_service_types AS (
  SELECT _id, any(name_es) AS name,
    any(multiIf(
      name_es IN ('Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor', 'Moto favor',
                  'Carga', 'Carga Carry', 'Carga Moto-Vagón', 'Carga NHR', 'Carga NKR', 'Carga NPR',
                  'Mensajería', 'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'), 'Pibox',
      name_es IN ('Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta', 'Taxi',
                  'Carro', 'Rapidín', 'Carro sin conductor', 'Moto VIP',
                  'Moto Económica', 'Carro Queen', 'Espero tranqui',
                  'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto',
                  'Grúa Carro', 'Grúa Moto'), 'Picap',
      'Other'
    )) AS type
  FROM picapmongoprod.service_types GROUP BY _id
),
q_transactions_filtered AS (
  SELECT
    wat._id AS _id, wat.booking_id, wat.account_id,
    wat._type AS txt_type, wat.created_at AS created_at, wat.amount,
    s.payment_method_cd,
    toFloat64OrZero(JSONExtractString(s.amount_charged_to_passenger_wallet, 'cents')) / 100 AS amount_charged_to_passenger_wallet,
    toFloat64OrZero(JSONExtractString(s.amount_charged_to_company_wallet, 'cents')) / 100   AS amount_charged_to_company_wallet,
    wa.passenger_id AS passenger_id,
    comp._id AS company_id, comp.name AS company_name,
    st.type AS service_type
  FROM picapmongoprod.wallet_account_transactions wat FINAL
  ANY LEFT JOIN picapmongoprod.bookings        s    FINAL ON s._id    = wat.booking_id
  ANY LEFT JOIN q_service_types                st         ON st._id   = s.requested_service_type_id
  ANY LEFT JOIN picapmongoprod.wallet_accounts wa   FINAL ON wa._id   = wat.account_id
  ANY LEFT JOIN picapmongoprod.companies       comp FINAL ON comp._id = s.company_id
  WHERE wat._type = 'WalletAccountCounterDeliveryPaymentTransaction'
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND st.type = 'Pibox'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
),
q_transactions AS (
  SELECT t.booking_id,
    JSONExtractString(t.amount, 'currency_iso') AS currency,
    SUM(IF(t._type = 'WalletAccountTransactionBookingDriverPayment',
           JSONExtractFloat(t.amount, 'cents') / 100, 0)) AS booking_driver_payment
  FROM picapmongoprod.wallet_account_transactions t FINAL
  INNER JOIN (SELECT DISTINCT booking_id FROM q_transactions_filtered) qtf ON t.booking_id = qtf.booking_id
  GROUP BY t.booking_id, currency
),
q_payment_methods AS (
  SELECT b._id AS booking_id,
    multiIf(b.payment_method_cd = '1', 'Cash',
            b.payment_method_cd = '2', 'Voucher',
            b.payment_method_cd = '3', 'Credit Card',
            'Other') AS txt_payment_method
  FROM picapmongoprod.bookings b FINAL
  WHERE b._id IN (SELECT booking_id FROM q_transactions_filtered)
  GROUP BY b._id, b.payment_method_cd
)
SELECT
  qtf.passenger_id AS passenger_id,
  qtf.company_id   AS company_id,
  qtf.company_name AS Company_name,
  qtf.txt_type     AS TXT_TYPE,
  toDate(toTimeZone(qtf.created_at, 'America/Bogota')) AS TMS_CREATED,
  qtf.booking_id   AS booking_id,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100 AS VAL_AMOUNT,
  qtf._id          AS _id,
  ifNull(t.booking_driver_payment, 0) AS VAL_AMOUNT_BOOKING_DRIVER_PAYMENT,
  multiIf((pm.txt_payment_method = 'Cash') AND (t.booking_driver_payment != 0), 'Company Wallet',
          pm.txt_payment_method != 'Cash', pm.txt_payment_method,
          qtf.amount_charged_to_company_wallet > 0, 'Company Wallet',
          'Cash') AS Payment_Type
FROM q_transactions_filtered qtf
LEFT JOIN q_transactions   t  ON t.booking_id  = qtf.booking_id
LEFT JOIN q_payment_methods pm ON pm.booking_id = qtf.booking_id
ORDER BY TMS_CREATED, qtf.booking_id
"""

# Query B: Comisión Recaudo (BookingCompanyCollectionFee) → Hoja 1
QUERY_COMISION = QUERY_RECAUDOS.replace(
    "WalletAccountCounterDeliveryPaymentTransaction",
    "WalletAccountTransactionBookingCompanyCollectionFee",
)


# Query C: collection_fee de companies → para Porcentaje Real en Hoja 3
QUERY_COMPANIES_FEE = """
SELECT _id, name,
  toFloat64OrZero(collection_fee) / 100.0 AS fee_decimal
FROM picapmongoprod.companies FINAL
WHERE collection_fee IS NOT NULL AND collection_fee != ''
"""


# Query D: Resumen Recaudo por User_Company → para Hoja 5
QUERY_RESUMEN_USER = """
WITH q_wat_filtered AS (
  SELECT wat.amount AS amount, p.company_id AS passenger_company_id
  FROM picapmongoprod.wallet_account_transactions AS wat FINAL
  INNER JOIN picapmongoprod.packages   AS pck FINAL ON pck._id = wat.package_id
  INNER JOIN picapmongoprod.bookings   AS b   FINAL ON b._id   = wat.booking_id
  INNER JOIN picapmongoprod.passengers AS p   FINAL ON p._id   = b.passenger_id
  INNER JOIN picapmongoprod.countries  AS c   FINAL ON c._id   = b.country_id
  WHERE JSONExtractString(c.name, 'es') = 'Colombia'
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND pck.counter_delivery = 'true'
    AND wat._type = 'WalletAccountCounterDeliveryPaymentTransaction'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
)
SELECT
  compp.name AS User_Company,
  SUM(toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100) AS Suma_Transaction_amount
FROM q_wat_filtered qtf
LEFT JOIN picapmongoprod.companies AS compp FINAL ON compp._id = qtf.passenger_company_id
GROUP BY compp.name
ORDER BY Suma_Transaction_amount DESC
"""


# ════════════════════════════════════════════════════════════════════════════
# Estilos
# ════════════════════════════════════════════════════════════════════════════

HDR_FILL  = PatternFill("solid", start_color="1F4E78")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="left", vertical="center")

PIVOT_HDR_FILL = PatternFill("solid", start_color="D9E1F2")
PIVOT_TOTAL_FILL = PatternFill("solid", start_color="FFE699")

THIN = Side(border_style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

INT_FMT   = "#,##0"
DEC_FMT   = "#,##0.00"
PCT_FMT   = "0.00%"
RED_FONT  = Font(color="C00000")

# Empresas a EXCLUIR en hojas Ida y Vuelta (matching case-insensitive sobre substring)
EXCLUSIONES_IDA_VUELTA = [
    "multipaquete",
    "surtitodo",       # incluye "Surtitodo express"
    "pibox admin",
    "testeo",
    "test",
    "qa",
    "prueba",
]


def empresa_excluida(nombre: str) -> bool:
    """True si el nombre matchea alguna keyword de exclusión Ida y Vuelta."""
    if not nombre:
        return False
    n = str(nombre).lower().strip()
    return any(kw in n for kw in EXCLUSIONES_IDA_VUELTA)


# ════════════════════════════════════════════════════════════════════════════
# Helpers de período
# ════════════════════════════════════════════════════════════════════════════

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def mes_a_rango(mes_str: str) -> tuple[str, str, str]:
    """'2026-04' → ('2026-04-01', '2026-04-30', '1 al 30 Abril 2026')"""
    año, mes = mes_str.split("-")
    año_i, mes_i = int(año), int(mes)
    _, last_day = calendar.monthrange(año_i, mes_i)
    desde = f"{año}-{mes}-01"
    hasta = f"{año}-{mes}-{last_day:02d}"
    periodo_txt = f"1 al {last_day} {MESES_ES[mes_i]} {año}"
    return desde, hasta, periodo_txt


# ════════════════════════════════════════════════════════════════════════════
# Escritura de hojas
# ════════════════════════════════════════════════════════════════════════════

HEADERS_QUERY = [
    "passenger_id", "company_id", "Company_name", "TXT_TYPE", "TMS_CREATED",
    "qtf.booking_id", "VAL_AMOUNT", "_id", "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT", "Payment_Type",
]
# Mapeo header CH → header Excel (sólo el qtf.booking_id cambia)
HEADER_CH_A_EXCEL = {
    "passenger_id": "passenger_id",
    "company_id": "company_id",
    "Company_name": "Company_name",
    "TXT_TYPE": "TXT_TYPE",
    "TMS_CREATED": "TMS_CREATED",
    "booking_id": "qtf.booking_id",
    "VAL_AMOUNT": "VAL_AMOUNT",
    "_id": "_id",
    "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT": "VAL_AMOUNT_BOOKING_DRIVER_PAYMENT",
    "Payment_Type": "Payment_Type",
}


def _aplicar_rojo_negativo(ws, col_letters: list[str], first_row: int, last_row: int):
    """Aplica formato condicional: fuente roja si valor < 0 en las columnas dadas."""
    if last_row < first_row:
        return
    for col in col_letters:
        rng = f"{col}{first_row}:{col}{last_row}"
        rule = CellIsRule(operator="lessThan", formula=["0"],
                          font=RED_FONT)
        ws.conditional_formatting.add(rng, rule)


def write_data_query_sheet(ws, headers_ch: list[str], rows: list[list[str]],
                            extra_cols: list[tuple[str, callable]] = None):
    """
    Escribe data cruda con headers azul. Reordena/renombra headers según HEADER_CH_A_EXCEL.
    extra_cols: lista de (header, fn_formula(row_idx) -> str) para columnas calculadas
                que se agregan a la derecha.
    """
    # Mapear índices CH → posición esperada en HEADERS_QUERY
    pos_ch = {h: i for i, h in enumerate(headers_ch)}
    headers_excel = HEADERS_QUERY[:]
    extras = extra_cols or []
    for hdr, _ in extras:
        headers_excel.append(hdr)

    # Escribir headers
    for c_idx, h in enumerate(headers_excel, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIGN

    # Mapeo de cada col Excel a su origen en headers_ch
    col_mapping = []
    for hdr_excel in HEADERS_QUERY:
        # encontrar el header CH equivalente
        hdr_ch = next((ch for ch, ex in HEADER_CH_A_EXCEL.items() if ex == hdr_excel), hdr_excel)
        col_mapping.append(pos_ch.get(hdr_ch))

    val_amount_col_letter = get_column_letter(HEADERS_QUERY.index("VAL_AMOUNT") + 1)

    for r_idx, r in enumerate(rows, start=2):
        # Cols originales
        for c_idx, src_idx in enumerate(col_mapping, start=1):
            if src_idx is None or src_idx >= len(r):
                ws.cell(row=r_idx, column=c_idx, value=None)
            else:
                ws.cell(row=r_idx, column=c_idx, value=to_num(r[src_idx]))
        # Cols extra calculadas
        for off, (hdr, fn) in enumerate(extras, start=len(HEADERS_QUERY) + 1):
            ws.cell(row=r_idx, column=off, value=fn(r_idx))

    # Formato moneda en VAL_AMOUNT
    last_row = 1 + len(rows)
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=HEADERS_QUERY.index("VAL_AMOUNT") + 1).number_format = DEC_FMT
        ws.cell(row=r, column=HEADERS_QUERY.index("VAL_AMOUNT_BOOKING_DRIVER_PAYMENT") + 1).number_format = DEC_FMT

    # Formato condicional rojo para negativos en VAL_AMOUNT
    _aplicar_rojo_negativo(ws, [val_amount_col_letter], 2, last_row)

    # Anchos
    widths = {"A": 22, "B": 22, "C": 28, "D": 35, "E": 12, "F": 22, "G": 14, "H": 22, "I": 14, "J": 16}
    for col_l, w in widths.items():
        ws.column_dimensions[col_l].width = w
    for off in range(len(HEADERS_QUERY) + 1, len(headers_excel) + 1):
        ws.column_dimensions[get_column_letter(off)].width = 16

    ws.freeze_panes = "A2"

    # AutoFilter sobre todo el rango de datos
    if len(rows) > 0:
        last_col_letter = get_column_letter(len(headers_excel))
        _add_auto_filter(ws, f"A1:{last_col_letter}{1 + len(rows)}")


# ════════════════════════════════════════════════════════════════════════════
# Hoja 1: Comisión Recaudo
# ════════════════════════════════════════════════════════════════════════════

def write_hoja1(wb, headers_ch, rows):
    ws = wb.create_sheet("1. Comisión Recaudo")
    write_data_query_sheet(ws, headers_ch, rows)


# ════════════════════════════════════════════════════════════════════════════
# Hoja 2: Recaudos + cols calculadas % y Comisión Manual
# ════════════════════════════════════════════════════════════════════════════

def write_hoja2(wb, headers_ch, rows):
    ws = wb.create_sheet("2. Recaudos")
    # Col K (%) → BUSCARV a hoja 3
    # Col L (Comisión Manual) → K * G  (G es VAL_AMOUNT)
    extras = [
        ("%",               lambda r_idx: f"=IFERROR(VLOOKUP(C{r_idx},'3. Cruce company'!A:C,3,0),0)"),
        ("Comisión Manual", lambda r_idx: f"=K{r_idx}*G{r_idx}"),
    ]
    write_data_query_sheet(ws, headers_ch, rows, extra_cols=extras)
    last_row = 1 + len(rows)
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=11).number_format = PCT_FMT
        ws.cell(row=r, column=12).number_format = DEC_FMT
    _aplicar_rojo_negativo(ws, ["L"], 2, last_row)


# ════════════════════════════════════════════════════════════════════════════
# Hoja 3: Cruce Company (2 pivotes lado a lado)
# ════════════════════════════════════════════════════════════════════════════

def write_hoja3(wb, recaudos_dicts, comision_dicts, fee_por_company_id):
    """
    A:G  → Pivote de Hoja 2 (Recaudos) por Company_name + cols calculadas.
    I:J  → Mini-pivote de Hoja 1 (Comisión) por Company_name.
    """
    ws = wb.create_sheet("3. Cruce company")

    # Pivote Hoja 2 (Recaudos) por Company_name
    sumas_recaudos = defaultdict(float)
    company_id_por_name = {}
    for r in recaudos_dicts:
        emp = (r.get("Company_name") or "").strip()
        if not emp:
            continue
        try:
            sumas_recaudos[emp] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
        if emp not in company_id_por_name:
            company_id_por_name[emp] = r.get("company_id") or ""
    sorted_recaudos = sorted(sumas_recaudos.items(), key=lambda kv: kv[0].lower())

    # Mini-pivote Hoja 1 (Comisión) por Company_name
    sumas_comision = defaultdict(float)
    for r in comision_dicts:
        emp = (r.get("Company_name") or "").strip()
        if not emp:
            continue
        try:
            sumas_comision[emp] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
    sorted_comision = sorted(sumas_comision.items(), key=lambda kv: kv[0].lower())

    # Títulos
    ws.cell(row=2, column=1, value="Query 2 Recaudos").font = Font(bold=True)
    ws.cell(row=2, column=9, value="1. Comisión Recaudo").font = Font(bold=True)

    # Headers fila 3
    hdrs_left  = ["CLIENTE", "Suma de VAL_AMOUNT", "Porcentaje Real", "Comisión Real",
                  "Comisión Trump", "Dif", "id company"]
    hdrs_right = ["Etiquetas de fila", "Suma de VAL_AMOUNT"]
    for c_idx, h in enumerate(hdrs_left, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL
    for c_idx, h in enumerate(hdrs_right, start=9):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    # Data izquierda (A:G)
    for i, (emp, monto) in enumerate(sorted_recaudos, start=4):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2)); c.number_format = INT_FMT; c.border = BORDER_ALL
        # Porcentaje Real (vía company_id → fee_por_company_id)
        cid = company_id_por_name.get(emp, "")
        pct = fee_por_company_id.get(cid, 0.0)
        c = ws.cell(row=i, column=3, value=pct); c.number_format = PCT_FMT; c.border = BORDER_ALL
        # Comisión Real = B*C
        c = ws.cell(row=i, column=4, value=f"=B{i}*C{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Comisión Trump = VLOOKUP a la mini-tabla I:J
        c = ws.cell(row=i, column=5, value=f"=IFERROR(VLOOKUP(A{i},I:J,2,0),0)")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Dif = E - D
        c = ws.cell(row=i, column=6, value=f"=E{i}-D{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        # id_company
        c = ws.cell(row=i, column=7, value=cid); c.border = BORDER_ALL

    total_left_row = 4 + len(sorted_recaudos)
    ws.cell(row=total_left_row, column=1, value="Total general").font = Font(bold=True)
    ws.cell(row=total_left_row, column=1).fill = PIVOT_TOTAL_FILL
    c = ws.cell(row=total_left_row, column=2,
                value=f"=SUM(B4:B{total_left_row - 1})")
    c.number_format = INT_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL
    for col in (4, 5, 6):
        c = ws.cell(row=total_left_row, column=col,
                    value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_left_row - 1})")
        c.number_format = DEC_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL

    # Data derecha (I:J) — mini-pivote
    for i, (emp, monto) in enumerate(sorted_comision, start=4):
        ws.cell(row=i, column=9, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=10, value=round(monto, 2)); c.number_format = DEC_FMT; c.border = BORDER_ALL

    total_right_row = 4 + len(sorted_comision)
    ws.cell(row=total_right_row, column=9, value="Total general").font = Font(bold=True)
    ws.cell(row=total_right_row, column=9).fill = PIVOT_TOTAL_FILL
    c = ws.cell(row=total_right_row, column=10,
                value=f"=SUM(J4:J{total_right_row - 1})")
    c.number_format = DEC_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL

    # Rojo negativos
    _aplicar_rojo_negativo(ws, ["D", "E", "F", "J"], 4, max(total_left_row, total_right_row))

    # Anchos
    ws.column_dimensions["A"].width = 35
    for col_l in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_l].width = 18
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 18

    # Devolver lookup: emp → (suma_recaudos, % real) para hoja 5 Resumen
    cruce_company = {}
    for emp, monto in sorted_recaudos:
        cid = company_id_por_name.get(emp, "")
        cruce_company[emp] = {
            "recaudos": round(monto, 2),
            "pct":      fee_por_company_id.get(cid, 0.0),
            "company_id": cid,
        }
    return cruce_company


# ════════════════════════════════════════════════════════════════════════════
# Hoja 4: Cruce Booking (2 pivotes)
# ════════════════════════════════════════════════════════════════════════════

def _add_auto_filter(ws, ref: str):
    """Habilita los íconos de auto-filtro en el rango dado.
    Excel los muestra al abrir, el usuario clickea y filtra."""
    try:
        ws.auto_filter.ref = ref
    except Exception:
        pass


def write_hoja4(wb, recaudos_dicts, comision_dicts):
    ws = wb.create_sheet("4. Cruce Booking")

    # Pivote izquierdo: booking_id + Company_name → Suma VAL_AMOUNT (de Hoja 2)
    sumas_book = defaultdict(float)
    comp_por_book = {}
    for r in recaudos_dicts:
        bid = r.get("booking_id") or ""
        if not bid:
            continue
        try:
            sumas_book[bid] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
        if bid not in comp_por_book:
            comp_por_book[bid] = r.get("Company_name") or ""
    sorted_book = sorted(sumas_book.items(), key=lambda kv: comp_por_book.get(kv[0], "").lower())

    # Pivote derecho: booking_id → Suma VAL_AMOUNT (de Hoja 1)
    sumas_book_com = defaultdict(float)
    for r in comision_dicts:
        bid = r.get("booking_id") or ""
        if not bid:
            continue
        try:
            sumas_book_com[bid] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
    sorted_book_com = sorted(sumas_book_com.items(), key=lambda kv: kv[0])

    # Títulos
    ws.cell(row=3, column=1, value="Query 2 Recaudos").font = Font(bold=True)
    ws.cell(row=3, column=10, value="1. Comisión Recaudo").font = Font(bold=True)

    # Headers fila 4
    hdrs_left  = ["qtf.booking_id", "Company_name", "Suma de VAL_AMOUNT", "%",
                  "Comisión Manual", "Cruce", "Diferencia", "Comentario"]
    hdrs_right = ["Etiquetas de fila", "Suma de VAL_AMOUNT"]
    for c_idx, h in enumerate(hdrs_left, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL
    for c_idx, h in enumerate(hdrs_right, start=10):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    # Data izquierda
    for i, (bid, monto) in enumerate(sorted_book, start=5):
        ws.cell(row=i, column=1, value=bid).border = BORDER_ALL
        ws.cell(row=i, column=2, value=comp_por_book.get(bid, "")).border = BORDER_ALL
        c = ws.cell(row=i, column=3, value=round(monto, 2)); c.number_format = INT_FMT; c.border = BORDER_ALL
        # % = BUSCARV a Hoja 2 col K
        c = ws.cell(row=i, column=4, value=f"=IFERROR(VLOOKUP(A{i},'2. Recaudos'!F:K,6,0),0)")
        c.number_format = PCT_FMT; c.border = BORDER_ALL
        # Comisión Manual = C*D
        c = ws.cell(row=i, column=5, value=f"=C{i}*D{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Cruce = VLOOKUP a mini-pivote J:K
        c = ws.cell(row=i, column=6, value=f"=IFERROR(VLOOKUP(A{i},J:K,2,0),0)")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Diferencia = E + F
        c = ws.cell(row=i, column=7, value=f"=E{i}+F{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Comentario vacío
        ws.cell(row=i, column=8, value=None).border = BORDER_ALL

    # Data derecha
    for i, (bid, monto) in enumerate(sorted_book_com, start=5):
        ws.cell(row=i, column=10, value=bid).border = BORDER_ALL
        c = ws.cell(row=i, column=11, value=round(monto, 2))
        c.number_format = DEC_FMT; c.border = BORDER_ALL

    # Rojo negativos
    last_row = max(4 + len(sorted_book), 4 + len(sorted_book_com))
    _aplicar_rojo_negativo(ws, ["E", "F", "G", "K"], 5, last_row)

    # AutoFilter sobre tabla izquierda (A4:H... headers en fila 4)
    _add_auto_filter(ws, f"A4:H{4 + len(sorted_book)}")

    # Anchos
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    for col_l in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_l].width = 16
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 18


# ════════════════════════════════════════════════════════════════════════════
# Hoja 5: Resumen
# ════════════════════════════════════════════════════════════════════════════

def write_hoja5(wb, cruce_company, recaudos_por_user_company, periodo_txt):
    """
    cruce_company: dict {emp_name: {recaudos, pct, company_id}}  (de Hoja 3)
    recaudos_por_user_company: dict {emp_name: monto_total}      (de Query D)
    """
    ws = wb.create_sheet("Resumen")

    # Headers en fila 2 (como en la plantilla, col B en adelante)
    headers = ["N°", "Cliente", "Recaudo", "%", "Comisión", "Periodo",
               "Anticipo", "Pendiente", "Estado", "Comentario"]
    for c_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = HDR_ALIGN

    # Filtrar: % > 0 + excluir "Cruz Verde" (cualquier variante)
    clientes = []
    for emp, info in cruce_company.items():
        if info["pct"] <= 0:
            continue
        if "cruz verde" in emp.lower():
            continue
        clientes.append((emp, info))
    # Ordenar por recaudos descendente
    clientes.sort(key=lambda kv: -kv[1]["recaudos"])

    for i, (emp, info) in enumerate(clientes, start=3):
        # N°
        ws.cell(row=i, column=2, value=i - 2).border = BORDER_ALL
        # Cliente
        ws.cell(row=i, column=3, value=emp).border = BORDER_ALL
        # Recaudo (de la query D)
        recaudo_real = recaudos_por_user_company.get(emp, 0.0)
        c = ws.cell(row=i, column=4, value=round(recaudo_real, 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        # %
        c = ws.cell(row=i, column=5, value=info["pct"])
        c.number_format = PCT_FMT; c.border = BORDER_ALL
        # Comisión = D * -E
        c = ws.cell(row=i, column=6, value=f"=D{i}*-E{i}")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Periodo
        ws.cell(row=i, column=7, value=periodo_txt).border = BORDER_ALL
        # Anticipo = F * -1
        c = ws.cell(row=i, column=8, value=f"=F{i}*-1")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Pendiente = H + F
        c = ws.cell(row=i, column=9, value=f"=H{i}+F{i}")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        # Estado
        c = ws.cell(row=i, column=10, value=f'=IF(I{i}=0,"Pagada","Pendiente")')
        c.border = BORDER_ALL
        # Comentario condicional
        c = ws.cell(row=i, column=11,
                    value=f'=IF(J{i}="Pendiente","Cliente con deuda vigente","")')
        c.border = BORDER_ALL

    # Negativos en rojo (cols F, H, I)
    last_row = 2 + len(clientes)
    _aplicar_rojo_negativo(ws, ["F", "H", "I"], 3, last_row)

    # Anchos
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 32
    for col_l in ["D", "F", "H", "I"]:
        ws.column_dimensions[col_l].width = 16
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 28


# ════════════════════════════════════════════════════════════════════════════
# Hojas 6 / 7: Ida y Vuelta (filtros)
# ════════════════════════════════════════════════════════════════════════════

def _filtrar_ida_vuelta(rows_dicts, val_amount_positivo=False):
    out = []
    for r in rows_dicts:
        emp = r.get("Company_name") or ""
        if empresa_excluida(emp):
            continue
        if val_amount_positivo:
            try:
                if float(r.get("VAL_AMOUNT") or 0) <= 0:
                    continue
            except (ValueError, TypeError):
                continue
        out.append(r)
    return out


def write_hoja6(wb, headers_ch, rows, dicts):
    """Comisión Recaudo Ida y Vuelta: Hoja 1 filtrada (sin exclusiones)."""
    ws = wb.create_sheet("Comisión Recaudo ida y vuelta")
    filtered_dicts = _filtrar_ida_vuelta(dicts, val_amount_positivo=False)
    # Reconvertir dicts → rows en el orden de headers_ch
    rows_f = [[d.get(h) for h in headers_ch] for d in filtered_dicts]
    write_data_query_sheet(ws, headers_ch, rows_f)


def write_hoja7(wb, headers_ch, rows, dicts):
    """Recaudos Ida y Vuelta: Hoja 2 filtrada + VAL_AMOUNT > 0 + cols % y Comisión Manual."""
    ws = wb.create_sheet("Recaudos ida y vuelta")
    filtered_dicts = _filtrar_ida_vuelta(dicts, val_amount_positivo=True)
    rows_f = [[d.get(h) for h in headers_ch] for d in filtered_dicts]
    extras = [
        ("%",               lambda r_idx: f"=IFERROR(VLOOKUP(C{r_idx},'3. Cruce company'!A:C,3,0),0)"),
        ("Comisión Manual", lambda r_idx: f"=K{r_idx}*G{r_idx}"),
    ]
    write_data_query_sheet(ws, headers_ch, rows_f, extra_cols=extras)
    last_row = 1 + len(rows_f)
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=11).number_format = PCT_FMT
        ws.cell(row=r, column=12).number_format = DEC_FMT
    _aplicar_rojo_negativo(ws, ["L"], 2, last_row)


# ════════════════════════════════════════════════════════════════════════════
# Hoja 8: TD Company Ida y Vuelta
# ════════════════════════════════════════════════════════════════════════════

def write_hoja8(wb, recaudos_iv_dicts, comision_dicts):
    """
    Pivote izquierdo: Hoja 7 (Recaudos Ida y Vuelta).
    Pivote derecho:   Hoja 1 COMPLETA (Comisión Recaudo, incluye Surtitodo).
    """
    ws = wb.create_sheet("TD company ida y vuelta")

    # Pivote izquierdo (Hoja 7 = recaudos_iv_dicts)
    sumas_left = defaultdict(float)
    for r in recaudos_iv_dicts:
        emp = (r.get("Company_name") or "").strip()
        if not emp:
            continue
        try:
            sumas_left[emp] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
    sorted_left = sorted(sumas_left.items(), key=lambda kv: kv[0].lower())

    # Pivote derecho (Hoja 1 COMPLETA = comision_dicts — sin filtros, incluye Surtitodo)
    sumas_right = defaultdict(float)
    for r in comision_dicts:
        emp = (r.get("Company_name") or "").strip()
        if not emp:
            continue
        try:
            sumas_right[emp] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
    sorted_right = sorted(sumas_right.items(), key=lambda kv: kv[0].lower())

    # Títulos
    ws.cell(row=2, column=1, value="2. Recaudos").font = Font(bold=True)
    ws.cell(row=2, column=8, value="1. Comisión Recaudo").font = Font(bold=True)

    # Headers fila 3
    hdrs_left  = ["Etiquetas de fila", "Suma de VAL_AMOUNT", "Porcentaje Real",
                  "Comisión Real", "Comisión Trump", "Dif"]
    hdrs_right = ["Etiquetas de fila", "Suma de VAL_AMOUNT"]
    for c_idx, h in enumerate(hdrs_left, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL
    for c_idx, h in enumerate(hdrs_right, start=8):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    # Data izquierda
    for i, (emp, monto) in enumerate(sorted_left, start=4):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2)); c.number_format = INT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=3, value=f"=IFERROR(VLOOKUP(A{i},'3. Cruce company'!A:C,3,0),0)")
        c.number_format = PCT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=4, value=f"=B{i}*C{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=5, value=f"=IFERROR(VLOOKUP(A{i},H:I,2,0),0)")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=6, value=f"=D{i}+E{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL

    r_total_l = 4 + len(sorted_left)
    ws.cell(row=r_total_l, column=1, value="Total general").font = Font(bold=True)
    ws.cell(row=r_total_l, column=1).fill = PIVOT_TOTAL_FILL
    for col, fmt in [(2, INT_FMT), (4, DEC_FMT), (5, DEC_FMT), (6, DEC_FMT)]:
        c = ws.cell(row=r_total_l, column=col,
                    value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{r_total_l - 1})")
        c.number_format = fmt; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL

    # Data derecha
    for i, (emp, monto) in enumerate(sorted_right, start=4):
        ws.cell(row=i, column=8, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=9, value=round(monto, 2)); c.number_format = DEC_FMT; c.border = BORDER_ALL

    r_total_r = 4 + len(sorted_right)
    ws.cell(row=r_total_r, column=8, value="Total general").font = Font(bold=True)
    ws.cell(row=r_total_r, column=8).fill = PIVOT_TOTAL_FILL
    c = ws.cell(row=r_total_r, column=9, value=f"=SUM(I4:I{r_total_r - 1})")
    c.number_format = DEC_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL

    _aplicar_rojo_negativo(ws, ["D", "E", "F", "I"], 4, max(r_total_l, r_total_r))

    # AutoFilter sobre tabla izquierda (A3:F... headers en fila 3)
    _add_auto_filter(ws, f"A3:F{3 + len(sorted_left)}")

    ws.column_dimensions["A"].width = 35
    for col_l in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_l].width = 18
    ws.column_dimensions["H"].width = 35
    ws.column_dimensions["I"].width = 18


# ════════════════════════════════════════════════════════════════════════════
# Hoja 9: TD Bookings Ida y Vuelta
# ════════════════════════════════════════════════════════════════════════════

def write_hoja9(wb, recaudos_iv_dicts, comision_dicts):
    """
    Pivote izquierdo: Hoja 7 (Recaudos Ida y Vuelta) por booking_id + Company_name.
    Pivote derecho:   Hoja 1 COMPLETA (Comisión Trump) por booking_id, incluye Surtitodo.
    """
    ws = wb.create_sheet("TD Bookings ida y vuelta")

    # Pivote izquierdo (Hoja 7 = recaudos_iv_dicts): booking_id + Company_name
    sumas_left = defaultdict(float)
    comp_por_book = {}
    for r in recaudos_iv_dicts:
        bid = r.get("booking_id") or ""
        if not bid:
            continue
        try:
            sumas_left[bid] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
        if bid not in comp_por_book:
            comp_por_book[bid] = r.get("Company_name") or ""
    sorted_left = sorted(sumas_left.items(), key=lambda kv: comp_por_book.get(kv[0], "").lower())

    # Pivote derecho (Hoja 1 COMPLETA = comision_dicts): sin filtros, incluye Surtitodo
    sumas_right = defaultdict(float)
    for r in comision_dicts:
        bid = r.get("booking_id") or ""
        if not bid:
            continue
        try:
            sumas_right[bid] += float(r.get("VAL_AMOUNT") or 0)
        except (ValueError, TypeError):
            pass
    sorted_right = sorted(sumas_right.items(), key=lambda kv: kv[0])

    # Títulos
    ws.cell(row=2, column=1, value="recaudos").font = Font(bold=True)
    ws.cell(row=2, column=2, value="comision manual").font = Font(bold=True)
    ws.cell(row=2, column=10, value="comision trump").font = Font(bold=True)

    # Headers fila 3
    hdrs_left  = ["booking_id", "Company_name", "Valor Recaudo", "%",
                  "Comision manual", "cruce % en sistema", "dif"]
    hdrs_right = ["Etiquetas de fila", "Suma de VAL_AMOUNT", "A", "B"]
    for c_idx, h in enumerate(hdrs_left, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL
    for c_idx, h in enumerate(hdrs_right, start=10):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = PIVOT_HDR_FILL; cell.font = Font(bold=True); cell.border = BORDER_ALL

    # Data izquierda
    for i, (bid, monto) in enumerate(sorted_left, start=4):
        ws.cell(row=i, column=1, value=bid).border = BORDER_ALL
        ws.cell(row=i, column=2, value=comp_por_book.get(bid, "")).border = BORDER_ALL
        c = ws.cell(row=i, column=3, value=round(monto, 2)); c.number_format = INT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=4, value=f"=IFERROR(VLOOKUP(A{i},'Recaudos ida y vuelta'!F:K,6,0),0)")
        c.number_format = PCT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=5, value=f"=(C{i}*D{i})*-1"); c.number_format = DEC_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=6, value=f"=IFERROR(VLOOKUP(A{i},J:K,2,0),0)")
        c.number_format = DEC_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=7, value=f"=F{i}-E{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL

    # Data derecha
    for i, (bid, monto) in enumerate(sorted_right, start=4):
        ws.cell(row=i, column=10, value=bid).border = BORDER_ALL
        c = ws.cell(row=i, column=11, value=round(monto, 2)); c.number_format = DEC_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=12, value=f"=IFERROR(VLOOKUP(J{i},A:D,4,0),0)")
        c.number_format = PCT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=13, value=f"=K{i}+L{i}"); c.number_format = DEC_FMT; c.border = BORDER_ALL

    last_row = max(3 + len(sorted_left), 3 + len(sorted_right))
    _aplicar_rojo_negativo(ws, ["E", "F", "G", "K", "M"], 4, last_row)

    # AutoFilter sobre tabla izquierda (A3:G... headers en fila 3) — habilita
    # los íconos en col D (%) y col G (dif). Excel los muestra al abrir, el
    # usuario clickea y filtra: en D excluir 0, en G mostrar > 0.
    _add_auto_filter(ws, f"A3:G{3 + len(sorted_left)}")

    # Anchos
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    for col_l in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_l].width = 18
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 14


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera el Excel 'Comisión Recaudos' (9 hojas) desde ClickHouse",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--mes", required=True, help="Mes a procesar en formato YYYY-MM (ej. 2026-04)")
    parser.add_argument("--output", help="Nombre del archivo .xlsx")
    parser.add_argument("--ch-host", default=os.environ.get("MINTIC_CH_HOST", ""))
    parser.add_argument("--ch-user", default=os.environ.get("MINTIC_CH_USER", ""))
    parser.add_argument("--ch-pass", default=os.environ.get("MINTIC_CH_PASS", ""))
    args = parser.parse_args()

    if not args.ch_host or not args.ch_user or not args.ch_pass:
        sys.exit(
            "❌ Credenciales ClickHouse no configuradas.\n"
            "   Setealas como env vars (MINTIC_CH_HOST/USER/PASS) o pasalas con\n"
            "   --ch-host --ch-user --ch-pass."
        )

    try:
        desde, hasta, periodo_txt = mes_a_rango(args.mes)
    except Exception as e:
        sys.exit(f"❌ --mes inválido: {args.mes} ({e}). Usá formato YYYY-MM (ej. 2026-04).")

    año, mes = args.mes.split("-")
    nombre_mes = MESES_ES[int(mes)]
    output = args.output or f"Comisión Recaudos {nombre_mes} {año}.xlsx"
    output_path = Path(output)

    print(f"🚀 Generando reporte Comisión Recaudos")
    print(f"   • Período  : {desde} → {hasta} ({periodo_txt})")
    print(f"   • Output   : {output_path}")
    print()

    # Query A: Recaudos
    print("📊 Query A: Recaudos (CounterDelivery)…")
    t0 = time.time()
    sql_a = QUERY_RECAUDOS.format(desde=desde, hasta=hasta)
    text_a = ch_request(sql_a, args.ch_host, args.ch_user, args.ch_pass)
    headers_a, rows_a = parse_tsv(text_a)
    print(f"   ✓ {len(rows_a)} filas en {time.time() - t0:.1f}s")

    # Query B: Comisión
    print("📊 Query B: Comisión (BookingCompanyCollectionFee)…")
    t0 = time.time()
    sql_b = QUERY_COMISION.format(desde=desde, hasta=hasta)
    text_b = ch_request(sql_b, args.ch_host, args.ch_user, args.ch_pass)
    headers_b, rows_b = parse_tsv(text_b)
    print(f"   ✓ {len(rows_b)} filas en {time.time() - t0:.1f}s")

    # Query C: companies.collection_fee
    print("📊 Query C: % de comisión por company (collection_fee)…")
    t0 = time.time()
    text_c = ch_request(QUERY_COMPANIES_FEE, args.ch_host, args.ch_user, args.ch_pass, timeout=120)
    headers_c, rows_c = parse_tsv(text_c)
    # collection_fee viene como "1" (= 1%). La query ya divide por 100 → 0.01.
    # No volver a dividir.
    fee_por_company_id = {}
    for r in rows_c:
        if len(r) >= 3:
            try:
                fee_por_company_id[r[0]] = float(r[2])  # decimal (0.01 = 1%)
            except (ValueError, IndexError):
                pass
    print(f"   ✓ {len(fee_por_company_id)} empresas con collection_fee en {time.time() - t0:.1f}s")

    # Query D: Recaudo total por User_Company (para Hoja 5)
    print("📊 Query D: Recaudo total por User_Company…")
    t0 = time.time()
    sql_d = QUERY_RESUMEN_USER.format(desde=desde, hasta=hasta)
    text_d = ch_request(sql_d, args.ch_host, args.ch_user, args.ch_pass)
    headers_d, rows_d = parse_tsv(text_d)
    recaudos_por_user_company = {}
    for r in rows_d:
        if len(r) >= 2:
            try:
                recaudos_por_user_company[r[0]] = float(r[1])
            except (ValueError, IndexError):
                pass
    print(f"   ✓ {len(recaudos_por_user_company)} empresas en {time.time() - t0:.1f}s")

    # Convertir rows → list[dict] para hojas que requieren lookups
    recaudos_dicts = [
        {h: to_num(v) if h not in ("Company_name", "TXT_TYPE", "Payment_Type",
                                    "passenger_id", "company_id", "booking_id", "_id", "TMS_CREATED")
              else v for h, v in zip(headers_a, r)}
        for r in rows_a
    ]
    comision_dicts = [
        {h: to_num(v) if h not in ("Company_name", "TXT_TYPE", "Payment_Type",
                                    "passenger_id", "company_id", "booking_id", "_id", "TMS_CREATED")
              else v for h, v in zip(headers_b, r)}
        for r in rows_b
    ]

    # ───── Generar Excel ─────
    print()
    print("📝 Generando Excel…")
    wb = Workbook()
    wb.remove(wb.active)

    write_hoja1(wb, headers_b, rows_b)
    print(f"   ✓ Hoja 1: '1. Comisión Recaudo' ({len(rows_b)} filas)")

    write_hoja2(wb, headers_a, rows_a)
    print(f"   ✓ Hoja 2: '2. Recaudos' ({len(rows_a)} filas)")

    cruce_company = write_hoja3(wb, recaudos_dicts, comision_dicts, fee_por_company_id)
    print(f"   ✓ Hoja 3: '3. Cruce company' ({len(cruce_company)} empresas)")

    write_hoja4(wb, recaudos_dicts, comision_dicts)
    print(f"   ✓ Hoja 4: '4. Cruce Booking'")

    write_hoja5(wb, cruce_company, recaudos_por_user_company, periodo_txt)
    print(f"   ✓ Hoja 5: 'Resumen'")

    write_hoja6(wb, headers_b, rows_b, comision_dicts)
    print(f"   ✓ Hoja 6: 'Comisión Recaudo ida y vuelta'")

    write_hoja7(wb, headers_a, rows_a, recaudos_dicts)
    print(f"   ✓ Hoja 7: 'Recaudos ida y vuelta'")

    # Recalcular dicts filtrados para hojas 8 y 9 (sólo el lado izquierdo).
    # El lado derecho de Hojas 8 y 9 usa comision_dicts COMPLETO para que aparezca
    # Surtitodo (que está en Hoja 1 pero NO en Hoja 6 por la exclusión Ida y Vuelta).
    recaudos_iv_dicts = _filtrar_ida_vuelta(recaudos_dicts, val_amount_positivo=True)
    write_hoja8(wb, recaudos_iv_dicts, comision_dicts)
    print(f"   ✓ Hoja 8: 'TD company ida y vuelta'")

    write_hoja9(wb, recaudos_iv_dicts, comision_dicts)
    print(f"   ✓ Hoja 9: 'TD Bookings ida y vuelta'")

    wb.save(output_path)
    print()
    print(f"✅ Listo: {output_path}")
    print(f"   Tamaño: {output_path.stat().st_size / 1024 / 1024:.1f} MB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
