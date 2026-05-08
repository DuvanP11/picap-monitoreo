"""
api.py — Picap Evasión de Comisiones (versión low-memory)
Toda la agregación se hace en ClickHouse. Python solo recibe
unas pocas filas de resumen — cabe en los 512 MB de Render Free.
"""

import uuid
from flask import Flask, jsonify, request, send_from_directory, send_file
import io
from flask_cors import CORS
import clickhouse_connect
from datetime import datetime, date, timedelta
import threading
import time
import math

app = Flask(__name__)
CORS(app)

from werkzeug.exceptions import HTTPException

def _is_api_path():
    """True si la request actual apunta a un endpoint /api/*."""
    try:
        return request.path.startswith("/api/")
    except Exception:
        return False

@app.errorhandler(HTTPException)
def handle_http_exception(e):
    """Toda excepción HTTP en rutas /api/* devuelve JSON; HTML solo para SPA."""
    from flask import jsonify as _jfy
    code = e.code or 500
    if _is_api_path():
        return _jfy({
            "ok": False,
            "error": e.description or e.name,
            "code": code,
            "path": request.path,
        }), code
    # Para rutas no-API: devolvemos el dashboard (SPA) en 404 para que el
    # router del frontend resuelva, en lugar de HTML por defecto de Flask.
    if code == 404:
        try:
            return send_from_directory(".", "dashboard.html")
        except Exception:
            pass
    return _jfy({"ok": False, "error": e.description or e.name, "code": code}), code

@app.errorhandler(Exception)
def handle_error(e):
    """Cualquier excepción no-HTTP retorna JSON con 500 (nunca HTML)."""
    from flask import jsonify as _jfy
    import traceback as _tb
    print(f"[handle_error] {type(e).__name__}: {e}\n{_tb.format_exc()}")
    return _jfy({
        "ok": False,
        "error": str(e),
        "code": 500,
        "type": type(e).__name__,
    }), 500

def limpiar_nan(obj):
    from datetime import datetime, date
    if isinstance(obj, dict):
        return {k: limpiar_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [limpiar_nan(v) for v in obj]
    elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    elif isinstance(obj, datetime):
        if obj.tzinfo is not None:
            obj = obj.replace(tzinfo=None)
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(obj, date):
        return obj.strftime('%Y-%m-%d')
    return obj

def get_client():
    # Credenciales de ClickHouse vía variables de entorno.
    # Para desarrollo local usa un .env (no incluido en git) o exporta:
    #   CLICKHOUSE_HOST, CLICKHOUSE_PORT, CLICKHOUSE_USER,
    #   CLICKHOUSE_PASSWORD, CLICKHOUSE_DATABASE
    import os as _os
    return clickhouse_connect.get_client(
        host=_os.environ.get("CLICKHOUSE_HOST", "clickhouse.picap.io"),
        port=int(_os.environ.get("CLICKHOUSE_PORT", "8443")),
        username=_os.environ.get("CLICKHOUSE_USER", "dperilla"),
        password=_os.environ.get("CLICKHOUSE_PASSWORD", ""),
        database=_os.environ.get("CLICKHOUSE_DATABASE", "picapmongoprod"),
        secure=True
    )

BASE_CTE = """
WITH base AS (
    SELECT
        toTimeZone(b.created_at, 'America/Bogota') AS creacion_servicio,
        b._id              AS booking_id,
        b.driver_id        AS id_driver,
        pd.name            AS name_driver,
        CASE
            WHEN b.g_adm_area_lv_1 = 'MN' THEN 'Managua'
            WHEN b.g_adm_area_lv_1 = 'Guatemala Department' THEN 'Guatemala'
            WHEN b.g_adm_area_lv_1 = '' THEN 'Sin ciudad'
            ELSE b.g_adm_area_lv_1
        END AS ciudad,
        JSONExtractString(b.final_cost, 'currency_iso') AS moneda,
        CASE
            WHEN b.g_country = 'CO' THEN 'Colombia'
            WHEN b.g_country = 'MX' THEN 'Mexico'
            WHEN b.g_country = 'NI' THEN 'Nicaragua'
            WHEN b.g_country = 'GT' THEN 'Guatemala'
            ELSE 'Otro'
        END AS pais,
        toFloat64OrNull(JSONExtractString(b.estimated_cost, 'cents')) / 100 AS costo_estimado,
        extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[\s*([+-]?\d+\.\d+)')     AS ev_cancel_lon_str,
        extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[.*?,\s*([+-]?\d+\.\d+)') AS ev_cancel_lat_str,
        extract(ifNull(b.events,''), 'event_cd":20.*?created_at":"([^"]+)')                       AS ev_accept,
        extract(ifNull(b.events,''), 'event_cd":26.*?created_at":"([^"]+)')                       AS ev_cancel,
        toFloat64OrNull(ev_cancel_lon_str) AS cancel_lon,
        toFloat64OrNull(ev_cancel_lat_str) AS cancel_lat,
        toFloat64(JSONExtractString(b.end_geojson, 'coordinates', 1)) AS end_lon,
        toFloat64(JSONExtractString(b.end_geojson, 'coordinates', 2)) AS end_lat,
        dateDiff('minute',
            parseDateTimeBestEffortOrNull(ev_accept),
            parseDateTimeBestEffortOrNull(ev_cancel)
        ) AS minutos_entre_eventos,
        JSONExtractString(st.name, 'es') AS type_service,
        b.company_id AS id_company,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    LEFT JOIN picapmongoprod.passengers pd ON b.driver_id = pd._id
    LEFT JOIN picapmongoprod.service_types st ON st._id = b.requested_service_type_id
    WHERE
        NOT empty(b.origin_geojson)
        AND NOT empty(b.end_geojson)
        AND b.status_cd IN (100, 102)
        AND b.created_at >= toDateTime('{fecha_desde} 00:00:00')
        AND b.created_at <= toDateTime('{fecha_hasta} 23:59:59')
),
clasificado AS (
    SELECT *,
        round(geoDistance(cancel_lon, cancel_lat, end_lon, end_lat), 2) AS distancia_cancel_destino,
        (cancel_lon IS NULL OR cancel_lat IS NULL) AS sin_gps,
        (minutos_entre_eventos > 5)                AS flag_tiempo,
        -- Colombia 450m, Mexico/Nicaragua 280m
        multiIf(
            pais = 'Colombia',              geoDistance(cancel_lon, cancel_lat, end_lon, end_lat) <= 450,
            pais IN ('Mexico','Nicaragua'),  geoDistance(cancel_lon, cancel_lat, end_lon, end_lat) <= 280,
            geoDistance(cancel_lon, cancel_lat, end_lon, end_lat) <= 450
        ) AS flag_distancia,
        multiIf(
            (minutos_entre_eventos > 5) AND multiIf(pais='Colombia',geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=450,pais IN ('Mexico','Nicaragua'),geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=280,geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=450), 3,
            (minutos_entre_eventos > 5) AND (cancel_lon IS NULL OR cancel_lat IS NULL), 2,
            (minutos_entre_eventos > 5) OR  multiIf(pais='Colombia',geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=450,pais IN ('Mexico','Nicaragua'),geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=280,geoDistance(cancel_lon,cancel_lat,end_lon,end_lat)<=450), 2,
            0
        ) AS nivel,
        multiIf(pais='Colombia', costo_estimado*0.12, pais IN ('Mexico','Nicaragua'), costo_estimado*0.10, pais='Guatemala', costo_estimado*0.15, costo_estimado*0.15) AS comision_servicio,
        multiIf(pais='Colombia', costo_estimado*0.12*1.05, pais IN ('Mexico','Nicaragua'), costo_estimado*0.10*1.05, pais='Guatemala', costo_estimado*0.15*1.05, costo_estimado*0.15*1.05) AS comision_mas_penalizacion
    FROM base
    WHERE rn = 1
      AND (type_service != 'Mensajería' OR type_service = '' OR type_service IS NULL)
      AND (id_company IS NULL OR id_company = '')
)
"""
Q_KPIS = BASE_CTE + """
SELECT
    count()                                              AS total,
    countIf(nivel = 3)                                   AS confirmadas,
    countIf(nivel = 2)                                   AS probables,
    countIf(nivel = 0)                                   AS ok,
    countIf(sin_gps = 1)                                 AS sin_gps,
    countIf(flag_tiempo = 1)                             AS flag_tiempo,
    countIf(flag_distancia = 1)                          AS flag_distancia,
    round(sumIf(comision_servicio,        nivel >= 2), 0) AS comision_evadida,
    round(sumIf(comision_mas_penalizacion, nivel = 3),  0) AS penalizacion_evadida,  -- solo confirmadas
    round(avgIf(minutos_entre_eventos,    nivel >= 2), 1) AS prom_minutos,
    round(avgIf(distancia_cancel_destino, nivel >= 2), 1) AS prom_distancia
FROM clasificado
"""

Q_TENDENCIA = BASE_CTE + """
SELECT
    toDate(creacion_servicio) AS fecha,
    countIf(nivel = 3)        AS conf,
    countIf(nivel = 2)        AS prob,
    countIf(nivel = 0)        AS ok
FROM clasificado
GROUP BY fecha ORDER BY fecha
"""

Q_CIUDADES = BASE_CTE + """
SELECT
    if(ciudad = '' OR ciudad IS NULL, 'Sin ciudad', ciudad) AS ciudad,
    countIf(nivel >= 2) AS evasiones
FROM clasificado
WHERE nivel >= 2
GROUP BY ciudad ORDER BY evasiones DESC LIMIT 8
"""

Q_TOP_DRIVERS = BASE_CTE + """
SELECT
    id_driver,
    any(name_driver)   AS nombre,
    countIf(nivel = 3) AS conf,
    countIf(nivel = 2) AS prob,
    count()            AS total
FROM clasificado
WHERE nivel >= 2
GROUP BY id_driver ORDER BY conf DESC, total DESC LIMIT 10
"""

# ══════════════════════════════════════════════════════════
# Query independiente — Estado de cobros en wallet
# No usa BASE_CTE porque cruza con wallet_account_transactions
# ══════════════════════════════════════════════════════════
Q_WALLET = """
WITH evasores AS (
    SELECT
        b._id AS booking_id,
        b.driver_id,
        CASE
            WHEN b.g_country = 'CO' THEN 'Colombia'
            WHEN b.g_country = 'MX' THEN 'Mexico'
            WHEN b.g_country = 'NI' THEN 'Nicaragua'
            WHEN b.g_country = 'GT' THEN 'Guatemala'
            ELSE 'Otro'
        END AS pais,
        toFloat64OrNull(JSONExtractString(b.estimated_cost,'cents')) / 100 AS costo_estimado,
        dateDiff('minute',
            parseDateTimeBestEffortOrNull(extract(ifNull(b.events,''), 'event_cd":20.*?created_at":"([^"]+)')),
            parseDateTimeBestEffortOrNull(extract(ifNull(b.events,''), 'event_cd":26.*?created_at":"([^"]+)'))
        ) AS minutos,
        geoDistance(
            toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[\s*([+-]?\d+\.\d+)')),
            toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[.*?,\s*([+-]?\d+\.\d+)')),
            toFloat64(JSONExtractString(b.end_geojson,'coordinates',1)),
            toFloat64(JSONExtractString(b.end_geojson,'coordinates',2))
        ) AS distancia
    FROM picapmongoprod.bookings b
    WHERE b.status_cd IN (100, 102)
      AND b.g_country IN ('CO','MX','NI','GT'){filtro_pais}
      AND b.created_at >= toDateTime('{fecha_desde} 00:00:00')
      AND b.created_at <= toDateTime('{fecha_hasta} 23:59:59')
      AND NOT empty(b.origin_geojson)
      AND NOT empty(b.end_geojson)
),
confirmados AS (
    SELECT
        e.booking_id,
        e.driver_id,
        multiIf(
            e.pais = 'Colombia',             e.costo_estimado * 0.12,
            e.pais IN ('Mexico','Nicaragua'), e.costo_estimado * 0.10,
            e.costo_estimado * 0.15
        ) AS comision_esperada
    FROM evasores e
    WHERE e.minutos > 5
      AND multiIf(
            e.pais = 'Colombia',             e.distancia <= 450,
            e.pais IN ('Mexico','Nicaragua'), e.distancia <= 280,
            e.distancia <= 450
          )
),
cobros AS (
    SELECT
        w.booking_id,
        abs(toFloat64OrNull(JSONExtractString(w.amount,'cents')) / 100)          AS cobrado,
        toFloat64OrNull(JSONExtractString(w.amount_after_transaction,'cents')) / 100 AS saldo_post
    FROM picapmongoprod.wallet_account_transactions w
    WHERE w._type = 'WalletAccountTransactionFraudCommission'
      AND w.created_at >= toDateTime('{fecha_desde} 00:00:00')
      AND w.created_at <= toDateTime('{fecha_hasta} 23:59:59')
      AND w.booking_id IN (SELECT booking_id FROM confirmados)
),
por_driver AS (
    SELECT
        c.driver_id,
        count()                                    AS n_evasiones,
        round(sum(c.comision_esperada), 0)         AS esperada,
        round(sum(ifNull(w.cobrado, 0)), 0)        AS cobrado,
        round(sum(c.comision_esperada) - sum(ifNull(w.cobrado, 0)), 0) AS deuda,
        countIf(ifNull(w.saldo_post, 0) >= 0)      AS n_pago,
        countIf(ifNull(w.saldo_post, -1) < 0)      AS n_negativo,
        multiIf(
            countIf(ifNull(w.saldo_post,-1) < 0) = 0, 'AL DÍA',
            countIf(ifNull(w.saldo_post, 0) >= 0) = 0, 'DEUDA TOTAL',
            'DEUDA PARCIAL'
        ) AS estado
    FROM confirmados c
    LEFT JOIN cobros w ON c.booking_id = w.booking_id
    GROUP BY c.driver_id
)
SELECT
    count()                                                      AS total_conductores,
    countIf(estado = 'AL DÍA')                                  AS conductores_pagaron,
    countIf(estado != 'AL DÍA')                                 AS conductores_no_pagaron,
    countIf(estado = 'DEUDA PARCIAL')                           AS deuda_parcial,
    countIf(estado = 'DEUDA TOTAL')                             AS deuda_total,
    round(sum(esperada), 0)                                      AS comision_esperada,
    round(sum(cobrado), 0)                                       AS cobrado_wallet,
    round(sum(deuda), 0)                                         AS brecha_no_cobrada,
    round(sumIf(cobrado, estado = 'AL DÍA'), 0)                 AS monto_pagado,
    round(sumIf(cobrado, estado != 'AL DÍA'), 0)                AS monto_cobrado_en_negativo,
    round(sumIf(deuda,   estado != 'AL DÍA'), 0)                AS monto_pendiente,
    round(sum(cobrado) / greatest(sum(esperada), 1) * 100, 1)   AS pct_recuperado
FROM por_driver
"""

Q_WALLET_DRIVERS = """
WITH evasores AS (
    SELECT
        b._id AS booking_id,
        b.driver_id,
        pd.name AS name_driver,
        CASE
            WHEN b.g_country = 'CO' THEN 'Colombia'
            WHEN b.g_country = 'MX' THEN 'Mexico'
            WHEN b.g_country = 'NI' THEN 'Nicaragua'
            WHEN b.g_country = 'GT' THEN 'Guatemala'
            ELSE 'Otro'
        END AS pais,
        toFloat64OrNull(JSONExtractString(b.estimated_cost,'cents')) / 100 AS costo_estimado,
        dateDiff('minute',
            parseDateTimeBestEffortOrNull(
                extract(ifNull(b.events,''), 'event_cd":20.*?created_at":"([^"]+)')),
            parseDateTimeBestEffortOrNull(
                extract(ifNull(b.events,''), 'event_cd":26.*?created_at":"([^"]+)'))
        ) AS minutos,
        geoDistance(
            toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[\s*([+-]?\d+\.\d+)')),
            toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd":26.*?coordinates":\[.*?,\s*([+-]?\d+\.\d+)')),
            toFloat64(JSONExtractString(b.end_geojson,'coordinates',1)),
            toFloat64(JSONExtractString(b.end_geojson,'coordinates',2))
        ) AS distancia
    FROM picapmongoprod.bookings b
    LEFT JOIN picapmongoprod.passengers pd ON b.driver_id = pd._id
    WHERE b.status_cd IN (100, 102)
      AND b.g_country IN ('CO','MX','NI','GT'){filtro_pais}
      AND b.created_at >= toDateTime('{fecha_desde} 00:00:00')
      AND b.created_at <= toDateTime('{fecha_hasta} 23:59:59')
      AND NOT empty(b.origin_geojson)
      AND NOT empty(b.end_geojson)
),
solo_confirmados AS (
    SELECT
        e.booking_id, e.driver_id, e.name_driver, e.pais,
        multiIf(
            e.pais = 'Colombia',             e.costo_estimado * 0.12,
            e.pais IN ('Mexico','Nicaragua'), e.costo_estimado * 0.10,
            e.costo_estimado * 0.15
        ) AS comision_evasion
    FROM evasores e
    WHERE e.minutos > 5
      AND multiIf(
            e.pais = 'Colombia',             e.distancia <= 450,
            e.pais IN ('Mexico','Nicaragua'), e.distancia <= 280,
            e.distancia <= 450
          )
),
cobros_wallet AS (
    SELECT
        w.booking_id,
        toFloat64OrNull(JSONExtractString(w.amount,'cents')) / 100 AS comision_cobrada,
        CASE
            WHEN toFloat64OrNull(JSONExtractString(w.amount_after_transaction,'cents')) >= 0
            THEN 'PAGÓ' ELSE 'SALDO NEGATIVO'
        END AS estado_cobro
    FROM picapmongoprod.wallet_account_transactions w
    WHERE w._type = 'WalletAccountTransactionFraudCommission'
      AND w.created_at >= toDateTime('{fecha_desde} 00:00:00')
      AND w.created_at <= toDateTime('{fecha_hasta} 23:59:59')
      AND w.booking_id IN (SELECT booking_id FROM solo_confirmados)
)
SELECT
    sc.driver_id,
    any(sc.name_driver)                              AS nombre,
    any(sc.pais)                                     AS pais,
    count()                                          AS evasiones,
    round(sum(sc.comision_evasion), 0)               AS comision_esperada,
    round(sum(abs(cw.comision_cobrada)), 0)           AS cobrado,
    round(sum(sc.comision_evasion) - sum(abs(cw.comision_cobrada)), 0) AS deuda,
    CASE
        WHEN countIf(cw.estado_cobro = 'SALDO NEGATIVO') = 0 THEN 'AL DÍA'
        WHEN countIf(cw.estado_cobro = 'PAGÓ') = 0           THEN 'DEUDA TOTAL'
        ELSE 'DEUDA PARCIAL'
    END AS estado
FROM solo_confirmados sc
LEFT JOIN cobros_wallet cw ON sc.booking_id = cw.booking_id
GROUP BY sc.driver_id
ORDER BY deuda DESC
LIMIT 10
"""

_cache = {
    "resumen": None, "updated_at": None,
    "loading": False, "error": None,
    "desde": None, "hasta": None,
}
CACHE_TTL = 3600

# Tasa de comisión real por país
TASAS_PAIS = {
    "Colombia":  0.12,
    "Mexico":    0.10,
    "Nicaragua": 0.10,
    "Guatemala": 0.15,   # pendiente confirmar tasa real
    "Peru":      0.15,
    "Ecuador":   0.15,
}
# Tasa por moneda (para filtro por moneda sin filtro de país)
TASAS_MONEDA = {
    "COP": 0.12,
    "MXN": 0.10,
    "NIO": 0.10,
    "PEN": 0.15,
    "USD": 0.15,
}
TASA_DEFAULT = 0.15

def tasa_para(pais=None, moneda=None):
    """Devuelve la tasa de comisión decimal. Prioridad: país > moneda > default."""
    if pais:   return TASAS_PAIS.get(pais,   TASA_DEFAULT)
    if moneda: return TASAS_MONEDA.get(moneda, TASA_DEFAULT)
    return TASA_DEFAULT

def run_query(client, sql, desde, hasta, tasa_comision=0.15):
    # tasa_comision ya no se usa en SQL (multiIf en ClickHouse), se mantiene por compatibilidad
    r = client.query(sql.format(fecha_desde=desde, fecha_hasta=hasta))
    return [dict(zip(r.column_names, row)) for row in r.result_rows]

def cargar_datos(desde, hasta, pais=None, moneda=None):
    """
    Función única para cargar datos. Acepta filtros opcionales de pais y moneda.
    Inyecta los filtros en la CTE de ClickHouse y usa la tasa correcta por país.
    """
    global _cache
    _cache["loading"] = True
    _cache["error"]   = None
    print(f"[{datetime.now():%H:%M:%S}] Cargando {desde}→{hasta} pais={pais} moneda={moneda}")
    tasa_com = tasa_para(pais, moneda)  # pais tiene prioridad sobre moneda

    # Construir filtros extra para la CTE
    # IMPORTANTE: b.g_country usa códigos ISO (CO, MX, NI), no nombres
    PAIS_A_ISO = {"Colombia": "CO", "Mexico": "MX", "Nicaragua": "NI",
                  "Guatemala": "GT", "Peru": "PE", "Ecuador": "EC"}
    extra = ""
    if pais:
        iso = PAIS_A_ISO.get(pais, pais)  # convierte Colombia→CO, etc.
        extra += f" AND b.g_country = '{iso}'"
    if moneda: extra += f" AND JSONExtractString(b.final_cost, 'currency_iso') = '{moneda}'"

    # Si hay filtros los inyectamos en BASE_CTE; si no, se usa la CTE original
    if extra:
        cte_activa = BASE_CTE.replace(
            "AND b.status_cd IN (100, 102)",
            f"AND b.status_cd IN (100, 102){extra}"
        )
        # Reconstruir las queries usando la CTE filtrada
        sfx_k = Q_KPIS[len(BASE_CTE):]
        sfx_t = Q_TENDENCIA[len(BASE_CTE):]
        sfx_c = Q_CIUDADES[len(BASE_CTE):]
        sfx_d = Q_TOP_DRIVERS[len(BASE_CTE):]
        q_kpis     = cte_activa + sfx_k
        q_tend     = cte_activa + sfx_t
        q_ciudad   = cte_activa + sfx_c
        q_drivers  = cte_activa + sfx_d
    else:
        q_kpis    = Q_KPIS
        q_tend    = Q_TENDENCIA
        q_ciudad  = Q_CIUDADES
        q_drivers = Q_TOP_DRIVERS

    try:
        client = get_client()
        k = (run_query(client, q_kpis,    desde, hasta, tasa_com) or [{}])[0]
        t =  run_query(client, q_tend,    desde, hasta, tasa_com)
        c =  run_query(client, q_ciudad,  desde, hasta, tasa_com)
        d =  run_query(client, q_drivers, desde, hasta, tasa_com)

        total        = int(k.get("total", 0))
        conf         = int(k.get("confirmadas", 0))
        prob         = int(k.get("probables", 0))
        tasa_evasion = round((conf + prob) / total * 100, 1) if total else 0

        _cache["resumen"] = {
            "kpis": {
                "total": total, "confirmadas": conf, "probables": prob,
                "ok": int(k.get("ok", 0)),
                "tasa_evasion": tasa_evasion,
                "comision_evadida_cop":     int(k.get("comision_evadida",    0) or 0),
                "penalizacion_evadida_cop": int(k.get("penalizacion_evadida", 0) or 0),
                "sin_gps": int(k.get("sin_gps", 0)),
                "tasa_comision_pct": round(tasa_com * 100),  # 12, 10 o 15
                "pais_filtro": pais or "",
                "moneda_filtro": moneda or "",
            },
            "operativo": {
                "prom_minutos_evasion":   float(k.get("prom_minutos",   0) or 0),
                "prom_distancia_evasion": float(k.get("prom_distancia", 0) or 0),
            },
            "funnel": {
                "total": total,
                "flag_tiempo":    int(k.get("flag_tiempo",    0)),
                "flag_distancia": int(k.get("flag_distancia", 0)),
                "confirmadas":    conf,
            },
            "tendencia":   [{"fecha": str(r["fecha"])[:10], "conf": int(r.get("conf",0)), "prob": int(r.get("prob",0)), "ok": int(r.get("ok",0))} for r in t],
            "ciudades":    [{"ciudad": r["ciudad"], "count": int(r.get("evasiones",0))} for r in c],
            "top_drivers": [{
                "id":           r["id_driver"],
                "nombre":       r.get("nombre") or "Sin nombre",
                "conf":         int(r.get("conf",0)),
                "prob":         int(r.get("prob",0)),
                "total":        int(r.get("total",0)),
                "conf_primera": int(r.get("conf_primera",0)),
                "conf_segunda": int(r.get("conf_segunda",0)),
            } for r in d],
        }
        _cache["updated_at"] = datetime.now()
        _cache["desde"]      = desde
        _cache["hasta"]      = hasta
        print(f"  OK — total={total}, conf={conf}, prob={prob}")
    except Exception as e:
        _cache["error"] = str(e)
        print(f"  ERROR: {e}")
    finally:
        _cache["loading"] = False

def _necesita_recarga(desde, hasta):
    return desde and hasta and (desde != _cache["desde"] or hasta != _cache["hasta"])

def _recargar_en_hilo(desde, hasta, pais=None, moneda=None):
    if not _cache["loading"]:
        threading.Thread(target=cargar_datos, args=(desde, hasta, pais, moneda), daemon=True).start()

# _cargar_con_filtros eliminado — ahora cargar_datos acepta pais y moneda directamente

def auto_refresh():
    time.sleep(2)
    while True:
        hoy   = datetime.now()
        hasta = hoy.strftime("%Y-%m-%d")
        desde = (hoy - timedelta(days=14)).strftime("%Y-%m-%d")
        cargar_datos(desde, hasta, pais=None)
        time.sleep(CACHE_TTL)

@app.route("/api/status")
def status():
    return jsonify({
        "ok": True, "loading": _cache["loading"], "error": _cache["error"],
        "updated_at": _cache["updated_at"].isoformat() if _cache["updated_at"] else None,
        "rango": {"desde": _cache["desde"], "hasta": _cache["hasta"]},
    })

@app.route("/api/buscar")
def buscar():
    """
    Búsqueda por booking_id o driver_id dentro del período cacheado.
    ?tipo=booking&q=<id>  → devuelve el registro exacto
    ?tipo=driver&q=<id>   → devuelve todos los servicios del conductor
    """
    tipo  = request.args.get("tipo", "booking")
    q     = (request.args.get("q") or "").strip()
    desde = request.args.get("desde")
    hasta = request.args.get("hasta")

    if _necesita_recarga(desde, hasta):
        cargar_datos(desde, hasta)

    if not q or _cache.get("resumen") is None:
        return jsonify({"resultado": None}), 200

    # Re-ejecutar query específica en ClickHouse para no cargar todo el detalle en memoria
    try:
        client = get_client()
        if tipo == "booking":
            sql = BASE_CTE + f"""
                SELECT booking_id, id_driver, name_driver, ciudad,
                       creacion_servicio, minutos_entre_eventos,
                       distancia_cancel_destino, costo_estimado, nivel
                FROM clasificado
                WHERE booking_id = '{q}'
                LIMIT 1
            """
            r = client.query(sql.format(
                fecha_desde=desde or _cache.get("desde","2026-01-01"),
                fecha_hasta=hasta or _cache.get("hasta","2026-12-31")
            ))
            rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
            if not rows:
                return jsonify({"resultado": None})
            row = rows[0]
            for k in ["creacion_servicio"]:
                if k in row: row[k] = str(row[k])
            return jsonify(limpiar_nan({"resultado": row}))

        else:  # driver
            sql = BASE_CTE + f"""
                SELECT booking_id, id_driver, name_driver, ciudad,
                       creacion_servicio, minutos_entre_eventos,
                       distancia_cancel_destino, costo_estimado, nivel
                FROM clasificado
                WHERE id_driver = '{q}'
                ORDER BY creacion_servicio DESC
                LIMIT 20
            """
            r = client.query(sql.format(
                fecha_desde=desde or _cache.get("desde","2026-01-01"),
                fecha_hasta=hasta or _cache.get("hasta","2026-12-31")
            ))
            rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
            for row in rows:
                for k in ["creacion_servicio"]:
                    if k in row: row[k] = str(row[k])
            return jsonify(limpiar_nan({"resultado": rows if rows else None}))

    except Exception as e:
        return jsonify({"error": str(e), "resultado": None}), 200


@app.route("/api/resumen")
def resumen():
    desde  = request.args.get("desde")
    hasta  = request.args.get("hasta")
    pais   = request.args.get("pais")
    moneda = request.args.get("moneda")

    # Recargar si cambió el rango O los filtros
    filtros_cambiaron = (
        pais   != (_cache.get("resumen") or {}).get("kpis", {}).get("pais_filtro",   "") or
        moneda != (_cache.get("resumen") or {}).get("kpis", {}).get("moneda_filtro", "")
    )
    if (_necesita_recarga(desde, hasta) or filtros_cambiaron) and not _cache.get("loading"):
        # Disparar en background SOLO si no hay ya una carga en curso
        threading.Thread(
            target=cargar_datos,
            args=(desde or _cache["desde"], hasta or _cache["hasta"], pais or None, moneda or None),
            daemon=True
        ).start()
    if _cache["resumen"] is None:
        return jsonify({
            "loading": True,
            "kpis": {"total":0,"confirmadas":0,"probables":0,"ok":0,"tasa_evasion":0,"comision_evadida_cop":0,"sin_gps":0},
            "operativo": {"prom_minutos_evasion":0,"prom_distancia_evasion":0},
            "funnel": {"total":0,"flag_tiempo":0,"flag_distancia":0,"confirmadas":0},
            "ciudades": [], "top_drivers": [], "tendencia": [],
        }), 200
    return jsonify(limpiar_nan({
        "updated_at": _cache["updated_at"].isoformat() if _cache["updated_at"] else None,
        "loading": _cache["loading"],
        **_cache["resumen"],
    }))


# ══════════════════════════════════════════════════════════════════════
# PANEL RECUPERACIÓN — Top 10 evasores cruzado con pagos en wallet
# Tercer panel del módulo de evasión
# ══════════════════════════════════════════════════════════════════════

# Q_RECUPERACION_TOP inlineada en el endpoint

Q_WALLET_BY_DRIVER = """
WITH confirmados AS (
    SELECT
        b._id      AS booking_id,
        b.driver_id,
        b.g_country,
        toFloat64OrNull(JSONExtractString(b.estimated_cost,'cents')) / 100 AS costo_est
    FROM picapmongoprod.bookings b
    WHERE b.status_cd IN (100, 102)
      AND b.driver_id IN ({ids})
      AND b.created_at >= toDateTime('{desde} 00:00:00')
      AND b.created_at <= toDateTime('{hasta} 23:59:59')
),
cobros AS (
    SELECT
        w.booking_id,
        abs(toFloat64OrNull(JSONExtractString(w.amount,'cents')) / 100) AS cobrado
    FROM picapmongoprod.wallet_account_transactions w
    WHERE w._type = 'WalletAccountTransactionFraudCommission'
      AND w.booking_id IN (SELECT booking_id FROM confirmados)
)
SELECT
    c.driver_id,
    round(sum(multiIf(c.g_country='CO', c.costo_est*0.12,
                      c.g_country IN ('MX','NI'), c.costo_est*0.10,
                      c.costo_est*0.15)), 0)          AS penalidad_conf,
    round(sum(ifNull(w.cobrado, 0)), 0)               AS pagado,
    round(sum(multiIf(c.g_country='CO', c.costo_est*0.12,
                      c.g_country IN ('MX','NI'), c.costo_est*0.10,
                      c.costo_est*0.15))
          - sum(ifNull(w.cobrado, 0)), 0)             AS deuda
FROM confirmados c
LEFT JOIN cobros w ON c.booking_id = w.booking_id
GROUP BY c.driver_id
"""

Q_RESUMEN_PERIODO = """
SELECT
    toDate(w.created_at)                                                       AS dia,
    round(sum(abs(toFloat64OrNull(JSONExtractString(w.amount,'cents'))/100)),0) AS cobrado_dia
FROM picapmongoprod.wallet_account_transactions w
WHERE w._type = 'WalletAccountTransactionFraudCommission'
  AND w.created_at >= toDateTime('{desde} 00:00:00')
  AND w.created_at <= toDateTime('{hasta} 23:59:59')
GROUP BY dia
ORDER BY dia
"""

@app.route("/api/recuperacion")
def recuperacion():
    desde  = request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d")
    hasta  = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais   = request.args.get("pais") or ""
    PAIS_A_ISO = {"Colombia":"CO","Mexico":"MX","Nicaragua":"NI","Guatemala":"GT","Peru":"PE","Ecuador":"EC"}
    iso    = PAIS_A_ISO.get(pais, pais)
    filtro_pais = f" AND b.g_country = '{iso}'" if iso else ""

    try:
        ch = get_client()

        # Construir CTE con filtro de país (igual que cargar_datos)
        cte_activa = BASE_CTE
        if iso:
            cte_activa = BASE_CTE.replace(
                "AND b.status_cd IN (100, 102)",
                f"AND b.status_cd IN (100, 102) AND b.g_country = '{iso}'"
            )
        q_top = (cte_activa + """
SELECT
    id_driver,
    any(name_driver)                          AS nombre,
    countIf(nivel = 3)                        AS conf,
    round(sum(comision_servicio), 0)          AS penalidad_total
FROM clasificado
WHERE nivel = 3
GROUP BY id_driver
ORDER BY conf DESC, penalidad_total DESC
LIMIT 10
""").format(fecha_desde=desde, fecha_hasta=hasta)
        top_rows = ch.query(q_top).result_rows

        if not top_rows:
            return jsonify({"ok": True, "top": [], "resumen": {}, "tendencia": []})

        # Cruzar con wallet por driver
        ids_list = ",".join([f"'{r[0]}'" for r in top_rows])
        wallet_rows = ch.query(Q_WALLET_BY_DRIVER.format(
            ids=ids_list, desde=desde, hasta=hasta
        )).result_rows
        wallet_map = {str(r[0]): {"penalidad": float(r[1] or 0),
                                   "pagado":    float(r[2] or 0),
                                   "deuda":     float(r[3] or 0)} for r in wallet_rows}

        top_final = []
        total_penalidad = 0
        total_pagado    = 0
        for r in top_rows:
            did, nombre, conf, pen = str(r[0]), r[1], int(r[2]), float(r[3] or 0)
            w = wallet_map.get(did, {"penalidad": pen, "pagado": 0, "deuda": pen})
            pagado   = w["pagado"]
            deuda    = w["deuda"]
            penalidad = w["penalidad"] if w["penalidad"] > 0 else pen
            pct      = round(pagado / penalidad * 100, 1) if penalidad > 0 else 0
            total_penalidad += penalidad
            total_pagado    += pagado
            top_final.append({
                "id":        did,
                "nombre":    nombre or "Sin nombre",
                "conf":      conf,
                "penalidad": round(penalidad, 0),
                "pagado":    round(pagado, 0),
                "deuda":     round(deuda, 0),
                "pct":       pct,
                "estado":    "AL DÍA" if deuda <= 0 else ("PARCIAL" if pagado > 0 else "SIN PAGO"),
            })

        # Tendencia diaria de cobros en el período
        tend_rows = ch.query(Q_RESUMEN_PERIODO.format(desde=desde, hasta=hasta)).result_rows

        # Resumen global
        pct_global = round(total_pagado / total_penalidad * 100, 1) if total_penalidad > 0 else 0
        resumen = {
            "total_penalidad": round(total_penalidad, 0),
            "total_pagado":    round(total_pagado, 0),
            "total_deuda":     round(total_penalidad - total_pagado, 0),
            "pct_recuperado":  pct_global,
        }

        return jsonify({
            "ok":       True,
            "top":      top_final,
            "resumen":  resumen,
            "tendencia": [{"dia": str(r[0]), "cobrado": float(r[1] or 0)} for r in tend_rows],
        })

    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/api/wallet")
def wallet():
    """
    Estado de cobros en wallet para evasiones confirmadas.
    Independiente del caché principal — consulta directa a ClickHouse.
    """
    desde  = request.args.get("desde") or _cache.get("desde") or (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    hasta  = request.args.get("hasta") or _cache.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais   = request.args.get("pais") or ""
    PAIS_A_ISO = {"Colombia":"CO","Mexico":"MX","Nicaragua":"NI","Guatemala":"GT","Peru":"PE","Ecuador":"EC"}
    filtro_pais = f" AND b.g_country = '{PAIS_A_ISO.get(pais, pais)}'" if pais else ""
    try:
        client = get_client()

        # KPIs globales de wallet
        r_kpi = client.query(Q_WALLET.format(fecha_desde=desde, fecha_hasta=hasta, filtro_pais=filtro_pais))
        k = dict(zip(r_kpi.column_names, r_kpi.result_rows[0])) if r_kpi.result_rows else {}

        # Top conductores con mayor deuda
        r_drv = client.query(Q_WALLET_DRIVERS.format(fecha_desde=desde, fecha_hasta=hasta, filtro_pais=filtro_pais))
        drivers = [dict(zip(r_drv.column_names, row)) for row in r_drv.result_rows]

        return jsonify(limpiar_nan({
            "desde": desde, "hasta": hasta,
            "kpis": {
                "total_conductores":  int(k.get("total_conductores", 0)),
                "al_dia":             int(k.get("al_dia", 0)),
                "deuda_parcial":      int(k.get("deuda_parcial", 0)),
                "deuda_total":        int(k.get("deuda_total", 0)),
                "comision_esperada":        float(k.get("comision_esperada", 0) or 0),
                "cobrado_wallet":           float(k.get("cobrado_wallet", 0) or 0),
                "brecha_no_cobrada":        float(k.get("brecha_no_cobrada", 0) or 0),
                "pct_recuperado":           float(k.get("pct_recuperado", 0) or 0),
                "conductores_pagaron":      int(k.get("conductores_pagaron", 0)),
                "conductores_no_pagaron":   int(k.get("conductores_no_pagaron", 0)),
                "monto_pagado":             float(k.get("monto_pagado", 0) or 0),
                "monto_cobrado_en_negativo":float(k.get("monto_cobrado_en_negativo", 0) or 0),
                "monto_pendiente":          float(k.get("monto_pendiente", 0) or 0),
            },
            "top_deuda": [
                {
                    "id":                r.get("driver_id",""),
                    "nombre":            r.get("nombre","Sin nombre"),
                    "pais":              r.get("pais",""),
                    "evasiones":         int(r.get("evasiones", 0)),
                    "comision_esperada": float(r.get("comision_esperada", 0) or 0),
                    "cobrado":           float(r.get("cobrado", 0) or 0),
                    "deuda":             float(r.get("deuda", 0) or 0),
                    "estado":            r.get("estado",""),
                }
                for r in drivers
            ]
        }))
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ══════════════════════════════════════════════════════════════
# Query Vista de Bloqueos — basada en passenger/driver_suspensions
# ══════════════════════════════════════════════════════════════
Q_BLOQUEOS = """
WITH ultimos_ps AS (
    SELECT
        starts_at   AS starts_block_p,
        ends_at     AS ends_block_p,
        passenger_id,
        created_at,
        updated_at  AS reactivado_en_p,
        ROW_NUMBER() OVER (
            PARTITION BY passenger_id
            ORDER BY created_at DESC NULLS LAST
        ) AS rn
    FROM picapmongoprod.passenger_suspensions
    WHERE created_at IS NOT NULL
),
ultimos_ds AS (
    SELECT
        starts_at   AS starts_block_d,
        ends_at     AS ends_block_d,
        driver_id,
        created_at,
        updated_at  AS reactivado_en_d,
        ROW_NUMBER() OVER (
            PARTITION BY driver_id
            ORDER BY created_at DESC NULLS LAST
        ) AS rn
    FROM picapmongoprod.driver_suspensions
    WHERE created_at IS NOT NULL
),
ps_final AS (SELECT * FROM ultimos_ps WHERE rn = 1),
ds_final AS (SELECT * FROM ultimos_ds WHERE rn = 1)
SELECT
    p._id                                        AS id_usuario,
    p.name                                       AS nombre,
    p.g_country                                  AS pais_codigo,
    p.g_adm_area_lv_1                            AS ciudad,
    ps.starts_block_p                            AS starts_block_user,
    ps.ends_block_p                              AS ends_block_user,
    ifNull(toString(p.suspended), '')            AS suspendido,
    ifNull(p.passenger_suspension_comment, '')   AS comentario_user,
    ifNull(p.passenger_expulsion_comment, '')    AS comentario_expulsion_user,
    dateDiff('day', toDate(ps.created_at), today()) AS dias_suspension_user,
    ds.starts_block_d                            AS starts_block_driver,
    ds.ends_block_d                              AS ends_block_driver,
    ifNull(toString(p.is_driver_suspended), '')  AS driver_suspendido,
    ifNull(p.driver_suspension_comment, '')      AS comentario_driver,
    dateDiff('day', toDate(ds.created_at), today()) AS dias_suspension_driver,
    ifNull(toString(p.expelled), '')             AS expulsado,
    -- Tipo de bloqueo: EXPULSADO tiene prioridad sobre SUSPENDIDO
    CASE
        WHEN lower(ifNull(toString(p.expelled),'')) = 'true' THEN 'EXPULSADO'
        WHEN lower(ifNull(toString(p.suspended),'')) = 'true'
          OR lower(ifNull(toString(p.is_driver_suspended),'')) = 'true' THEN 'SUSPENDIDO'
        ELSE 'ACTIVO'
    END AS tipo_bloqueo,
    CASE
        WHEN p.driver_enrollment_status_cd = 3 THEN 'PILOTO'
        ELSE 'USUARIO'
    END AS tipo_usuario,
    formatDateTime(
        toTimeZone(
            coalesce(
                if(ps.created_at IS NOT NULL AND ds.created_at IS NOT NULL,
                   greatest(ps.created_at, ds.created_at), NULL),
                ps.created_at, ds.created_at
            ), 'America/Bogota'
        ), '%Y-%m-%d %H:%M'
    ) AS fecha_ultima_suspension,
    dateDiff('day', toDate(coalesce(
        if(ps.created_at IS NOT NULL AND ds.created_at IS NOT NULL,
           greatest(ps.created_at, ds.created_at), NULL),
        ps.created_at, ds.created_at
    )), today()) AS dias_bloqueado_total,
    -- ── DÍAS REALES BLOQUEADO ────────────────────────────────────
    -- Toma la suspensión más reciente entre passenger_suspensions y
    -- driver_suspensions (sin importar el tipo de cuenta). updated_at del
    -- registro se actualiza cuando el admin reactiva, así que esa diff
    -- con starts_at = duración real del bloqueo.
    --
    -- Nota: la tabla donde está el registro NO siempre coincide con el
    -- tipo de cuenta — un USUARIO puede tener record en driver_suspensions.
    -- Por eso elegimos por created_at más reciente, no por tipo_usuario.
    multiIf(
        -- Caso 1: hay registro en driver_suspensions y es el más reciente
        ds.starts_block_d IS NOT NULL
            AND (ps.created_at IS NULL OR ds.created_at >= ps.created_at),
        greatest(0, dateDiff('day',
            toDate(ds.starts_block_d),
            -- Fecha de fin del bloqueo: updated_at SOLO si:
            --   • es posterior a starts_at (no antes del inicio)
            --   • no está en el futuro
            --   • la cuenta está reactivada (suspended=false, expelled=false)
            if(
                ds.reactivado_en_d IS NOT NULL
                AND toDate(ds.reactivado_en_d) > toDate(ds.starts_block_d)
                AND toDate(ds.reactivado_en_d) <= today()
                AND lower(ifNull(toString(p.expelled),'')) != 'true'
                AND lower(ifNull(toString(p.suspended),'')) IN ('false','0','')
                AND lower(ifNull(toString(p.is_driver_suspended),'')) IN ('false','0',''),
                toDate(ds.reactivado_en_d),
                today()
            )
        )),
        -- Caso 2: hay registro en passenger_suspensions
        ps.starts_block_p IS NOT NULL,
        greatest(0, dateDiff('day',
            toDate(ps.starts_block_p),
            if(
                ps.reactivado_en_p IS NOT NULL
                AND toDate(ps.reactivado_en_p) > toDate(ps.starts_block_p)
                AND toDate(ps.reactivado_en_p) <= today()
                AND lower(ifNull(toString(p.expelled),'')) != 'true'
                AND lower(ifNull(toString(p.suspended),'')) IN ('false','0','')
                AND lower(ifNull(toString(p.is_driver_suspended),'')) IN ('false','0',''),
                toDate(ps.reactivado_en_p),
                today()
            )
        )),
        -- Fallback (sin starts_at confiable): días desde la suspensión más reciente hasta hoy
        dateDiff('day', toDate(coalesce(
            if(ps.created_at IS NOT NULL AND ds.created_at IS NOT NULL,
               greatest(ps.created_at, ds.created_at), NULL),
            ps.created_at, ds.created_at
        )), today())
    ) AS dias_bloqueo_real,
    -- estado_suspension: expulsados son PERMANENTE, suspendidos evalúan 30 días
    CASE
        WHEN lower(ifNull(toString(p.expelled),'')) = 'true' THEN 'Permanente (expulsión)'
        WHEN dateDiff('day', toDate(coalesce(
            if(ps.created_at IS NOT NULL AND ds.created_at IS NOT NULL,
               greatest(ps.created_at, ds.created_at), NULL),
            ps.created_at, ds.created_at
        )), today()) > 30 THEN 'Más de 30 días'
        ELSE 'Menos de 30 días'
    END AS estado_suspension,
    CASE
        WHEN lower(ifNull(toString(p.expelled),'')) = 'true'
          OR lower(ifNull(toString(p.suspended),'')) = 'true' THEN 'bloqueado'
        WHEN lower(ifNull(toString(p.expelled),'')) = 'false'
         AND lower(ifNull(toString(p.suspended),'')) = 'false' THEN 'activo'
        ELSE ''
    END AS esta_activo
FROM picapmongoprod.passengers AS p
LEFT JOIN ps_final AS ps ON p._id = ps.passenger_id
LEFT JOIN ds_final AS ds ON p._id = ds.driver_id
WHERE coalesce(
    if(ps.created_at IS NOT NULL AND ds.created_at IS NOT NULL,
       greatest(ps.created_at, ds.created_at), NULL),
    ps.created_at, ds.created_at
) BETWEEN toDateTime('{fecha_desde} 00:00:00')
      AND toDateTime('{fecha_hasta} 23:59:59')
ORDER BY fecha_ultima_suspension DESC
LIMIT 10000
"""

# ── Mapeo oficial de motivos ─────────────────────────────────
MOTIVOS_OFICIALES = [
    "Cobrar dos o más veces el mismo servicio",
    "Realizar cobros adicionales no acordados",
    "No cumplir con la totalidad del servicio",
    "Modificar ubicación de destino (Fake GPS)",
    "No devolver vehículo al arrendatario",
    "Cobrar en TC/Pica$h y no realizar el servicio",
    "No prestar servicio por método de pago",
    "Cobro excesivo en tarifa",
    "Servicio con vehículo diferente al registrado",
    "Preguntar destino y negarse a prestar servicio",
    "Vocabulario inadecuado con agente SAC",
    "No devolver dinero sobrante al usuario",
    "Solicitar al usuario cancelar para no prestar",
    "No pagar valor del servicio al prestador",
    "Cancelar servicios para evadir comisión",
    "Insultar al usuario prestador",
    "Solicitar cancelación para generar bonificación",
    "Cancelar servicios de forma reiterada",
    "Comercializar saldos Pica$h",
    "Insultar o amenazar por chat",
    "Registrar documento que no corresponde",
    "Crear cuenta nueva con cuenta cancelada",
    "Fraude dentro de la APP",
    "Estafa dentro de la APP",
    "Alterar documentos o datos en la APP",
    "Prestar o alquilar cuenta personal",
    "Hurtar pertenencias del usuario",
    "Amenazar contra la vida del usuario",
    "Portar armas al prestar el servicio",
    "Entregar paquete en mal estado",
    "No entregar paquete el mismo día",
    "Entregar paquete incompleto",
    "Entregar paquete en dirección equivocada",
    "No entregar dinero recaudado a Pibox",
    "No entregar paquete al destinatario",
    "Apropiarse del paquete",
    "Solicitar envío de sustancias ilícitas",
    "Presentar comparendo D12 fraudulento",
    "No finalizar servicio en destino",
    "Negarse a presentar SOAT en accidente",
    "Malas prácticas en prestación del servicio",
    "No cumplir normas de seguridad (casco/chaleco)",
    "Conducción peligrosa o sin respetar tránsito",
    "Hostigar/molestar con comportamiento vulgar",
    "No tomar evidencias del servicio",
    "Incumplir acción de mejora propuesta",
    "Incumplir horario establecido",
    "No asistir al servicio programado",
    "Insultar al cliente",
    "No entregar paquete",
    "No entregar recaudo del paquete",
    "Malas prácticas en servicios corporativos",
    "Tener antecedentes penales",
]

KEYWORDS_MOTIVOS = [
    (["cobrar dos", "doble cobro", "dos veces", "diferentes medios de pago"],      "Cobrar dos o más veces el mismo servicio"),
    (["cobros adicionales", "cobro adicional"],                                    "Realizar cobros adicionales no acordados"),
    (["no cumplir con la totalidad"],                                              "No cumplir con la totalidad del servicio"),
    (["fake gps", "modificar la ubicacion", "finaliza el servicio", "modificar ubicacion"], "Modificar ubicación de destino (Fake GPS)"),
    (["no devolver el vehiculo", "arrendataria"],                                  "No devolver vehículo al arrendatario"),
    (["cobrar el valor del servicio en tarjeta", "cobrar en tc", "picash y no realizarlo"], "Cobrar en TC/Pica$h y no realizar el servicio"),
    (["no prestar el servicio por no aceptar el metodo de pago", "no aceptar el metodo de pago"], "No prestar servicio por método de pago"),
    (["cobro excesivo", "tarifa excesiva", "cobro excesivo en la tarifa"],         "Cobro excesivo en tarifa"),
    (["vehiculo diferente", "vehiculo registrado"],                                "Servicio con vehículo diferente al registrado"),
    (["preguntar el lugar de destino y negarse"],                                  "Preguntar destino y negarse a prestar servicio"),
    (["vocabulario adecuado", "agente de servicio al cliente", "no manejar un vocabulario"], "Vocabulario inadecuado con agente SAC"),
    (["no devolver", "dinero que sobre", "efectivo"],                              "No devolver dinero sobrante al usuario"),
    (["solicitar al usuario", "cancelar el servicio para no prestarlo"],           "Solicitar al usuario cancelar para no prestar"),
    (["no pagar el valor del servicio al usuario prestador"],                      "No pagar valor del servicio al prestador"),
    (["cancelar los servicios para evitar pagar la comision", "evadir comision", "cancelar servicios para"],  "Cancelar servicios para evadir comisión"),
    (["insultar al usuario prestador"],                                            "Insultar al usuario prestador"),
    (["bonificacion por cancelacion", "solicitar al usuario consumidor cancelar el servicio y si realizarlo"], "Solicitar cancelación para generar bonificación"),
    (["cancelar servicios de forma reiterada"],                                    "Cancelar servicios de forma reiterada"),
    (["comercializar o vender saldos", "picash entre"],                            "Comercializar saldos Pica$h"),
    (["insultar o amenazar", "chat del servicio"],                                 "Insultar o amenazar por chat"),
    (["documento que no corresponda"],                                             "Registrar documento que no corresponde"),
    (["cuenta nueva cuando se haya cancelado", "crear una cuenta nueva"],          "Crear cuenta nueva con cuenta cancelada"),
    (["fraude dentro de la app", "realizar cualquier fraude"],                     "Fraude dentro de la APP"),
    (["estafa dentro de la app", "realizar cualquier estafa"],                     "Estafa dentro de la APP"),
    (["alterar cualquier documento", "datos registrados en la app"],               "Alterar documentos o datos en la APP"),
    (["prestar o alquilar su cuenta", "cuenta de otro usuario"],                   "Prestar o alquilar cuenta personal"),
    (["hurtar las pertenencias", "hurtar"],                                        "Hurtar pertenencias del usuario"),
    (["amenazar o atentar contra", "atentar contra de la vida"],                   "Amenazar contra la vida del usuario"),
    (["arma de fuego", "arma blanca", "elemento que pueda atentar"],               "Portar armas al prestar el servicio"),
    (["mal estado al destinatario"],                                               "Entregar paquete en mal estado"),
    (["no entregar el paquete al destinatario el mismo dia"],                      "No entregar paquete el mismo día"),
    (["paquete incompleto"],                                                       "Entregar paquete incompleto"),
    (["direccion equivocada"],                                                     "Entregar paquete en dirección equivocada"),
    (["no entregar el dinero recaudado a pibox"],                                  "No entregar dinero recaudado a Pibox"),
    (["no entregar el paquete al destinatario"],                                   "No entregar paquete al destinatario"),
    (["apropiarse del paquete"],                                                   "Apropiarse del paquete"),
    (["sustancia ilicita", "peligrosa"],                                           "Solicitar envío de sustancias ilícitas"),
    (["comparendo d12"],                                                           "Presentar comparendo D12 fraudulento"),
    (["no finalizar el servicio en el lugar de destino"],                          "No finalizar servicio en destino"),
    (["negarse a prestar el soat", "soat"],                                        "Negarse a presentar SOAT en accidente"),
    (["malas practicas en la prestacion del servicio"],                            "Malas prácticas en prestación del servicio"),
    (["normas basicas de seguridad", "casco", "chaleco reflectivo"],               "No cumplir normas de seguridad (casco/chaleco)"),
    (["normas de transito", "maniobras peligrosas"],                               "Conducción peligrosa o sin respetar tránsito"),
    (["hostigar", "comportamiento vulgar", "obsceno"],                             "Hostigar/molestar con comportamiento vulgar"),
    (["evidencias de haber realizado el servicio", "no tomar correctamente las evidencias"], "No tomar evidencias del servicio"),
    (["accion de mejora"],                                                         "Incumplir acción de mejora propuesta"),
    (["horario establecido"],                                                      "Incumplir horario establecido"),
    (["no asistir al servicio programado"],                                        "No asistir al servicio programado"),
    (["insultar al cliente"],                                                      "Insultar al cliente"),
    (["no entregar el paquete"],                                                   "No entregar paquete"),
    (["no entregar del recaudo"],                                                  "No entregar recaudo del paquete"),
    (["servicios corporativos"],                                                   "Malas prácticas en servicios corporativos"),
    (["antecedentes penales"],                                                     "Tener antecedentes penales"),
]

PAISES_MAP = {
    'CO': 'Colombia', 'MX': 'México', 'NI': 'Nicaragua',
    'GT': 'Guatemala', 'PE': 'Perú', 'EC': 'Ecuador',
}

def mapear_motivo(texto):
    """Mapea texto libre de motivo al motivo oficial más cercano."""
    if not texto or not texto.strip():
        return None
    t = texto.lower().strip()
    # Eliminar tildes para comparación
    import unicodedata
    t_norm = ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')
    for keywords, motivo_oficial in KEYWORDS_MOTIVOS:
        for kw in keywords:
            kw_norm = ''.join(c for c in unicodedata.normalize('NFD', kw) if unicodedata.category(c) != 'Mn')
            if kw_norm in t_norm:
                return motivo_oficial
    # Si no coincide, devolver primeros 80 chars del texto original
    return texto[:80] + ('…' if len(texto) > 80 else '')

def top10_stats(rows, campo_tipo='tipo_usuario'):
    """
    Calcula Top 10 de motivos + país + ciudad.
    Segmenta por PILOTO/USUARIO/TODOS e incluye breakdown por tipo_bloqueo.
    Fuente de motivos: comentario_user + comentario_driver + comentario_expulsion_user.
    """
    from collections import Counter
    result = {}
    for tipo in ['PILOTO', 'USUARIO', 'TODOS']:
        subset = rows if tipo == 'TODOS' else [r for r in rows if r.get(campo_tipo) == tipo]
        total  = len(subset)

        motivos  = Counter()
        paises   = Counter()
        ciudades = Counter()
        por_tipo_blq = Counter()

        for r in subset:
            # Motivo mapeado — ya calculado upstream con prioridad correcta
            m = (r.get('motivo_mapeado') or '').strip()
            # Si motivo_mapeado está vacío, intentar extraer de cualquier campo disponible
            if not m:
                candidatos = [
                    r.get('comentario_driver',''),
                    r.get('comentario_user',''),
                    r.get('comentario_expulsion_user',''),
                ]
                for c in candidatos:
                    if c and c.strip():
                        m = mapear_motivo(c.strip()) or ''
                        if m:
                            break
            if m:
                motivos[m] += 1

            p = r.get('pais_nombre','')
            if p: paises[p] += 1

            ciudad = r.get('ciudad','')
            if ciudad: ciudades[ciudad] += 1

            tb = r.get('tipo_bloqueo','')
            if tb: por_tipo_blq[tb] += 1

        result[tipo] = {
            'total': total,
            'top_motivos':      [{'motivo':k,'count':v,'pct':round(v/total*100,1) if total else 0} for k,v in motivos.most_common(10)],
            'top_paises':       [{'pais':k,'count':v,'pct':round(v/total*100,1) if total else 0} for k,v in paises.most_common(5)],
            'top_ciudades':     [{'ciudad':k,'count':v,'pct':round(v/total*100,1) if total else 0} for k,v in ciudades.most_common(10)],
            'por_tipo_bloqueo': [{'tipo':k,'count':v,'pct':round(v/total*100,1) if total else 0} for k,v in por_tipo_blq.most_common()],
        }
    return result

@app.route("/api/bloqueos")
def bloqueos():
    """Vista de Bloqueos — alertas tiempo de bloqueo, activos y reactivaciones."""
    desde = request.args.get("desde") or _cache.get("desde") or "2026-03-01"
    hasta = request.args.get("hasta") or _cache.get("hasta") or "2026-03-31"
    try:
        client = get_client()
        r = client.query(Q_BLOQUEOS.format(fecha_desde=desde, fecha_hasta=hasta))
        rows = [dict(zip(r.column_names, row)) for row in r.result_rows]

        # Convertir datetime a string, limpiar None
        for row in rows:
            for k, v in list(row.items()):
                if hasattr(v, 'isoformat'):
                    row[k] = str(v)[:16] if v else None
                elif v is None:
                    row[k] = None
            # Mapear país
            row['pais_nombre'] = PAISES_MAP.get(row.get('pais_codigo',''), row.get('pais_codigo',''))
            # Mapear motivo oficial — prioridad según tipo de bloqueo
            # Para PILOTO: driver > user > expulsion
            # Para USUARIO: user > expulsion > driver
            # Incluye siempre el comentario de expulsion que antes se ignoraba
            tipo_usr = row.get('tipo_usuario', '')
            if tipo_usr == 'PILOTO':
                motivo_raw = (
                    (row.get('comentario_driver') or '').strip() or
                    (row.get('comentario_user') or '').strip() or
                    (row.get('comentario_expulsion_user') or '').strip()
                )
            else:
                motivo_raw = (
                    (row.get('comentario_user') or '').strip() or
                    (row.get('comentario_expulsion_user') or '').strip() or
                    (row.get('comentario_driver') or '').strip()
                )
            row['motivo_mapeado'] = mapear_motivo(motivo_raw)

        # Clasificar
        alertas, bloqueados, reactivados = [], [], []
        for row in rows:
            dias       = row.get('dias_bloqueado_total') or 0
            activo     = row.get('esta_activo', '')
            tipo_blq   = row.get('tipo_bloqueo', '')
            es_expulsado = tipo_blq == 'EXPULSADO'

            # Regla 30 días: solo aplica a SUSPENDIDOS, no a EXPULSADOS (permanentes)
            if es_expulsado:
                row['veredicto']    = 'EXPULSIÓN PERMANENTE'
                row['alerta_30dias'] = False  # no genera alerta temporal
            else:
                row['veredicto']    = 'ALERTA DE TIEMPO' if dias > 30 else 'TODO OK'
                row['alerta_30dias'] = dias > 30

            alertas.append(row)
            if activo == 'bloqueado':
                bloqueados.append(row)
            if activo == 'activo':
                reactivados.append(row)

        n_alerta     = sum(1 for x in alertas if x['veredicto'] == 'ALERTA DE TIEMPO')
        n_ok         = sum(1 for x in alertas if x['veredicto'] == 'TODO OK')
        n_expulsados = sum(1 for x in bloqueados if x.get('tipo_bloqueo') == 'EXPULSADO')
        n_suspendidos= sum(1 for x in bloqueados if x.get('tipo_bloqueo') == 'SUSPENDIDO')
        n_susp_30    = sum(1 for x in bloqueados if x.get('tipo_bloqueo') == 'SUSPENDIDO' and (x.get('dias_bloqueado_total') or 0) > 30)
        total        = len(alertas)

        # Ordenar por días DESC. Reactivados usan dias_bloqueo_real (la duración
        # efectiva del bloqueo, que es la métrica que el frontend muestra).
        # Bloqueados también usan dias_bloqueo_real (cae a today()-starts cuando
        # siguen activos, equivalente a dias_bloqueado_total).
        reactivados.sort(key=lambda x: x.get('dias_bloqueo_real') or x.get('dias_bloqueado_total') or 0, reverse=True)
        bloqueados.sort(key=lambda x: x.get('dias_bloqueo_real') or x.get('dias_bloqueado_total') or 0, reverse=True)

        # Sample por sección: el frontend pinta tablas, así que entregamos
        # una muestra amplia (3000) suficiente para visualización. Los conteos
        # del resumen vienen de las listas COMPLETAS (sin slice) para que el
        # usuario vea cuántos hay aunque la tabla solo muestre los primeros.
        SAMPLE = 3000
        muestra_truncada = (
            len(alertas)     > SAMPLE
            or len(bloqueados)  > SAMPLE
            or len(reactivados) > SAMPLE
        )

        return jsonify(limpiar_nan({
            "desde": desde, "hasta": hasta,
            "alertas":     alertas[:SAMPLE],
            "bloqueados":  bloqueados[:SAMPLE],
            "reactivados": reactivados[:SAMPLE],
            "resumen": {
                "total":            total,
                "alerta":           n_alerta,
                "ok":               n_ok,
                "bloqueados":       len(bloqueados),
                "reactivados":      len(reactivados),
                "expulsados":       n_expulsados,
                "suspendidos":      n_suspendidos,
                "susp_mas30":       n_susp_30,
                "pct_alerta":       round(n_alerta/total*100) if total else 0,
                "pct_ok":           round(n_ok/total*100) if total else 0,
                "muestra_size":     SAMPLE,
                "muestra_truncada": muestra_truncada,
            },
            "stats_bloqueados":  top10_stats(bloqueados),
            "stats_reactivados": top10_stats(reactivados),
        }))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/")
def index():
    return send_from_directory(".", "dashboard.html")

@app.route("/dashboard.html")
def dashboard_html():
    return send_from_directory(".", "dashboard.html")

threading.Thread(target=auto_refresh, daemon=True).start()

# ══════════════════════════════════════════════════════════════════════
# SCHEDULER DE RECORDATORIOS DEL CALENDARIO
# Corre en background cada 60 segundos, revisa tareas pendientes
# y envía recordatorio por email a la hora programada
# ══════════════════════════════════════════════════════════════════════

_RECORDATORIO_TABLE = """
CREATE TABLE IF NOT EXISTS picapmongoprod.calendario_recordatorios (
    id           String,
    titulo       String,
    fecha        String,
    hora         String,
    email        String,
    creado_por   String,
    enviado      UInt8 DEFAULT 0,
    creado_en    DateTime DEFAULT now()
) ENGINE = MergeTree()
ORDER BY (fecha, hora, id)
"""

def _init_recordatorios_table():
    """Crear la tabla de recordatorios si no existe."""
    try:
        ch = get_client()
        ch.command(_RECORDATORIO_TABLE)
    except Exception as e:
        print(f"[recordatorios] Error creando tabla: {e}")

def _scheduler_recordatorios():
    """Thread que corre cada 60s y envía recordatorios pendientes."""
    import time as _time
    _time.sleep(10)  # Esperar que el servidor arranque
    _init_recordatorios_table()

    while True:
        try:
            now    = datetime.now()
            hoy    = now.strftime("%Y-%m-%d")
            minuto = now.strftime("%H:%M")

            ch = get_client()
            # Buscar recordatorios de hoy a este minuto que no se han enviado
            rows = ch.query(f"""
                SELECT id, titulo, fecha, hora, email, creado_por
                FROM picapmongoprod.calendario_recordatorios
                WHERE fecha = '{hoy}'
                  AND hora  = '{minuto}'
                  AND enviado = 0
                  AND email != ''
            """).result_rows

            for row in rows:
                rid, titulo, fecha, hora, email, creado_por = row
                fecha_parts = fecha.split("-")
                MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
                try:
                    fecha_display = f"{int(fecha_parts[2])} de {MESES[int(fecha_parts[1])]} de {fecha_parts[0]}"
                except:
                    fecha_display = fecha

                cuerpo = f"""
                <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;
                            background:#f9fafb;border-radius:12px;overflow:hidden;">
                  <div style="background:linear-gradient(135deg,#4c1d95,#7c3aed);
                              padding:24px 32px;text-align:center;">
                    <h1 style="color:#fff;margin:0;font-size:20px">⏰ Recordatorio</h1>
                    <p style="color:rgba(255,255,255,.8);margin:4px 0 0;font-size:12px">
                      Picap Monitoreo · Calendario
                    </p>
                  </div>
                  <div style="padding:24px 32px;background:#fff;">
                    <h2 style="color:#1e1b4b;font-size:15px;margin:0 0 16px">
                      Es hora de tu tarea programada
                    </h2>
                    <div style="background:#f5f3ff;border-left:4px solid #7c3aed;
                                border-radius:0 8px 8px 0;padding:14px 16px;margin-bottom:16px;">
                      <div style="font-size:20px;font-weight:700;color:#1e1b4b;margin-bottom:10px">
                        {titulo}
                      </div>
                      <div style="font-size:13px;color:#374151;margin-bottom:4px;">
                        📆 <strong>Fecha:</strong> {fecha_display}
                      </div>
                      <div style="font-size:14px;color:#7c3aed;font-weight:700;">
                        ⏰ <strong>Hora:</strong> {hora}
                      </div>
                    </div>
                    <hr style="border:none;border-top:1px solid #e5e7eb;margin:14px 0;">
                    <p style="color:#9ca3af;font-size:11px;margin:0">
                      Picap Monitoreo · Sistema interno · No responder este correo.
                    </p>
                  </div>
                </div>"""

                ok, err = _enviar_email(
                    email,
                    f"⏰ Recordatorio: {titulo} — {hora}",
                    cuerpo
                )

                if ok:
                    # Marcar como enviado (insert con enviado=1 para ReplacingMergeTree)
                    ch.command(f"""
                        ALTER TABLE picapmongoprod.calendario_recordatorios
                        UPDATE enviado = 1
                        WHERE id = '{rid}'
                    """)
                    print(f"[recordatorio] ✓ Enviado a {email}: {titulo} a las {hora}")
                else:
                    print(f"[recordatorio] ✗ Error enviando a {email}: {err}")

        except Exception as e:
            print(f"[recordatorio] Error en scheduler: {e}")

        _time.sleep(60)  # Revisar cada minuto

# Iniciar scheduler de recordatorios
threading.Thread(target=_scheduler_recordatorios, daemon=True).start()

# ══════════════════════════════════════════════════════════════════════
# CRONOGRAMA DE FUNCIONES — tareas recurrentes con email push
# A diferencia de calendario_recordatorios (eventos puntuales),
# cronograma_tareas son tareas que se repiten en días específicos
# de la semana a una hora fija (ej: "Cada lunes y miércoles a las 9:00").
# ══════════════════════════════════════════════════════════════════════

_CRONOGRAMA_TABLE = """
CREATE TABLE IF NOT EXISTS picapmongoprod.cronograma_tareas (
    id           String,
    titulo       String,
    descripcion  String,
    dias_semana  String,           -- 'lun,mar,mie,jue,vie,sab,dom' separados por coma
    hora         String,           -- HH:MM (24h)
    email        String,
    creado_por   String,
    activo       UInt8 DEFAULT 1,
    ultima_ejecucion String DEFAULT '',  -- YYYY-MM-DD última vez enviado (anti-duplicado)
    creado_en    DateTime DEFAULT now(),
    actualizado_en DateTime DEFAULT now()
) ENGINE = ReplacingMergeTree(actualizado_en)
ORDER BY id
"""

# Mapeo entre el índice de Python (Monday=0) y la abreviatura usada en la tabla.
_DIAS_ABREV = ['lun','mar','mie','jue','vie','sab','dom']

def _init_cronograma_table():
    try:
        ch = get_client()
        ch.command(_CRONOGRAMA_TABLE)
    except Exception as e:
        print(f"[cronograma] Error creando tabla: {e}")

def _scheduler_cronograma():
    """Cada 60s: revisar tareas activas que coincidan con día y hora actual."""
    import time as _time
    _time.sleep(15)  # esperar a que el server arranque
    _init_cronograma_table()

    while True:
        try:
            now    = datetime.now()
            hoy    = now.strftime("%Y-%m-%d")
            minuto = now.strftime("%H:%M")
            dia_idx = now.weekday()  # 0=lunes, 6=domingo
            dia_abr = _DIAS_ABREV[dia_idx]

            ch = get_client()
            rows = ch.query(f"""
                SELECT id, titulo, descripcion, dias_semana, hora, email, creado_por
                FROM picapmongoprod.cronograma_tareas FINAL
                WHERE activo = 1
                  AND hora = '{minuto}'
                  AND email != ''
                  AND ultima_ejecucion != '{hoy}'
                  AND (dias_semana = '*' OR positionCaseInsensitive(dias_semana, '{dia_abr}') > 0)
            """).result_rows

            for row in rows:
                rid, titulo, descripcion, dias_semana, hora, email, creado_por = row
                # Render del email
                desc_html = ''
                if descripcion:
                    descripcion_safe = (descripcion[:500]
                                        .replace('<', '&lt;').replace('>', '&gt;'))
                    desc_html = f"""
                      <div style="margin-top:10px;font-size:13px;color:#374151;
                                  background:#f9fafb;padding:10px 14px;border-radius:8px;
                                  border-left:3px solid #a78bfa;line-height:1.5">
                        {descripcion_safe}
                      </div>"""
                cuerpo = f"""
                <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;
                            background:#f9fafb;border-radius:12px;overflow:hidden;">
                  <div style="background:linear-gradient(135deg,#4c1d95,#7c3aed);
                              padding:24px 32px;text-align:center;">
                    <h1 style="color:#fff;margin:0;font-size:20px">🗓️ Cronograma — Picap Monitoreo</h1>
                    <p style="color:rgba(255,255,255,.85);margin:4px 0 0;font-size:12px">
                      Recordatorio programado de tu cronograma de funciones
                    </p>
                  </div>
                  <div style="padding:24px 32px;background:#fff;">
                    <h2 style="color:#1e1b4b;font-size:15px;margin:0 0 12px">
                      Es hora de tu tarea programada
                    </h2>
                    <div style="background:#f5f3ff;border-left:4px solid #7c3aed;
                                border-radius:0 8px 8px 0;padding:14px 16px;margin-bottom:14px;">
                      <div style="font-size:18px;font-weight:700;color:#1e1b4b;margin-bottom:8px">
                        {titulo}
                      </div>
                      <div style="font-size:13px;color:#374151;margin-bottom:4px;">
                        ⏰ <strong>Hora:</strong> {hora}
                      </div>
                      <div style="font-size:13px;color:#374151">
                        📅 <strong>Días:</strong> {dias_semana if dias_semana != '*' else 'Todos los días'}
                      </div>
                      {desc_html}
                    </div>
                    <p style="color:#9ca3af;font-size:11px;margin:14px 0 0">
                      Picap Monitoreo · Cronograma de funciones · No responder este correo.
                    </p>
                  </div>
                </div>"""

                ok, err = _enviar_email(
                    email,
                    f"🗓️ Cronograma: {titulo} — {hora}",
                    cuerpo
                )

                if ok:
                    # Marcar última ejecución para no enviar otro hoy mismo
                    ch.command(f"""
                        ALTER TABLE picapmongoprod.cronograma_tareas
                        UPDATE ultima_ejecucion = '{hoy}',
                               actualizado_en = now()
                        WHERE id = '{rid}'
                    """)
                    print(f"[cronograma] ✓ {email}: {titulo} a las {hora}")
                else:
                    print(f"[cronograma] ✗ {email}: {err}")

        except Exception as e:
            print(f"[cronograma] Error scheduler: {e}")

        _time.sleep(60)

threading.Thread(target=_scheduler_cronograma, daemon=True).start()


# ── CRUD del cronograma ──────────────────────────────────────────────

def _es_admin(token):
    """Helper: True si el token corresponde a un usuario con rol admin."""
    s = _verificar_sesion(None, token)
    return bool(s and s.get('rol') == 'admin')

@app.route("/api/admin/diag")
def admin_diag():
    """Endpoint de diagnóstico solo-admin: muestra qué envs detectó el server.
    Útil para verificar la configuración en Render sin tocar el código."""
    token = request.headers.get("X-Token", "")
    s = _verificar_sesion(None, token)
    if not s or s.get('rol') != 'admin':
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    def _mask(v):
        if not v: return None
        v = str(v)
        return v[:3] + '***' + v[-2:] if len(v) > 6 else '***'
    # Determinar qué proveedor de email se usará
    if BREVO_API_KEY:    email_provider = 'BREVO (HTTP)'
    elif RESEND_API_KEY: email_provider = 'RESEND (HTTP)'
    elif SMTP_EMAIL and SMTP_PASSWORD: email_provider = f'SMTP ({SMTP_HOST}:{SMTP_PORT})'
    else:                email_provider = 'NINGUNO — emails no se enviarán'

    return jsonify({
        "ok": True,
        "envs_detectadas": {
            "CLICKHOUSE_HOST":     _mask(os.environ.get("CLICKHOUSE_HOST")),
            "CLICKHOUSE_USER":     _mask(os.environ.get("CLICKHOUSE_USER")),
            "CLICKHOUSE_PASSWORD": _mask(os.environ.get("CLICKHOUSE_PASSWORD")),
            "TOKEN_SECRET":        _mask(os.environ.get("TOKEN_SECRET")),
            "BREVO_API_KEY":       _mask(os.environ.get("BREVO_API_KEY")),
            "RESEND_API_KEY":      _mask(os.environ.get("RESEND_API_KEY")),
            "RESEND_FROM":         os.environ.get("RESEND_FROM"),
            "BREVO_FROM":          os.environ.get("BREVO_FROM"),
            "SMTP_EMAIL":          _mask(os.environ.get("SMTP_EMAIL")),
            "SMTP_USER":           _mask(os.environ.get("SMTP_USER")),
            "SMTP_PASSWORD":       _mask(os.environ.get("SMTP_PASSWORD")),
            "SMTP_PASS":           _mask(os.environ.get("SMTP_PASS")),
            "SMTP_HOST":           os.environ.get("SMTP_HOST"),
            "SMTP_PORT":           os.environ.get("SMTP_PORT"),
        },
        "email_provider_activo": email_provider,
        "smtp_resuelto": {
            "SMTP_EMAIL_efectivo":    _mask(SMTP_EMAIL),
            "SMTP_PASSWORD_definido": bool(SMTP_PASSWORD),
            "SMTP_HOST_efectivo":     SMTP_HOST,
            "SMTP_PORT_efectivo":     SMTP_PORT,
        },
    })


@app.route("/api/cronograma", methods=["GET"])
def cronograma_list():
    """Lista todas las tareas del cronograma. Solo admin."""
    token = request.headers.get("X-Token", "")
    if not _es_admin(token):
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    try:
        ch = get_client()
        _init_cronograma_table()
        r = ch.query("""
            SELECT id, titulo, descripcion, dias_semana, hora, email, creado_por,
                   activo, ultima_ejecucion, creado_en, actualizado_en
            FROM picapmongoprod.cronograma_tareas FINAL
            ORDER BY hora, titulo
        """)
        cols = r.column_names
        tareas = []
        for row in r.result_rows:
            t = dict(zip(cols, row))
            # Normalizar tipos
            t['activo']         = int(t.get('activo', 0) or 0)
            t['creado_en']      = str(t.get('creado_en', ''))[:19]
            t['actualizado_en'] = str(t.get('actualizado_en', ''))[:19]
            tareas.append(t)
        return jsonify({"ok": True, "tareas": tareas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/cronograma", methods=["POST"])
def cronograma_create():
    """Crea una nueva tarea recurrente. Solo admin."""
    token = request.headers.get("X-Token", "")
    s = _verificar_sesion(None, token)
    if not s or s.get('rol') != 'admin':
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    data = request.get_json() or {}
    titulo      = (data.get("titulo") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()
    dias_semana = (data.get("dias_semana") or "").strip().lower()
    hora        = (data.get("hora") or "").strip()
    email       = (data.get("email") or "").strip()
    if not titulo or not dias_semana or not hora or not email:
        return jsonify({"ok": False, "error": "Faltan campos: titulo, dias_semana, hora, email"}), 400
    # Validaciones simples
    if not _validar_hora(hora):
        return jsonify({"ok": False, "error": "Hora inválida (HH:MM, 24h)"}), 400
    if not _validar_dias(dias_semana):
        return jsonify({"ok": False, "error": "Días inválidos (usar abreviaturas: lun,mar,mie,jue,vie,sab,dom o '*')"}), 400
    try:
        ch = get_client()
        _init_cronograma_table()
        new_id = str(uuid.uuid4())
        ch.insert("picapmongoprod.cronograma_tareas",
            [[new_id, titulo, descripcion, dias_semana, hora, email,
              s.get('usuario','admin'), 1, '',
              datetime.utcnow(), datetime.utcnow()]],
            column_names=["id","titulo","descripcion","dias_semana","hora","email",
                          "creado_por","activo","ultima_ejecucion",
                          "creado_en","actualizado_en"])
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/cronograma/<tid>", methods=["PUT"])
def cronograma_update(tid):
    """Edita una tarea. Solo admin."""
    token = request.headers.get("X-Token", "")
    s = _verificar_sesion(None, token)
    if not s or s.get('rol') != 'admin':
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    data = request.get_json() or {}
    titulo      = (data.get("titulo") or "").strip()
    descripcion = (data.get("descripcion") or "").strip()
    dias_semana = (data.get("dias_semana") or "").strip().lower()
    hora        = (data.get("hora") or "").strip()
    email       = (data.get("email") or "").strip()
    activo      = 1 if data.get("activo", True) else 0
    if not titulo or not dias_semana or not hora or not email:
        return jsonify({"ok": False, "error": "Faltan campos"}), 400
    if not _validar_hora(hora):
        return jsonify({"ok": False, "error": "Hora inválida"}), 400
    if not _validar_dias(dias_semana):
        return jsonify({"ok": False, "error": "Días inválidos"}), 400
    # Sanitizar tid (prevenir SQL injection en el WHERE)
    if not all(c in '0123456789abcdef-' for c in tid.lower()):
        return jsonify({"ok": False, "error": "ID inválido"}), 400
    try:
        ch = get_client()
        # Leer fila actual para preservar campos no modificables
        r = ch.query(f"""
            SELECT creado_por, ultima_ejecucion, creado_en
            FROM picapmongoprod.cronograma_tareas FINAL
            WHERE id = '{tid}' LIMIT 1
        """).result_rows
        if not r:
            return jsonify({"ok": False, "error": "Tarea no encontrada"}), 404
        creado_por, ultima_ejecucion, creado_en = r[0]
        # Insert con mismo id para que ReplacingMergeTree haga el merge
        ch.insert("picapmongoprod.cronograma_tareas",
            [[tid, titulo, descripcion, dias_semana, hora, email,
              creado_por, activo, ultima_ejecucion,
              creado_en, datetime.utcnow()]],
            column_names=["id","titulo","descripcion","dias_semana","hora","email",
                          "creado_por","activo","ultima_ejecucion",
                          "creado_en","actualizado_en"])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/cronograma/<tid>", methods=["DELETE"])
def cronograma_delete(tid):
    """Borra una tarea (hard delete). Solo admin."""
    token = request.headers.get("X-Token", "")
    s = _verificar_sesion(None, token)
    if not s or s.get('rol') != 'admin':
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    if not all(c in '0123456789abcdef-' for c in tid.lower()):
        return jsonify({"ok": False, "error": "ID inválido"}), 400
    try:
        ch = get_client()
        ch.command(f"ALTER TABLE picapmongoprod.cronograma_tareas DELETE WHERE id = '{tid}'")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/cronograma/<tid>/test", methods=["POST"])
def cronograma_test(tid):
    """Envía un email de prueba inmediato con la tarea. Solo admin."""
    token = request.headers.get("X-Token", "")
    s = _verificar_sesion(None, token)
    if not s or s.get('rol') != 'admin':
        return jsonify({"ok": False, "error": "Solo admins"}), 403
    if not all(c in '0123456789abcdef-' for c in tid.lower()):
        return jsonify({"ok": False, "error": "ID inválido"}), 400
    try:
        ch = get_client()
        r = ch.query(f"""
            SELECT titulo, descripcion, dias_semana, hora, email
            FROM picapmongoprod.cronograma_tareas FINAL
            WHERE id = '{tid}' LIMIT 1
        """).result_rows
        if not r:
            return jsonify({"ok": False, "error": "Tarea no encontrada"}), 404
        titulo, descripcion, dias_semana, hora, email = r[0]
        desc_safe = (descripcion[:500].replace('<','&lt;').replace('>','&gt;')) if descripcion else ''
        desc_html = f"""<div style="margin-top:10px;font-size:13px;color:#374151;background:#f9fafb;padding:10px 14px;border-radius:8px;border-left:3px solid #a78bfa;line-height:1.5">{desc_safe}</div>""" if desc_safe else ''
        cuerpo = f"""
        <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;background:#f9fafb;border-radius:12px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#16a34a,#22c55e);padding:24px 32px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:20px">🧪 Prueba de Cronograma</h1>
            <p style="color:rgba(255,255,255,.85);margin:4px 0 0;font-size:12px">
              Este es un envío de prueba — la tarea se enviará automáticamente según su programación
            </p>
          </div>
          <div style="padding:24px 32px;background:#fff;">
            <div style="background:#f0fdf4;border-left:4px solid #16a34a;border-radius:0 8px 8px 0;padding:14px 16px;margin-bottom:14px;">
              <div style="font-size:18px;font-weight:700;color:#1e1b4b;margin-bottom:8px">{titulo}</div>
              <div style="font-size:13px;color:#374151;margin-bottom:4px;">⏰ <strong>Hora:</strong> {hora}</div>
              <div style="font-size:13px;color:#374151">📅 <strong>Días:</strong> {dias_semana if dias_semana != '*' else 'Todos los días'}</div>
              {desc_html}
            </div>
          </div>
        </div>"""
        ok, err = _enviar_email(email, f"🧪 [Prueba] {titulo}", cuerpo)
        if ok:
            return jsonify({"ok": True, "mensaje": f"Email de prueba enviado a {email}"})
        else:
            return jsonify({"ok": False, "error": err or "Falló el envío"}), 500
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def _validar_hora(hora):
    try:
        if len(hora) != 5 or hora[2] != ':': return False
        h = int(hora[0:2]); m = int(hora[3:5])
        return 0 <= h <= 23 and 0 <= m <= 59
    except Exception:
        return False

def _validar_dias(dias):
    if not dias: return False
    if dias.strip() == '*': return True
    valid = set(_DIAS_ABREV)
    parts = [p.strip() for p in dias.split(',') if p.strip()]
    return bool(parts) and all(p in valid for p in parts)


# ══════════════════════════════════════════════════════════════
# MÓDULO RECONOCIMIENTO FACIAL — aislado, sin tocar lo anterior
# Los datos los genera reconocimiento_facial_v3.py localmente
# y los persiste en ClickHouse. Este endpoint solo lee.
# ══════════════════════════════════════════════════════════════
Q_RF_RESUMEN = """
SELECT
    count()                    AS total_alertas,
    countIf(nivel='ALERTA')    AS total_alerta,
    countIf(nivel='REVISAR')   AS total_revisar,
    countIf(nivel='POSIBLE')   AS total_posible,
    round(max(similitud),4)    AS sim_max,
    round(avg(similitud),4)    AS sim_avg,
    count(DISTINCT user_id_a)  AS pilotos_unicos
FROM picapmongoprod.alertas_reconocimiento
WHERE procesado_en BETWEEN toDateTime('{desde} 00:00:00')
                       AND toDateTime('{hasta} 23:59:59')
"""

Q_RF_ALERTAS = """
SELECT nivel, similitud,
       user_id_a, url_a, created_at_a,
       user_id_b, url_b, created_at_b,
       procesado_en
FROM picapmongoprod.alertas_reconocimiento
WHERE procesado_en BETWEEN toDateTime('{desde} 00:00:00')
                       AND toDateTime('{hasta} 23:59:59')
ORDER BY similitud DESC
LIMIT 300
"""

@app.route("/api/reconocimiento")
def reconocimiento():
    # Rango amplio por defecto para cubrir datos recientes
    from datetime import date
    hoy = date.today().strftime("%Y-%m-%d")
    desde = request.args.get("desde") or "2024-01-01"
    hasta  = request.args.get("hasta") or hoy

    # Umbral de "misma persona". Modelos de face embeddings cosine: 0.92+ ≈
    # mismo individuo con alta confianza. Below that, hay alta probabilidad
    # de que sean personas distintas que casualmente comparten rasgos.
    # Reglas de negocio:
    #   - similitud >= umbral  → MISMA_PERSONA → genera alerta
    #   - similitud <  umbral  → PERSONA_DIFERENTE → NO genera alerta
    #   - mismo_imei = 'SÍ'    → siempre alerta (mismo dispositivo, señal
    #                             independiente del rostro)
    #
    # Modos preset (para que el operador no tenga que pensar en numeros):
    #   - alta:        umbral 0.96 — solo casos muy seguros (defecto)
    #   - equilibrado: umbral 0.93 — balance entre cobertura y precision
    #   - auditoria:   umbral 0.85 — vista amplia para revision manual
    PRESETS = {
        "alta":        0.96,
        "equilibrado": 0.93,
        "auditoria":   0.85,
    }
    modo = (request.args.get("modo") or "alta").lower()
    umbral_default = PRESETS.get(modo, PRESETS["alta"])

    try:
        umbral = float(request.args.get("umbral", str(umbral_default)))
    except Exception:
        umbral = umbral_default
    # Acotar a rango razonable para evitar consultas inútiles
    umbral = max(0.50, min(1.00, umbral))

    # Cuando el operador activa "solo con apellido coincidente" exigimos que
    # los nombres compartan al menos un token (>=3 chars) además del umbral.
    # Esto filtra falsos positivos del estilo:
    #   "Nicolás Esteban Pérez"  vs  "Gabriel Hernández Suárez"  con sim=0.95
    # donde el modelo se equivoca pero los apellidos son claramente distintos.
    solo_con_apellido = (request.args.get("solo_con_apellido", "0") in ("1", "true", "True"))

    try:
        ch = get_client()

        # Verificar tabla
        tabla_ok = ch.query(
            "SELECT count() FROM system.tables "
            "WHERE database='picapmongoprod' AND name='alertas_reconocimiento'"
        ).result_rows[0][0] > 0

        if not tabla_ok:
            return jsonify({"tabla_existe": False, "alertas": [], "resumen": {
                "total_alertas":0,"total_alerta":0,"total_revisar":0,
                "total_posible":0,"pilotos_unicos":0}})

        # Helper SQL: tokens normalizados (>=3 chars) de cada nombre.
        # arrayCount(t -> has(...), ...) > 0  → al menos un token compartido.
        # Esto se usa como señal de "apellido / nombre coincide" — sirve como
        # filtro secundario contra falsos positivos del modelo facial.
        TOKENS_SQL = """
            arrayFilter(
                tk -> length(tk) >= 3,
                arrayMap(s -> lowerUTF8(s), splitByChar(' ', toString({col})))
            )
        """

        # Expresion booleana: nombres comparten al menos un token significativo.
        APELLIDO_OK_SQL = (
            "(length(" + TOKENS_SQL.format(col="nombre_a") + ") > 0 "
            "AND length(" + TOKENS_SQL.format(col="nombre_b") + ") > 0 "
            "AND arrayCount(t -> has("
                + TOKENS_SQL.format(col="nombre_b") + ", t), "
                + TOKENS_SQL.format(col="nombre_a") + ") > 0)"
        )

        # Filtro WHERE compuesto: misma persona si pasa el umbral.
        # Con solo_con_apellido=1 exigimos también nombres compartidos
        # (excepto cuando hay coincidencia de IMEI — eso siempre alerta).
        if solo_con_apellido:
            FILTRO_ALERTA = (
                "((toFloat64(similitud) >= {u} AND " + APELLIDO_OK_SQL + ") "
                " OR ifNull(mismo_imei,'NO')='SÍ')"
            )
        else:
            FILTRO_ALERTA = (
                "(toFloat64(similitud) >= {u} OR ifNull(mismo_imei,'NO')='SÍ')"
            )

        # Resumen completo (TODAS las filas en el rango — incluye descartadas)
        # y resumen filtrado (solo MISMA_PERSONA o mismo IMEI). Esto permite
        # mostrar tanto "alertas reales" como "descartadas por baja similitud"
        # en el front, en una sola consulta.
        Q_simple = ("""
            SELECT
                count()                                                AS total_filas,
                countIf(""" + FILTRO_ALERTA + """)                     AS total_alertas,
                count() - countIf(""" + FILTRO_ALERTA + """)           AS total_descartadas,

                countIf(tipo_alerta='RF + IMEI' AND """ + FILTRO_ALERTA + """) AS n_rf_imei,
                countIf(tipo_alerta='RF'        AND """ + FILTRO_ALERTA + """) AS n_rf,
                countIf(tipo_alerta='IMEI'      AND ifNull(mismo_imei,'NO')='SÍ') AS n_imei,
                countIf(nivel='FOTO_DUPLICADA' AND """ + FILTRO_ALERTA + """) AS n_duplicada,
                countIf(nivel='ALERTA'         AND """ + FILTRO_ALERTA + """) AS n_alerta,
                countIf(nivel='REVISAR'        AND """ + FILTRO_ALERTA + """) AS n_revisar,
                countIf(nivel='POSIBLE'        AND """ + FILTRO_ALERTA + """) AS n_posible,

                round(maxIf(similitud, toFloat64(similitud) > 0), 4)   AS sim_max,
                round(avgIf(similitud, toFloat64(similitud) > 0), 4)   AS sim_avg,
                count(DISTINCT user_id_a)                              AS pilotos,
                countIf(NOT """ + APELLIDO_OK_SQL + """ AND """ + FILTRO_ALERTA + """) AS n_apellido_distinto
            FROM picapmongoprod.alertas_reconocimiento
            WHERE procesado_en >= toDateTime('{d} 00:00:00')
              AND procesado_en <= toDateTime('{h} 23:59:59')
        """).format(d=desde, h=hasta, u=umbral)

        r_res = ch.query(Q_simple)
        res = {}
        if r_res.result_rows:
            row = r_res.result_rows[0]
            cols = r_res.column_names
            res = dict(zip(cols, row))

        # Alertas detalladas — aplican mismo FILTRO_ALERTA y exponen
        # apellido_coincide para que el front pueda marcar visualmente
        # las coincidencias sospechosas (rostros parecidos pero apellidos
        # distintos = posible falso positivo).
        Q_det = ("""
            SELECT
                ifNull(tipo_alerta,'RF')   AS tipo_alerta,
                nivel,
                toFloat64(similitud)       AS similitud,
                ifNull(mismo_imei,'NO')    AS mismo_imei,
                toString(nombre_a)         AS nombre_a,
                toString(user_id_a)        AS user_id_a,
                toString(url_a)            AS url_a,
                toString(created_at_a)     AS created_at_a,
                toString(nombre_b)         AS nombre_b,
                toString(user_id_b)        AS user_id_b,
                toString(url_b)            AS url_b,
                toString(created_at_b)     AS created_at_b,
                procesado_en,
                IF(""" + APELLIDO_OK_SQL + """, 1, 0) AS apellido_coincide,
                multiIf(
                    toFloat64(similitud) >= {u}, 'MISMA_PERSONA',
                    ifNull(mismo_imei,'NO') = 'SÍ', 'MISMO_DISPOSITIVO',
                    'PERSONA_DIFERENTE'
                ) AS clasificacion_final
            FROM picapmongoprod.alertas_reconocimiento
            WHERE procesado_en >= toDateTime('{d} 00:00:00')
              AND procesado_en <= toDateTime('{h} 23:59:59')
              AND """ + FILTRO_ALERTA + """
            ORDER BY
                multiIf(tipo_alerta='RF + IMEI',0,tipo_alerta='RF',1,2),
                similitud DESC
            LIMIT 300
        """).format(d=desde, h=hasta, u=umbral)

        r_det = ch.query(Q_det)
        alertas = []
        for row in r_det.result_rows:
            d_row = dict(zip(r_det.column_names, row))
            try:
                d_row["similitud"] = round(float(d_row.get("similitud", 0)), 4)
            except Exception:
                d_row["similitud"] = 0.0
            alertas.append(d_row)

        return jsonify({
            "tabla_existe": True,
            "desde": desde,
            "hasta": hasta,
            "modo":                modo,
            "umbral_aplicado":     umbral,
            "solo_con_apellido":   solo_con_apellido,
            "resumen": {
                "total_filas":          int(res.get("total_filas", 0)),
                "total_alertas":        int(res.get("total_alertas", 0)),
                "total_descartadas":    int(res.get("total_descartadas", 0)),
                "total_rf_imei":        int(res.get("n_rf_imei", 0)),
                "total_rf":             int(res.get("n_rf", 0)),
                "total_imei":           int(res.get("n_imei", 0)),
                "total_duplicada":      int(res.get("n_duplicada", 0)),
                "total_alerta":         int(res.get("n_alerta", 0)),
                "total_revisar":        int(res.get("n_revisar", 0)),
                "total_posible":        int(res.get("n_posible", 0)),
                "total_apellido_distinto": int(res.get("n_apellido_distinto", 0)),
                "sim_max":              float(res.get("sim_max") or 0),
                "sim_avg":              float(res.get("sim_avg") or 0),
                "pilotos_unicos":       int(res.get("pilotos", 0)),
                "umbral":               umbral,
                "modo":                 modo,
            },
            "alertas": alertas,
        })
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc(),
            "tabla_existe": False,
            "alertas": [],
            "resumen": {"total_alertas": 0}
        }), 500



# ══════════════════════════════════════════════════════════════
# MÓDULO ALERTA CÉDULA
# Compara la cédula que extrajo Rekognition (foto del documento) con la
# cédula que aparece en el reporte de antecedentes (people_police_records).
# Si NO coinciden → 'alerta' (la persona pudo registrar documento ajeno).
# ══════════════════════════════════════════════════════════════

# Mapeo: nombre de país visible (en la UI) → código ISO en passengers.g_country
_CEDULA_PAIS_ISO = {
    "Colombia":  "CO",
    "Mexico":    "MX",
    "Nicaragua": "NI",
    "Guatemala": "GT",
    "Peru":      "PE",
    "Ecuador":   "EC",
}

def _cedula_filtro_pais(pais):
    """Cláusula AND para filtrar la query de cédula por país."""
    if not pais:
        return ""
    iso = _CEDULA_PAIS_ISO.get(pais, pais)
    return f"AND p.g_country = '{iso}'"

# Query base (sin LIMIT) — el endpoint la usa para detalle (con LIMIT) y
# para agregación.
_Q_CEDULA = r"""
SELECT
    creacion_cuenta, id_user, name_user, pais_codigo,
    rekognition_cc, cc_antecedentes, nombre_antecedentes, cc_igual
FROM (
    SELECT
        toTimeZone(p.created_at, 'America/Bogota') AS creacion_cuenta,
        toString(pwd._id)                           AS id_user,
        toString(p.name)                            AS name_user,
        toString(p.g_country)                       AS pais_codigo,
        JSONExtractString(pwd.rekognition_metadata, 'fiscal_number')
                                                    AS rekognition_cc,
        extract(pwd.people_police_records, 'Cédula de Ciudadanía Nº\\s*([0-9]+)')
                                                    AS cc_antecedentes,
        trim(extract(pwd.people_police_records, 'Apellidos y Nombres:\\s*([^\\\\]+)'))
                                                    AS nombre_antecedentes,
        CASE
            WHEN JSONExtractString(pwd.rekognition_metadata, 'fiscal_number')
               = extract(pwd.people_police_records, 'Cédula de Ciudadanía Nº\\s*([0-9]+)')
            THEN 'ok' ELSE 'alerta'
        END                                          AS cc_igual,
        ROW_NUMBER() OVER (PARTITION BY p._id ORDER BY p.created_at DESC) AS rn
    FROM picapmongoprod.passengers_w_data pwd
    LEFT JOIN picapmongoprod.passengers p ON pwd._id = p._id
    WHERE p.created_at BETWEEN toDateTime('{desde} 00:00:00')
                           AND toDateTime('{hasta} 23:59:59')
      {filtro_pais}
)
WHERE rn = 1
  AND rekognition_cc != ''
  AND cc_antecedentes != ''
ORDER BY creacion_cuenta DESC
LIMIT {limit_filas}
"""

# Query de agregación (totales por día y conteo total) sin LIMIT, para que
# los KPIs reflejen el universo completo aunque la tabla muestre solo un sample.
_Q_CEDULA_AGG = r"""
SELECT
    toDate(creacion_cuenta)               AS dia,
    count()                               AS total_dia,
    countIf(cc_igual = 'alerta')          AS alertas_dia
FROM (
    SELECT
        toTimeZone(p.created_at, 'America/Bogota') AS creacion_cuenta,
        CASE
            WHEN JSONExtractString(pwd.rekognition_metadata, 'fiscal_number')
               = extract(pwd.people_police_records, 'Cédula de Ciudadanía Nº\\s*([0-9]+)')
            THEN 'ok' ELSE 'alerta'
        END                                          AS cc_igual,
        JSONExtractString(pwd.rekognition_metadata, 'fiscal_number') AS rk_cc,
        extract(pwd.people_police_records, 'Cédula de Ciudadanía Nº\\s*([0-9]+)') AS pr_cc,
        ROW_NUMBER() OVER (PARTITION BY p._id ORDER BY p.created_at DESC) AS rn
    FROM picapmongoprod.passengers_w_data pwd
    LEFT JOIN picapmongoprod.passengers p ON pwd._id = p._id
    WHERE p.created_at BETWEEN toDateTime('{desde} 00:00:00')
                           AND toDateTime('{hasta} 23:59:59')
      {filtro_pais}
)
WHERE rn = 1
  AND rk_cc != ''
  AND pr_cc != ''
GROUP BY dia
ORDER BY dia
"""

@app.route("/api/cedula-alertas")
def cedula_alertas():
    """Alertas de Cédula: compara CC en foto vs CC en antecedentes."""
    # Verificar autenticación
    token = request.headers.get("X-Token", "")
    sesion = _verificar_sesion(None, token)
    if not sesion:
        return jsonify({"ok": False, "error": "Sesión expirada"}), 401

    desde   = request.args.get("desde") or (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    hasta   = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais    = request.args.get("pais", "").strip()
    LIMIT_DETALLE = 5000

    filtro_pais = _cedula_filtro_pais(pais)

    try:
        ch = get_client()

        # 1) Agregación: totales reales (no limitada)
        sql_agg = _Q_CEDULA_AGG.format(desde=desde, hasta=hasta, filtro_pais=filtro_pais)
        r_agg   = ch.query(sql_agg)
        agg_rows = [dict(zip(r_agg.column_names, row)) for row in r_agg.result_rows]

        total_real    = sum(int(r.get('total_dia', 0)   or 0) for r in agg_rows)
        alertas_real  = sum(int(r.get('alertas_dia', 0) or 0) for r in agg_rows)
        ok_real       = total_real - alertas_real

        trend = [
            {
                'fecha':   str(r['dia']),
                'alertas': int(r.get('alertas_dia', 0) or 0),
                'ok':      int(r.get('total_dia', 0) or 0) - int(r.get('alertas_dia', 0) or 0),
            }
            for r in agg_rows
        ]

        # 2) Sample detallado para la tabla
        sql_det = _Q_CEDULA.format(desde=desde, hasta=hasta, filtro_pais=filtro_pais,
                                   limit_filas=LIMIT_DETALLE)
        r_det = ch.query(sql_det)
        rows = [dict(zip(r_det.column_names, row)) for row in r_det.result_rows]

        # Normalización: fechas a string, mapeo de país visible
        for row in rows:
            fc = row.get('creacion_cuenta')
            if hasattr(fc, 'isoformat'):
                row['creacion_cuenta'] = str(fc)[:19]
            row['pais_nombre'] = {
                'CO':'Colombia','MX':'México','NI':'Nicaragua','GT':'Guatemala',
                'PE':'Perú','EC':'Ecuador'
            }.get(row.get('pais_codigo',''), row.get('pais_codigo',''))

        muestra_truncada = total_real > len(rows)

        return jsonify(limpiar_nan({
            "ok": True,
            "desde": desde, "hasta": hasta, "pais": pais,
            "resumen": {
                "total":        total_real,
                "alertas":      alertas_real,
                "ok":           ok_real,
                "pct_alertas":  round(alertas_real / total_real * 100, 1) if total_real else 0,
                "pct_ok":       round(ok_real     / total_real * 100, 1) if total_real else 0,
                "muestra_size":     len(rows),
                "muestra_truncada": muestra_truncada,
            },
            "trend":   trend,
            "alertas": rows,
        }))
    except Exception as e:
        import traceback
        return jsonify({
            "ok": False,
            "error": str(e),
            "detalle": traceback.format_exc()[:500],
        }), 500


@app.route("/api/cedula-alertas/exportar")
def cedula_alertas_exportar():
    """Exporta las alertas (todas, sin filtros client-side) a Excel."""
    # Auth: aceptar token en header o query param (para descargas directas)
    token = request.headers.get("X-Token", "") or request.args.get("token", "")
    sesion = _verificar_sesion(None, token)
    if not sesion:
        return jsonify({"ok": False, "error": "Sesión expirada"}), 401

    desde = request.args.get("desde") or (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    hasta = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais  = request.args.get("pais", "").strip()
    filtro_pais = _cedula_filtro_pais(pais)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ch  = get_client()
        sql = _Q_CEDULA.format(desde=desde, hasta=hasta, filtro_pais=filtro_pais,
                               limit_filas=20000)
        r = ch.query(sql)
        rows = [dict(zip(r.column_names, row)) for row in r.result_rows]

        wb  = Workbook()
        ws  = wb.active
        ws.title = "Alertas Cédula"

        # Cabecera
        headers = [
            "Creación cuenta", "ID Usuario", "Nombre", "País",
            "CC Rekognition", "CC Antecedentes", "Nombre antecedentes", "Resultado"
        ]
        purple_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.cell(row=1, column=1, value=f"Alertas de Cédula  ·  {desde} → {hasta}  ·  {pais or 'Todos los países'}")
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = purple_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 28

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = purple_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        # Filas
        for i, row in enumerate(rows, start=3):
            fc = row.get('creacion_cuenta')
            fc_str = str(fc)[:19] if fc else ''
            pais_iso = row.get('pais_codigo', '')
            pais_nombre = {
                'CO':'Colombia','MX':'México','NI':'Nicaragua','GT':'Guatemala',
                'PE':'Perú','EC':'Ecuador'
            }.get(pais_iso, pais_iso)
            cc_igual = row.get('cc_igual', 'ok')
            datos = [
                fc_str,
                row.get('id_user', ''),
                row.get('name_user', ''),
                pais_nombre,
                row.get('rekognition_cc', ''),
                row.get('cc_antecedentes', ''),
                row.get('nombre_antecedentes', ''),
                cc_igual.upper(),
            ]
            for col, val in enumerate(datos, start=1):
                c = ws.cell(row=i, column=col, value=val)
                c.border = border
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Color por resultado
            res_cell = ws.cell(row=i, column=len(headers))
            if cc_igual == 'alerta':
                res_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                res_cell.font = Font(bold=True, color="991B1B")
            else:
                res_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                res_cell.font = Font(bold=True, color="166534")

        # Anchos
        widths = [22, 30, 28, 12, 18, 18, 30, 12]
        for col, w in enumerate(widths, start=1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = ws.cell(row=3, column=1)

        # Salida en memoria
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"alerta_cedula_{desde}_{hasta}_{ts}.xlsx",
        )
    except Exception as e:
        import traceback
        return jsonify({"ok": False, "error": str(e), "detalle": traceback.format_exc()[:500]}), 500


# ══════════════════════════════════════════════════════════════
# MÓDULO PAGOS — Tarjeta de Crédito + PromoCode
# Clasificación GPS igual que el script Python:
#   OK           → driver recibió pago wallet
#   Mala práctica → pagado=0 AND geoDistance(cancel→dest) ≤ radio país
#   Fraude        → pagado=0 AND (sin GPS OR geoDistance > radio)
# Radio: Colombia 450m | México/Nicaragua 280m | resto 450m
# Deduplicación: ROW_NUMBER() OVER (PARTITION BY b._id)
# ══════════════════════════════════════════════════════════════

def _pagos_filtro(pais, ciudad):
    PAIS_A_ISO = {
        "Colombia":"CO","Mexico":"MX","Nicaragua":"NI",
        "Guatemala":"GT","Peru":"PE","Ecuador":"EC"
    }
    parts = []
    if pais:
        iso = PAIS_A_ISO.get(pais, pais)
        parts.append(f"AND b.g_country=\'{iso}\'")
    if ciudad:
        parts.append(f"AND b.g_adm_area_lv_1=\'{ciudad}\'")
    return " ".join(parts)

# ── Expresiones GPS (igual que BASE_CTE del evasion module) ──
_CANCEL_LON_EXPR = r"toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd\":26.*?coordinates\":\[\s*([+-]?\d+\.\d+)'))"
_CANCEL_LAT_EXPR = r"toFloat64OrNull(extract(ifNull(b.events,''), 'event_cd\":26.*?coordinates\":\[.*?,\s*([+-]?\d+\.\d+)'))"
_END_LON_EXPR    = "toFloat64(JSONExtractString(b.end_geojson,'coordinates',1))"
_END_LAT_EXPR    = "toFloat64(JSONExtractString(b.end_geojson,'coordinates',2))"
_RADIO_EXPR      = "multiIf(b.g_country='CO',450,b.g_country IN ('MX','NI'),280,450)"

# ── CTE reutilizable TC base — deduplica + pre-computa GPS ────
_TC_BASE_CTE = """
WITH b_raw AS (
    SELECT
        b._id AS booking_id,
        b.driver_id,
        b.passenger_id,
        b.status_cd,
        b.g_country,
        b.g_adm_area_lv_1  AS ciudad_raw,
        toDate(toTimeZone(b.created_at,'America/Bogota')) AS fecha,
        toInt64(JSONExtractFloat(b.final_cost,'cents'))/100 AS monto,
        {cancel_lon} AS cancel_lon,
        {cancel_lat} AS cancel_lat,
        {end_lon} AS end_lon,
        {end_lat} AS end_lat,
        {radio}   AS radio,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    WHERE b.payment_method_cd='3'
      AND b.verification_required='false'
      AND b.status_cd IN(4,107,108)
      AND b.created_at>=toDateTime('{{desde}} 00:00:00')
      AND b.created_at<=toDateTime('{{hasta}} 23:59:59')
      AND toInt64(JSONExtractFloat(b.final_cost,'cents'))>0
      AND b.events IS NOT NULL
      AND b.end_geojson IS NOT NULL
      {{filtro}}
),
b AS (SELECT * FROM b_raw WHERE rn=1),
pd AS (
    SELECT booking_id,
           sum(toInt64(JSONExtractFloat(amount,'cents'))) AS pagado
    FROM picapmongoprod.wallet_account_transactions
    WHERE _type='WalletAccountTransactionBookingDriverPayment'
      AND toInt64(JSONExtractFloat(amount,'cents'))>0
      AND created_at>=toDateTime('{{desde}} 00:00:00')
      AND created_at<=toDateTime('{{hasta}} 23:59:59')
    GROUP BY booking_id
)""".format(
    cancel_lon=_CANCEL_LON_EXPR,
    cancel_lat=_CANCEL_LAT_EXPR,
    end_lon=_END_LON_EXPR,
    end_lat=_END_LAT_EXPR,
    radio=_RADIO_EXPR,
)

# Expresiones de clasificación sobre campos pre-computados del CTE b
_TC_CLS = """
    countIf(coalesce(pd.pagado,0)>0) AS ok,
    countIf(coalesce(pd.pagado,0)=0
        AND b.cancel_lon IS NOT NULL AND b.cancel_lat IS NOT NULL
        AND geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)<=b.radio) AS mala_practica,
    countIf(coalesce(pd.pagado,0)=0
        AND (b.cancel_lon IS NULL OR b.cancel_lat IS NULL
             OR geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)>b.radio)) AS fraude"""

_TC_MONTO = """
    round(sumIf(b.monto, coalesce(pd.pagado,0)=0
        AND b.cancel_lon IS NOT NULL AND b.cancel_lat IS NOT NULL
        AND geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)<=b.radio),0) AS monto_mp,
    round(sumIf(b.monto, coalesce(pd.pagado,0)=0
        AND (b.cancel_lon IS NULL OR b.cancel_lat IS NULL
             OR geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)>b.radio)),0) AS monto_fraude,
    round(sum(b.monto),0) AS monto_total"""

_TC_JOIN = """FROM b
LEFT JOIN pd ON b.booking_id=pd.booking_id"""

_TC_WHERE_EXTRA = ""  # bookings ya filtrados en CTE

_PAIS_CASE = "CASE b.g_country WHEN 'CO' THEN 'Colombia' WHEN 'MX' THEN 'Mexico' WHEN 'NI' THEN 'Nicaragua' WHEN 'GT' THEN 'Guatemala' ELSE b.g_country END"
_CIUDAD_CASE = "multiIf(b.ciudad_raw='' OR b.ciudad_raw IS NULL,'Sin ciudad',b.ciudad_raw='MN','Managua',b.ciudad_raw='Guatemala Department','Guatemala',b.ciudad_raw)"

# ── Queries TC finales ────────────────────────────────────────
Q_TC_KPIS = _TC_BASE_CTE + """
SELECT count() AS total,{cls},{monto}
{join}""".format(cls=_TC_CLS, monto=_TC_MONTO, join=_TC_JOIN)

Q_TC_TREND = _TC_BASE_CTE + """
SELECT b.fecha AS fecha,{cls}
{join}
GROUP BY b.fecha ORDER BY b.fecha""".format(cls=_TC_CLS, join=_TC_JOIN)

Q_TC_CIUDADES = _TC_BASE_CTE + """
SELECT
    {ciudad} AS ciudad,
    {pais}   AS pais,
    count() AS total,{cls}
{join}
GROUP BY ciudad,pais ORDER BY total DESC LIMIT 10""".format(
    ciudad=_CIUDAD_CASE, pais=_PAIS_CASE, cls=_TC_CLS, join=_TC_JOIN)

Q_TC_DUO = _TC_BASE_CTE + """
SELECT
    b.driver_id, b.passenger_id,
    count() AS servicios,
    round(sum(b.monto),0) AS monto_total,
    countIf(coalesce(pd.pagado,0)=0
        AND (b.cancel_lon IS NULL OR b.cancel_lat IS NULL
             OR geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)>b.radio)) AS n_fraude,
    countIf(coalesce(pd.pagado,0)=0
        AND b.cancel_lon IS NOT NULL AND b.cancel_lat IS NOT NULL
        AND geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)<=b.radio) AS n_mp
{join}
WHERE coalesce(pd.pagado,0)=0
GROUP BY b.driver_id,b.passenger_id
HAVING servicios>=2
ORDER BY servicios DESC,monto_total DESC LIMIT 20""".format(join=_TC_JOIN)

# ── CTE reutilizable Promo — deduplica + pre-computa GPS ──────
_PROMO_BASE_CTE = """
WITH promo AS (
    SELECT DISTINCT booking_id
    FROM picapmongoprod.wallet_account_transactions
    WHERE _type IN(
        'WalletAccountTransactionPromoCodeMultipleUse',
        'WalletAccountTransactionPromoCodeReferral',
        'WalletAccountTransactionExpirePromoBalance'
    )
    AND created_at>=toDateTime('{{desde}} 00:00:00')
    AND created_at<=toDateTime('{{hasta}} 23:59:59')
),
b_raw AS (
    SELECT
        b._id AS booking_id,
        b.driver_id,
        b.passenger_id,
        b.status_cd,
        b.g_country,
        b.g_adm_area_lv_1  AS ciudad_raw,
        toDate(toTimeZone(b.created_at,'America/Bogota')) AS fecha,
        toInt64(JSONExtractFloat(b.final_cost,'cents'))/100 AS monto,
        {cancel_lon} AS cancel_lon,
        {cancel_lat} AS cancel_lat,
        {end_lon} AS end_lon,
        {end_lat} AS end_lat,
        {radio}   AS radio,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    INNER JOIN promo ON b._id=promo.booking_id
    WHERE b.status_cd IN(4,107,108)
      AND b.created_at>=toDateTime('{{desde}} 00:00:00')
      AND b.created_at<=toDateTime('{{hasta}} 23:59:59')
      AND toInt64(JSONExtractFloat(b.final_cost,'cents'))>0
      AND b.events IS NOT NULL
      AND b.end_geojson IS NOT NULL
      {{filtro}}
),
b AS (SELECT * FROM b_raw WHERE rn=1),
pd AS (
    SELECT booking_id,
           sum(toInt64(JSONExtractFloat(amount,'cents'))) AS pagado
    FROM picapmongoprod.wallet_account_transactions
    WHERE _type='WalletAccountTransactionBookingDriverPayment'
      AND toInt64(JSONExtractFloat(amount,'cents'))>0
      AND created_at>=toDateTime('{{desde}} 00:00:00')
      AND created_at<=toDateTime('{{hasta}} 23:59:59')
    GROUP BY booking_id
)""".format(
    cancel_lon=_CANCEL_LON_EXPR,
    cancel_lat=_CANCEL_LAT_EXPR,
    end_lon=_END_LON_EXPR,
    end_lat=_END_LAT_EXPR,
    radio=_RADIO_EXPR,
)

# Clasificación promo idéntica a TC (mismos campos en CTE)
_PROMO_CLS  = _TC_CLS
_PROMO_MONTO = _TC_MONTO
_PROMO_JOIN = _TC_JOIN

# ── Queries Promo finales ─────────────────────────────────────
Q_PROMO_KPIS = _PROMO_BASE_CTE + """
SELECT count() AS total,{cls},{monto}
{join}""".format(cls=_PROMO_CLS, monto=_PROMO_MONTO, join=_PROMO_JOIN)

Q_PROMO_TREND = _PROMO_BASE_CTE + """
SELECT b.fecha AS fecha,{cls}
{join}
GROUP BY b.fecha ORDER BY b.fecha""".format(cls=_PROMO_CLS, join=_PROMO_JOIN)

Q_PROMO_CIUDADES = _PROMO_BASE_CTE + """
SELECT
    {ciudad} AS ciudad,
    {pais}   AS pais,
    count() AS total,{cls}
{join}
GROUP BY ciudad,pais ORDER BY total DESC LIMIT 10""".format(
    ciudad=_CIUDAD_CASE, pais=_PAIS_CASE, cls=_PROMO_CLS, join=_PROMO_JOIN)

Q_PROMO_DUO = _PROMO_BASE_CTE + """
SELECT
    b.driver_id, b.passenger_id,
    count() AS servicios,
    round(sum(b.monto),0) AS monto_total,
    countIf(coalesce(pd.pagado,0)=0
        AND (b.cancel_lon IS NULL OR b.cancel_lat IS NULL
             OR geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)>b.radio)) AS n_fraude,
    countIf(coalesce(pd.pagado,0)=0
        AND b.cancel_lon IS NOT NULL AND b.cancel_lat IS NOT NULL
        AND geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)<=b.radio) AS n_mp
{join}
WHERE coalesce(pd.pagado,0)=0
GROUP BY b.driver_id,b.passenger_id
HAVING servicios>=2
ORDER BY servicios DESC,monto_total DESC LIMIT 20""".format(join=_PROMO_JOIN)


def _exec_pagos_queries(ch, query_map, desde, hasta, pais, ciudad):
    filtro = _pagos_filtro(pais, ciudad)
    out = {}
    for key, q in query_map.items():
        try:
            r = ch.query(q.format(desde=desde, hasta=hasta, filtro=filtro))
            out[key] = [dict(zip(r.column_names, row)) for row in r.result_rows]
        except Exception as e:
            out[key] = []
            print(f"[pagos/{key}] ERROR: {e}")
    return out

def _kpis_row(data):
    row = data.get("kpis", [{}])[0] if data.get("kpis") else {}
    return {
        "total":         int(row.get("total", 0)),
        "ok":            int(row.get("ok", 0)),
        "mala_practica": int(row.get("mala_practica", 0)),
        "fraude":        int(row.get("fraude", 0)),
        "monto_mp":      float(row.get("monto_mp", 0) or 0),
        "monto_fraude":  float(row.get("monto_fraude", 0) or 0),
        "monto_total":   float(row.get("monto_total", 0) or 0),
    }

def _trend_rows(data):
    return [{"fecha": str(r["fecha"])[:10],
             "ok":           int(r.get("ok", 0)),
             "mala_practica":int(r.get("mala_practica", 0)),
             "fraude":       int(r.get("fraude", 0))}
            for r in data.get("trend", [])]

def _ciudad_rows(data):
    return [{"ciudad":        r.get("ciudad", ""),
             "pais":          r.get("pais", ""),
             "total":         int(r.get("total", 0)),
             "mala_practica": int(r.get("mala_practica", 0)),
             "fraude":        int(r.get("fraude", 0))}
            for r in data.get("ciudades", [])]

def _duo_rows(data):
    return [{"driver_id":   r.get("driver_id", ""),
             "passenger_id":r.get("passenger_id", ""),
             "servicios":   int(r.get("servicios", 0)),
             "monto_total": float(r.get("monto_total", 0) or 0),
             "n_fraude":    int(r.get("n_fraude", 0)),
             "n_mp":        int(r.get("n_mp", 0))}
            for r in data.get("duo", [])]

@app.route("/api/pagos/tc")
def pagos_tc():
    desde  = request.args.get("desde") or (date.today()-timedelta(days=14)).strftime("%Y-%m-%d")
    hasta  = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais   = request.args.get("pais", "")
    ciudad = request.args.get("ciudad", "")
    try:
        ch   = get_client()
        data = _exec_pagos_queries(ch, {
            "kpis":    Q_TC_KPIS,
            "trend":   Q_TC_TREND,
            "ciudades":Q_TC_CIUDADES,
            "duo":     Q_TC_DUO,
        }, desde, hasta, pais, ciudad)
        return jsonify(limpiar_nan({
            "desde": desde, "hasta": hasta,
            "kpis":    _kpis_row(data),
            "trend":   _trend_rows(data),
            "ciudades":_ciudad_rows(data),
            "duo":     _duo_rows(data),
        }))
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500

@app.route("/api/pagos/promo")
def pagos_promo():
    desde  = request.args.get("desde") or (date.today()-timedelta(days=14)).strftime("%Y-%m-%d")
    hasta  = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais   = request.args.get("pais", "")
    ciudad = request.args.get("ciudad", "")
    try:
        ch   = get_client()
        data = _exec_pagos_queries(ch, {
            "kpis":    Q_PROMO_KPIS,
            "trend":   Q_PROMO_TREND,
            "ciudades":Q_PROMO_CIUDADES,
            "duo":     Q_PROMO_DUO,
        }, desde, hasta, pais, ciudad)
        return jsonify(limpiar_nan({
            "desde": desde, "hasta": hasta,
            "kpis":    _kpis_row(data),
            "trend":   _trend_rows(data),
            "ciudades":_ciudad_rows(data),
            "duo":     _duo_rows(data),
        }))
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500

# ══════════════════════════════════════════════════════════════
# MÓDULO SERVICIOS ESTAFA
# Query optimizada: todos los CTEs filtran por fecha para evitar
# full scans en bookings/passengers/sessions
# Clasificación: ESTAFA | OK   (sólo dos categorías)
#
# Nota: la categoría intermedia "POSIBLE_ESTAFA" fue retirada porque
# generaba falsos positivos: palabras como "celular", "número", "envío"
# o nombres de bancos aparecían en denuncias normales (no fraudulentas).
# Solo cuentan como estafa los términos que apuntan directamente al
# patrón financiero conocido (KW_ESTAFA). Todo lo demás es OK.
# ══════════════════════════════════════════════════════════════

# ── Palabras clave ESTAFA (del query original) ────────────────
KW_ESTAFA = [
    'abono','administracion','administración','bancaria','bancario',
    'bono','cajero multifuncional','compra','con base','con una base',
    'conbase','convenio','copago','cuota moderadora','datafono',
    'despacho','disponibilidad','disponible','económica','farmaceutica',
    'farmacéutica','farmacéutico','farmacia','fotocopia','fotocopias',
    'gratifica','gratificación','insulina','mande','multifuncional',
    'multifuncional de bogotá','multivitaminicos','orden medica','picap',
    'plante','reembolsado','sancion','sanción','serio',
    'servicio al cliente','soporte picap','soporte tecnico de picap',
    'soporte técnico de picap','soporte tecnico pibox',
    'sporte tecnico de pibox','tirilla','transaccion','transacción',
    'transfiera el dinero','transfiere','vase','wasab',
]

# Lista mantenida solo por compatibilidad/debug; YA NO se usa para clasificar.
# Si en el futuro alguien quiere reactivarla, debería ser un panel separado de
# "señales informativas", no una categoría de alerta.
KW_POSIBLE = []  # vacía intencionalmente; ver nota arriba.

def _detectar_palabras(texto):
    """Clasifica un texto como ESTAFA u OK según KW_ESTAFA.
    No genera 'POSIBLE_ESTAFA': aquello que antes caía allí ahora es OK."""
    if not texto or str(texto).strip() in ('','None','null'):
        return 'OK', []
    t = texto.lower()
    hits_estafa = [kw for kw in KW_ESTAFA if kw in t]
    if hits_estafa:
        return 'ESTAFA', hits_estafa
    return 'OK', []

def _kw_sql_array(kws):
    """Convierte una lista de keywords Python en un literal de array SQL para
    ClickHouse, escapando comillas simples. Usado en multiSearchAnyCaseInsensitive."""
    return "[" + ",".join("'" + kw.replace("'", "''") + "'" for kw in kws) + "]"

# Lista lista para inyectar en SQL — debe quedar idéntica a KW_ESTAFA.
_KW_ESTAFA_SQL = _kw_sql_array(KW_ESTAFA)

def _estafa_filtro_pais(pais):
    """Retorna cláusula AND para filtrar por país en bookings (g_country = ISO)."""
    PAIS_A_ISO = {
        "Colombia":"CO","Mexico":"MX","Nicaragua":"NI",
        "Guatemala":"GT","Peru":"PE","Ecuador":"EC"
    }
    if not pais:
        return ""
    iso = PAIS_A_ISO.get(pais, pais)
    return f"AND b.g_country = \'{iso}\'"

# ── Query optimizada ─────────────────────────────────────────
# Todos los CTEs filtran por fecha: sin full table scans
_Q_ESTAFA_BASE = """
WITH
-- ID del tipo de servicio "Mensajería" (donde mas se ven estafas reales segun
-- revision manual). Lo incluimos como vía adicional incluso si el booking NO
-- se canceló por denuncia, porque a menudo el usuario paga sin denunciar.
mensajeria_type AS (SELECT '5c71b03a58b9ba10fa6393cf' AS id),

-- CTE 1: bookings del período. Dos vías:
--   A) Cancelados por denuncia (21) o seguridad (13)  +  evento 26 confirmado
--   B) Servicios de mensajería (cualquier estado) cuyas indicaciones traen
--      al menos una palabra del patrón conocido — pre-filtro en SQL para
--      no traer millones de filas limpias.
bk_raw AS (
    SELECT
        b._id,
        b.driver_id,
        b.passenger_id,
        b.company_id,
        toString(b.requested_service_type_id) AS service_type_id,
        toTimeZone(coalesce(b.scheduled_at, b.created_at), 'America/Bogota') AS fecha_servicio,
        CASE toInt64OrZero(b.cancelation_reason_cd)
            WHEN 21 THEN 'user_denounce'
            WHEN 13 THEN 'Security_issues'
            WHEN 0  THEN 'completed/active'
            ELSE concat('other_cancel_', toString(toInt64OrZero(b.cancelation_reason_cd)))
        END AS cancelation_reason,
        JSONExtractInt(
            arrayFirst(
                x -> JSONExtractInt(x, 'event_cd') = 26,
                JSONExtract(ifNull(b.events, '[]'), 'Array(String)')
            ),
            'event_cd'
        ) AS estado_cancelacion,
        toInt64OrZero(b.cancelation_reason_cd) AS cancel_reason_cd,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    WHERE (
        -- Vía A: denuncia / seguridad
        toInt64OrZero(b.cancelation_reason_cd) IN (21, 13)
        OR
        -- Vía B: mensajería (cualquier estado)
        toString(b.requested_service_type_id) = (SELECT id FROM mensajeria_type)
    )
      AND b.created_at >= toDateTime('{desde} 00:00:00')
      AND b.created_at <= toDateTime('{hasta} 23:59:59')
      {filtro_pais}
),
-- CTE 2: aplicar la regla específica por vía
--   Vía A (denuncia/seguridad)  → exigimos evento 26 confirmado
--   Vía B (mensajería)          → no necesita evento 26
bk AS (
    SELECT * FROM bk_raw
    WHERE rn = 1
      AND (
          (cancel_reason_cd IN (21, 13) AND estado_cancelacion = 26)
          OR
          service_type_id = (SELECT id FROM mensajeria_type)
      )
),
-- CTE 3: pasajeros — solo los que aparecen en bk (no full scan)
pax AS (
    SELECT
        p._id AS id_user,
        p.name AS name_user,
        toString(p.is_driver_suspended) AS status_driver_suspend,
        toString(p.suspended)           AS status_user_suspend,
        toString(p.expelled)            AS status_expelled,
        p.g_country AS pais,
        p.g_adm_area_lv_1 AS departamento,
        p.g_adm_area_lv_2 AS city
    FROM picapmongoprod.passengers p
    WHERE p._id IN (SELECT passenger_id FROM bk)
),
-- CTE 4: sesión más reciente por pasajero — solo de usuarios en bk
sess AS (
    SELECT
        passenger_id AS id_user,
        imei,
        active AS status_imei,
        ROW_NUMBER() OVER (PARTITION BY passenger_id ORDER BY created_at DESC) AS rn
    FROM picapmongoprod.sessions
    WHERE passenger_id IN (SELECT passenger_id FROM bk)
    QUALIFY rn = 1
)
-- Query principal
-- Para los servicios de mensajería (vía B) que NO fueron denunciados, exigimos
-- que las indicaciones contengan al menos una palabra del patrón conocido.
-- Así evitamos traer cientos de miles de mensajerías limpias.
SELECT
    bk.fecha_servicio,
    bk._id             AS booking_id,
    bk.driver_id,
    bk.passenger_id    AS user_id,
    bk.service_type_id,
    bk.cancel_reason_cd,
    c.name             AS name_company,
    pax.name_user,
    pax.status_driver_suspend,
    pax.status_user_suspend,
    pax.status_expelled,
    sess.status_imei,
    sess.imei          AS imei_sesion,
    pax.pais,
    pax.departamento,
    pax.city,
    bk.cancelation_reason,
    pd.indications
FROM bk
LEFT JOIN pax  ON bk.passenger_id = pax.id_user
LEFT JOIN sess ON bk.passenger_id = sess.id_user
LEFT JOIN picapmongoprod.companies c ON bk.company_id = c._id
INNER JOIN picapmongoprod.packages pd ON pd.booking_id = bk._id
WHERE (
    bk.cancel_reason_cd IN (21, 13)
    OR multiSearchAnyCaseInsensitive(coalesce(pd.indications, ''), {kws_estafa}) > 0
)
ORDER BY bk.fecha_servicio DESC
LIMIT 1 BY bk._id   -- ClickHouse: una fila por booking (evita fan-out de JOINs)
LIMIT {limit_filas}
"""

# Query de agregación: cuenta y desagrega por día sin LIMIT — entrega los
# totales VERDADEROS aunque el detalle se trunque para no colapsar la red.
_Q_ESTAFA_AGREGADO = """
WITH
mensajeria_type AS (SELECT '5c71b03a58b9ba10fa6393cf' AS id),
bk_raw AS (
    SELECT
        b._id,
        b.passenger_id,
        toTimeZone(coalesce(b.scheduled_at, b.created_at), 'America/Bogota') AS fecha_servicio,
        toString(b.requested_service_type_id) AS service_type_id,
        toInt64OrZero(b.cancelation_reason_cd) AS cancel_reason_cd,
        JSONExtractInt(
            arrayFirst(
                x -> JSONExtractInt(x, 'event_cd') = 26,
                JSONExtract(ifNull(b.events, '[]'), 'Array(String)')
            ),
            'event_cd'
        ) AS estado_cancelacion,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    WHERE (
        toInt64OrZero(b.cancelation_reason_cd) IN (21, 13)
        OR toString(b.requested_service_type_id) = (SELECT id FROM mensajeria_type)
    )
      AND b.created_at >= toDateTime('{desde} 00:00:00')
      AND b.created_at <= toDateTime('{hasta} 23:59:59')
      {filtro_pais}
),
bk AS (
    SELECT * FROM bk_raw
    WHERE rn = 1
      AND (
          (cancel_reason_cd IN (21, 13) AND estado_cancelacion = 26)
          OR service_type_id = (SELECT id FROM mensajeria_type)
      )
)
SELECT
    dia,
    count() AS total_dia,
    countIf(has_kw) AS estafa_dia
FROM (
    -- Subquery: una fila por booking_id (evita fan-out por packages duplicados)
    SELECT
        bk._id AS booking_id,
        toDate(bk.fecha_servicio) AS dia,
        multiSearchAnyCaseInsensitive(coalesce(pd.indications, ''), {kws_estafa}) > 0 AS has_kw
    FROM bk
    INNER JOIN picapmongoprod.packages pd ON pd.booking_id = bk._id
    WHERE (
        bk.cancel_reason_cd IN (21, 13)
        OR multiSearchAnyCaseInsensitive(coalesce(pd.indications, ''), {kws_estafa}) > 0
    )
    LIMIT 1 BY bk._id
)
GROUP BY dia
ORDER BY dia
"""

# Conteo de cuentas (passenger_id) únicas: con al menos un servicio clasificado
# como estafa, vs sin ninguno. Misma estructura de filtros que el agregado por
# día — comparte la definición del universo.
_Q_ESTAFA_CUENTAS = """
WITH
mensajeria_type AS (SELECT '5c71b03a58b9ba10fa6393cf' AS id),
bk_raw AS (
    SELECT
        b._id,
        b.passenger_id,
        toString(b.requested_service_type_id) AS service_type_id,
        toInt64OrZero(b.cancelation_reason_cd) AS cancel_reason_cd,
        JSONExtractInt(
            arrayFirst(
                x -> JSONExtractInt(x, 'event_cd') = 26,
                JSONExtract(ifNull(b.events, '[]'), 'Array(String)')
            ),
            'event_cd'
        ) AS estado_cancelacion,
        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn
    FROM picapmongoprod.bookings b
    WHERE (
        toInt64OrZero(b.cancelation_reason_cd) IN (21, 13)
        OR toString(b.requested_service_type_id) = (SELECT id FROM mensajeria_type)
    )
      AND b.created_at >= toDateTime('{desde} 00:00:00')
      AND b.created_at <= toDateTime('{hasta} 23:59:59')
      {filtro_pais}
),
bk AS (
    SELECT * FROM bk_raw
    WHERE rn = 1
      AND (
          (cancel_reason_cd IN (21, 13) AND estado_cancelacion = 26)
          OR service_type_id = (SELECT id FROM mensajeria_type)
      )
),
-- Una fila por (booking_id, passenger_id) sin fan-out
booking_user AS (
    SELECT
        bk._id AS booking_id,
        bk.passenger_id AS user_id,
        multiSearchAnyCaseInsensitive(coalesce(pd.indications, ''), {kws_estafa}) > 0 AS has_kw
    FROM bk
    INNER JOIN picapmongoprod.packages pd ON pd.booking_id = bk._id
    WHERE (
        bk.cancel_reason_cd IN (21, 13)
        OR multiSearchAnyCaseInsensitive(coalesce(pd.indications, ''), {kws_estafa}) > 0
    )
    LIMIT 1 BY bk._id
),
-- Por usuario: ¿tuvo al menos un servicio con keyword?
por_user AS (
    SELECT
        user_id,
        max(has_kw) AS user_has_estafa
    FROM booking_user
    WHERE user_id != ''
    GROUP BY user_id
)
SELECT
    count()                     AS total_cuentas,
    countIf(user_has_estafa)    AS cuentas_estafa,
    count() - countIf(user_has_estafa) AS cuentas_ok
FROM por_user
"""

def _procesar_fila_estafa(row):
    """Clasifica una fila y retorna el dict enriquecido."""
    indications   = row.get('indications', '') or ''
    clasificacion, kws = _detectar_palabras(indications)
    fs = row.get('fecha_servicio', '')
    fecha_str = str(fs)[:16] if fs else '—'
    return {
        'booking_id':             row.get('booking_id', ''),
        'driver_id':              row.get('driver_id', ''),
        'user_id':                row.get('user_id', ''),
        'name_user':              row.get('name_user', ''),
        'pais':                   row.get('pais', ''),
        'departamento':           row.get('departamento', ''),
        'city':                   row.get('city', ''),
        'fecha_servicio':         fecha_str,
        'cancelation_reason':     row.get('cancelation_reason', ''),
        'status_driver_suspend':  str(row.get('status_driver_suspend', '')),
        'status_user_suspend':    str(row.get('status_user_suspend', '')),
        'status_expelled':        str(row.get('status_expelled', '')),
        'imei_sesion':            row.get('imei_sesion', ''),
        'indications':            indications[:500],
        'clasificacion':          clasificacion,
        'palabras_detectadas':    kws[:10],
    }

@app.route("/api/estafa")
def estafa():
    desde  = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta  = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    pais   = request.args.get("pais", "")
    q_id   = request.args.get("q", "").strip()
    q_tipo = request.args.get("tipo", "booking")

    filtro_pais = _estafa_filtro_pais(pais)
    LIMIT_DETALLE = 5000  # Sample para la tabla; los conteos vienen del agregado.

    try:
        ch = get_client()

        # ── 1) Conteo real (sin LIMIT) y desglose diario ─────────
        sql_agg = _Q_ESTAFA_AGREGADO.format(
            desde=desde, hasta=hasta,
            filtro_pais=filtro_pais,
            kws_estafa=_KW_ESTAFA_SQL,
        )
        agg = ch.query(sql_agg)
        agg_cols = agg.column_names
        agg_rows = [dict(zip(agg_cols, r)) for r in agg.result_rows]

        total_real    = sum(int(r.get('total_dia', 0) or 0)  for r in agg_rows)
        estafa_real   = sum(int(r.get('estafa_dia', 0) or 0) for r in agg_rows)
        ok_real       = total_real - estafa_real

        # Trend diario directamente del agregado
        trend = [
            {
                'fecha':  str(r['dia']),
                'estafa': int(r.get('estafa_dia', 0) or 0),
                'ok':     int(r.get('total_dia', 0) or 0) - int(r.get('estafa_dia', 0) or 0),
            }
            for r in agg_rows
        ]

        # ── 1.5) Conteo de cuentas únicas (con/sin estafa) ────────
        try:
            sql_cuentas = _Q_ESTAFA_CUENTAS.format(
                desde=desde, hasta=hasta,
                filtro_pais=filtro_pais,
                kws_estafa=_KW_ESTAFA_SQL,
            )
            r_cu = ch.query(sql_cuentas)
            if r_cu.result_rows:
                _row = dict(zip(r_cu.column_names, r_cu.result_rows[0]))
                total_cuentas   = int(_row.get('total_cuentas', 0)   or 0)
                cuentas_estafa  = int(_row.get('cuentas_estafa', 0)  or 0)
                cuentas_ok      = int(_row.get('cuentas_ok', 0)      or 0)
            else:
                total_cuentas = cuentas_estafa = cuentas_ok = 0
        except Exception as _ce:
            # Si la query de cuentas falla, no bloqueamos el resto del endpoint.
            print(f"[estafa] error en query de cuentas: {_ce}")
            total_cuentas = cuentas_estafa = cuentas_ok = 0

        # ── 2) Sample detallado (LIMIT) para la tabla y top palabras ──
        sql_det = _Q_ESTAFA_BASE.format(
            desde=desde, hasta=hasta,
            filtro_pais=filtro_pais,
            kws_estafa=_KW_ESTAFA_SQL,
            limit_filas=LIMIT_DETALLE,
        )
        r_det    = ch.query(sql_det)
        rows_raw = [dict(zip(r_det.column_names, row)) for row in r_det.result_rows]

        # Dedup por booking_id: las tablas passengers/companies pueden repetir
        # registros (no son ReplacingMergeTree con FINAL aquí), causando fan-out
        # en los LEFT JOIN. Conservamos la primera ocurrencia.
        _seen = set()
        rows_dedup = []
        for row in rows_raw:
            bid = row.get('booking_id')
            if bid and bid in _seen:
                continue
            if bid:
                _seen.add(bid)
            rows_dedup.append(row)
        rows = [_procesar_fila_estafa(row) for row in rows_dedup]

        # Filtro por ID (client-side, sobre el sample)
        if q_id and len(q_id) >= 4:
            q_low = q_id.lower()
            campo = {'booking':'booking_id','driver':'driver_id','user':'user_id'}.get(q_tipo,'booking_id')
            rows  = [row for row in rows if q_low in (row.get(campo,'') or '').lower()]

        # Top palabras clave: del sample (suficiente para visión cualitativa)
        from collections import Counter
        kw_counter = Counter()
        for row in rows:
            for kw in row.get('palabras_detectadas', []):
                kw_counter[kw] += 1
        top_kw = [{'kw': k, 'count': v} for k, v in kw_counter.most_common(15)]

        muestra_truncada = total_real > len(rows)

        return jsonify(limpiar_nan({
            'desde': desde, 'hasta': hasta,
            'resumen': {
                'total':       total_real,
                'estafa':      estafa_real,
                'ok':          ok_real,
                'pct_estafa':  round(estafa_real / total_real * 100, 1) if total_real else 0,
                'pct_ok':      round(ok_real     / total_real * 100, 1) if total_real else 0,
                # Cuentas (passenger_id) únicas: cuántas distintas están detrás
                # de los servicios analizados, y cuántas tuvieron al menos un
                # servicio clasificado como estafa.
                'total_cuentas':   total_cuentas,
                'cuentas_estafa':  cuentas_estafa,
                'cuentas_ok':      cuentas_ok,
                'pct_cuentas_estafa': round(cuentas_estafa / total_cuentas * 100, 1) if total_cuentas else 0,
                'muestra_size':      len(rows),
                'muestra_truncada':  muestra_truncada,
            },
            'trend':   trend,
            'top_kw':  top_kw,
            'alertas': rows,
        }))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


# ══════════════════════════════════════════════════════════════
# MÓDULO VALIDACIÓN DE RECAUDOS
# Basado en WalletAccountCounterDeliveryTransaction
# Clasificación:
#   Está Correcto       → balance_neto = 0
#   Se pagó demás       → balance_neto > 0
#   Se debe dinero      → balance_neto < 0
#   Revisar             → balance_neto IS NULL o inconsistencia
# ══════════════════════════════════════════════════════════════

_Q_RECAUDOS = """
WITH deduplicated AS (
    SELECT
        toTimeZone(wat.created_at, 'America/Bogota') AS fecha_tx,
        wat.booking_id AS id_booking,
        wat._id,
        wat._type AS tipo_tx,
        JSONExtractString(wat.amount, 'currency_iso') AS moneda,
        toFloat64OrNull(JSONExtractString(wat.amount, 'cents')) / 100 AS valor,
        ROW_NUMBER() OVER (PARTITION BY wat._id ORDER BY wat.created_at DESC) AS rn
    FROM picapmongoprod.wallet_account_transactions wat
    WHERE wat._type = 'WalletAccountCounterDeliveryTransaction'
      AND wat.created_at >= toDateTime('{desde} 00:00:00')
      AND wat.created_at <= toDateTime('{hasta} 23:59:59')
      {filtro_moneda}
),
base AS (
    SELECT fecha_tx, id_booking, _id, tipo_tx, moneda, valor
    FROM deduplicated
    WHERE rn = 1
),
bookings_en_rango AS (
    SELECT DISTINCT id_booking
    FROM base
    WHERE fecha_tx >= toDateTime('{desde} 00:00:00')
      AND fecha_tx <= toDateTime('{hasta} 23:59:59')
),
agregado AS (
    SELECT
        b.id_booking,
        any(b.fecha_tx)                                              AS fecha_tx,
        b.tipo_tx,
        b.moneda,
        sumIf(b.valor, b.valor < 0)                                  AS suma_negativos,
        sumIf(b.valor, b.valor > 0)                                  AS suma_positivos,
        sumIf(b.valor, b.valor > 0) + sumIf(b.valor, b.valor < 0)   AS balance_neto,
        countIf(b.valor < 0)                                         AS cnt_negativos,
        countIf(b.valor > 0)                                         AS cnt_positivos,
        count()                                                       AS cnt_total
    FROM base b
    INNER JOIN bookings_en_rango br ON b.id_booking = br.id_booking
    GROUP BY b.id_booking, b.tipo_tx, b.moneda
)
SELECT
    id_booking,
    toString(fecha_tx) AS fecha_tx,
    tipo_tx,
    moneda,
    round(suma_negativos, 2)  AS suma_negativos,
    round(suma_positivos, 2)  AS suma_positivos,
    round(balance_neto, 2)    AS balance_neto,
    cnt_negativos,
    cnt_positivos,
    cnt_total,
    CASE
        WHEN balance_neto IS NULL OR (cnt_positivos > 0 AND cnt_negativos > 0
             AND abs(suma_positivos + suma_negativos) < 0.01
             AND cnt_total > 2)                          THEN 'Revisar'
        WHEN balance_neto = 0                            THEN 'Correcto'
        WHEN balance_neto > 0                            THEN 'Pagado_demas'
        WHEN balance_neto < 0                            THEN 'Debe_dinero'
        ELSE 'Revisar'
    END AS clasificacion
FROM agregado
ORDER BY abs(balance_neto) DESC
LIMIT 3000
"""

@app.route("/api/recaudos")
def recaudos():
    desde   = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta   = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    moneda  = request.args.get("moneda", "")
    q_id    = request.args.get("q", "").strip()
    q_tipo  = request.args.get("tipo", "booking")

    filtro_moneda = f"AND JSONExtractString(wat.amount,'currency_iso')=\'{moneda}\'" if moneda else ""

    sql = _Q_RECAUDOS.format(
        desde=desde, hasta=hasta, filtro_moneda=filtro_moneda
    )
    try:
        ch   = get_client()
        r    = ch.query(sql)
        rows = [dict(zip(r.column_names, row)) for row in r.result_rows]

        # Convertir tipos para JSON
        for row in rows:
            for k in ['suma_negativos','suma_positivos','balance_neto']:
                try: row[k] = round(float(row[k] or 0), 2)
                except: row[k] = 0.0
            for k in ['cnt_negativos','cnt_positivos','cnt_total']:
                try: row[k] = int(row[k] or 0)
                except: row[k] = 0
            # Fecha legible
            row['fecha_tx'] = str(row.get('fecha_tx',''))[:16]

        # Filtro por ID (client-side)
        if q_id and len(q_id) >= 4:
            campo = {'booking':'id_booking','driver':'driver_id'}.get(q_tipo,'id_booking')
            q_low = q_id.lower()
            rows  = [r for r in rows if q_low in (r.get(campo,'') or '').lower()]

        # ── Estadísticas agregadas ────────────────────────────
        total        = len(rows)
        n_correcto   = sum(1 for r in rows if r['clasificacion'] == 'Correcto')
        n_demas      = sum(1 for r in rows if r['clasificacion'] == 'Pagado_demas')
        n_deuda      = sum(1 for r in rows if r['clasificacion'] == 'Debe_dinero')
        n_revisar    = sum(1 for r in rows if r['clasificacion'] == 'Revisar')

        v_correcto   = sum(abs(r['balance_neto']) for r in rows if r['clasificacion'] == 'Correcto')
        v_demas      = sum(r['balance_neto']      for r in rows if r['clasificacion'] == 'Pagado_demas')
        v_deuda      = sum(abs(r['balance_neto'])  for r in rows if r['clasificacion'] == 'Debe_dinero')
        v_revisar    = sum(abs(r['balance_neto'])  for r in rows if r['clasificacion'] == 'Revisar')

        # Tendencia por día
        trend_map = {}
        for row in rows:
            fecha = (row.get('fecha_tx') or '')[:10]
            if not fecha or fecha == '—': continue
            if fecha not in trend_map:
                trend_map[fecha] = {'correcto':0,'demas':0,'deuda':0,'revisar':0}
            cls = row['clasificacion']
            if   cls == 'Correcto':      trend_map[fecha]['correcto'] += 1
            elif cls == 'Pagado_demas':  trend_map[fecha]['demas']    += 1
            elif cls == 'Debe_dinero':   trend_map[fecha]['deuda']    += 1
            else:                        trend_map[fecha]['revisar']  += 1
        trend = [{'fecha':k, **v} for k,v in sorted(trend_map.items())]

        # Distribución por moneda
        monedas_map = {}
        for row in rows:
            m = row.get('moneda','') or 'N/A'
            if m not in monedas_map:
                monedas_map[m] = {'total':0,'correcto':0,'demas':0,'deuda':0,'revisar':0,
                                  'v_demas':0,'v_deuda':0}
            monedas_map[m]['total'] += 1
            cls = row['clasificacion']
            if   cls == 'Correcto':     monedas_map[m]['correcto'] += 1
            elif cls == 'Pagado_demas': monedas_map[m]['demas']   += 1; monedas_map[m]['v_demas'] += row['balance_neto']
            elif cls == 'Debe_dinero':  monedas_map[m]['deuda']   += 1; monedas_map[m]['v_deuda'] += abs(row['balance_neto'])
            else:                       monedas_map[m]['revisar'] += 1
        por_moneda = [{'moneda':k, **v} for k,v in
                      sorted(monedas_map.items(), key=lambda x:-x[1]['total'])]

        return jsonify(limpiar_nan({
            'desde': desde, 'hasta': hasta,
            'resumen': {
                'total':      total,
                'correcto':   n_correcto,
                'pagado_demas': n_demas,
                'debe_dinero':  n_deuda,
                'revisar':    n_revisar,
                'v_correcto': round(v_correcto, 2),
                'v_demas':    round(v_demas, 2),
                'v_deuda':    round(v_deuda, 2),
                'v_revisar':  round(v_revisar, 2),
                'pct_correcto':  round(n_correcto/total*100,1) if total else 0,
                'pct_demas':     round(n_demas/total*100,1)    if total else 0,
                'pct_deuda':     round(n_deuda/total*100,1)    if total else 0,
                'pct_revisar':   round(n_revisar/total*100,1)  if total else 0,
            },
            'trend':     trend,
            'por_moneda': por_moneda,
            'filas':     rows,  # completo para tabla + alertas
        }))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


# ══════════════════════════════════════════════════════════════
# MÓDULO AUDITORÍAS PIBOX COMERCIAL
# Submódulos: Comisiones + Créditos (Tarifas: estructura futura)
# Filtro anti-test integrado en SQL
# ══════════════════════════════════════════════════════════════

# ── Filtro anti-test (excluye datos de prueba desde la query) ─
_ANTI_TEST_EXPR = """NOT multiSearchAnyCaseInsensitive(
        lowerUTF8(concat(
            ifNull(com.name,''), ' ',
            ifNull(d.name,''), ' ',
            ifNull(com.commercial_manager_id,'')
        )),
        ['test','testeo','pruebas','qa','dummy','sandbox',
         'demo','ejemplo','testing','user test','internal',
         'liliana peña','liliana pena']
    )"""

# ── Flag de calidad de dato (solo monitoreo, no se muestra) ──
_DATA_QUALITY_FLAG = """CASE
        WHEN multiSearchAnyCaseInsensitive(
            lowerUTF8(concat(ifNull(com.name,''),' ',ifNull(d.name,''))),
            ['test','testeo','pruebas','qa','dummy','sandbox',
             'demo','ejemplo','testing','user test','internal',
             'liliana peña','liliana pena']
        ) THEN 'posible_test'
        ELSE 'real'
    END"""

# ── Query base unificada (Comisiones y Créditos usan la misma) ─
_Q_AUDITORIA_BASE = """
WITH bookings_filtered AS (
    SELECT
        company_id,
        created_at,
        _id AS booking_id,
        toFloat64OrNull(JSONExtractString(final_cost,'cents')) / 100 AS final_cost
    FROM picapmongoprod.bookings
    WHERE status_cd = 4                            -- Solo finalizados (Fix 4)
      AND g_country = 'CO'                         -- Solo Colombia (Fix 1)
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
),
last_booking AS (
    SELECT company_id, max(created_at) AS last_created_at
    FROM bookings_filtered
    GROUP BY company_id
),
fare_configs_ext AS (
    SELECT DISTINCT
        _id,
        company_parent_id  AS company_parent_id_extracted,
        service_type_id    AS service_type_id_extracted,
        geo_fence_id       AS geo_fence_id_extracted,
        toFloat64OrNull(JSONExtractString(base_fare,'cents'))         / 100 AS base_fare,
        toFloat64OrNull(JSONExtractString(minimum_fare,'cents'))       / 100 AS minimum_fare,
        toFloat64OrNull(JSONExtractString(distance_fare,'cents'))      / 100 AS distance_fare,
        toFloat64OrNull(JSONExtractString(hour_fare,'cents'))          / 100 AS hour_fare,
        toFloat64OrNull(JSONExtractString(extra_stop_fare,'cents'))    / 100 AS extra_stop_fare,
        toFloat64OrNull(JSONExtractString(package_fare,'cents'))       / 100 AS package_fare,
        toFloat64OrNull(JSONExtractString(hour_base_fare,'cents'))     / 100 AS hour_base_fare,
        toFloat64OrNull(JSONExtractString(hour_standby_fare,'cents'))  / 100 AS hour_standby_fare,
        company_commission_percentage AS utilidad_corporativa
    FROM picapmongoprod.fare_configs FINAL
    {filtro_tarifa}
),
company_auth_ext AS (
    SELECT company_id AS company_id_extracted,
           active     AS active_extracted
    FROM picapmongoprod.company_authorization_logs
),
base AS (
    SELECT
        CASE
            WHEN a.active_extracted = 'false' THEN 'inactivo'
            ELSE 'activo'            -- desconocido → activo (Fix 3)
        END AS estado,
        com._id                                                       AS id_company,
        lb.last_created_at                                            AS last_service,
        com.name                                                      AS linea_de_negocio,
        com.commercial_manager_id                                     AS commercial_manager,
        d.name                                                        AS name_manager,
        fc._id                                                        AS tarifa_id,
        JSONExtractString(tst.name,'es')                              AS type_service,
        if(JSONExtractString(ci.name,'es') IS NULL,'Colombia',
           JSONExtractString(ci.name,'es'))                           AS ciudad,
        JSONExtractString(com.max_wallet_negative,'currency_iso')     AS moneda,
        fc.base_fare, fc.minimum_fare, fc.distance_fare,
        fc.hour_fare, fc.extra_stop_fare, fc.package_fare,
        fc.hour_base_fare, fc.hour_standby_fare,
        toFloat64OrNull(com.commission_percentage)                    AS comission,
        fc.utilidad_corporativa,
        toFloat64OrNull(JSONExtractString(com.max_wallet_negative,'cents')) / 100 AS credit,
        toFloat64OrNull(JSONExtractString(com.max_declared_value,'cents'))  / 100 AS valor_declarado,
        ROW_NUMBER() OVER (PARTITION BY fc._id ORDER BY lb.last_created_at DESC) AS rn
    FROM picapmongoprod.companies com
    INNER JOIN picapmongoprod.countries co ON co._id = com.geo_fence_id
    LEFT JOIN fare_configs_ext fc ON fc.company_parent_id_extracted = com._id
    LEFT JOIN picapmongoprod.service_types tst ON tst._id = fc.service_type_id_extracted
    LEFT JOIN last_booking lb ON com._id = lb.company_id
    LEFT JOIN company_auth_ext a ON com._id = a.company_id_extracted
    LEFT JOIN picapmongoprod.cities ci ON ci._id = fc.geo_fence_id_extracted
    LEFT JOIN picapmongoprod.passengers d ON com.commercial_manager_id = d._id
    WHERE com._type = 'Company'
      AND fc._id IS NOT NULL
      {filtro_company}
      {filtro_moneda}
      {anti_test}
)
SELECT
    estado, id_company, last_service, linea_de_negocio,
    commercial_manager, name_manager, tarifa_id, type_service,
    ciudad, moneda, base_fare, minimum_fare, distance_fare,
    hour_fare, extra_stop_fare, package_fare, hour_base_fare,
    hour_standby_fare, comission, utilidad_corporativa, credit,
    valor_declarado
FROM base
WHERE rn = 1
  {filtro_last_service}
ORDER BY last_service DESC, name_manager
LIMIT 5000
"""

def _aud_filtros(company_id='', tarifa_id='', moneda='', anti_test=True,
                 last_desde='', last_hasta=''):
    """Construye filtros dinámicos para la query de auditoría."""
    fc = f"AND com._id = \'{company_id}\'" if company_id else ""
    ft = f"WHERE _id = \'{tarifa_id}\'" if tarifa_id else ""
    fm = f"AND JSONExtractString(com.max_wallet_negative,'currency_iso') = \'{moneda}\'" if moneda else ""
    at = f"AND {_ANTI_TEST_EXPR}" if anti_test else ""
    # Filtro de último servicio (segundo período)
    fls_parts = []
    if last_desde:
        fls_parts.append(f"AND toTimeZone(last_service,'America/Bogota') >= '{last_desde} 00:00:00'")
    if last_hasta:
        fls_parts.append(f"AND toTimeZone(last_service,'America/Bogota') <= '{last_hasta} 23:59:59'")
    fls = ' '.join(fls_parts)
    return fc, ft, fm, at, fls

# ── Clasificación Comisiones ──────────────────────────────────
def _clasificar_comision(row):
    """Retorna lista de alertas para una fila de comisiones."""
    comision  = float(row.get('comission') or 0)
    utilidad  = float(row.get('utilidad_corporativa') or 0)
    credit    = float(row.get('credit') or 0)
    alertas   = []

    # Fix 2: si todo es 0 no hay inconsistencia real → Correcto
    if comision == 0 and utilidad == 0 and credit == 0:
        return ['Correcto']

    # Regla 1: Utilidad errada solo cuando comision > 0 pero < 2%
    if 0 < comision < 2:
        alertas.append('Utilidad errada')

    # Regla 2: Sin crédito con utilidad
    if credit == 0 and utilidad > 0:
        alertas.append('Sin crédito y con utilidad')

    # Regla 3: Crédito sin utilidad
    if credit > 0 and utilidad == 0:
        alertas.append('Con crédito y sin utilidad')

    # Regla 4: Diferencia entre comisión y utilidad (solo si ambas > 0)
    if abs(comision - utilidad) > 0.01 and comision > 0 and utilidad > 0:
        alertas.append('Utilidades diferentes')

    return alertas if alertas else ['Correcto']

# ── Clasificación Créditos ────────────────────────────────────
def _clasificar_credito(row):
    """Retorna lista de alertas para una fila de créditos."""
    credit   = float(row.get('credit') or 0)
    utilidad = float(row.get('utilidad_corporativa') or 0)
    alertas  = []

    if credit == 9999:
        alertas.append('Validación de crédito')
    if credit == 0 and utilidad > 0:
        alertas.append('Sin crédito con utilidad')
    if credit > 0 and utilidad == 0:
        alertas.append('Con crédito sin utilidad')

    return alertas if alertas else ['Correcto']

def _run_auditoria(desde, hasta, company_id='', tarifa_id='', moneda='',
                   last_desde='', last_hasta=''):
    """Ejecuta la query y retorna filas raw enriquecidas."""
    fc, ft, fm, at, fls = _aud_filtros(company_id, tarifa_id, moneda,
                                       last_desde=last_desde, last_hasta=last_hasta)
    sql = _Q_AUDITORIA_BASE.format(
        desde=desde, hasta=hasta,
        filtro_company=fc, filtro_tarifa=ft,
        filtro_moneda=fm, anti_test=at,
        filtro_last_service=fls,
    )
    ch = get_client()
    r  = ch.query(sql)
    rows = [dict(zip(r.column_names, row)) for row in r.result_rows]

    # Normalizar tipos y agregar clasificaciones
    for row in rows:
        for k in ['comission','utilidad_corporativa','credit',
                  'base_fare','minimum_fare','distance_fare',
                  'hour_fare','extra_stop_fare','package_fare',
                  'hour_base_fare','hour_standby_fare','valor_declarado']:
            try: row[k] = round(float(row[k] or 0), 4)
            except: row[k] = 0.0
        row['last_service'] = str(row.get('last_service',''))[:16]
        # Fix 6: name_manager vacío/None → "Sin comercial"
        if not (row.get('name_manager') or '').strip():
            row['name_manager'] = 'Sin comercial'
        row['alertas_comision'] = _clasificar_comision(row)
        row['alertas_credito']  = _clasificar_credito(row)
        row['ok_comision']      = row['alertas_comision'] == ['Correcto']
        row['ok_credito']       = row['alertas_credito']  == ['Correcto']
    return rows

def _resumen_alertas(rows, campo_alertas):
    """Genera resumen estadístico de alertas incluyendo distribución por ciudad."""
    from collections import Counter
    total      = len(rows)
    correctos  = sum(1 for r in rows if r[campo_alertas] == ['Correcto'])
    con_error  = total - correctos
    dist = Counter()
    for r in rows:
        for a in r[campo_alertas]:
            if a != 'Correcto':
                dist[a] += 1
    # Segmentación por KAM (Fix 6: usa 'Sin comercial' ya aplicado upstream)
    kam_map = {}
    for r in rows:
        kam = r.get('name_manager') or 'Sin comercial'
        if kam not in kam_map:
            kam_map[kam] = {'total':0,'error':0,'correctos':0}
        kam_map[kam]['total'] += 1
        if r[campo_alertas] != ['Correcto']:
            kam_map[kam]['error'] += 1
        else:
            kam_map[kam]['correctos'] += 1
    por_kam = sorted(
        [{'kam':k,'total':v['total'],'error':v['error'],'correctos':v['correctos']}
         for k,v in kam_map.items()],
        key=lambda x: -x['error']
    )[:20]
    # Fix 5: Distribución por ciudades (Colombia — ya filtrado en query)
    ciudad_map = Counter()
    ciudad_error_map = Counter()
    for r in rows:
        c = r.get('ciudad','Sin ciudad') or 'Sin ciudad'
        ciudad_map[c] += 1
        if r[campo_alertas] != ['Correcto']:
            ciudad_error_map[c] += 1
    por_ciudad = [
        {'ciudad':c, 'total':ciudad_map[c],
         'error':ciudad_error_map.get(c,0),
         'correctos':ciudad_map[c]-ciudad_error_map.get(c,0)}
        for c in [x for x,_ in ciudad_map.most_common(20)]
    ]
    return {
        'total':     total,
        'correctos': correctos,
        'con_error': con_error,
        'pct_error': round(con_error/total*100,1) if total else 0,
        'distribucion': [{'alerta':k,'count':v} for k,v in dist.most_common()],
        'por_kam':    por_kam,
        'por_ciudad': por_ciudad,   # Fix 5: nuevo campo
    }

@app.route("/api/auditoria/comisiones")
def auditoria_comisiones():
    desde       = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta       = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    last_desde  = request.args.get("last_desde") or ""
    last_hasta  = request.args.get("last_hasta", "")
    company_id  = request.args.get("company_id","").strip()
    tarifa_id   = request.args.get("tarifa_id","").strip()
    moneda      = request.args.get("moneda","")
    q_id        = request.args.get("q","").strip()
    q_tipo      = request.args.get("tipo","company")

    try:
        rows = _run_auditoria(desde, hasta, company_id, tarifa_id, moneda, last_desde, last_hasta)

        # Filtro client-side por ID
        if q_id and len(q_id) >= 4:
            q_low = q_id.lower()
            campo = {'booking':'booking_id','company':'id_company',
                     'tarifa':'tarifa_id'}.get(q_tipo,'id_company')
            rows = [r for r in rows if q_low in (r.get(campo,'') or '').lower()]

        resumen   = _resumen_alertas(rows, 'alertas_comision')
        alertas   = [r for r in rows if not r['ok_comision']]
        alertas.sort(key=lambda x: len(x['alertas_comision']), reverse=True)

        return jsonify(limpiar_nan({
            'desde': desde, 'hasta': hasta,
            'resumen': resumen,
            'alertas': alertas[:500],
            'total_filas': len(rows),
        }))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500

@app.route("/api/auditoria/creditos")
def auditoria_creditos():
    desde       = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta       = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    last_desde  = request.args.get("last_desde") or ""
    last_hasta  = request.args.get("last_hasta", "")
    company_id  = request.args.get("company_id","").strip()
    tarifa_id   = request.args.get("tarifa_id","").strip()
    moneda      = request.args.get("moneda","")
    q_id        = request.args.get("q","").strip()
    q_tipo      = request.args.get("tipo","company")

    try:
        rows = _run_auditoria(desde, hasta, company_id, tarifa_id, moneda, last_desde, last_hasta)

        # Filtro client-side
        if q_id and len(q_id) >= 4:
            q_low = q_id.lower()
            campo = {'company':'id_company','tarifa':'tarifa_id'}.get(q_tipo,'id_company')
            rows = [r for r in rows if q_low in (r.get(campo,'') or '').lower()]

        resumen = _resumen_alertas(rows, 'alertas_credito')
        alertas = [r for r in rows if not r['ok_credito']]
        alertas.sort(key=lambda x: abs(x.get('credit',0)), reverse=True)

        # Distribución de créditos
        dist_credito = {
            'credit_9999':  sum(1 for r in rows if r.get('credit',0) == 9999),
            'credit_0':     sum(1 for r in rows if r.get('credit',0) == 0),
            'credit_pos':   sum(1 for r in rows if r.get('credit',0) > 0 and r.get('credit',0) != 9999),
            'v_total_credito': round(sum(r.get('credit',0) for r in rows if r.get('credit',0) not in (0,9999)), 2),
        }

        return jsonify(limpiar_nan({
            'desde': desde, 'hasta': hasta,
            'resumen': resumen,
            'alertas': alertas[:500],
            'dist_credito': dist_credito,
            'total_filas': len(rows),
        }))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


# ══════════════════════════════════════════════════════════════
# ENDPOINT: Exportar Auditoría a Excel (formato profesional)
# Usa openpyxl para aplicar estilos completos: colores, freeze, autofit
# ══════════════════════════════════════════════════════════════
import io
from flask import send_file

# Paleta corporativa Picap
_PURPLE     = "6B21A8"
_PURPLE_LT  = "EDE9F5"
_GREEN      = "16A34A"
_GREEN_LT   = "DCFCE7"
_RED        = "DC2626"
_RED_LT     = "FEE2E2"
_AMBER      = "D97706"
_AMBER_LT   = "FEF9C3"
_GRAY_LT    = "F3F0FA"
_WHITE      = "FFFFFF"
_DARK       = "1E1333"

def _xl_header_style(ws, row, cols, bg=_PURPLE, fg=_WHITE, bold=True, sz=10):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=bold, color=fg, size=sz, name='Arial')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
        cell.border = border

def _xl_row_style(ws, row, cols, bg, fg=_DARK, bold=False, sz=9, wrap=False):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=bg) if bg else None
    font = Font(bold=bold, color=fg, size=sz, name='Arial')
    align = Alignment(horizontal='left', vertical='center', wrap_text=wrap)
    for c in range(1, cols+1):
        cell = ws.cell(row, c)
        if fill: cell.fill = fill
        cell.font  = font
        cell.alignment = align
        cell.border = border

def _xl_autofit(ws, min_w=10, max_w=45):
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell):
                continue
            try:
                v = str(cell.value or '')
                max_len = max(max_len, len(v))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def _xl_set_cell(ws, row, col, value, bg=None, fg=_DARK, bold=False, sz=9,
                 align='left', wrap=False, num_fmt=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style='thin', color="E5E7EB")
    cell = ws.cell(row, col)
    cell.value = value
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font  = Font(bold=bold, color=fg, size=sz, name='Arial')
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = Border(left=Side(style='thin',color="E5E7EB"),
                         right=Side(style='thin',color="E5E7EB"),
                         top=Side(style='thin',color="E5E7EB"),
                         bottom=Side(style='thin',color="E5E7EB"))
    if num_fmt:
        cell.number_format = num_fmt

def _build_alertas_sheet(ws, rows, campos, titulo):
    """Construye hoja de alertas con encabezados y datos formateados."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from collections import Counter
    import re

    # ── Fila 1: Banner de título ──────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(campos))
    cell = ws.cell(1, 1)
    cell.value = titulo.upper()
    cell.fill  = PatternFill("solid", fgColor=_PURPLE)
    cell.font  = Font(bold=True, color=_WHITE, size=13, name='Arial')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Fila 2: Metadatos ─────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(campos))
    from datetime import date as _date
    cell2 = ws.cell(2, 1)
    cell2.value = f"Exportado el {_date.today().strftime('%d/%m/%Y')}  ·  Total registros: {len(rows)}  ·  Picap — Solo datos Colombia"
    cell2.fill  = PatternFill("solid", fgColor="EDE9F5")
    cell2.font  = Font(color=_PURPLE, size=9, italic=True, name='Arial')
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Encabezados en MAYÚSCULAS ─────────────────────
    headers = [label.upper() for label, _ in campos]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c).value = h
    _xl_header_style(ws, 3, len(campos), bg=_PURPLE, sz=9)
    ws.row_dimensions[3].height = 22

    # ── Datos (fila 4 en adelante) ────────────────────────────
    ALERT_COLORS = {
        'Utilidad errada':            (_RED_LT, _RED),
        'Sin crédito y con utilidad': (_AMBER_LT, _AMBER),
        'Con crédito y sin utilidad': (_AMBER_LT, _AMBER),
        'Utilidades diferentes':      (_AMBER_LT, _AMBER),
        'Validación de crédito':      (_RED_LT, _RED),
        'Sin crédito con utilidad':   (_AMBER_LT, _AMBER),
        'Con crédito sin utilidad':   (_AMBER_LT, _AMBER),
    }

    for r_idx, row in enumerate(rows):
        xl_row = r_idx + 4
        alertas = row.get('alertas_comision') or row.get('alertas_credito') or []
        alerta_str = '; '.join(alertas) if isinstance(alertas, list) else str(alertas)

        # Color de fila según alerta más grave
        row_bg = _GREEN_LT if alertas == ['Correcto'] else None
        for alert in alertas:
            if alert in ALERT_COLORS:
                row_bg = ALERT_COLORS[alert][0]
                break

        row_bg_use = row_bg or (_GRAY_LT if r_idx % 2 == 0 else _WHITE)

        for c_idx, (label, key) in enumerate(campos, 1):
            val = row.get(key)
            if isinstance(val, list):
                val = '; '.join(val)
            elif val is None:
                val = ''

            # Formato numérico para comisión/utilidad
            num_fmt = None
            if 'comission' in key or 'utilidad' in key:
                if isinstance(val, (int, float)):
                    num_fmt = '0.00"%"'
            elif 'credit' in key:
                if isinstance(val, (int, float)):
                    num_fmt = '#,##0.00'

            _xl_set_cell(ws, xl_row, c_idx, val, bg=row_bg_use, sz=9,
                         wrap=True, num_fmt=num_fmt)

        ws.row_dimensions[xl_row].height = 15

    # ── Freeze: primera fila de datos visible ─────────────────
    ws.freeze_panes = ws.cell(4, 1)

    # ── Autofit columnas ──────────────────────────────────────
    _xl_autofit(ws)

    # Forzar mínimo en columna de IDs
    for c_idx, (label, key) in enumerate(campos, 1):
        if 'id' in key.lower():
            ws.column_dimensions[get_column_letter(c_idx)].width = 26

def _build_resumen_sheet(ws, data_com, data_cred):
    """Construye hoja de Resumen Ejecutivo con análisis completo."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from collections import Counter

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    row = 1

    def titulo_seccion(texto, r, ncols=5, bg=_PURPLE):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(r, 1)
        c.value = texto.upper()
        c.fill  = PatternFill("solid", fgColor=bg)
        c.font  = Font(bold=True, color=_WHITE, size=10, name='Arial')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[r].height = 20

    def kpi_row(r, label, val, pct=None, bg=_WHITE, fg=_DARK, bold=False):
        _xl_set_cell(ws, r, 1, label, sz=9, bold=bold, bg=bg, fg=fg)
        _xl_set_cell(ws, r, 2, val, sz=9, bold=bold, bg=bg, fg=fg, align='right')
        if pct is not None:
            _xl_set_cell(ws, r, 3, pct, sz=9, bg=bg, fg=fg, align='right', num_fmt='0.0"%"')

    def header_row(r, cols, bg=_PURPLE_LT):
        for c_idx, txt in enumerate(cols, 1):
            _xl_set_cell(ws, r, c_idx, txt.upper(), bold=True, sz=8, bg=bg, fg=_PURPLE, align='center')
        ws.row_dimensions[r].height = 16

    # ── BANNER ────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    c = ws.cell(row, 1)
    c.value = "RESUMEN EJECUTIVO — AUDITORÍAS PIBOX COMERCIAL"
    c.fill  = PatternFill("solid", fgColor=_PURPLE)
    c.font  = Font(bold=True, color=_WHITE, size=14, name='Arial')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 32
    row += 1

    from datetime import date as _date
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    sub = ws.cell(row, 1)
    sub.value = f"Generado: {_date.today().strftime('%d/%m/%Y')}  ·  Solo Colombia  ·  Datos limpios (anti-test activo)"
    sub.fill  = PatternFill("solid", fgColor=_PURPLE_LT)
    sub.font  = Font(italic=True, color=_PURPLE, size=9, name='Arial')
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 16
    row += 2

    # ── SECCIÓN COMISIONES ────────────────────────────────────
    titulo_seccion("💰 Comisiones", row); row += 1
    com_rows = data_com or []
    total_com = len(com_rows)
    correctos_com = sum(1 for r in com_rows if r.get('alertas_comision') == ['Correcto'])
    error_com     = total_com - correctos_com

    header_row(row, ['Métrica', 'Cantidad', '% del total', '', '']); row += 1
    kpi_row(row, 'Total líneas auditadas',  total_com, 100.0, bold=True); row += 1
    kpi_row(row, 'Sin alertas (correctos)', correctos_com,
            round(correctos_com/total_com*100,1) if total_com else 0,
            bg=_GREEN_LT, fg=_GREEN, bold=True); row += 1
    kpi_row(row, 'Con alertas (errores)',   error_com,
            round(error_com/total_com*100,1) if total_com else 0,
            bg=_RED_LT, fg=_RED, bold=True); row += 2

    # Distribución por tipo de alerta
    titulo_seccion("  Distribución de alertas — Comisiones", row, bg="9333EA"); row += 1
    header_row(row, ['Tipo de alerta', 'Cantidad', '% del total', '', '']); row += 1
    dist_com = Counter()
    for r_d in com_rows:
        for a in (r_d.get('alertas_comision') or []):
            if a != 'Correcto': dist_com[a] += 1
    ALERTA_BGS = {
        'Utilidad errada':            _RED_LT,
        'Sin crédito y con utilidad': _AMBER_LT,
        'Con crédito y sin utilidad': _AMBER_LT,
        'Utilidades diferentes':      _AMBER_LT,
    }
    for alerta, cnt in dist_com.most_common():
        bg = ALERTA_BGS.get(alerta, _GRAY_LT)
        fg = _RED if 'errada' in alerta else _AMBER
        kpi_row(row, alerta, cnt,
                round(cnt/total_com*100,1) if total_com else 0,
                bg=bg, fg=fg); row += 1
    row += 1

    # Por KAM
    titulo_seccion("  Por Comercial (KAM) — Comisiones", row, bg="9333EA"); row += 1
    header_row(row, ['KAM', 'Total', 'Con error', 'Correctos', '% error']); row += 1
    kam_com = {}
    for r_d in com_rows:
        k = r_d.get('name_manager') or 'Sin comercial'
        if k not in kam_com: kam_com[k] = {'total':0,'error':0}
        kam_com[k]['total'] += 1
        if r_d.get('alertas_comision') != ['Correcto']:
            kam_com[k]['error'] += 1
    for k, v in sorted(kam_com.items(), key=lambda x:-x[1]['error'])[:20]:
        corr = v['total'] - v['error']
        pct_e = round(v['error']/v['total']*100,1) if v['total'] else 0
        bg = _RED_LT if pct_e > 50 else (_AMBER_LT if pct_e > 20 else _GREEN_LT)
        _xl_set_cell(ws, row, 1, k, sz=9, bg=bg)
        _xl_set_cell(ws, row, 2, v['total'], sz=9, bg=bg, align='right')
        _xl_set_cell(ws, row, 3, v['error'],  sz=9, bg=bg, align='right', fg=_RED)
        _xl_set_cell(ws, row, 4, corr,        sz=9, bg=bg, align='right', fg=_GREEN)
        _xl_set_cell(ws, row, 5, pct_e,       sz=9, bg=bg, align='right',
                     num_fmt='0.0"%"')
        row += 1
    row += 1

    # Por Ciudad
    titulo_seccion("  Por Ciudad — Comisiones", row, bg="9333EA"); row += 1
    header_row(row, ['Ciudad', 'Total', 'Con error', 'Correctos', '% error']); row += 1
    ciudad_com = {}
    for r_d in com_rows:
        c = r_d.get('ciudad') or 'Sin ciudad'
        if c not in ciudad_com: ciudad_com[c] = {'total':0,'error':0}
        ciudad_com[c]['total'] += 1
        if r_d.get('alertas_comision') != ['Correcto']:
            ciudad_com[c]['error'] += 1
    for c, v in sorted(ciudad_com.items(), key=lambda x:-x[1]['total'])[:20]:
        corr = v['total'] - v['error']
        pct_e = round(v['error']/v['total']*100,1) if v['total'] else 0
        bg = _RED_LT if pct_e > 50 else (_AMBER_LT if pct_e > 20 else _GREEN_LT)
        _xl_set_cell(ws, row, 1, c,         sz=9, bg=bg)
        _xl_set_cell(ws, row, 2, v['total'], sz=9, bg=bg, align='right')
        _xl_set_cell(ws, row, 3, v['error'],  sz=9, bg=bg, align='right', fg=_RED)
        _xl_set_cell(ws, row, 4, corr,        sz=9, bg=bg, align='right', fg=_GREEN)
        _xl_set_cell(ws, row, 5, pct_e,       sz=9, bg=bg, align='right', num_fmt='0.0"%"')
        row += 1
    row += 2

    # ── SECCIÓN CRÉDITOS ──────────────────────────────────────
    titulo_seccion("💳 Créditos", row); row += 1
    cred_rows = data_cred or []
    total_cred    = len(cred_rows)
    correctos_cred = sum(1 for r in cred_rows if r.get('alertas_credito') == ['Correcto'])
    error_cred     = total_cred - correctos_cred
    n_9999         = sum(1 for r in cred_rows if (r.get('credit') or 0) == 9999)
    n_0            = sum(1 for r in cred_rows if (r.get('credit') or 0) == 0)

    header_row(row, ['Métrica', 'Cantidad', '% del total', '', '']); row += 1
    kpi_row(row, 'Total líneas auditadas',   total_cred, 100.0, bold=True); row += 1
    kpi_row(row, 'Sin alertas (correctos)',  correctos_cred,
            round(correctos_cred/total_cred*100,1) if total_cred else 0,
            bg=_GREEN_LT, fg=_GREEN, bold=True); row += 1
    kpi_row(row, 'Con alertas (errores)',    error_cred,
            round(error_cred/total_cred*100,1) if total_cred else 0,
            bg=_RED_LT, fg=_RED, bold=True); row += 1
    kpi_row(row, 'Crédito = 9999 (pendiente)', n_9999,
            round(n_9999/total_cred*100,1) if total_cred else 0,
            bg=_RED_LT, fg=_RED); row += 1
    kpi_row(row, 'Sin crédito (= 0)',         n_0,
            round(n_0/total_cred*100,1) if total_cred else 0,
            bg=_AMBER_LT, fg=_AMBER); row += 2

    # Distribución créditos por KAM
    titulo_seccion("  Por Comercial (KAM) — Créditos", row, bg="7C3AED"); row += 1
    header_row(row, ['KAM', 'Total', 'Con error', 'Correctos', '% error']); row += 1
    kam_cred = {}
    for r_d in cred_rows:
        k = r_d.get('name_manager') or 'Sin comercial'
        if k not in kam_cred: kam_cred[k] = {'total':0,'error':0}
        kam_cred[k]['total'] += 1
        if r_d.get('alertas_credito') != ['Correcto']:
            kam_cred[k]['error'] += 1
    for k, v in sorted(kam_cred.items(), key=lambda x:-x[1]['error'])[:20]:
        corr = v['total'] - v['error']
        pct_e = round(v['error']/v['total']*100,1) if v['total'] else 0
        bg = _RED_LT if pct_e > 50 else (_AMBER_LT if pct_e > 20 else _GREEN_LT)
        _xl_set_cell(ws, row, 1, k, sz=9, bg=bg)
        _xl_set_cell(ws, row, 2, v['total'], sz=9, bg=bg, align='right')
        _xl_set_cell(ws, row, 3, v['error'],  sz=9, bg=bg, align='right', fg=_RED)
        _xl_set_cell(ws, row, 4, corr,        sz=9, bg=bg, align='right', fg=_GREEN)
        _xl_set_cell(ws, row, 5, pct_e,       sz=9, bg=bg, align='right', num_fmt='0.0"%"')
        row += 1

    ws.freeze_panes = ws.cell(3, 1)

@app.route("/api/auditoria/exportar")
def auditoria_exportar():
    desde      = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta      = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    last_desde = request.args.get("last_desde") or ""
    last_hasta = request.args.get("last_hasta", "")
    company_id = request.args.get("company_id","").strip()
    tarifa_id  = request.args.get("tarifa_id","").strip()
    moneda     = request.args.get("moneda","")
    tipo       = request.args.get("tipo","ambos")   # comisiones | creditos | ambos

    try:
        from openpyxl import Workbook
        from datetime import date as _date

        rows = _run_auditoria(desde, hasta, company_id, tarifa_id, moneda,
                              last_desde, last_hasta)

        # Clasificar todas las filas
        rows_com  = [r for r in rows if not r.get('ok_comision', True)]
        rows_cred = [r for r in rows if not r.get('ok_credito', True)]

        CAMPOS_COM = [
            ('Estado',            'estado'),
            ('Company ID',        'id_company'),
            ('Tarifa ID',         'tarifa_id'),
            ('Línea de negocio',  'linea_de_negocio'),
            ('KAM',               'name_manager'),
            ('Tipo servicio',     'type_service'),
            ('Ciudad',            'ciudad'),
            ('Moneda',            'moneda'),
            ('Comisión (%)',      'comission'),
            ('Utilidad corp. (%)','utilidad_corporativa'),
            ('Crédito',           'credit'),
            ('Alertas',           'alertas_comision'),
            ('Último servicio',   'last_service'),
        ]
        CAMPOS_CRED = [
            ('Estado',            'estado'),
            ('Company ID',        'id_company'),
            ('Tarifa ID',         'tarifa_id'),
            ('Línea de negocio',  'linea_de_negocio'),
            ('KAM',               'name_manager'),
            ('Tipo servicio',     'type_service'),
            ('Ciudad',            'ciudad'),
            ('Moneda',            'moneda'),
            ('Crédito',           'credit'),
            ('Utilidad corp. (%)','utilidad_corporativa'),
            ('Comisión (%)',      'comission'),
            ('Alertas crédito',   'alertas_credito'),
            ('Último servicio',   'last_service'),
        ]

        wb = Workbook()
        wb.remove(wb.active)  # quitar hoja default

        # Hoja 1: Alertas Comisiones (TODAS las filas)
        ws_com = wb.create_sheet("Comisiones")
        _build_alertas_sheet(ws_com, rows, CAMPOS_COM,
                             f"Auditoría Pibox — Comisiones · {desde} → {hasta}")
        ws_com.sheet_view.showGridLines = False
        ws_com.sheet_properties.tabColor = "6B21A8"

        # Hoja 2: Alertas Créditos (TODAS las filas)
        ws_cred = wb.create_sheet("Créditos")
        _build_alertas_sheet(ws_cred, rows, CAMPOS_CRED,
                             f"Auditoría Pibox — Créditos · {desde} → {hasta}")
        ws_cred.sheet_view.showGridLines = False
        ws_cred.sheet_properties.tabColor = "7C3AED"

        # Hoja 3: Resumen Ejecutivo
        ws_res = wb.create_sheet("Resumen Ejecutivo")
        _build_resumen_sheet(ws_res, rows, rows)
        ws_res.sheet_view.showGridLines = False
        ws_res.sheet_properties.tabColor = "22C55E"

        # Guardar en buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"Auditoria_Pibox_{_date.today().strftime('%Y-%m-%d')}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname,
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500


# ══════════════════════════════════════════════════════════════
# HELPERS GENÉRICOS PARA EXPORTACIÓN EXCEL (compartidos)
# ══════════════════════════════════════════════════════════════
def _xl_make_workbook():
    """Crea un Workbook nuevo con estilos base."""
    from openpyxl import Workbook as WB
    wb = WB()
    wb.remove(wb.active)
    return wb

def _xl_banner(ws, titulo, subtitulo, ncols):
    """Fila 1: banner morado + fila 2: metadatos."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from datetime import date as _d
    # Fila 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1)
    c.value = titulo.upper()
    c.fill  = PatternFill("solid", fgColor="6B21A8")
    c.font  = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # Fila 2
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1)
    c2.value = f"{subtitulo}  ·  Exportado: {_d.today().strftime('%d/%m/%Y')}  ·  Picap Monitoreo"
    c2.fill  = PatternFill("solid", fgColor="EDE9F5")
    c2.font  = Font(italic=True, color="6B21A8", size=9, name="Arial")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

def _xl_headers(ws, row, headers, bg="6B21A8"):
    """Fila de encabezados en MAYÚSCULAS sobre fondo morado."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=bg)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c)
        cell.value = h.upper()
        cell.fill  = fill
        cell.font  = Font(bold=True, color="FFFFFF", size=9, name="Arial")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[row].height = 20

def _xl_data_row(ws, row_idx, values, bg=None, aligns=None):
    """Escribe una fila de datos con estilo alternado."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bg_use = bg or ("F3F0FA" if row_idx % 2 == 0 else "FFFFFF")
    fill   = PatternFill("solid", fgColor=bg_use)
    for c, v in enumerate(values, 1):
        cell = ws.cell(row_idx, c)
        cell.value = v if not isinstance(v, (list, dict)) else str(v)
        cell.fill  = fill
        cell.font  = Font(size=9, name="Arial", color="1E1333")
        ha = (aligns or {}).get(c, "left")
        cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row_idx].height = 14

def _xl_autofit_safe(ws, min_w=10, max_w=50):
    """Autofit que maneja MergedCell sin errores."""
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    for col_idx in range(1, ws.max_column + 1):
        mx = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell, MergedCell): continue
            try: mx = max(mx, len(str(cell.value or "")))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, mx + 2))

def _xl_color_cell(ws, row, col, value, cls):
    """Aplica color semáforo según clasificación."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    thin = Side(style="thin", color="E5E7EB")
    COLORS = {
        "OK":               ("DCFCE7", "166534"),
        "CORRECTO":         ("DCFCE7", "166534"),
        "EVASION CONFIRMADA": ("FEE2E2", "DC2626"),
        "FRAUDE":           ("FEE2E2", "DC2626"),
        "ESTAFA":           ("FEE2E2", "DC2626"),
        "EVASION PROBABLE": ("FEF9C3", "D97706"),
        "MALA_PRACTICA":    ("FEF9C3", "D97706"),
        "POSIBLE_ESTAFA":   ("FEF9C3", "D97706"),
        "DEBE_DINERO":      ("FEE2E2", "DC2626"),
        "PAGADO_DEMAS":     ("FEF9C3", "D97706"),
        "REVISAR":          ("EDE9F5", "6B21A8"),
    }
    key = str(cls).upper().replace(" ", "_")
    bg, fg = COLORS.get(key, ("F3F0FA", "1E1333"))
    cell = ws.cell(row, col)
    cell.value = value
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.font  = Font(bold=True, size=9, name="Arial", color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

def _xl_kpi_section(ws, start_row, titulo, kpis, ncols=4):
    """Sección de KPIs con fila de título + filas de datos."""
    from openpyxl.styles import PatternFill, Font, Alignment
    # título de sección
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncols)
    c = ws.cell(start_row, 1)
    c.value = titulo.upper()
    c.fill  = PatternFill("solid", fgColor="9333EA")
    c.font  = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 18
    row = start_row + 1
    for label, val, *rest in kpis:
        bg = rest[0] if rest else None
        _xl_data_row(ws, row, [label, val], bg=bg, aligns={2: "right"})
        row += 1
    return row + 1  # una fila vacía entre secciones

def _xl_finalize(ws, freeze_row=4):
    """Congela encabezados, quita cuadrícula, aplica autofit."""
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    ws.freeze_panes = ws.cell(freeze_row, 1)
    ws.sheet_view.showGridLines = False
    _xl_autofit_safe(ws)

def _xl_to_response(wb, filename):
    """Convierte el workbook a respuesta Flask descargable."""
    from datetime import date as _d
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{filename}_{_d.today().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )

# ══════════════════════════════════════════════════════════════
# EXPORT 1: EVASIÓN DE COMISIONES
# ══════════════════════════════════════════════════════════════
@app.route("/api/exportar/evasion")
def exportar_evasion():
    desde = request.args.get("desde", _cache.get("desde") or "2026-04-01")
    hasta  = request.args.get("hasta", _cache.get("hasta")  or "2026-04-17")
    pais   = request.args.get("pais", "")
    moneda = request.args.get("moneda", "")
    try:
        if _necesita_recarga(desde, hasta):
            cargar_datos(desde, hasta, pais or None, moneda or None)
        resumen = _cache.get("resumen") or {}
        kpis    = resumen.get("kpis", {})
        ch      = get_client()
        # Obtener detalle completo
        # Usar datos ya en caché + clasificación Python (evita re-query problemática)
        # El cache ya contiene todos los datos clasificados del período
        cached = _cache.get("resumen") or {}
        top_drv = cached.get("top_drivers", [])

        # Hacer una query directa con solo las columnas que SÍ existen en BASE_CTE
        sql_det = BASE_CTE.format(
            fecha_desde=desde, fecha_hasta=hasta
        ) + """
SELECT
    creacion_servicio, booking_id, id_driver, name_driver,
    id_company, type_service, moneda, pais, ciudad,
    costo_estimado, minutos_entre_eventos,
    cancel_lon, cancel_lat, end_lon, end_lat,
    distancia_cancel_destino, nivel, comision_servicio,
    comision_mas_penalizacion
FROM base
WHERE rn = 1
ORDER BY nivel DESC, creacion_servicio DESC
LIMIT 5000
"""
        r_det = ch.query(sql_det)
        rows  = [dict(zip(r_det.column_names, row)) for row in r_det.result_rows]
        # Enriquecer con clasificación Python
        for row in rows:
            n = row.get('nivel', 0)
            row['veredicto'] = (
                'EVASION CONFIRMADA' if n == 3 else
                'EVASION PROBABLE'   if n == 2 else 'OK'
            )

        wb = _xl_make_workbook()

        # ── Hoja 1: Resumen Ejecutivo ─────────────────────────
        ws1 = wb.create_sheet("Resumen Ejecutivo")
        ws1.sheet_properties.tabColor = "6B21A8"
        _xl_banner(ws1, f"Evasión de Comisiones — {pais or 'Todos los países'}",
                   f"Período: {desde} → {hasta}", 4)
        row = 4
        row = _xl_kpi_section(ws1, row, "Métricas generales", [
            ("Total servicios auditados", kpis.get("total", 0)),
            ("Evasión confirmada (nivel 3)", kpis.get("confirmadas", 0), "FEE2E2"),
            ("Evasión probable (nivel 2)",  kpis.get("probables", 0),   "FEF9C3"),
            ("OK — Sin evasión",            kpis.get("ok", 0),          "DCFCE7"),
            ("Sin GPS",                     kpis.get("sin_gps", 0)),
            ("Tasa de evasión (%)",         f"{kpis.get('tasa_evasion', 0):.1f}%"),
            ("Comisión evadida estimada",
             f"$ {kpis.get('comision_evadida_cop', 0):,.0f}"),
        ], ncols=4)

        # Distribución por ciudad
        ciudades = resumen.get("ciudades", [])
        if ciudades:
            row = _xl_kpi_section(ws1, row, "Top ciudades", [
                (c.get("ciudad","?"),
                 f"{c.get('confirmadas',0)} conf / {c.get('probables',0)} prob")
                for c in ciudades[:15]
            ], ncols=4)

        # Top drivers
        drivers = resumen.get("top_drivers", [])
        if drivers:
            _xl_kpi_section(ws1, row, "Top conductores con evasión", [
                (f"{d.get('name_driver','?')} ({d.get('id_driver','')[:10]}…)",
                 f"{d.get('total_evasiones',0)} evasiones")
                for d in drivers[:15]
            ], ncols=4)

        _xl_autofit_safe(ws1)
        ws1.freeze_panes = ws1.cell(4, 1)
        ws1.sheet_view.showGridLines = False

        # ── Hoja 2: Detalle servicio por servicio ─────────────
        ws2 = wb.create_sheet("Detalle")
        ws2.sheet_properties.tabColor = "9333EA"
        headers = ["Fecha", "Booking ID", "Driver ID", "Conductor",
                   "Empresa", "Tipo servicio", "Moneda", "País", "Ciudad",
                   "Costo Est. (COP)", "Minutos entre eventos",
                   "Distancia (m)", "Veredicto", "Nivel",
                   "Comisión svc", "Comisión + Penalización"]
        _xl_banner(ws2, "Detalle de Servicios — Evasión de Comisiones",
                   f"Período: {desde} → {hasta}  ·  Registros: {len(rows)}", len(headers))
        _xl_headers(ws2, 3, headers)
        for i, row_d in enumerate(rows):
            xl_row = i + 4
            vals = [
                str(row_d.get("creacion_servicio",""))[:16],
                row_d.get("booking_id",""),
                row_d.get("id_driver",""),
                row_d.get("name_driver",""),
                row_d.get("id_company",""),
                row_d.get("type_service",""),
                row_d.get("moneda",""),
                row_d.get("pais",""),
                row_d.get("ciudad",""),
                round(float(row_d.get("costo_estimado") or 0), 2),
                row_d.get("minutos_entre_eventos") or 0,
                round(float(row_d.get("distancia_cancel_destino") or 0), 1),
                row_d.get("veredicto","OK"),
                row_d.get("nivel", 0),
                round(float(row_d.get("comision_servicio") or 0), 2),
                round(float(row_d.get("comision_mas_penalizacion") or 0), 2),
            ]
            _xl_data_row(ws2, xl_row, vals, aligns={10:"right",12:"right",15:"right",16:"right"})
            _xl_color_cell(ws2, xl_row, 13, vals[12], vals[12])

        _xl_finalize(ws2, freeze_row=4)
        return _xl_to_response(wb, "Picap_Evasion_Comisiones")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500

# ══════════════════════════════════════════════════════════════
# EXPORT 2: SERVICIOS ESTAFA
# ══════════════════════════════════════════════════════════════
@app.route("/api/exportar/estafa")
def exportar_estafa():
    desde  = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta  = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    pais   = request.args.get("pais", "")
    try:
        filtro_pais = _estafa_filtro_pais(pais)
        sql = _Q_ESTAFA_BASE.format(
            desde=desde, hasta=hasta, filtro_pais=filtro_pais,
            kws_estafa=_KW_ESTAFA_SQL, limit_filas=20000,
        )
        ch   = get_client()
        r    = ch.query(sql)
        rows_raw = [dict(zip(r.column_names, row)) for row in r.result_rows]
        rows = [_procesar_fila_estafa(row) for row in rows_raw]

        n_e = sum(1 for r in rows if r["clasificacion"] == "ESTAFA")
        total = len(rows)
        n_o = total - n_e

        wb = _xl_make_workbook()

        # ── Hoja 1: Estadística ───────────────────────────────
        ws1 = wb.create_sheet("Estadística")
        ws1.sheet_properties.tabColor = "DC2626"
        _xl_banner(ws1, "Servicios Estafa — Estadística",
                   f"Período: {desde} → {hasta}  ·  País: {pais or 'Todos'}", 3)
        row = 4
        row = _xl_kpi_section(ws1, row, "Resumen de clasificación", [
            ("Total servicios analizados",  total),
            ("Estafa confirmada",            n_e, "FEE2E2"),
            ("OK — Sin indicadores",         n_o, "DCFCE7"),
            ("% Estafa",  f"{round(n_e/total*100,1) if total else 0}%"),
            ("% OK",      f"{round(n_o/total*100,1) if total else 0}%"),
        ], ncols=3)

        # Top palabras clave
        from collections import Counter
        kw_counter = Counter()
        for r2 in rows:
            for kw in r2.get("palabras_detectadas", []):
                kw_counter[kw] += 1
        if kw_counter:
            _xl_kpi_section(ws1, row, "Top palabras clave detectadas",
                [(kw, cnt) for kw, cnt in kw_counter.most_common(20)], ncols=3)
        _xl_autofit_safe(ws1)
        ws1.freeze_panes = ws1.cell(4, 1)
        ws1.sheet_view.showGridLines = False

        # ── Hoja 2: Alertas ────────────────────────────────────
        alertas = [r2 for r2 in rows if r2["clasificacion"] != "OK"]
        headers = ["Clasificación", "Booking ID", "Driver ID", "User ID",
                   "Usuario", "País", "Ciudad", "Fecha", "Motivo cancelación",
                   "Palabras detectadas", "Indicaciones"]
        ws2 = wb.create_sheet("Alertas")
        ws2.sheet_properties.tabColor = "9333EA"
        _xl_banner(ws2, "Servicios Estafa — Alertas",
                   f"Período: {desde} → {hasta}  ·  {len(alertas)} alertas", len(headers))
        _xl_headers(ws2, 3, headers)
        for i, a in enumerate(alertas):
            xl_row = i + 4
            vals = [
                a["clasificacion"],
                a.get("booking_id",""), a.get("driver_id",""), a.get("user_id",""),
                a.get("name_user",""), a.get("pais",""), a.get("city",""),
                a.get("fecha_servicio",""), a.get("cancelation_reason",""),
                ", ".join(a.get("palabras_detectadas",[])),
                a.get("indications","")[:300],
            ]
            _xl_data_row(ws2, xl_row, vals)
            _xl_color_cell(ws2, xl_row, 1, vals[0], vals[0])
        _xl_finalize(ws2, freeze_row=4)
        return _xl_to_response(wb, "Picap_Servicios_Estafa")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500

# ══════════════════════════════════════════════════════════════
# EXPORT 3: BLOQUEOS
# ══════════════════════════════════════════════════════════════
@app.route("/api/exportar/bloqueos")
def exportar_bloqueos():
    desde   = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta   = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    pais    = request.args.get("pais", "")
    subtipo = request.args.get("subtipo", "todos")
    try:
        import json as _json
        # Ejecutar la query de bloqueos directamente
        ch = get_client()
        r  = ch.query(Q_BLOQUEOS.format(fecha_desde=desde, fecha_hasta=hasta))
        rows_raw = [dict(zip(r.column_names, row)) for row in r.result_rows]

        # Misma lógica de enriquecimiento que bloqueos()
        for row in rows_raw:
            for k, v in list(row.items()):
                if hasattr(v, 'isoformat'):
                    row[k] = str(v)[:16] if v else None
            row['pais_nombre'] = PAISES_MAP.get(row.get('pais_codigo',''), row.get('pais_codigo',''))
            tipo_usr = row.get('tipo_usuario', '')
            if tipo_usr == 'PILOTO':
                motivo_raw = (row.get('comentario_driver','') or '').strip() or                              (row.get('comentario_user','') or '').strip()
            else:
                motivo_raw = (row.get('comentario_user','') or '').strip() or                              (row.get('comentario_expulsion_user','') or '').strip()
            row['motivo_mapeado'] = mapear_motivo(motivo_raw) if 'mapear_motivo' in dir() or callable(globals().get('mapear_motivo')) else motivo_raw
            tipo_blq = row.get('tipo_bloqueo', '')
            dias     = row.get('dias_bloqueado_total') or 0
            if tipo_blq == 'EXPULSADO':
                row['veredicto'] = 'EXPULSIÓN PERMANENTE'
            else:
                row['veredicto'] = 'ALERTA DE TIEMPO' if dias > 30 else 'TODO OK'

        # Filtrar por subtipo
        FILTROS = {
            'alertas':        lambda r: r.get('veredicto') == 'ALERTA DE TIEMPO',
            'actuales':       lambda r: r.get('esta_activo') == 'bloqueado',
            'reactivaciones': lambda r: r.get('esta_activo') == 'activo',
            'todos':          lambda r: True,
        }
        fn_filtro = FILTROS.get(subtipo, FILTROS['todos'])
        # Filtro de país
        if pais:
            iso_map = {"Colombia":"CO","Mexico":"MX","Nicaragua":"NI","Guatemala":"GT"}
            iso = iso_map.get(pais, pais)
            registros = [row for row in rows_raw if fn_filtro(row) and row.get('pais_codigo','') == iso]
        else:
            registros = [row for row in rows_raw if fn_filtro(row)]

        wb = _xl_make_workbook()
        subtipo_label = {"alertas":"Alertas","actuales":"Actuales",
                         "reactivaciones":"Reactivaciones","todos":"Todos"}.get(subtipo,"Todos")

        headers = ["Tipo bloqueo","Tipo usuario","User ID","Nombre","País","Ciudad",
                   "Veredicto","Motivo oficial","Estado actual",
                   "Días bloqueados","Inicio suspensión","Fin suspensión","Comentario"]
        ws1 = wb.create_sheet(subtipo_label[:31])
        ws1.sheet_properties.tabColor = "22C55E"
        _xl_banner(ws1, f"Vista de Bloqueos — {subtipo_label}",
                   f"Período: {desde} → {hasta}  ·  País: {pais or 'Todos'}  ·  {len(registros)} registros",
                   len(headers))
        _xl_headers(ws1, 3, headers)
        for i, reg in enumerate(registros):
            xl_row = i + 4
            tipo_blq  = reg.get("tipo_bloqueo","")
            veredicto = reg.get("veredicto","")
            cls_color = ("ESTAFA"          if tipo_blq == "EXPULSADO" else
                         "EVASION PROBABLE" if "ALERTA" in veredicto else "OK")
            vals = [
                tipo_blq,
                reg.get("tipo_usuario",""),
                reg.get("id_usuario",""),
                reg.get("nombre",""),
                reg.get("pais_nombre","") or reg.get("pais_codigo",""),
                reg.get("ciudad",""),
                veredicto,
                reg.get("motivo_mapeado",""),
                reg.get("esta_activo","") or reg.get("estado_suspension",""),
                reg.get("dias_bloqueado_total","") or "",
                str(reg.get("starts_block_user","") or reg.get("starts_block_driver","") or "")[:10],
                str(reg.get("ends_block_user","")   or reg.get("ends_block_driver","")   or "")[:10],
                (reg.get("comentario_user","") or
                 reg.get("comentario_expulsion_user","") or
                 reg.get("comentario_driver","") or "")[:300],
            ]
            _xl_data_row(ws1, xl_row, vals)
            _xl_color_cell(ws1, xl_row, 7, vals[6], cls_color)
        _xl_finalize(ws1, freeze_row=4)

        # Hoja resumen
        total     = len(registros)
        n_exp     = sum(1 for r in registros if r.get("tipo_bloqueo") == "EXPULSADO")
        n_susp    = total - n_exp
        n_alerta  = sum(1 for r in registros if r.get("veredicto") == "ALERTA DE TIEMPO")
        n_react   = sum(1 for r in registros if r.get("esta_activo") == "activo")
        ws2 = wb.create_sheet("Resumen")
        ws2.sheet_properties.tabColor = "16A34A"
        _xl_banner(ws2, "Resumen de Bloqueos",
                   f"Período: {desde} → {hasta}  ·  Subtipo: {subtipo_label}", 3)
        _xl_kpi_section(ws2, 4, "Totales", [
            ("Total registros",      total),
            ("Expulsados",           n_exp,    "FEE2E2"),
            ("Suspendidos",          n_susp,   "FEF9C3"),
            ("Alertas (>30 días)",   n_alerta, "FEE2E2"),
            ("Reactivados",          n_react,  "DCFCE7"),
        ], ncols=3)
        _xl_autofit_safe(ws2)
        ws2.freeze_panes = ws2.cell(4, 1)
        ws2.sheet_view.showGridLines = False

        return _xl_to_response(wb, f"Picap_Bloqueos_{subtipo_label}")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500


@app.route("/api/exportar/pagos")
def exportar_pagos():
    desde = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta  = request.args.get("hasta",  "2026-04-17")
    pais   = request.args.get("pais", "")
    tipo   = request.args.get("tipo", "tc")  # tc | promo
    try:
        ch    = get_client()
        label = "Tarjeta de Crédito" if tipo == "tc" else "Promocode"
        q_map = {"kpis":    Q_TC_KPIS,    "trend":   Q_TC_TREND,
                 "ciudades":Q_TC_CIUDADES, "duo":     Q_TC_DUO} if tipo=="tc"            else {"kpis":    Q_PROMO_KPIS,  "trend":   Q_PROMO_TREND,
                 "ciudades":Q_PROMO_CIUDADES, "duo":  Q_PROMO_DUO}

        data   = _exec_pagos_queries(ch, q_map, desde, hasta, pais, "")
        kpis   = _kpis_row(data)
        trend  = _trend_rows(data)
        cities = _ciudad_rows(data)
        duos   = _duo_rows(data)

        # Detalle fila por fila: usar la query de ciudades con detalle por servicio
        # Se reutiliza la CTE con formato correcto vía _exec_pagos_queries
        filtro = _pagos_filtro(pais, "")
        Q_DET  = (_TC_BASE_CTE if tipo=="tc" else _PROMO_BASE_CTE) + """
SELECT
    b.fecha, b.booking_id, b.driver_id, b.passenger_id,
    b.g_country AS pais,
    CASE b.g_country
        WHEN 'CO' THEN 'Colombia' WHEN 'MX' THEN 'Mexico'
        WHEN 'NI' THEN 'Nicaragua' WHEN 'GT' THEN 'Guatemala'
        ELSE b.g_country END AS pais_nombre,
    b.monto     AS costo_final,
    CASE
        WHEN coalesce(pd.pagado,0)>0 THEN 'OK'
        WHEN b.cancel_lon IS NOT NULL AND b.cancel_lat IS NOT NULL
             AND geoDistance(b.cancel_lon,b.cancel_lat,b.end_lon,b.end_lat)<=b.radio
        THEN 'Mala Práctica'
        ELSE 'Fraude'
    END AS clasificacion
FROM b
LEFT JOIN pd ON b.booking_id=pd.booking_id
ORDER BY b.fecha DESC
LIMIT 3000
"""
        r_det  = ch.query(Q_DET.format(desde=desde, hasta=hasta, filtro=filtro))
        rows   = [dict(zip(r_det.column_names, row)) for row in r_det.result_rows]

        wb = _xl_make_workbook()

        # ── Hoja Resumen ─────────────────────────────────────
        ws1 = wb.create_sheet("Resumen")
        ws1.sheet_properties.tabColor = "7C3AED"
        _xl_banner(ws1, f"Pagos — {label}",
                   f"Período: {desde} → {hasta}  ·  País: {pais or 'Todos'}", 4)
        row = 4
        row = _xl_kpi_section(ws1, row, "Métricas de pagos", [
            ("Total servicios",  kpis.get("total",0)),
            ("OK",               kpis.get("ok",0),            "DCFCE7"),
            ("Mala práctica",    kpis.get("mala_practica",0), "FEF9C3"),
            ("Fraude",           kpis.get("fraude",0),        "FEE2E2"),
            ("Monto total",      f"$ {kpis.get('monto_total',0):,.0f}"),
        ], ncols=4)
        if cities:
            _xl_kpi_section(ws1, row, "Top ciudades", [
                (c.get("ciudad","?"),
                 f"OK:{c.get('ok',0)} MP:{c.get('mala_practica',0)} F:{c.get('fraude',0)}")
                for c in cities[:15]
            ], ncols=4)
        _xl_autofit_safe(ws1)
        ws1.freeze_panes = ws1.cell(4, 1)
        ws1.sheet_view.showGridLines = False

        # ── Hoja Detalle ─────────────────────────────────────
        headers = ["Fecha","Booking ID","Driver ID","User ID","País","Costo Final","Clasificación"]
        ws2 = wb.create_sheet("Detalle")
        ws2.sheet_properties.tabColor = "9333EA"
        _xl_banner(ws2, f"Detalle Pagos — {label}",
                   f"{len(rows)} registros  ·  {desde} → {hasta}", len(headers))
        _xl_headers(ws2, 3, headers)
        CLS_MAP = {"OK":"OK","Mala Práctica":"MALA_PRACTICA","Fraude":"FRAUDE"}
        for i, r in enumerate(rows):
            xl_row = i + 4
            vals = [
                str(r.get("fecha",""))[:10],
                r.get("booking_id",""), r.get("driver_id",""),
                r.get("passenger_id",""), r.get("pais_nombre",""),
                r.get("costo_final") or 0,
                r.get("clasificacion",""),
            ]
            _xl_data_row(ws2, xl_row, vals, aligns={6:"right"})
            _xl_color_cell(ws2, xl_row, 7, vals[6], CLS_MAP.get(vals[6], vals[6]))
        _xl_finalize(ws2, freeze_row=4)
        return _xl_to_response(wb, f"Picap_Pagos_{tipo.upper()}")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500


@app.route("/api/exportar/recaudos")
def exportar_recaudos():
    desde  = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta  = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    moneda = request.args.get("moneda", "")
    try:
        filtro_moneda = (f"AND JSONExtractString(wat.amount,'currency_iso')=\'{moneda}\'"
                         if moneda else "")
        sql = _Q_RECAUDOS.format(desde=desde, hasta=hasta, filtro_moneda=filtro_moneda)
        ch   = get_client()
        r    = ch.query(sql)
        rows_raw = [dict(zip(r.column_names, row)) for row in r.result_rows]
        # normalizar
        for row_r in rows_raw:
            for k in ["suma_negativos","suma_positivos","balance_neto"]:
                try: row_r[k] = round(float(row_r[k] or 0), 2)
                except: row_r[k] = 0.0
            row_r["fecha_tx"] = str(row_r.get("fecha_tx",""))[:16]

        total    = len(rows_raw)
        n_corr   = sum(1 for r2 in rows_raw if r2.get("clasificacion")=="Correcto")
        n_demas  = sum(1 for r2 in rows_raw if r2.get("clasificacion")=="Pagado_demas")
        n_deuda  = sum(1 for r2 in rows_raw if r2.get("clasificacion")=="Debe_dinero")
        n_rev    = sum(1 for r2 in rows_raw if r2.get("clasificacion")=="Revisar")

        wb = _xl_make_workbook()

        # Hoja Estadística
        ws1 = wb.create_sheet("Estadística")
        ws1.sheet_properties.tabColor = "16A34A"
        _xl_banner(ws1, "Validación de Recaudos — Estadística",
                   f"Período: {desde} → {hasta}  ·  Moneda: {moneda or 'Todas'}", 3)
        _xl_kpi_section(ws1, 4, "Resumen de recaudos", [
            ("Total transacciones",    total),
            ("Está correcto",          n_corr,  "DCFCE7"),
            ("Se pagó demás",          n_demas, "FEF9C3"),
            ("Se debe dinero",         n_deuda, "FEE2E2"),
            ("Revisar inconsistencia", n_rev,   "EDE9F5"),
            ("Valor total deuda",
             f"$ {sum(abs(r2['balance_neto']) for r2 in rows_raw if r2.get('clasificacion')=='Debe_dinero'):,.2f}"),
            ("Valor total exceso",
             f"$ {sum(r2['balance_neto'] for r2 in rows_raw if r2.get('clasificacion')=='Pagado_demas'):,.2f}"),
        ], ncols=3)
        _xl_autofit_safe(ws1)
        ws1.freeze_panes = ws1.cell(4, 1)
        ws1.sheet_view.showGridLines = False

        # Hoja Alertas — solo casos anómalos
        alertas = [r2 for r2 in rows_raw
                   if r2.get("clasificacion") in ("Pagado_demas","Debe_dinero","Revisar")]
        alertas.sort(key=lambda x: abs(x.get("balance_neto",0)), reverse=True)
        headers = ["Clasificación", "Booking ID", "Tipo TX", "Moneda",
                   "Negativos", "Positivos", "Balance neto", "Fecha"]
        ws2 = wb.create_sheet("Alertas")
        ws2.sheet_properties.tabColor = "DC2626"
        _xl_banner(ws2, "Validación de Recaudos — Alertas",
                   f"{len(alertas)} alertas  ·  {desde} → {hasta}", len(headers))
        _xl_headers(ws2, 3, headers)
        CLS_MAP = {"Pagado_demas":"PAGADO_DEMAS","Debe_dinero":"DEBE_DINERO",
                   "Revisar":"REVISAR","Correcto":"CORRECTO"}
        for i, a in enumerate(alertas):
            xl_row = i + 4
            vals = [
                a.get("clasificacion",""),
                a.get("id_booking",""), a.get("tipo_tx",""), a.get("moneda",""),
                a.get("suma_negativos",0), a.get("suma_positivos",0),
                a.get("balance_neto",0), a.get("fecha_tx",""),
            ]
            _xl_data_row(ws2, xl_row, vals, aligns={5:"right",6:"right",7:"right"})
            _xl_color_cell(ws2, xl_row, 1, vals[0], CLS_MAP.get(vals[0], vals[0]))
        _xl_finalize(ws2, freeze_row=4)
        return _xl_to_response(wb, "Picap_Recaudos")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500


# ══════════════════════════════════════════════════════════════
# Admin: Editar usuario (nombre y email)
# ══════════════════════════════════════════════════════════════
@app.route("/api/admin/editar_usuario", methods=["POST"])
def admin_editar_usuario():
    token = request.headers.get("X-Token","")
    try:
        ch = get_client()
        sesion = _verificar_sesion(ch, token)
        if not sesion or sesion.get("rol") != "admin":
            return jsonify({"ok":False,"error":"Solo admins"}), 403
    except Exception as e:
        return jsonify({"ok":False,"error":f"Sin permisos: {e}"}), 403

    data    = request.get_json() or {}
    usuario = (data.get("usuario") or "").strip()
    nombre  = (data.get("nombre")  or "").strip()
    email   = (data.get("email")   or "").strip()
    if not usuario:
        return jsonify({"ok":False,"error":"usuario requerido"}), 400
    try:
        ch = get_client()
        # Leer fila actual con todos sus campos (incluyendo password_hash y activo)
        r = ch.query(
            f"SELECT usuario, password_hash, nombre, email, rol, creado_en, activo "
            f"FROM picapmongoprod.dashboard_users FINAL "
            f"WHERE usuario='{usuario}' AND activo=1 LIMIT 1"
        )
        rows = r.result_rows
        if not rows:
            return jsonify({"ok":False,"error":"Usuario no encontrado"}), 404
        cur = rows[0]
        # Re-insertar con nombre/email actualizados — ReplacingMergeTree deduplica
        from datetime import datetime as _dt2
        ch.insert(
            "picapmongoprod.dashboard_users",
            [[cur[0],
              cur[1],                    # password_hash sin cambio
              nombre or cur[2],          # nombre nuevo o actual
              email  or cur[3],          # email nuevo o actual
              cur[4],                    # rol sin cambio
              cur[5],                    # creado_en sin cambio
              1]],                       # activo=1
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"]
        )
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# Admin: Eliminar usuario (activo=0, soft delete)
# ══════════════════════════════════════════════════════════════
@app.route("/api/admin/eliminar_usuario", methods=["POST"])
def admin_eliminar_usuario():
    token = request.headers.get("X-Token","")
    try:
        ch = get_client()
        sesion = _verificar_sesion(ch, token)
        if not sesion or sesion.get("rol") != "admin":
            return jsonify({"ok":False,"error":"Solo admins"}), 403
    except Exception as e:
        return jsonify({"ok":False,"error":f"Sin permisos: {e}"}), 403

    data    = request.get_json() or {}
    usuario = (data.get("usuario") or "").strip()
    if not usuario:
        return jsonify({"ok":False,"error":"usuario requerido"}), 400
    if usuario == sesion.get("sub",""):
        return jsonify({"ok":False,"error":"No puedes eliminarte a ti mismo"}), 400
    try:
        ch = get_client()
        # Soft delete: insertar con activo=0 (ReplacingMergeTree usará la versión más nueva)
        r = ch.query(
            f"SELECT usuario, password_hash, nombre, email, rol, creado_en "
            f"FROM picapmongoprod.dashboard_users FINAL "
            f"WHERE usuario='{usuario}' AND activo=1 LIMIT 1"
        )
        if not r.result_rows:
            return jsonify({"ok":False,"error":"Usuario no encontrado"}), 404
        cur = r.result_rows[0]
        from datetime import datetime as _dt3
        ch.insert(
            "picapmongoprod.dashboard_users",
            [[cur[0], cur[1], cur[2], cur[3], cur[4], cur[5], 0]],
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"]
        )
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500


# NOTA: el arranque de Flask se mueve al final del archivo para que TODAS
# las rutas (incluyendo /api/login, /api/me, /api/admin/*, /api/pibox/*)
# queden registradas antes de bloquear con app.run().

# ══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
# Para ejecutar la ubicación del archivo api.py --> cd C:\Users\Picap\Documents\AUTOMATIZACIONES\AUTOMATIZACIONES\dashboards\picap_evasion_dashboard
# ══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
# Para ejecutar el script: python api.py
# ══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
# Para correrlo online en web: python -m http.server 8080
# ══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
# link: https://picap-monitoreo.onrender.com/
# ══════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# SISTEMA DE USUARIOS Y ROLES
# Configuración centralizada — sin BD externa
# ══════════════════════════════════════════════════════════════
import os
import hashlib, secrets, time
from datetime import datetime as _dt

# ── Auth viejo (dict-based) eliminado — solo se usa el sistema CH con HMAC ──

# ══════════════════════════════════════════════════════════════
# ENDPOINT: Estadísticas consolidadas de bloqueos + reactivaciones
# ══════════════════════════════════════════════════════════════
Q_STATS_BLOQUEOS = """
WITH
suspensiones AS (
    SELECT
        passenger_id AS id,
        created_at,
        ends_at,
        'suspension_usuario' AS tipo
    FROM picapmongoprod.passenger_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
    UNION ALL
    SELECT
        driver_id AS id,
        created_at,
        ends_at,
        'suspension_piloto' AS tipo
    FROM picapmongoprod.driver_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
),
enriquecido AS (
    SELECT
        s.id,
        s.created_at,
        s.ends_at,
        s.tipo,
        p.driver_enrollment_status_cd,
        CASE WHEN p.driver_enrollment_status_cd = 3 THEN 'PILOTO' ELSE 'USUARIO' END AS tipo_cuenta,
        CASE WHEN lower(ifNull(toString(p.expelled),'')) = 'true' THEN 'EXPULSADO' ELSE 'SUSPENDIDO' END AS clase_bloqueo,
        -- Reactivado = actualmente activo
        CASE
            WHEN lower(ifNull(toString(p.suspended),'')) = 'false'
             AND lower(ifNull(toString(p.expelled),''))  = 'false'
             AND ifNull(p.is_driver_suspended, 0) = 0
            THEN 1 ELSE 0
        END AS fue_reactivado,
        -- ends_at ficticia (año > 2100) = bloqueo indefinido
        CASE WHEN toYear(s.ends_at) >= 2100 OR s.ends_at IS NULL THEN 0 ELSE 1 END AS tiene_fecha_fin,
        CASE
            WHEN p.g_country = 'CO' THEN 'Colombia'
            WHEN p.g_country = 'MX' THEN 'Mexico'
            WHEN p.g_country = 'NI' THEN 'Nicaragua'
            WHEN p.g_country = 'GT' THEN 'Guatemala'
            ELSE ifNull(p.g_country, 'Otro')
        END AS pais
    FROM suspensiones s
    LEFT JOIN picapmongoprod.passengers p ON p._id = s.id
)
SELECT
    -- Totales
    count(DISTINCT id)                                          AS total_bloqueados,
    countDistinctIf(id, tipo_cuenta = 'PILOTO')                AS pilotos_bloqueados,
    countDistinctIf(id, tipo_cuenta = 'USUARIO')               AS usuarios_bloqueados,
    -- Clasificación
    countDistinctIf(id, clase_bloqueo = 'SUSPENDIDO')          AS total_suspendidos,
    countDistinctIf(id, clase_bloqueo = 'EXPULSADO')           AS total_expulsados,
    -- Reactivaciones (de los suspendidos)
    countDistinctIf(id, clase_bloqueo = 'SUSPENDIDO' AND fue_reactivado = 1) AS reactivados,
    countDistinctIf(id, clase_bloqueo = 'SUSPENDIDO' AND fue_reactivado = 0) AS siguen_bloqueados,
    -- Por país
    groupArray((pais, toString(countDistinct(id)))) AS por_pais_raw
FROM enriquecido
GROUP BY pais
"""

Q_STATS_BLOQUEOS_RESUMEN = """
SELECT
    count()                                                        AS total_bloqueados,
    countIf(driver_enrollment_status_cd = 3)                      AS pilotos_bloqueados,
    countIf(driver_enrollment_status_cd != 3)                     AS usuarios_bloqueados,
    countIf(lower(ifNull(toString(expelled),'')) != 'true')        AS total_suspendidos,
    countIf(lower(ifNull(toString(expelled),'')) = 'true')         AS total_expulsados,
    countIf(
        lower(ifNull(toString(expelled),'')) != 'true'
        AND lower(ifNull(toString(suspended),'')) = 'false'
        AND lower(ifNull(toString(expelled),''))  = 'false'
        AND ifNull(is_driver_suspended, 0) = 0
    )                                                              AS reactivados,
    countIf(
        lower(ifNull(toString(expelled),'')) != 'true'
        AND NOT (
            lower(ifNull(toString(suspended),'')) = 'false'
            AND lower(ifNull(toString(expelled),''))  = 'false'
            AND ifNull(is_driver_suspended, 0) = 0
        )
    )                                                              AS siguen_bloqueados
FROM picapmongoprod.passengers p
WHERE p._id IN (
    SELECT passenger_id FROM picapmongoprod.passenger_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
    UNION DISTINCT
    SELECT driver_id FROM picapmongoprod.driver_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
)
"""

Q_STATS_BLOQUEOS_PAIS = """
SELECT
    CASE
        WHEN p.g_country = 'CO' THEN 'Colombia'
        WHEN p.g_country = 'MX' THEN 'Mexico'
        WHEN p.g_country = 'NI' THEN 'Nicaragua'
        WHEN p.g_country = 'GT' THEN 'Guatemala'
        ELSE ifNull(p.g_country, 'Otro')
    END                                                            AS pais,
    count()                                                        AS total,
    countIf(p.driver_enrollment_status_cd = 3)                    AS pilotos,
    countIf(p.driver_enrollment_status_cd != 3)                   AS usuarios,
    countIf(lower(ifNull(toString(p.expelled),'')) != 'true')      AS suspendidos,
    countIf(lower(ifNull(toString(p.expelled),'')) = 'true')       AS expulsados,
    countIf(
        lower(ifNull(toString(p.suspended),'')) = 'false'
        AND lower(ifNull(toString(p.expelled),''))  = 'false'
        AND ifNull(p.is_driver_suspended, 0) = 0
    )                                                              AS reactivados,
    round(countIf(
        lower(ifNull(toString(p.suspended),'')) = 'false'
        AND lower(ifNull(toString(p.expelled),''))  = 'false'
        AND ifNull(p.is_driver_suspended, 0) = 0
    ) / count() * 100, 1)                                          AS pct_reactivados
FROM picapmongoprod.passengers p
WHERE p._id IN (
    SELECT passenger_id FROM picapmongoprod.passenger_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
    UNION DISTINCT
    SELECT driver_id FROM picapmongoprod.driver_suspensions
    WHERE created_at IS NOT NULL
      AND created_at >= toDateTime('{desde} 00:00:00')
      AND created_at <= toDateTime('{hasta} 23:59:59')
)
{filtro_driver}
GROUP BY pais
ORDER BY total DESC
"""

@app.route("/api/estadisticas_bloqueos")
def estadisticas_bloqueos():
    desde      = (request.args.get("desde") or (date.today()-timedelta(days=30)).strftime("%Y-%m-%d"))
    hasta      = (request.args.get("hasta") or date.today().strftime("%Y-%m-%d"))
    pais_fil   = request.args.get("pais", "")
    driver_id  = request.args.get("driver_id", "").strip()
    try:
        ch = get_client()
        # Resumen global
        r = ch.query(Q_STATS_BLOQUEOS_RESUMEN.format(desde=desde, hasta=hasta))
        res = dict(zip(r.column_names, r.result_rows[0])) if r.result_rows else {}
        total = int(res.get("total_bloqueados", 0)) or 1

        # Por país
        # filtro_driver: AND en el WHERE de la subquery IN
        filtro_driver = f"AND p._id = '{driver_id}'" if driver_id else ""
        rp = ch.query(Q_STATS_BLOQUEOS_PAIS.format(
            desde=desde, hasta=hasta, filtro_driver=filtro_driver))
        paises = [dict(zip(rp.column_names, row)) for row in rp.result_rows]
        if pais_fil:
            paises = [p for p in paises if p.get("pais") == pais_fil]

        reactivados = int(res.get("reactivados", 0))
        suspendidos = int(res.get("total_suspendidos", 0))

        return jsonify({
            "desde": desde, "hasta": hasta,
            "resumen": {
                "total_bloqueados":  int(res.get("total_bloqueados", 0)),
                "pilotos_bloqueados":int(res.get("pilotos_bloqueados", 0)),
                "usuarios_bloqueados":int(res.get("usuarios_bloqueados", 0)),
                "total_suspendidos": int(res.get("total_suspendidos", 0)),
                "total_expulsados":  int(res.get("total_expulsados", 0)),
                "reactivados":       reactivados,
                "siguen_bloqueados": int(res.get("siguen_bloqueados", 0)),
                "pct_reactivados":   round(reactivados / suspendidos * 100, 1) if suspendidos else 0,
                "pct_pilotos":       round(int(res.get("pilotos_bloqueados",0))/total*100,1),
                "pct_usuarios":      round(int(res.get("usuarios_bloqueados",0))/total*100,1),
            },
            "por_pais": paises,
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detalle": traceback.format_exc()}), 500

# ══════════════════════════════════════════════════════════════
# SISTEMA DE USUARIOS PERSISTENTE EN CLICKHOUSE
# Los usuarios se registran y almacenan en una tabla CH.
# El admin asigna roles desde el panel.
# ══════════════════════════════════════════════════════════════
import os
import hashlib, secrets, time
from datetime import datetime as _dt

ROLES_ACCESO = {
    # admin: TODO incluyendo Panel Admin
    "admin":      ["monitoreo","conversor","bloqueos","wallet","pagos",
                   "retencion","cashout","reconocimiento","pibox",
                   "estafa","recaudos","auditoria","cedula",
                   "admin_panel","home"],
    # monitoreo: TODO excepto Panel Admin
    "monitoreo":  ["monitoreo","conversor","bloqueos","wallet","pagos",
                   "retencion","cashout","reconocimiento","pibox",
                   "estafa","recaudos","auditoria","cedula","home"],
    # financiero: módulos financieros
    "financiero": ["pagos","wallet","retencion","cashout",
                   "recaudos","auditoria","pibox","home"],
    # sac: bloqueos, reconocimiento, estafa y alerta de cédula
    "sac":        ["bloqueos","reconocimiento","estafa","cedula","home"],
    # pibox: solo pibox, recaudos y auditoría
    "pibox":      ["pibox","recaudos","auditoria","home"],
    "pendiente":  ["home"],
}

# ── Tokens firmados HMAC (sin estado en servidor) ────────────
# Sobreviven reinicios de Render porque la verificación es matemática,
# no depende de ningún dict ni tabla externa.
import hmac, base64, json as _json

# IMPORTANTE: en producción definir TOKEN_SECRET vía variable de entorno.
# El default solo sirve para desarrollo local con datos no sensibles.
TOKEN_SECRET = os.environ.get('TOKEN_SECRET', 'dev_only_change_me_in_production')
SESSION_TTL  = 24 * 3600  # 24 horas

# ── Diagnóstico de configuración al arranque ──
# Imprime una matriz clara de qué env vars están definidas. Útil para
# debuggear en Render cuando algo no logra conectar.
def _config_diagnostico():
    flags = [
        ("CLICKHOUSE_HOST",          bool(os.environ.get("CLICKHOUSE_HOST"))),
        ("CLICKHOUSE_PORT",          bool(os.environ.get("CLICKHOUSE_PORT"))),
        ("CLICKHOUSE_USER",          bool(os.environ.get("CLICKHOUSE_USER"))),
        ("CLICKHOUSE_PASSWORD",      bool(os.environ.get("CLICKHOUSE_PASSWORD"))),
        ("CLICKHOUSE_DATABASE",      bool(os.environ.get("CLICKHOUSE_DATABASE"))),
        ("TOKEN_SECRET",             bool(os.environ.get("TOKEN_SECRET"))),
        ("ADMIN_INITIAL_PASSWORD",   bool(os.environ.get("ADMIN_INITIAL_PASSWORD"))),
        ("ADMIN_EMERGENCY_PASSWORD", bool(os.environ.get("ADMIN_EMERGENCY_PASSWORD"))),
        ("SMTP_EMAIL/SMTP_USER",     bool(os.environ.get("SMTP_EMAIL") or os.environ.get("SMTP_USER"))),
        ("SMTP_PASSWORD/SMTP_PASS",  bool(os.environ.get("SMTP_PASSWORD") or os.environ.get("SMTP_PASS"))),
    ]
    print("=" * 60)
    print("  CONFIG DIAGNOSTICO — variables de entorno detectadas:")
    for name, ok in flags:
        print(f"    {'OK ' if ok else 'NO '} {name}")
    if not any(ok for n, ok in flags if n.startswith("CLICKHOUSE")):
        print("  AVISO: ClickHouse no esta configurado. La autenticacion via CH")
        print("  va a fallar. Define CLICKHOUSE_PASSWORD (al menos) en el panel")
        print("  de Render -> Environment, y los usuarios podran iniciar sesion.")
    print("=" * 60)
try: _config_diagnostico()
except Exception as _diag_err: print(f"[diag] error: {_diag_err}")

def _guardar_sesion(ch, token, usuario, rol):
    pass  # No necesario — el token ya lleva la info firmada

def _crear_token(usuario, rol):
    payload = _json.dumps({'u': usuario, 'r': rol, 'ts': time.time()})
    payload_b64 = base64.urlsafe_b64encode(payload.encode()).decode()
    sig = hmac.new(TOKEN_SECRET.encode(), payload_b64.encode(), 'sha256').hexdigest()
    return f"{payload_b64}.{sig}"

def _verificar_sesion(ch, token):
    """Verifica firma HMAC y expiración. No necesita BD ni memoria."""
    if not token: return None
    try:
        parts = token.split('.')
        if len(parts) != 2:
            return None
        payload_b64, sig = parts[0], parts[1]
        # Verificar firma
        expected = hmac.new(TOKEN_SECRET.encode(), payload_b64.encode(), 'sha256').hexdigest()
        if not hmac.compare_digest(sig, expected):
            return None
        # Decodificar payload
        payload = _json.loads(base64.urlsafe_b64decode(payload_b64 + '=='))
        # Verificar expiración
        if time.time() - payload.get('ts', 0) > SESSION_TTL:
            return None
        return {'usuario': payload['u'], 'rol': payload['r']}
    except Exception:
        return None

def _ch_init_users():
    """Crear tabla de usuarios si no existe e insertar admin por defecto."""
    try:
        ch = get_client()
        ch.command("""
            CREATE TABLE IF NOT EXISTS picapmongoprod.dashboard_users (
                usuario      String,
                password_hash String,
                nombre       String,
                email        String,
                rol          String DEFAULT 'pendiente',
                creado_en    DateTime DEFAULT now(),
                activo       UInt8 DEFAULT 1
            ) ENGINE = ReplacingMergeTree(creado_en)
            ORDER BY usuario
        """)
        # FINAL fuerza deduplicacion en ReplacingMergeTree
        r = ch.query("SELECT count() FROM picapmongoprod.dashboard_users FINAL WHERE usuario='admin'")
        if r.result_rows[0][0] == 0:
            # Solo seedeamos el admin si el operador define una contraseña inicial
            # vía ADMIN_INITIAL_PASSWORD. Sin esa env var, el admin debe crearse
            # manualmente desde el panel.
            seed_pwd = os.environ.get('ADMIN_INITIAL_PASSWORD', '')
            if seed_pwd:
                ch.insert("picapmongoprod.dashboard_users",
                    [["admin",
                      hashlib.sha256(seed_pwd.encode()).hexdigest(),
                      "Administrador",
                      os.environ.get('ADMIN_INITIAL_EMAIL', 'admin@picap.com'),
                      "admin", _dt.utcnow(), 1]],
                    column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
                print("[users init] Admin creado en dashboard_users")
            else:
                print("[users init] No hay admin y ADMIN_INITIAL_PASSWORD no está definida — créalo manualmente.")
        else:
            print("[users init] Admin ya existe en dashboard_users")
    except Exception as e:
        print(f"[users init] ERROR: {e}")

# Inicializar tabla al arrancar
try: _ch_init_users()
except: pass

@app.route("/api/register", methods=["POST"])
def register():
    data     = request.get_json() or {}
    usuario  = data.get("usuario","").strip().lower()
    password = data.get("password","").strip()
    nombre   = data.get("nombre","").strip()
    email    = data.get("email","").strip()
    if not all([usuario, password, nombre]):
        return jsonify({"ok":False,"error":"Usuario, nombre y contraseña son obligatorios"}), 400
    if len(password) < 6:
        return jsonify({"ok":False,"error":"La contraseña debe tener al menos 6 caracteres"}), 400
    try:
        ch = get_client()
        r  = ch.query(f"SELECT count() FROM picapmongoprod.dashboard_users WHERE usuario='{usuario}'")
        if r.result_rows[0][0] > 0:
            return jsonify({"ok":False,"error":"El usuario ya existe"}), 409
        pwd_hash = hashlib.sha256(password.encode()).hexdigest()
        ch.insert("picapmongoprod.dashboard_users",
            [[usuario, pwd_hash, nombre, email, "pendiente", _dt.utcnow(), 1]],
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
        return jsonify({"ok":True,"mensaje":"Usuario creado. Espera que un administrador asigne tu rol."})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

# ── Credenciales admin de emergencia (siempre funcionan aunque falle CH) ──────
# La contraseña real se setea con la variable de entorno ADMIN_EMERGENCY_PASSWORD.
# Si no está definida, NO se permite login de emergencia (password vacío -> hash inalcanzable).
_admin_emergency_pwd = os.environ.get('ADMIN_EMERGENCY_PASSWORD', '')
_ADMIN_EMERGENCY = {
    "admin": {
        "password_hash": hashlib.sha256(_admin_emergency_pwd.encode()).hexdigest() if _admin_emergency_pwd else '',
        "nombre": "Administrador",
        "rol": "admin",
        "email": os.environ.get('ADMIN_EMERGENCY_EMAIL', 'admin@picap.com'),
    }
} if _admin_emergency_pwd else {}

@app.route("/api/login", methods=["POST"])
def login_ch():
    data     = request.get_json() or {}
    usuario  = data.get("usuario","").strip().lower()
    password = data.get("password","").strip()
    if not usuario or not password:
        return jsonify({"ok":False,"error":"Credenciales requeridas"}), 400

    pwd_hash = hashlib.sha256(password.encode()).hexdigest()

    # Trackear el estado de la consulta a ClickHouse para distinguir entre:
    #   - usuario realmente no encontrado
    #   - CH inaccesible (probablemente env vars mal configuradas)
    ch_error  = None
    ch_buscado = False

    # 1) Intentar login via ClickHouse (usuarios registrados + admin en tabla)
    try:
        ch = get_client()
        # FINAL fuerza deduplicacion en ReplacingMergeTree — evita filas fantasma
        r  = ch.query(f"""
            SELECT usuario, password_hash, nombre, email, rol
            FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario=\'{usuario}\' AND activo=1
            LIMIT 1
        """)
        ch_buscado = True
        if r.result_rows:
            row     = r.result_rows[0]
            db_hash = row[1]
            if pwd_hash != db_hash:
                return jsonify({"ok":False,"error":"Contraseña incorrecta"}), 401
            rol    = row[4] or "pendiente"
            acceso = ROLES_ACCESO.get(rol, ["home"])
            token  = _crear_token(usuario, rol)
            return jsonify({
                "ok":True, "token":token,
                "nombre":row[2], "rol":rol,
                "acceso":acceso, "email":row[3]
            })
    except Exception as e:
        ch_error = str(e)
        print(f"[login] CH error: {e}")

    # 2) Fallback: credenciales de emergencia configuradas vía env var
    em = _ADMIN_EMERGENCY.get(usuario)
    if em and em.get("password_hash") and pwd_hash == em["password_hash"]:
        rol    = em["rol"]
        acceso = ROLES_ACCESO.get(rol, ["home"])
        token  = _crear_token(usuario, rol)
        # Intentar sincronizar admin a CH en segundo plano
        def _sync():
            try:
                ch2 = get_client()
                ch2.insert("picapmongoprod.dashboard_users",
                    [[usuario, em["password_hash"], em["nombre"],
                      em["email"], rol, _dt.utcnow(), 1]],
                    column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
            except: pass
        threading.Thread(target=_sync, daemon=True).start()
        return jsonify({
            "ok":True, "token":token,
            "nombre":em["nombre"], "rol":rol,
            "acceso":acceso, "email":em["email"]
        })

    # 3) No se pudo autenticar. Distinguir el caso CH muerto del caso usuario inexistente.
    if not ch_buscado:
        # No alcanzamos a consultar CH — los usuarios existen pero no podemos validar.
        return jsonify({
            "ok": False,
            "error": "Servicio de autenticación no disponible. Contacta al administrador.",
            "code": "auth_backend_unavailable",
            "detalle": (ch_error or 'sin detalle')[:200],
        }), 503

    return jsonify({"ok":False,"error":"Usuario no encontrado"}), 401

@app.route("/api/logout", methods=["POST"])
def logout_ch():
    # Con tokens HMAC no hay estado en servidor que borrar
    # El frontend borra el token de localStorage
    return jsonify({"ok":True})

@app.route("/api/me")
def me_ch():
    token = request.headers.get("X-Token","")
    # 1) Verificar el token PRIMERO (no requiere CH). Si la firma o expiración
    #    fallan → 401 (la sesión es inválida y el frontend debe cerrar).
    s = _verificar_sesion(None, token)
    if not s:
        return jsonify({"ok":False,"error":"Sesión expirada","code":"session_invalid"}), 401

    # 2) Token válido — leer datos del usuario en CH. Si CH falla, devolvemos
    #    503 (service unavailable). El frontend NO debe cerrar la sesión por
    #    un 503: es problema del backend, no del usuario.
    try:
        ch = get_client()
        r  = ch.query(f"""
            SELECT nombre, rol, email FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario='{s["usuario"]}' AND activo=1 LIMIT 1
        """)
        if not r.result_rows:
            # Usuario no encontrado o inactivo → trata como sesión inválida
            return jsonify({"ok":False,"error":"Usuario inactivo","code":"user_inactive"}), 401
        row = r.result_rows[0]
        rol = row[1] or "pendiente"
        return jsonify({"ok":True,"nombre":row[0],"rol":rol,
                        "acceso":ROLES_ACCESO.get(rol,["home"]),"email":row[2]})
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": "Backend de datos no disponible",
            "code": "backend_unavailable",
            "detalle": str(e)[:200],
        }), 503


# ══════════════════════════════════════════════════════════════════════════════
# CAMBIO Y RECUPERACIÓN DE CONTRASEÑA
# ══════════════════════════════════════════════════════════════════════════════
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Variables de entorno para el email (configurar en Render)
# Acepta SMTP_EMAIL (canónico) o SMTP_USER (alias) para retrocompatibilidad.
# Igual con la contraseña: SMTP_PASSWORD o SMTP_PASS.
SMTP_HOST     = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT     = int(os.environ.get('SMTP_PORT', '465'))
SMTP_EMAIL    = (os.environ.get('SMTP_EMAIL', '') or
                 os.environ.get('SMTP_USER', '')).strip()       # ej: picap.monitoreo@gmail.com
SMTP_PASSWORD = (os.environ.get('SMTP_PASSWORD', '') or
                 os.environ.get('SMTP_PASS', '')).strip()        # contraseña de app Gmail
APP_URL        = os.environ.get('APP_URL', 'https://picap-monitoreo.onrender.com')

def _crear_reset_token(usuario, email):
    """Token HMAC válido 1 hora para reset de contraseña."""
    payload = _json.dumps({'u': usuario, 'e': email, 'ts': time.time(), 'tipo': 'reset'})
    payload_b64 = base64.urlsafe_b64encode(payload.encode()).decode()
    sig = hmac.new(TOKEN_SECRET.encode(), payload_b64.encode(), 'sha256').hexdigest()
    return f"{payload_b64}.{sig}"

def _verificar_reset_token(token):
    """Verifica token de reset. Retorna dict con usuario/email o None."""
    try:
        parts = token.split('.')
        if len(parts) != 2:
            return None
        payload_b64, sig = parts
        expected = hmac.new(TOKEN_SECRET.encode(), payload_b64.encode(), 'sha256').hexdigest()
        if not hmac.compare_digest(sig, expected):
            return None
        payload = _json.loads(base64.urlsafe_b64decode(payload_b64).decode())
        if payload.get('tipo') != 'reset':
            return None
        if time.time() - payload.get('ts', 0) > 3600:  # 1 hora
            return None
        return payload
    except:
        return None

RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '').strip()
BREVO_API_KEY  = os.environ.get('BREVO_API_KEY', '').strip()

def _enviar_email_brevo(destinatario, asunto, cuerpo_html):
    """Envía email vía Brevo (antes SendinBlue) HTTP API.
    Free tier: 300 emails/día sin verificación de dominio. Funciona en
    Render Free porque es HTTPS estándar (no SMTP)."""
    import urllib.request, urllib.error, json as _j
    UA = "PicapMonitoreo/1.0 (+https://picap-monitoreo.onrender.com)"
    # Remitente: usa el SMTP_EMAIL configurado, o uno por defecto.
    # Brevo permite cualquier email como remitente siempre que esté verificado
    # en su panel (te llega un correo de verificación al registrarte).
    sender_email = (os.environ.get('BREVO_FROM', '').strip() or
                    SMTP_EMAIL or
                    'noreply@picap-monitoreo.onrender.com')
    payload = _j.dumps({
        "sender":      {"email": sender_email, "name": "Picap Monitoreo"},
        "to":          [{"email": destinatario}],
        "subject":     asunto,
        "htmlContent": cuerpo_html,
    }).encode('utf-8')
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email",
        data=payload,
        headers={
            "api-key":      BREVO_API_KEY,
            "Content-Type": "application/json",
            "Accept":       "application/json",
            "User-Agent":   UA,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
            print(f"[brevo] OK from={sender_email} → {destinatario}: {body[:200]}")
            return True, None
    except urllib.error.HTTPError as e:
        body = ''
        try: body = e.read().decode('utf-8', errors='replace')[:400]
        except Exception: pass
        if e.code == 401:
            err = f"Brevo rechazó la API KEY (HTTP 401). Verifica BREVO_API_KEY en Render. Detalle: {body}"
        elif e.code == 400:
            err = (f"Brevo rechazó el envío (400). Causa común: el remitente '{sender_email}' "
                   f"no está verificado en Brevo. Ve a https://app.brevo.com/senders y verifícalo. "
                   f"Detalle: {body}")
        elif e.code == 403:
            err = f"Brevo devolvió 403 (Forbidden). Verifica permisos de tu API KEY. Detalle: {body}"
        else:
            err = f"HTTP {e.code} de Brevo: {body or e.reason}"
        print(f"[brevo] HTTP_ERROR: {err}")
        return False, err
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"[brevo] UNEXPECTED: {err}")
        return False, err

def _enviar_email_resend(destinatario, asunto, cuerpo_html):
    """Envía email vía Resend HTTP API. Funciona en Render Free (no bloqueado).
    Sign up gratis en https://resend.com — 100 emails/día sin tarjeta.

    Usa el dominio sandbox de Resend (onboarding@resend.dev) si no hay un
    dominio verificado configurado vía RESEND_FROM. Para producción,
    verificar dominio propio en https://resend.com/domains y setear:
        RESEND_FROM = 'Picap Monitoreo <noreply@tudominio.com>'
    """
    import urllib.request, urllib.error, json as _j
    # Cloudflare bloquea el UA por defecto de urllib (error 1010). Usar uno
    # que parezca un cliente HTTP normal evita el bloqueo.
    UA = "PicapMonitoreo/1.0 (+https://picap-monitoreo.onrender.com)"

    # Remitente: por defecto el sandbox público de Resend (siempre funciona).
    # Si tienes un dominio verificado, define RESEND_FROM en Render.
    from_addr = (os.environ.get('RESEND_FROM', '').strip()
                 or 'Picap Monitoreo <onboarding@resend.dev>')

    payload = _j.dumps({
        "from":    from_addr,
        "to":      [destinatario],
        "subject": asunto,
        "html":    cuerpo_html,
    }).encode('utf-8')
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type":  "application/json",
            "Accept":        "application/json",
            "User-Agent":    UA,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
            print(f"[resend] OK from={from_addr} → {destinatario}: {body[:200]}")
            return True, None
    except urllib.error.HTTPError as e:
        body = ''
        try: body = e.read().decode('utf-8', errors='replace')[:400]
        except Exception: pass
        # Detección de errores comunes
        if e.code == 401:
            err = f"Resend rechazó la API KEY (HTTP 401). Verifica RESEND_API_KEY en Render. Detalle: {body}"
        elif e.code == 403:
            err = (f"Resend devolvió 403 (Forbidden). Posibles causas: "
                   f"(1) la cuenta de Resend bloqueó la IP de Render — escribir a soporte; "
                   f"(2) el dominio del 'from' no está verificado — usar onboarding@resend.dev "
                   f"o verificar tu dominio. Detalle: {body}")
        elif e.code == 422:
            err = f"Resend rechazó el envío (422 — datos inválidos). Detalle: {body}"
        else:
            err = f"HTTP {e.code} de Resend: {body or e.reason}"
        print(f"[resend] HTTP_ERROR: {err}")
        return False, err
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"[resend] UNEXPECTED: {err}")
        return False, err

def _enviar_email_smtp(destinatario, asunto, cuerpo_html):
    """Envía email vía SMTP (Gmail/otros). Soporta puerto 465 (SSL) y 587 (STARTTLS).
    En Render Free el puerto 465 suele estar bloqueado a nivel de red — usar 587."""
    pwd_norm = SMTP_PASSWORD.replace(' ', '')  # Gmail App Password puede traer espacios
    print(f"[smtp] intentando — host={SMTP_HOST}:{SMTP_PORT} from={SMTP_EMAIL} to={destinatario} pwd_len={len(pwd_norm)}")

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = asunto
        msg['From']    = f'Picap Monitoreo <{SMTP_EMAIL}>'
        msg['To']      = destinatario
        msg.attach(MIMEText(cuerpo_html, 'html', 'utf-8'))
        ctx = ssl.create_default_context()

        if int(SMTP_PORT) == 587:
            # STARTTLS — el server de Render Free suele permitir este puerto
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
                server.ehlo()
                server.starttls(context=ctx)
                server.ehlo()
                server.login(SMTP_EMAIL, pwd_norm)
                server.sendmail(SMTP_EMAIL, destinatario, msg.as_string())
        else:
            # SSL directo (puerto 465) — puede estar bloqueado en Render Free
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx, timeout=15) as server:
                server.login(SMTP_EMAIL, pwd_norm)
                server.sendmail(SMTP_EMAIL, destinatario, msg.as_string())
        print(f"[smtp] OK → {destinatario}")
        return True, None
    except smtplib.SMTPAuthenticationError as e:
        err = f"Autenticación rechazada. Si usas Gmail, debe ser una APP PASSWORD (no la contraseña normal). Detalle: {e}"
        print(f"[smtp] AUTH_ERROR: {err}")
        return False, err
    except (TimeoutError, OSError) as e:
        # Error típico cuando Render Free bloquea el puerto SMTP
        err = (f"Conexión SMTP bloqueada o timeout ({type(e).__name__}: {e}). "
               f"Render Free suele bloquear el puerto {SMTP_PORT}. "
               f"Probá cambiar SMTP_PORT a 587 (STARTTLS), o usar Resend "
               f"(definir RESEND_API_KEY en Environment).")
        print(f"[smtp] NETWORK_BLOCKED: {err}")
        return False, err
    except smtplib.SMTPException as e:
        err = f"Error SMTP: {type(e).__name__}: {e}"
        print(f"[smtp] SMTP_ERROR: {err}")
        return False, err
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"[smtp] UNEXPECTED: {err}")
        return False, err

def _enviar_email(destinatario, asunto, cuerpo_html):
    """Despachador de email. Prioridad:
       1) Brevo (HTTP) si BREVO_API_KEY — permite enviar a cualquier destinatario.
       2) Resend (HTTP) si RESEND_API_KEY — requiere dominio verificado para
          enviar a no-propietarios.
       3) SMTP — último recurso, suele estar bloqueado en Render Free.
    """
    if not destinatario or '@' not in destinatario:
        return False, f"Destinatario inválido: {destinatario}"

    # 1) Brevo (recomendado en Render Free): HTTP, libre de DNS/dominio.
    if BREVO_API_KEY:
        return _enviar_email_brevo(destinatario, asunto, cuerpo_html)

    # 2) Resend
    if RESEND_API_KEY:
        return _enviar_email_resend(destinatario, asunto, cuerpo_html)

    # 3) Fallback a SMTP
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        falta = []
        if not SMTP_EMAIL:    falta.append('SMTP_EMAIL (o SMTP_USER)')
        if not SMTP_PASSWORD: falta.append('SMTP_PASSWORD (o SMTP_PASS)')
        msg = ('Email no configurado. Falta: ' + ', '.join(falta) +
               '. Alternativas HTTP (mejores en Render Free): BREVO_API_KEY '
               '(https://brevo.com — 300/día sin dominio) o RESEND_API_KEY '
               '(https://resend.com — 100/día con dominio verificado).')
        print(f"[email] CONFIG_MISSING: {msg}")
        return False, msg
    return _enviar_email_smtp(destinatario, asunto, cuerpo_html)

@app.route("/api/cambiar_password", methods=["POST"])
def cambiar_password():
    """Cambia contraseña estando logueado."""
    token = request.headers.get("X-Token", "")
    data  = request.get_json() or {}
    pwd_actual  = data.get("password_actual", "").strip()
    pwd_nueva   = data.get("password_nueva",  "").strip()
    try:
        ch = get_client()
        s  = _verificar_sesion(ch, token)
        if not s:
            return jsonify({"ok": False, "error": "Sesión expirada"}), 401
        usuario = s["usuario"]
        if len(pwd_nueva) < 6:
            return jsonify({"ok": False, "error": "La nueva contraseña debe tener al menos 6 caracteres"})
        # Verificar contraseña actual
        hash_actual = hashlib.sha256(pwd_actual.encode()).hexdigest()
        r = ch.query(f"""
            SELECT password_hash, nombre, email, rol FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario=\'{usuario}\' AND activo=1 LIMIT 1
        """)
        if not r.result_rows:
            return jsonify({"ok": False, "error": "Usuario no encontrado"})
        row = r.result_rows[0]
        if row[0] != hash_actual:
            return jsonify({"ok": False, "error": "La contraseña actual es incorrecta"})
        # Actualizar contraseña
        hash_nueva = hashlib.sha256(pwd_nueva.encode()).hexdigest()
        ch.insert("picapmongoprod.dashboard_users",
            [[usuario, hash_nueva, row[1], row[2], row[3], _dt.utcnow(), 1]],
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
        return jsonify({"ok": True, "mensaje": "Contraseña actualizada correctamente"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/solicitar_reset", methods=["POST"])
def solicitar_reset():
    """Solicita reset de contraseña por email."""
    data    = request.get_json() or {}
    usuario = data.get("usuario", "").strip().lower()
    if not usuario:
        return jsonify({"ok": False, "error": "Ingresa tu usuario"})
    try:
        ch = get_client()
        r  = ch.query(f"""
            SELECT usuario, nombre, email FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario=\'{usuario}\' AND activo=1 LIMIT 1
        """)
        if not r.result_rows:
            # Por seguridad, no revelar si el usuario existe o no
            return jsonify({"ok": True, "mensaje": "Si el usuario existe, recibirás un correo en los próximos minutos."})
        row    = r.result_rows[0]
        email  = row[2]
        nombre = row[1]
        if not email:
            return jsonify({"ok": False, "error": "Este usuario no tiene email registrado. Contacta al administrador."})
        reset_token = _crear_reset_token(usuario, email)
        reset_url   = f"{APP_URL}?reset_token={reset_token}"
        cuerpo = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;background:#f9fafb;border-radius:12px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#4c1d95,#7c3aed);padding:28px 32px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:22px">🏍️ Picap Monitoreo</h1>
            <p style="color:rgba(255,255,255,.8);margin:6px 0 0;font-size:13px">Sistema de Monitoreo de Evasión</p>
          </div>
          <div style="padding:28px 32px;background:#fff;">
            <h2 style="color:#1e1b4b;font-size:18px;margin:0 0 12px">Restablece tu contraseña</h2>
            <p style="color:#374151;font-size:14px;">Hola <strong>{nombre}</strong>,</p>
            <p style="color:#374151;font-size:14px;">Recibimos una solicitud para restablecer la contraseña de tu cuenta <strong>{usuario}</strong>.</p>
            <div style="text-align:center;margin:24px 0;">
              <a href="{reset_url}" style="background:linear-gradient(135deg,#7c3aed,#4c1d95);color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700;font-size:14px;">
                🔐 Restablecer contraseña
              </a>
            </div>
            <p style="color:#6b7280;font-size:12px;">Este enlace expira en <strong>1 hora</strong>. Si no solicitaste este cambio, ignora este mensaje.</p>
            <hr style="border:none;border-top:1px solid #e5e7eb;margin:20px 0;">
            <p style="color:#9ca3af;font-size:11px;margin:0">Picap Monitoreo · Sistema interno · No responder este correo.</p>
          </div>
        </div>"""
        ok, err = _enviar_email(email, "Restablece tu contraseña — Picap Monitoreo", cuerpo)
        if ok:
            email_masked = email[:2] + "***" + email[email.index('@'):]
            return jsonify({"ok": True, "mensaje": f"Correo enviado a {email_masked}"})
        else:
            print(f"[reset] Error email: {err}")
            # Si el email falla por config, igual retornar ok para no exponer info
            return jsonify({"ok": True, "mensaje": "Si el usuario existe, recibirás un correo en los próximos minutos."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/reset_password", methods=["POST"])
def reset_password():
    """Aplica nueva contraseña usando token del email."""
    data      = request.get_json() or {}
    token     = data.get("reset_token", "").strip()
    pwd_nueva = data.get("password_nueva", "").strip()
    if not token or not pwd_nueva:
        return jsonify({"ok": False, "error": "Datos incompletos"})
    if len(pwd_nueva) < 6:
        return jsonify({"ok": False, "error": "La contraseña debe tener al menos 6 caracteres"})
    payload = _verificar_reset_token(token)
    if not payload:
        return jsonify({"ok": False, "error": "El enlace expiró o es inválido. Solicita uno nuevo."})
    usuario = payload["u"]
    try:
        ch     = get_client()
        r      = ch.query(f"""
            SELECT nombre, email, rol FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario=\'{usuario}\' AND activo=1 LIMIT 1
        """)
        if not r.result_rows:
            return jsonify({"ok": False, "error": "Usuario no encontrado"})
        row        = r.result_rows[0]
        hash_nueva = hashlib.sha256(pwd_nueva.encode()).hexdigest()
        ch.insert("picapmongoprod.dashboard_users",
            [[usuario, hash_nueva, row[0], row[1], row[2], _dt.utcnow(), 1]],
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
        return jsonify({"ok": True, "mensaje": "Contraseña restablecida. Ya puedes iniciar sesión."})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════
# CALENDARIO — Notificación por email al agregar tarea
# ══════════════════════════════════════════════════════════════════════
@app.route("/api/calendario/notificar", methods=["POST"])
def calendario_notificar():
    token = request.headers.get("X-Token", "")
    data  = request.get_json() or {}
    try:
        ch = get_client()
        s  = _verificar_sesion(ch, token)
        if not s:
            return jsonify({"ok": False, "error": "Sesión expirada"}), 401

        titulo    = data.get("titulo", "").strip()
        fecha     = data.get("fecha", "").strip()
        hora      = data.get("hora", "").strip()
        destinatario = data.get("email", "").strip()

        if not titulo or not fecha or not destinatario:
            return jsonify({"ok": False, "error": "Datos incompletos"})

        # Obtener nombre del usuario que crea la tarea
        usuario = s.get("usuario", "")
        r = ch.query(f"""
            SELECT nombre FROM picapmongoprod.dashboard_users FINAL
            WHERE usuario=\'{usuario}\' AND activo=1 LIMIT 1
        """)
        nombre_creador = r.result_rows[0][0] if r.result_rows else usuario

        hora_display = hora if hora else "Sin hora definida"
        fecha_parts  = fecha.split("-")
        MESES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        try:
            fecha_display = f"{int(fecha_parts[2])} de {MESES[int(fecha_parts[1])]} de {fecha_parts[0]}"
        except:
            fecha_display = fecha

        cuerpo = f"""
        <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;background:#f9fafb;border-radius:12px;overflow:hidden;">
          <div style="background:linear-gradient(135deg,#4c1d95,#7c3aed);padding:24px 32px;text-align:center;">
            <h1 style="color:#fff;margin:0;font-size:20px">📅 Picap Monitoreo</h1>
            <p style="color:rgba(255,255,255,.8);margin:4px 0 0;font-size:12px">Nueva tarea en el calendario</p>
          </div>
          <div style="padding:24px 32px;background:#fff;">
            <h2 style="color:#1e1b4b;font-size:16px;margin:0 0 16px">Tienes una nueva tarea programada</h2>
            <div style="background:#f5f3ff;border-left:4px solid #7c3aed;border-radius:0 8px 8px 0;padding:14px 16px;margin-bottom:16px;">
              <div style="font-size:18px;font-weight:700;color:#1e1b4b;margin-bottom:8px">{titulo}</div>
              <div style="font-size:13px;color:#374151;margin-bottom:4px;">
                📆 <strong>Fecha:</strong> {fecha_display}
              </div>
              <div style="font-size:13px;color:#374151;">
                🕐 <strong>Hora:</strong> {hora_display}
              </div>
            </div>
            <p style="font-size:13px;color:#6b7280;margin:0">
              Agregado por <strong>{nombre_creador}</strong> desde el portal de monitoreo.
            </p>
            <hr style="border:none;border-top:1px solid #e5e7eb;margin:16px 0;">
            <p style="color:#9ca3af;font-size:11px;margin:0">Picap Monitoreo · Sistema interno · No responder este correo.</p>
          </div>
        </div>"""

        # Enviar email de confirmación inmediata
        ok, err = _enviar_email(destinatario, f"📅 Nueva tarea: {titulo} — {fecha_display}", cuerpo)

        # Guardar recordatorio programado si tiene hora
        recordatorio_guardado = False
        if hora:
            try:
                rid = str(uuid.uuid4())[:16]
                ch.insert(
                    "picapmongoprod.calendario_recordatorios",
                    [[rid, titulo, fecha, hora, destinatario, usuario, 0, datetime.now()]],
                    column_names=["id","titulo","fecha","hora","email","creado_por","enviado","creado_en"]
                )
                recordatorio_guardado = True
            except Exception as re_err:
                print(f"[recordatorio] Error guardando: {re_err}")

        if ok:
            msg = f"Notificación enviada"
            if recordatorio_guardado:
                msg += f" · Recordatorio programado para las {hora}"
            return jsonify({"ok": True, "mensaje": msg})
        else:
            # Aunque falle el email inmediato, el recordatorio se programó
            if recordatorio_guardado:
                return jsonify({"ok": True, "mensaje": f"Recordatorio programado para las {hora} (email inmediato falló)"})
            return jsonify({"ok": False, "error": f"Error al enviar email: {err}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/admin/usuarios")
def admin_usuarios():
    token = request.headers.get("X-Token","")
    try:
        _ch0 = get_client()
        _s0  = _verificar_sesion(_ch0, token)
        if not _s0 or _s0.get("rol") != "admin":
            return jsonify({"error":"Sin permisos"}), 403
    except Exception as _e0:
        return jsonify({"error": f"Sin permisos: {_e0}"}), 403
    try:
        ch = get_client()
        r  = ch.query("""
            SELECT usuario, nombre, email, rol,
                   formatDateTime(creado_en,'%Y-%m-%d %H:%M') AS creado_en
            FROM picapmongoprod.dashboard_users FINAL
            WHERE activo=1
            ORDER BY creado_en DESC
        """)
        usuarios = [dict(zip(r.column_names, row)) for row in r.result_rows]
        return jsonify({"ok":True,"usuarios":usuarios})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/api/admin/asignar_rol", methods=["POST"])
def asignar_rol():
    token = request.headers.get("X-Token","")
    try:
        _ch0 = get_client()
        _s0  = _verificar_sesion(_ch0, token)
        if not _s0 or _s0.get("rol") != "admin":
            return jsonify({"error":"Sin permisos"}), 403
    except Exception as _e0:
        return jsonify({"error": f"Sin permisos: {_e0}"}), 403
    data    = request.get_json() or {}
    usuario = data.get("usuario","").strip()
    rol     = data.get("rol","").strip()
    roles_validos = list(ROLES_ACCESO.keys())
    if rol not in roles_validos:
        return jsonify({"ok":False,"error":f"Rol inválido. Válidos: {roles_validos}"}), 400
    try:
        ch = get_client()
        # ReplacingMergeTree: reinsertamos con el nuevo rol para que reemplace
        r  = ch.query(f"""
            SELECT usuario, password_hash, nombre, email
            FROM picapmongoprod.dashboard_users
            WHERE usuario='{usuario}' AND activo=1 LIMIT 1
        """)
        if not r.result_rows:
            return jsonify({"ok":False,"error":"Usuario no encontrado"}), 404
        row = r.result_rows[0]
        ch.insert("picapmongoprod.dashboard_users",
            [[row[0], row[1], row[2], row[3], rol, _dt.utcnow(), 1]],
            column_names=["usuario","password_hash","nombre","email","rol","creado_en","activo"])
        return jsonify({"ok":True,"mensaje":f"Rol '{rol}' asignado a {usuario}"})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)}), 500

# ══════════════════════════════════════════════════════════════
# ENDPOINT PAGOS TC + PROMO — versión dashboard (resumen)
# ══════════════════════════════════════════════════════════════
Q_PAGOS_STATS = """
SELECT
    CASE b.g_country
        WHEN 'CO' THEN 'Colombia'
        WHEN 'MX' THEN 'Mexico'
        WHEN 'NI' THEN 'Nicaragua'
        WHEN 'GT' THEN 'Guatemala'
        ELSE b.g_country
    END                                             AS pais,
    CASE
        WHEN b.payment_method_cd = '3' THEN 'Tarjeta de Credito'
        ELSE 'PromoCode'
    END                                             AS medio_pago,
    count()                                         AS total_servicios,
    countIf(b.fraud_booking = 'true')               AS marcados_fraude,
    countIf(b.status_cd = 4)                        AS finalizados,
    countIf(b.status_cd = 107)                      AS sin_pago,
    countIf(b.status_cd = 108)                      AS cancelados_con_cobro,
    round(sum(toFloat64OrNull(
        JSONExtractString(b.final_cost,'cents')
    )/100), 0)                                      AS monto_total_cop
FROM picapmongoprod.bookings b
WHERE b.status_cd IN (4, 107, 108)
  AND b.payment_method_cd IN ('3')
  AND b.created_at >= toDateTime('{desde} 00:00:00')
  AND b.created_at <= toDateTime('{hasta} 23:59:59')
  AND b.events IS NOT NULL
  AND b.end_geojson IS NOT NULL
GROUP BY pais, medio_pago
ORDER BY total_servicios DESC
"""

@app.route("/api/pagos_stats")
def pagos_stats():
    desde  = request.args.get("desde") or (date.today()-timedelta(days=14)).strftime("%Y-%m-%d")
    hasta  = request.args.get("hasta") or date.today().strftime("%Y-%m-%d")
    pais   = request.args.get("pais", "")
    try:
        ch = get_client()
        q  = Q_PAGOS_STATS.format(desde=desde, hasta=hasta)
        if pais:
            q = q.replace("ORDER BY", f"HAVING pais='{pais}'\nORDER BY")
        r  = ch.query(q)
        rows = [dict(zip(r.column_names, row)) for row in r.result_rows]
        total  = sum(x.get("total_servicios",0) for x in rows)
        fraude = sum(x.get("marcados_fraude",0) for x in rows)
        monto  = sum(x.get("monto_total_cop",0) or 0 for x in rows)
        return jsonify({"ok":True, "desde":desde, "hasta":hasta,
            "resumen":{"total":total,"fraude":fraude,"monto":monto,
                       "pct_fraude":round(fraude/total*100,1) if total else 0},
            "por_pais_medio": rows})
    except Exception as e:
        import traceback
        return jsonify({"error":str(e),"detalle":traceback.format_exc()}), 500

# ══════════════════════════════════════════════════════════════
# CONSOLE SEGURA — solo admin, comandos predefinidos
# ══════════════════════════════════════════════════════════════
COMANDOS_PERMITIDOS = {
    "status":        lambda ch: {"status":"OK","tablas":ch.query("SELECT count() FROM system.tables WHERE database='picapmongoprod'").result_rows[0][0]},
    "contar_users":  lambda ch: {"usuarios":ch.query("SELECT count() FROM picapmongoprod.dashboard_users WHERE activo=1").result_rows[0][0]},
    "contar_alertas_rf": lambda ch: {"alertas_rf":ch.query("SELECT count() FROM system.tables WHERE database='picapmongoprod' AND name='alertas_reconocimiento'").result_rows[0][0]},
    "ver_roles":     lambda ch: {"roles": [r[0] for r in ch.query("SELECT DISTINCT rol FROM picapmongoprod.dashboard_users").result_rows]},
    "ping_ch":       lambda ch: {"ping":"OK","version":ch.query("SELECT version()").result_rows[0][0]},
    "help":          lambda ch: {"comandos":list(COMANDOS_PERMITIDOS.keys())},
}

@app.route("/api/console", methods=["POST"])
def console():
    token = request.headers.get("X-Token","")
    try:
        _ch0 = get_client()
        _s0  = _verificar_sesion(_ch0, token)
        if not _s0 or _s0.get("rol") != "admin":
            return jsonify({"error":"Sin permisos — solo admin"}), 403
    except Exception as _e0:
        return jsonify({"error": f"Sin permisos: {_e0}"}), 403
    data = request.get_json() or {}
    cmd  = data.get("cmd","").strip().lower()
    fn   = COMANDOS_PERMITIDOS.get(cmd)
    if not fn:
        return jsonify({"error":f"Comando '{cmd}' no reconocido. Escribe 'help' para ver los disponibles."})
    try:
        ch     = get_client()
        result = fn(ch)
        return jsonify({"ok":True,"cmd":cmd,"resultado":result})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})


# ══════════════════════════════════════════════════════════════════════════════════════
# MÓDULO PIBOX B2B - ROBOT DE AUDITORÍA AUTOMATIZADA
# Sistema de análisis de servicios corporativos con detección de alertas
# ══════════════════════════════════════════════════════════════════════════════════════

# ────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE REGLAS DE NEGOCIO
# ────────────────────────────────────────────────────────────────────────────
PIBOX_CONFIG = {
    'TIEMPO_MINIMO_SERVICIO': 5,  # minutos
    'TOLERANCIA_GPS': 0.001,  # grados (≈ 100 metros)
    'MONTOS_ALERTA': {
        'mensajeria': 400000,  # COP - Mensajería normal
        'carga_carry': 800000,  # COP - Carga Carry (1.5 toneladas)
        'carga_moto': 600000,  # COP - Carga Moto-Vagón (750 kg)
        'cruz_verde_mostrador': 80000  # COP - Cruz Verde Mostrador específico
    },
    'CLIENTES_EXCLUIDOS': ['tada'],
    'KEYWORDS_EXCLUIR': ['test', 'prueba', 'qa', 'onboarding', 'demo'],
    'EXCEPCION_CLIENTE': 'cruz verde integración',
    # IDs reales de requested_service_type_id
    'SERVICE_TYPE_IDS': {
        'mensajeria': '5c71b03a58b9ba10fa6393cf',
        'carga_carry': '62e2ae08790a6a0004ab0a3b',
        'carga_moto': '62e2ae08790a6a0004ab0a3a'
    },
    # Mapeo de IDs a nombres legibles
    'SERVICE_TYPE_NAMES': {
        '5c71b03a58b9ba10fa6393cf': 'Mensajería',
        '62e2ae08790a6a0004ab0a3b': 'Carga Carry',
        '62e2ae08790a6a0004ab0a3a': 'Carga Moto-Vagón',
        '57b28033f0350b00035d0ade': 'Moto Mensajería',
        '62e2ae08790a6a0004ab0a3c': 'Carga NHR',
        '57b27f84f0350b00035d0ad9': 'Otro tipo'
    },
    # Pilotos que requieren revisión adicional
    'PILOTOS_REVISION': {
        '67bb692f4623a92a61b4e1c1': 'Guilio Rene Velandia Suarez',
        '597c0cbc53bd7c0004e5d58f': 'Yonattan Camilo Galeano Moreno',
        '64c3e4891262d800573e6b12': 'Jairo Reyes',
        '634cb1add50da600442ea6f7': 'Anderson Gutierrez Mendez',
        '5c2723e43eb16b0030a160fd': 'Miguel Angel Galvis Guerrero',
        '65f4d3619b0bac0062fa0277': 'Carlos Augusto Hernandez Higuita',
        '64b2b4999bef87004d5ea234': 'Anyelo David Mendoza',
        '67899fc21ed419d3cb491b76': 'Alfredo Goez Ibarra',
        '662087a3ee8f1f0046b398b2': 'Ender Armando Pinzon Gonzalez',
        '6610408315a5e60062358606': 'Heber Méndez Santos'
    }
}

# ────────────────────────────────────────────────────────────────────────────
# QUERY BASE: Obtener servicios Pibox B2B para análisis
# ────────────────────────────────────────────────────────────────────────────
Q_PIBOX_BASE = """
WITH servicios_pibox AS (
    SELECT
        b._id AS booking_id,
        b.driver_id,
        b.passenger_id,
        b.company_id,
        toTimeZone(b.created_at, 'America/Bogota') AS fecha_servicio,
        b.status_cd,
        b.requested_service_type_id,

        -- Información del piloto
        CONCAT(p.name, ' ', COALESCE(p.last_name, '')) AS piloto_nombre,
        p.email AS piloto_email,
        p.phone AS piloto_telefono,

        -- Información del cliente/empresa
        c.name AS cliente_nombre,

        -- Tipo de vehículo
        JSONExtractString(vt.name, 'es') AS tipo_vehiculo,

        -- País
        CASE b.g_country
            WHEN 'CO' THEN 'Colombia'
            WHEN 'MX' THEN 'México'
            WHEN 'NI' THEN 'Nicaragua'
            WHEN 'GT' THEN 'Guatemala'
            ELSE b.g_country
        END AS pais,

        CASE
            WHEN b.g_adm_area_lv_1 = 'MN' THEN 'Managua'
            WHEN b.g_adm_area_lv_1 = 'Guatemala Department' THEN 'Guatemala'
            WHEN b.g_adm_area_lv_1 = '' THEN 'Sin ciudad'
            ELSE b.g_adm_area_lv_1
        END AS ciudad,

        -- Costos
        toFloat64OrZero(JSONExtractString(b.company_final_cost, 'cents')) / 100 AS monto_pagado,
        JSONExtractString(b.company_final_cost, 'currency_iso') AS moneda,

        -- Coordenadas extraídas de JSON
        JSONExtractFloat(b.origin_geojson, 'coordinates', 1) AS origin_longitude,
        JSONExtractFloat(b.origin_geojson, 'coordinates', 2) AS origin_latitude,
        JSONExtractFloat(b.end_geojson, 'coordinates', 1) AS destination_longitude,
        JSONExtractFloat(b.end_geojson, 'coordinates', 2) AS destination_latitude,

        -- Extraer eventos corregidos (event_cd: 22=recogido, 24=entregado)
        extract(COALESCE(b.events,''), 'event_cd":22.*?created_at":"([^"]+)') AS ev_recogido,
        extract(COALESCE(b.events,''), 'event_cd":24.*?created_at":"([^"]+)') AS ev_finalizado,

        -- Campos para reglas de negocio anti-falsos-positivos:
        -- return_to_origin = TRUE  → es un retorno a origen valido (NO alertar)
        -- original_booking_reservation_id NOT NULL → servicio nacio de una reserva (NO alertar)
        --
        -- Normalizamos via toString() porque las cargas desde Mongo pueden traer
        -- el bool como String ("true"/"false"), UInt8, Bool o Nullable(*).
        -- toString() funciona para cualquier tipo y maneja NULL como ''.
        IF(
            lower(COALESCE(toString(b.return_to_origin), '')) IN ('true', '1', 't'),
            1, 0
        ) AS return_to_origin,
        COALESCE(toString(b.original_booking_reservation_id), '') AS original_booking_reservation_id,
        IF(COALESCE(toString(b.original_booking_reservation_id), '') = '', 0, 1) AS tiene_reserva,

        b.updated_at,

        ROW_NUMBER() OVER (PARTITION BY b._id ORDER BY b.created_at DESC) AS rn

    FROM picapmongoprod.bookings b FINAL
    
    LEFT JOIN picapmongoprod.passengers p FINAL 
        ON b.driver_id = p._id
    
    LEFT JOIN picapmongoprod.companies c FINAL 
        ON b.company_id = c._id
    
    LEFT JOIN picapmongoprod.driver_vehicle_enrollments dve FINAL
        ON b.driver_id = dve.driver_id 
        AND dve.enrollment_status_cd = 3
    
    LEFT JOIN picapmongoprod.vehicles v FINAL
        ON dve.vehicle_id = v._id
    
    LEFT JOIN picapmongoprod.vehicle_types vt FINAL
        ON v.vehicle_type_id = vt._id
    
    WHERE 
        -- FILTROS OBLIGATORIOS: Solo servicios finalizados con company_id
        b.status_cd IN (4, 107, 108)
        AND b.company_id IS NOT NULL
        AND b.company_id != ''
        AND toDate(b.created_at) BETWEEN '{fecha_desde}' AND '{fecha_hasta}'
        
        -- EXCLUIR servicios con "Reserva iniciada" (event_cd: 103)
        AND NOT (b.events LIKE '%"event_cd":103%')
        
        -- FILTROS DE EXCLUSIÓN (nombres test, prueba, etc)
        AND LOWER(c.name) NOT LIKE '%tada%'
        AND LOWER(c.name) NOT LIKE '%test%'
        AND LOWER(c.name) NOT LIKE '%prueba%'
        AND LOWER(c.name) NOT LIKE '%qa%'
        AND LOWER(c.name) NOT LIKE '%onboarding%'
        AND LOWER(c.name) NOT LIKE '%demo%'
        {filtros_adicionales}
),
con_tiempos AS (
    SELECT
        sp.*,
        -- Calcular tiempo de servicio
        dateDiff('minute',
            parseDateTimeBestEffortOrNull(ev_recogido),
            parseDateTimeBestEffortOrNull(ev_finalizado)
        ) AS minutos_servicio,

        -- Señal pura GPS: origen y destino casi iguales (tolerancia ≈ 100m)
        IF(
            abs(origin_latitude - destination_latitude) < {tolerancia_gps}
            AND abs(origin_longitude - destination_longitude) < {tolerancia_gps},
            1, 0
        ) AS flag_mismo_punto,

        -- REGLA DE NEGOCIO (alerta de fraude por mismo punto):
        --   Disparar alerta SOLO si:
        --     1) origen y destino son el mismo punto (GPS), Y
        --     2) return_to_origin = false (no es un retorno a origen válido), Y
        --     3) original_booking_reservation_id IS NULL (no tiene reserva)
        --   Cualquier otra combinación NO genera alerta (filtra falsos positivos).
        --   `return_to_origin` y `tiene_reserva` ya vienen normalizados a UInt8 desde la CTE previa.
        IF(
            abs(origin_latitude - destination_latitude) < {tolerancia_gps}
            AND abs(origin_longitude - destination_longitude) < {tolerancia_gps}
            AND return_to_origin = 0
            AND tiene_reserva = 0,
            1, 0
        ) AS flag_alerta_mismo_punto,

        -- Calcular distancia origen-destino
        round(geoDistance(origin_longitude, origin_latitude,
                         destination_longitude, destination_latitude), 2) AS distancia_recorrido

    FROM servicios_pibox sp
    WHERE sp.rn = 1
)
SELECT * FROM con_tiempos
"""

# ────────────────────────────────────────────────────────────────────────────
# ENDPOINT: Obtener servicios Pibox filtrados
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/pibox/servicios")
def pibox_servicios():
    """Obtiene servicios Pibox B2B según filtros"""
    # Verificar autenticación
    token = request.headers.get("X-Token", "")
    try:
        ch_auth = get_client()
        sesion = _verificar_sesion(ch_auth, token)
        if not sesion:
            return jsonify({"error": "Sin autenticación"}), 401
    except Exception as e:
        return jsonify({"error": f"Error de autenticación: {e}"}), 401
    
    # Obtener parámetros
    fecha_desde = request.args.get("desde", (date.today() - timedelta(days=7)).isoformat())
    fecha_hasta = request.args.get("hasta", date.today().isoformat())
    pais = request.args.get("pais", "")
    cliente_id = request.args.get("cliente_id", "")
    piloto_id = request.args.get("piloto_id", "")
    
    # Construir filtros adicionales
    filtros = []
    if pais:
        filtros.append(f"AND b.g_country = '{pais[:2].upper()}'")
    if cliente_id:
        filtros.append(f"AND b.company_id = '{cliente_id}'")
    if piloto_id:
        filtros.append(f"AND b.driver_id = '{piloto_id}'")
    
    filtros_str = " ".join(filtros)
    
    try:
        ch = get_client()
        query = Q_PIBOX_BASE.format(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            tolerancia_gps=PIBOX_CONFIG['TOLERANCIA_GPS'],
            filtros_adicionales=filtros_str
        )
        
        result = ch.query(query)
        servicios = [dict(zip(result.column_names, row)) for row in result.result_rows]
        
        # Limpiar NaN
        servicios = limpiar_nan(servicios)
        
        return jsonify({
            "ok": True,
            "total": len(servicios),
            "servicios": servicios
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc()
        }), 500

# ────────────────────────────────────────────────────────────────────────────
# ENDPOINT: Ejecutar robot de auditoría y obtener alertas
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/pibox/alertas")
def pibox_alertas():
    """Ejecuta el robot de auditoría y retorna alertas detectadas"""
    # Verificar autenticación
    token = request.headers.get("X-Token", "")
    try:
        ch_auth = get_client()
        sesion = _verificar_sesion(ch_auth, token)
        if not sesion:
            return jsonify({"error": "Sin autenticación"}), 401
    except Exception as e:
        return jsonify({"error": f"Error de autenticación: {e}"}), 401
    
    # Obtener parámetros
    fecha_desde = request.args.get("desde", (date.today() - timedelta(days=7)).isoformat())
    fecha_hasta = request.args.get("hasta", date.today().isoformat())
    pais = request.args.get("pais", "")
    cliente_id = request.args.get("cliente_id", "")
    piloto_id = request.args.get("piloto_id", "")
    
    # Construir filtros adicionales
    filtros = []
    if pais:
        filtros.append(f"AND b.g_country = '{pais[:2].upper()}'")
    if cliente_id:
        filtros.append(f"AND b.company_id = '{cliente_id}'")
    if piloto_id:
        filtros.append(f"AND b.driver_id = '{piloto_id}'")
    
    filtros_str = " ".join(filtros)
    
    try:
        ch = get_client()
        
        # Obtener servicios
        query = Q_PIBOX_BASE.format(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            tolerancia_gps=PIBOX_CONFIG['TOLERANCIA_GPS'],
            filtros_adicionales=filtros_str
        )
        
        result = ch.query(query)
        servicios = [dict(zip(result.column_names, row)) for row in result.result_rows]
        
        # Ejecutar robot de auditoría
        alertas = []
        estadisticas = {
            'total_servicios': 0,
            'total_alertas': 0,
            'alertas_tiempo': 0,
            'alertas_recorrido': 0,
            'alertas_evidencia': 0,
            'alertas_pago': 0,
            'tipos_alerta': {},
            # Falsos positivos descartados (visibilidad operativa)
            'descartados_retorno_origen': 0,
            'descartados_con_reserva': 0,
        }

        for servicio in servicios:
            # Verificar exclusiones
            debe_excluir, razon = _pibox_debe_excluirse(servicio)
            if debe_excluir:
                continue

            estadisticas['total_servicios'] += 1

            # Contabilizar descartes por reglas de negocio sobre "mismo punto":
            # cuando hay coincidencia GPS pero el servicio es válido por retorno
            # a origen o por venir de una reserva.
            if servicio.get('flag_mismo_punto') == 1 and servicio.get('flag_alerta_mismo_punto') != 1:
                if (servicio.get('return_to_origin') or 0):
                    estadisticas['descartados_retorno_origen'] += 1
                elif (servicio.get('tiene_reserva') or 0):
                    estadisticas['descartados_con_reserva'] += 1
            
            # Analizar servicio y generar alertas
            alertas_servicio = _pibox_analizar_servicio(servicio, ch)
            
            for alerta in alertas_servicio:
                alertas.append(alerta)
                tipo = alerta.get('tipo_alerta', '')
                
                # Contar por tipo
                if tipo == 'Tiempo':
                    estadisticas['alertas_tiempo'] += 1
                elif tipo == 'GPS':
                    estadisticas['alertas_recorrido'] += 1
                elif tipo == 'Fotos':
                    estadisticas['alertas_evidencia'] += 1
                elif tipo == 'Pagos':
                    estadisticas['alertas_pago'] += 1
                
                # Contar por tipo general
                if tipo not in estadisticas['tipos_alerta']:
                    estadisticas['tipos_alerta'][tipo] = 0
                estadisticas['tipos_alerta'][tipo] += 1
        
        estadisticas['total_alertas'] = len(alertas)
        
        # Limpiar NaN
        alertas = limpiar_nan(alertas)
        estadisticas = limpiar_nan(estadisticas)
        
        return jsonify({
            "ok": True,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
            "alertas": alertas,
            "estadisticas": estadisticas
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc()
        }), 500

# ────────────────────────────────────────────────────────────────────────────
# FUNCIONES AUXILIARES DEL ROBOT
# ────────────────────────────────────────────────────────────────────────────

def _pibox_debe_excluirse(servicio):
    """
    Determina si un servicio debe excluirse del análisis de alertas.

    Returns:
        (debe_excluir: bool, razon: str)
    """
    cliente_nombre = str(servicio.get('cliente_nombre', '')).lower()

    # EXCEPCIÓN: Cruz Verde Integración siempre se analiza
    if PIBOX_CONFIG['EXCEPCION_CLIENTE'] in cliente_nombre:
        return False, ''

    # Excluir clientes específicos (TADA)
    for excl in PIBOX_CONFIG['CLIENTES_EXCLUIDOS']:
        if excl in cliente_nombre:
            return True, f'Cliente excluido ({excl.upper()})'

    # Excluir test/prueba/qa/onboarding
    for kw in PIBOX_CONFIG['KEYWORDS_EXCLUIR']:
        if kw in cliente_nombre:
            return True, 'Cliente de pruebas'

    # NOTA: la exclusión por servicio con reserva (original_booking_reservation_id)
    # y por retorno a origen (return_to_origin) NO debe vetar el servicio entero del
    # análisis — afectan solo a la alerta de "mismo punto". Las demás reglas
    # (tiempo, montos, fotos) siguen aplicándose. Por eso esa lógica vive dentro
    # de _pibox_analizar_servicio sobre el flag `flag_alerta_mismo_punto`.

    return False, ''

def _pibox_analizar_servicio(servicio, ch):
    """
    Analiza un servicio y genera alertas si corresponde
    
    Returns:
        List[Dict]: Lista de alertas generadas
    """
    alertas = []
    
    # Obtener tipo de servicio legible
    service_type_id = servicio.get('requested_service_type_id', '')
    tipo_servicio_nombre = PIBOX_CONFIG['SERVICE_TYPE_NAMES'].get(
        service_type_id,
        'Desconocido'
    )
    
    # Verificar si el piloto está en lista de revisión
    driver_id = servicio.get('driver_id', '')
    flag_revision = 'Sí' if driver_id in PIBOX_CONFIG['PILOTOS_REVISION'] else 'No'
    
    # 1. Validar tiempo de servicio (< 5 minutos)
    minutos = servicio.get('minutos_servicio')
    if minutos is not None and minutos > 0 and minutos < PIBOX_CONFIG['TIEMPO_MINIMO_SERVICIO']:
        alertas.append({
            'booking_id': servicio.get('booking_id', ''),
            'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
            'piloto_id': servicio.get('driver_id', ''),
            'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
            'tipo_servicio': servicio.get('requested_service_type_id', 'N/A'),
            'tipo_servicio_nombre': tipo_servicio_nombre,
            'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
            'tipo_alerta': 'Tiempo',
            'observacion': f'Servicio completado en {minutos} minutos (menos de {PIBOX_CONFIG["TIEMPO_MINIMO_SERVICIO"]} min)',
            'severidad': 'ALTA',
            'monto': servicio.get('monto_pagado', 0),
            'fecha_servicio': servicio.get('fecha_servicio', ''),
            'pais': servicio.get('pais', ''),
            'ciudad': servicio.get('ciudad', ''),
            'posible_revision': flag_revision
        })
    
    # 2. Validar recorrido (mismo punto origen/destino)
    #    Reglas de negocio (anti-falsos-positivos):
    #      - flag_alerta_mismo_punto ya consolida en SQL las 3 condiciones:
    #          mismo punto GPS  ∧  return_to_origin = false  ∧  original_booking_reservation_id IS NULL
    #      - Si return_to_origin = true → es retorno legítimo, NO se alerta.
    #      - Si tiene reserva (tiene_reserva = 1) → NO se alerta.
    if servicio.get('flag_alerta_mismo_punto') == 1:
        alertas.append({
            'booking_id': servicio.get('booking_id', ''),
            'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
            'piloto_id': servicio.get('driver_id', ''),
            'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
            'tipo_servicio': servicio.get('requested_service_type_id', 'N/A'),
            'tipo_servicio_nombre': tipo_servicio_nombre,
            'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
            'tipo_alerta': 'GPS',
            'observacion': 'Mismo punto de inicio y finalización (sin retorno a origen ni reserva asociada)',
            'severidad': 'ALTA',
            'monto': servicio.get('monto_pagado', 0),
            'fecha_servicio': servicio.get('fecha_servicio', ''),
            'pais': servicio.get('pais', ''),
            'ciudad': servicio.get('ciudad', ''),
            'posible_revision': flag_revision,
            'return_to_origin': bool(servicio.get('return_to_origin') or 0),
            'tiene_reserva': bool(servicio.get('tiene_reserva') or 0)
        })
    
    # 3. Validar evidencias fotográficas con robot Trump
    # TEMPORALMENTE DESACTIVADO - El robot tarda mucho y cuelga el servidor
    # TODO: Implementar procesamiento en background (Celery/Redis)
    # try:
    #     from trump_foto_validator import TrumpFotoValidator
    #     ... código del robot ...
    # except Exception as e:
    #     ... manejo de errores ...
    
    # 4. Validar pagos (montos excesivos)
    cliente_nombre = str(servicio.get('cliente_nombre', '')).lower()
    monto = servicio.get('monto_pagado', 0) or 0
    moneda = servicio.get('moneda', 'COP')
    
    # Solo validar montos en COP
    if moneda == 'COP':
        # Mensajería > 400,000
        if service_type_id == PIBOX_CONFIG['SERVICE_TYPE_IDS']['mensajeria']:
            if monto > PIBOX_CONFIG['MONTOS_ALERTA']['mensajeria']:
                alertas.append({
                    'booking_id': servicio.get('booking_id', ''),
                    'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
                    'piloto_id': servicio.get('driver_id', ''),
                    'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
                    'tipo_servicio': service_type_id,
                    'tipo_servicio_nombre': tipo_servicio_nombre,
                    'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
                    'tipo_alerta': 'Pagos',
                    'observacion': f'Monto excesivo ${monto:,.0f} COP para Mensajería (umbral ${PIBOX_CONFIG["MONTOS_ALERTA"]["mensajeria"]:,.0f})',
                    'severidad': 'MEDIA',
                    'monto': monto,
                    'fecha_servicio': servicio.get('fecha_servicio', ''),
                    'pais': servicio.get('pais', ''),
                    'ciudad': servicio.get('ciudad', ''),
                    'posible_revision': flag_revision
                })
        
        # Carga Carry > 800,000
        if service_type_id == PIBOX_CONFIG['SERVICE_TYPE_IDS']['carga_carry']:
            if monto > PIBOX_CONFIG['MONTOS_ALERTA']['carga_carry']:
                alertas.append({
                    'booking_id': servicio.get('booking_id', ''),
                    'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
                    'piloto_id': servicio.get('driver_id', ''),
                    'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
                    'tipo_servicio': service_type_id,
                    'tipo_servicio_nombre': tipo_servicio_nombre,
                    'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
                    'tipo_alerta': 'Pagos',
                    'observacion': f'Monto excesivo ${monto:,.0f} COP para Carga Carry (umbral ${PIBOX_CONFIG["MONTOS_ALERTA"]["carga_carry"]:,.0f})',
                    'severidad': 'MEDIA',
                    'monto': monto,
                    'fecha_servicio': servicio.get('fecha_servicio', ''),
                    'pais': servicio.get('pais', ''),
                    'ciudad': servicio.get('ciudad', ''),
                    'posible_revision': flag_revision
                })
        
        # Carga Moto > 600,000
        if service_type_id == PIBOX_CONFIG['SERVICE_TYPE_IDS']['carga_moto']:
            if monto > PIBOX_CONFIG['MONTOS_ALERTA']['carga_moto']:
                alertas.append({
                    'booking_id': servicio.get('booking_id', ''),
                    'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
                    'piloto_id': servicio.get('driver_id', ''),
                    'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
                    'tipo_servicio': service_type_id,
                    'tipo_servicio_nombre': tipo_servicio_nombre,
                    'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
                    'tipo_alerta': 'Pagos',
                    'observacion': f'Monto excesivo ${monto:,.0f} COP para Carga Moto (umbral ${PIBOX_CONFIG["MONTOS_ALERTA"]["carga_moto"]:,.0f})',
                    'severidad': 'MEDIA',
                    'monto': monto,
                    'fecha_servicio': servicio.get('fecha_servicio', ''),
                    'pais': servicio.get('pais', ''),
                    'ciudad': servicio.get('ciudad', ''),
                    'posible_revision': flag_revision
                })
    
        # Cruz Verde Mostrador con monto > 80,000
        if 'cruz verde mostrador' in cliente_nombre:
            if monto > PIBOX_CONFIG['MONTOS_ALERTA']['cruz_verde_mostrador']:
                alertas.append({
                    'booking_id': servicio.get('booking_id', ''),
                    'piloto_nombre': servicio.get('piloto_nombre', 'N/A'),
                    'piloto_id': servicio.get('driver_id', ''),
                    'cliente_nombre': servicio.get('cliente_nombre', 'N/A'),
                    'tipo_servicio': service_type_id,
                    'tipo_servicio_nombre': tipo_servicio_nombre,
                    'tipo_vehiculo': servicio.get('tipo_vehiculo', 'N/A'),
                    'tipo_alerta': 'Pagos',
                    'observacion': f'Monto excesivo ${monto:,.0f} COP para Cruz Verde Mostrador (umbral ${PIBOX_CONFIG["MONTOS_ALERTA"]["cruz_verde_mostrador"]:,.0f})',
                    'severidad': 'MEDIA',
                    'monto': monto,
                    'fecha_servicio': servicio.get('fecha_servicio', ''),
                    'pais': servicio.get('pais', ''),
                    'ciudad': servicio.get('ciudad', ''),
                    'posible_revision': flag_revision
                })
    
    return alertas

# ────────────────────────────────────────────────────────────────────────────
# ENDPOINT: Exportar alertas a Excel
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/pibox/export")
def pibox_export():
    """Exporta alertas Pibox a Excel"""
    # Verificar autenticación - aceptar token en header o parámetro URL
    token = request.headers.get("X-Token", "") or request.args.get("token", "")
    try:
        ch_auth = get_client()
        sesion = _verificar_sesion(ch_auth, token)
        if not sesion:
            return jsonify({"error": "Sin autenticación"}), 401
    except Exception as e:
        return jsonify({"error": f"Error de autenticación: {e}"}), 401
    
    # Obtener parámetros
    fecha_desde = request.args.get("desde", (date.today() - timedelta(days=7)).isoformat())
    fecha_hasta = request.args.get("hasta", date.today().isoformat())
    tipo = request.args.get("tipo", "alertas")  # 'alertas' o 'estadisticas'
    
    try:
        # Reutilizar lógica del endpoint de alertas
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        
        # Obtener datos (mismo código que pibox_alertas)
        ch = get_client()
        pais = request.args.get("pais", "")
        cliente_id = request.args.get("cliente_id", "")
        piloto_id = request.args.get("piloto_id", "")
        
        filtros = []
        if pais:
            filtros.append(f"AND b.g_country = '{pais[:2].upper()}'")
        if cliente_id:
            filtros.append(f"AND b.company_id = '{cliente_id}'")
        if piloto_id:
            filtros.append(f"AND b.driver_id = '{piloto_id}'")
        
        query = Q_PIBOX_BASE.format(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            tolerancia_gps=PIBOX_CONFIG['TOLERANCIA_GPS'],
            filtros_adicionales=" ".join(filtros)
        )
        
        result = ch.query(query)
        servicios = [dict(zip(result.column_names, row)) for row in result.result_rows]
        
        # Generar alertas
        alertas = []
        for servicio in servicios:
            debe_excluir, _ = _pibox_debe_excluirse(servicio)
            if not debe_excluir:
                alertas.extend(_pibox_analizar_servicio(servicio, ch))
        
        if tipo == "alertas":
            # Exportar alertas
            if not alertas:
                return jsonify({"error": "No hay alertas para exportar"}), 404
            
            df = pd.DataFrame(alertas)
            
            # Ordenar columnas
            columnas_orden = [
                'booking_id', 'fecha_servicio', 'pais', 'ciudad',
                'piloto_nombre', 'piloto_id', 'cliente_nombre',
                'tipo_servicio', 'tipo_vehiculo', 'tipo_alerta',
                'observacion', 'monto'
            ]
            df = df[[col for col in columnas_orden if col in df.columns]]
            
            # Fix: Quitar timezone de fechas (Excel no lo soporta)
            if 'fecha_servicio' in df.columns:
                df['fecha_servicio'] = pd.to_datetime(df['fecha_servicio']).dt.tz_localize(None)
            
            # Crear Excel en memoria
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Alertas Pibox')
                
                # Aplicar formato
                workbook = writer.book
                worksheet = writer.sheets['Alertas Pibox']
                
                # Header
                header_fill = PatternFill(start_color="6b21a8", end_color="6b21a8", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Autofit columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"alertas_pibox_{timestamp}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        
        else:
            # Exportar estadísticas
            estadisticas = {
                'Total servicios analizados': len([s for s in servicios if not _pibox_debe_excluirse(s)[0]]),
                'Total alertas detectadas': len(alertas),
                'Alertas de tiempo': len([a for a in alertas if a['tipo_alerta'] == 'Tiempo de servicio']),
                'Alertas de recorrido': len([a for a in alertas if a['tipo_alerta'] == 'Recorrido GPS']),
                'Alertas de evidencia': len([a for a in alertas if a['tipo_alerta'] == 'Evidencia fotográfica']),
                'Alertas de pago': len([a for a in alertas if a['tipo_alerta'] == 'Validación de pago'])
            }
            
            df = pd.DataFrame(list(estadisticas.items()), columns=['Métrica', 'Valor'])
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Estadísticas')
                
                workbook = writer.book
                worksheet = writer.sheets['Estadísticas']
                
                # Formato
                header_fill = PatternFill(start_color="6b21a8", end_color="6b21a8", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"estadisticas_pibox_{timestamp}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc()
        }), 500


if __name__ == "__main__":
    print("=" * 56)
    print("  Picap Evasion API  ->  http://localhost:5050")
    print("  GET /api/status")
    print("  GET /api/resumen?desde=YYYY-MM-DD&hasta=YYYY-MM-DD")
    print("=" * 56)
    app.run(host="0.0.0.0", port=5050, debug=False)
