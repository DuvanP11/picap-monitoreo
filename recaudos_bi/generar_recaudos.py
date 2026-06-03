"""
Generador del reporte "Saldo Recaudos" para BI.

Construye un Excel con 5 hojas siguiendo la plantilla operativa:
  1. Query Recaudos        — bookings con contraentrega cobrada (data cruda)
  2. TD Recaudos           — pivote por User_Company (Σ Transaction_amount)
  3. Query Transacciones   — wallet transactions Pibox del período + cols DIA/Recaudos/Servicios
  4. Mensual               — pivote por Company_name con: Recaudos, Servicios,
                              DIF, Comisión Recaudo, Retenciones, TOTAL, % comisión
  5. Control               — desde TD Recaudos: User Company, Suma, Pendiente, TOTAL, Comentario

El Porcentaje comisión se trae directo de `picapmongoprod.companies.collection_fee`
(es el % de comisión por pago contra-entrega, ej. ORTOPEDICOS FUTURO = 1).

Las fórmulas se escriben como literales de Excel (=SUM, =DIA, =SI.ERROR, etc.)
para que se preserven al abrir el archivo y los cálculos sean inspeccionables.

Las "tablas dinámicas" se pre-calculan con pandas (groupby) en vez de PivotTable
nativa de Excel — openpyxl no maneja bien PivotTables, y para ~500 empresas
una tabla estática + slicer manual funciona mejor que un pivot vivo.

Uso:
    # Mes completo (abril 2026)
    python generar_recaudos.py --desde 2026-04-01 --hasta 2026-04-30

    # Rangos distintos para cada query
    python generar_recaudos.py \\
        --recaudos-desde 2026-04-11 --recaudos-hasta 2026-04-18 \\
        --tx-desde 2026-04-01 --tx-hasta 2026-04-25

    # Output custom
    python generar_recaudos.py --desde 2026-04-01 --hasta 2026-04-30 \\
        --output "Saldo Recaudos al 30 abril 2026.xlsx"

Credenciales CH se leen de env vars: MINTIC_CH_HOST/USER/PASS (las mismas que
usaste para los reportes MINTIC).
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# UTF-8 en Windows
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ════════════════════════════════════════════════════════════════════════════
# Conexión a ClickHouse
# ════════════════════════════════════════════════════════════════════════════

def ch_request(query: str, host: str, user: str, password: str, timeout: int = 600) -> str:
    """Ejecuta query y devuelve el cuerpo como string en formato TabSeparatedWithNames."""
    url = host.rstrip("/") + "/"
    params = {"query": query + "\nFORMAT TabSeparatedWithNames"}
    r = requests.get(url, params=params, auth=(user, password), timeout=timeout)
    if r.status_code != 200:
        raise RuntimeError(f"CH HTTP {r.status_code}: {r.text[:500]}")
    return r.text


def parse_tsv(text: str) -> tuple[list[str], list[list[str]]]:
    """Parsea TabSeparatedWithNames → (headers, rows)."""
    lines = text.rstrip("\n").split("\n")
    if not lines:
        return [], []
    headers = lines[0].split("\t")
    rows = [ln.split("\t") for ln in lines[1:]] if len(lines) > 1 else []
    return headers, rows


def to_num(v: str):
    """Convierte string CH a número (float) o devuelve string si no es numérico.
    Trata '\\N' (null CH) como None."""
    if v is None or v == "" or v == "\\N":
        return None
    try:
        if "." in v or "e" in v.lower():
            return float(v)
        return int(v)
    except ValueError:
        return v


# ════════════════════════════════════════════════════════════════════════════
# Queries SQL
# ════════════════════════════════════════════════════════════════════════════

QUERY_RECAUDOS = """
WITH
  q_service_types AS (
    SELECT
      _id,
      name_es AS name,
      CASE
        WHEN name_es IN (
          'Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor', 'Moto favor',
          'Carga', 'Carga Carry', 'Carga Moto-Vagón', 'Carga NHR', 'Carga NKR', 'Carga NPR',
          'Mensajería', 'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'
        ) THEN 'Pibox'
        WHEN name_es IN (
          'Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta','Carro Subasta','Taxi', 'Carro',
          'Carro sin conductor', 'Moto VIP', 'Moto Económica', 'Carro Queen','Rapidín','Espero tranqui',
          'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto', 'Grúa Carro', 'Grúa Moto'
        ) THEN 'Picap'
        ELSE 'Other'
      END AS type
    FROM (SELECT * FROM picapmongoprod.service_types FINAL)
  ),
  q_wat_filtered AS (
    SELECT
      wat._id AS transaction_id,
      wat.booking_id AS booking_id,
      wat.package_id AS package_id,
      wat.account_id AS account_id,
      wat.amount AS amount,
      wat.normalized_amount_after_transaction AS normalized_amount_after_transaction,
      wat.transaction_state_cd AS transaction_state_cd,
      wat._type AS tx_type,
      wat.created_at AS created_at,
      b.created_at AS booking_created_at,
      pck.reference AS package_reference,
      pck.declared_value AS package_declared_value,
      pck.counter_delivery AS package_counter_delivery,
      b.passenger_id AS passenger_id,
      b.driver_id AS driver_id,
      b.served_vehicle_type_id AS served_vehicle_type_id,
      b.city_id AS city_id,
      b.requested_service_type_id AS requested_service_type_id,
      b.country_id AS country_id,
      wa._id AS wallet_account_id,
      p.company_id AS passenger_company_id,
      p.name AS passenger_name,
      d.name AS driver_name
    FROM picapmongoprod.wallet_account_transactions AS wat FINAL
    INNER JOIN picapmongoprod.packages AS pck FINAL ON pck._id = wat.package_id
    INNER JOIN picapmongoprod.bookings AS b FINAL ON b._id = wat.booking_id
    INNER JOIN picapmongoprod.wallet_accounts AS wa FINAL ON wa._id = wat.account_id
    INNER JOIN picapmongoprod.passengers AS p FINAL ON p._id = b.passenger_id
    INNER JOIN picapmongoprod.passengers AS d FINAL ON d._id = b.driver_id
    INNER JOIN picapmongoprod.countries AS c FINAL ON c._id = b.country_id
    WHERE
      JSONExtractString(c.name, 'es') = 'Colombia'
      AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
      AND pck.counter_delivery = 'true'
      AND wat._type = 'WalletAccountCounterDeliveryPaymentTransaction'
      AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
          BETWEEN toDate('{desde}') AND toDate('{hasta}')
  )
SELECT
  toDate(toTimeZone(qtf.created_at, 'America/Bogota'))                        AS Date_transaction,
  JSONExtractString(qtf.amount, 'currency_iso')                               AS Transaction_currency,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100               AS Transaction_amount,
  qtf.transaction_id                                                          AS Transaction_ID,
  toFloat64OrZero(JSONExtractString(qtf.normalized_amount_after_transaction, 'cents')) / 100 AS Normalized_Amount_After_Transaction,
  toDate(toTimeZone(qtf.booking_created_at, 'America/Bogota'))                AS Date_booking,
  qtf.booking_id                                                              AS ID_Booking,
  qtf.package_id                                                              AS ID_Package,
  qtf.package_reference                                                       AS Reference,
  qtf.passenger_id                                                            AS ID_User,
  compp.name                                                                  AS User_Company,
  qtf.passenger_name                                                          AS User_Name,
  toFloat64OrZero(JSONExtractString(qtf.package_declared_value, 'cents')) / 100 AS Declared_Value,
  qtf.transaction_state_cd                                                    AS transaction_state_cd,
  qtf.driver_id                                                               AS ID_Driver,
  qtf.driver_name                                                             AS Driver_Name,
  st.type                                                                     AS type,
  st.name                                                                     AS service_type_name,
  JSONExtractString(cit.name, 'es')                                           AS Ciudad,
  JSONExtractString(vt.name, 'es')                                            AS name_vehicle,
  multiIf(sf.new_final_score_rent < 0, 0, sf.new_final_score_rent >= 5, 5, sf.new_final_score_rent) AS score_rent_fixed,
  multiIf(sf.new_final_score_pibox < 0, 0, sf.new_final_score_pibox >= 5, 5, sf.new_final_score_pibox) AS score_pibox_fixed
FROM q_wat_filtered qtf
LEFT JOIN picapmongoprod.companies                          AS compp FINAL ON compp._id = qtf.passenger_company_id
LEFT JOIN picapmongoprod.vehicle_types                      AS vt    FINAL ON vt._id    = qtf.served_vehicle_type_id
LEFT JOIN picapmongoprod.cities                             AS cit   FINAL ON cit._id   = qtf.city_id
LEFT JOIN q_service_types                                   AS st          ON st._id    = qtf.requested_service_type_id
LEFT JOIN picapmongoprod.vw_atr_driver_scoring_with_frauds  AS sf    FINAL ON sf.driver_id = qtf.driver_id
ORDER BY Date_transaction ASC
"""


QUERY_TRANSACCIONES = """
WITH q_service_types AS (
  SELECT
    _id,
    any(name_es) AS name,
    any(multiIf(
      name_es IN (
        'Pibox (Mensajería)', 'Mensajería en bicicleta', 'Moto Favor',
        'Moto favor', 'Carga', 'Carga Carry', 'Carga Moto-Vagón',
        'Carga NHR', 'Carga NKR', 'Carga NPR', 'Mensajería',
        'Carro Mensajeria', 'Carga Trailer', 'NHR Refrigerada'
      ), 'Pibox',
      name_es IN (
        'Moto', 'Mototaxi', 'Moto sin conductor', 'Subasta', 'Taxi',
        'Carro', 'Rapidín', 'Carro sin conductor', 'Moto VIP',
        'Moto Económica', 'Carro Queen', 'Espero tranqui',
        'Moto lite', 'Moto Queen', 'Picap Carro', 'Picap Moto',
        'Grúa Carro', 'Grúa Moto'
      ), 'Picap',
      'Other'
    )) AS type
  FROM picapmongoprod.service_types
  GROUP BY _id
),
q_transactions_filtered AS (
  SELECT
    wat._id AS _id,
    wat.booking_id,
    wat.account_id,
    wat._type AS txt_type,
    wat.created_at AS created_at,
    wat.amount,
    s.payment_method_cd,
    toFloat64OrZero(JSONExtractString(s.amount_charged_to_passenger_wallet, 'cents')) / 100 AS amount_charged_to_passenger_wallet,
    toFloat64OrZero(JSONExtractString(s.amount_charged_to_company_wallet, 'cents')) / 100   AS amount_charged_to_company_wallet,
    wa.passenger_id AS passenger_id,
    comp._id   AS company_id,
    comp.name  AS company_name,
    st.type    AS service_type
  FROM picapmongoprod.wallet_account_transactions wat FINAL
  ANY LEFT JOIN picapmongoprod.bookings        s    FINAL ON s._id    = wat.booking_id
  ANY LEFT JOIN q_service_types                st         ON st._id   = s.requested_service_type_id
  ANY LEFT JOIN picapmongoprod.wallet_accounts wa   FINAL ON wa._id   = wat.account_id
  ANY LEFT JOIN picapmongoprod.companies       comp FINAL ON comp._id = s.company_id
  WHERE
    wat._type IN (
      'WalletAccountCounterDeliveryPaymentTransaction',
      'WalletAccountTransactionBookingCompanyCharge',
      'WalletAccountTransactionCommissionCompanyPayment'
    )
    AND JSONExtractString(wat.amount, 'currency_iso') = 'COP'
    AND st.type = 'Pibox'
    AND toDate(toTimeZone(wat.created_at, 'America/Bogota'))
        BETWEEN toDate('{desde}') AND toDate('{hasta}')
),
q_transactions AS (
  SELECT
    t.booking_id,
    JSONExtractString(t.amount, 'currency_iso') AS currency,
    SUM(IF(t._type = 'WalletAccountTransactionBookingDriverPayment',
           JSONExtractFloat(t.amount, 'cents') / 100, 0)) AS booking_driver_payment
  FROM picapmongoprod.wallet_account_transactions t FINAL
  INNER JOIN (SELECT DISTINCT booking_id FROM q_transactions_filtered) qtf ON t.booking_id = qtf.booking_id
  GROUP BY t.booking_id, currency
),
q_payment_methods AS (
  SELECT
    b._id AS booking_id,
    multiIf(
      b.payment_method_cd = '1', 'Cash',
      b.payment_method_cd = '2', 'Voucher',
      b.payment_method_cd = '3', 'Credit Card',
      'Other'
    ) AS txt_payment_method
  FROM picapmongoprod.bookings b FINAL
  WHERE b._id IN (SELECT booking_id FROM q_transactions_filtered)
  GROUP BY b._id, b.payment_method_cd
)
SELECT
  qtf.passenger_id,
  qtf.company_id,
  qtf.company_name                                                     AS Company_name,
  qtf.txt_type                                                         AS TXT_TYPE,
  toDate(toTimeZone(qtf.created_at, 'America/Bogota'))                 AS TMS_CREATED,
  qtf.booking_id,
  toFloat64OrZero(JSONExtractString(qtf.amount, 'cents')) / 100        AS VAL_AMOUNT,
  qtf._id,
  ifNull(t.booking_driver_payment, 0)                                  AS VAL_AMOUNT_BOOKING_DRIVER_PAYMENT,
  multiIf(
    (pm.txt_payment_method = 'Cash') AND (t.booking_driver_payment != 0), 'Company Wallet',
    pm.txt_payment_method != 'Cash', pm.txt_payment_method,
    qtf.amount_charged_to_company_wallet > 0, 'Company Wallet',
    'Cash'
  ) AS Payment_Type
FROM q_transactions_filtered qtf
LEFT JOIN q_transactions  t  ON t.booking_id  = qtf.booking_id
LEFT JOIN q_payment_methods pm ON pm.booking_id = qtf.booking_id
ORDER BY TMS_CREATED, qtf.booking_id
"""


def query_company_collection_fee(company_ids: list[str], host, user, password) -> dict[str, float]:
    """Devuelve dict {company_id: collection_fee_numero} para los IDs dados.
    collection_fee_numero está como porcentaje (ej. 1 = 1%) en formato decimal (0.01)."""
    if not company_ids:
        return {}
    ids_str = ",".join(f"'{cid}'" for cid in company_ids if cid)
    sql = f"""
      SELECT _id, toFloat64OrZero(collection_fee) / 100.0 AS fee
      FROM picapmongoprod.companies FINAL
      WHERE _id IN ({ids_str})
        AND collection_fee IS NOT NULL AND collection_fee != ''
    """
    text = ch_request(sql, host, user, password, timeout=120)
    _hdrs, rows = parse_tsv(text)
    out = {}
    for r in rows:
        if len(r) >= 2:
            try:
                out[r[0]] = float(r[1])
            except ValueError:
                pass
    return out


# ════════════════════════════════════════════════════════════════════════════
# Generación Excel
# ════════════════════════════════════════════════════════════════════════════

# Paletas / estilos comunes
HDR_FILL = PatternFill("solid", start_color="1F4E78")  # azul corporativo
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="left", vertical="center")

PIVOT_HDR_FILL = PatternFill("solid", start_color="D9E1F2")
PIVOT_TOTAL_FILL = PatternFill("solid", start_color="FFE699")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEDA_FMT = '_-"$"* #,##0.00_-;-"$"* #,##0.00_-;_-"$"* "-"_-;_-@_-'
INT_FMT = "#,##0"


def write_query_sheet(ws, headers: list[str], rows: list[list[str]]):
    """Escribe data cruda con headers azul + freeze pane en fila 2."""
    # Headers
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
    # Data
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, v in enumerate(r, start=1):
            ws.cell(row=row_idx, column=col_idx, value=to_num(v))
    ws.freeze_panes = "A2"
    # Anchos auto razonables
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def build_td_recaudos(wb, query_recaudos_rows: list[dict]) -> None:
    """Hoja 2: pivote User_Company → Σ Transaction_amount, ordenado DESC."""
    por_emp = defaultdict(float)
    for r in query_recaudos_rows:
        emp = r.get("User_Company") or ""
        if emp == "":
            continue
        try:
            por_emp[emp] += float(r.get("Transaction_amount") or 0)
        except (ValueError, TypeError):
            pass
    sorted_emp = sorted(por_emp.items(), key=lambda kv: -kv[1])

    ws = wb.create_sheet("TD Recaudos")
    # Header pivote
    c = ws.cell(row=3, column=1, value="Etiquetas de fila")
    c.font = Font(bold=True); c.fill = PIVOT_HDR_FILL; c.border = BORDER_ALL
    c = ws.cell(row=3, column=2, value="Suma de Transaction_amount")
    c.font = Font(bold=True); c.fill = PIVOT_HDR_FILL; c.border = BORDER_ALL

    for i, (emp, monto) in enumerate(sorted_emp, start=4):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2))
        c.number_format = INT_FMT
        c.border = BORDER_ALL

    total = sum(v for _, v in sorted_emp)
    r_total = 4 + len(sorted_emp)
    c = ws.cell(row=r_total, column=1, value="Total general")
    c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL; c.border = BORDER_ALL
    c = ws.cell(row=r_total, column=2, value=round(total, 2))
    c.number_format = INT_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL; c.border = BORDER_ALL

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 24


def write_query_transacciones(ws, headers: list[str], rows: list[list[str]]):
    """
    Hoja 3: data cruda con cols calculadas usando referencias DIRECTAS.
      - DIA entre TMS_CREATED (E) y booking_id (G)  → =DAY(E{row})
      - Recaudos al final  → =IF(D{row}="WalletAccountCounter...", H{row}, 0)
      - Servicios al final → =IF anidado con los 3 tipos

    NOTA HISTORICA: una versión previa usaba Excel Tables + refs estructuradas
    [@[col]]. La combinación de openpyxl + tabla + IFS generaba archivos XLSX
    que openpyxl podía leer pero Excel rechazaba ("Error en el método Open").
    Las refs directas funcionan en CUALQUIER versión de Excel sin riesgo de
    corrupción. Si el usuario quiere refs [@[col]], puede convertir el rango
    a Tabla manualmente desde Excel (Insertar > Tabla) después de abrir.
    """
    # Limpiar headers por las dudas (no esencial sin tabla, pero buena práctica)
    headers_clean = [h.replace(".", "_") for h in headers]

    # Reordenar headers: insertar "DIA" en pos 6 (después de TMS_CREATED que es col 5)
    new_headers = headers_clean[:5] + ["DIA"] + headers_clean[5:] + ["Recaudos", "Servicios"]

    # Headers con estilo
    for col_idx, h in enumerate(new_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN

    n_cols_orig = len(headers_clean)

    for row_idx, r in enumerate(rows, start=2):
        # Cols 1..5 (passenger_id, company_id, Company_name, TXT_TYPE, TMS_CREATED)
        for orig_idx in range(5):
            v = r[orig_idx] if orig_idx < len(r) else None
            ws.cell(row=row_idx, column=orig_idx + 1, value=to_num(v))
        # Col 6 = DIA con ref directa a E{row}
        ws.cell(row=row_idx, column=6, value=f"=DAY(E{row_idx})").number_format = "0"
        # Cols 7..(6+n_orig-5) = orig 5..N-1 (booking_id, VAL_AMOUNT, _id, VAL_AMOUNT_BDP, Payment_Type)
        for orig_idx in range(5, n_cols_orig):
            v = r[orig_idx] if orig_idx < len(r) else None
            ws.cell(row=row_idx, column=orig_idx + 2, value=to_num(v))
        # Recaudos = SI con ref directa
        col_recaudos = n_cols_orig + 2
        ws.cell(
            row=row_idx, column=col_recaudos,
            value=f'=IF(D{row_idx}="WalletAccountCounterDeliveryPaymentTransaction",H{row_idx},0)',
        ).number_format = INT_FMT
        # Servicios = IF anidado (más compatible que IFS en archivos generados por openpyxl)
        col_servicios = n_cols_orig + 3
        ws.cell(
            row=row_idx, column=col_servicios,
            value=(
                f'=IF(D{row_idx}="WalletAccountTransactionBookingCompanyCharge",H{row_idx},'
                f'IF(D{row_idx}="WalletAccountTransactionCommissionCompanyPayment",H{row_idx},'
                f'IF(D{row_idx}="WalletAccountCounterDeliveryPaymentTransaction",0,0)))'
            ),
        ).number_format = INT_FMT

    ws.freeze_panes = "A2"
    for c in range(1, len(new_headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def build_mensual(wb, query_tx_rows: list[dict], comision_por_company: dict[str, float]) -> None:
    """
    Hoja 4 - Mensual: pivote por Company_name con cols calculadas.
    Headers:
      A Etiquetas de fila | B Suma de Recaudos | C Suma de Servicios | D DIF
      E Comisión Recaudo  | F Retenciones    | G TOTAL              | H Facturas pendientes
      I Porcentaje comisión
    """
    # Agrupar por Company_name (y trackear company_id para lookup de comisión)
    por_emp = defaultdict(lambda: {"recaudos": 0.0, "servicios": 0.0, "company_id": ""})
    for r in query_tx_rows:
        emp = r.get("Company_name") or ""
        if not emp:
            continue
        txt_type = r.get("TXT_TYPE") or ""
        val = float(r.get("VAL_AMOUNT") or 0)
        if txt_type == "WalletAccountCounterDeliveryPaymentTransaction":
            por_emp[emp]["recaudos"] += val
        elif txt_type in ("WalletAccountTransactionBookingCompanyCharge",
                          "WalletAccountTransactionCommissionCompanyPayment"):
            por_emp[emp]["servicios"] += val
        por_emp[emp]["company_id"] = r.get("company_id") or por_emp[emp]["company_id"]

    sorted_emp = sorted(por_emp.items(), key=lambda kv: kv[0].lower())

    ws = wb.create_sheet("Mensual")

    # Filtros (mock — solo display, openpyxl no soporta pivot filters)
    ws.cell(row=1, column=1, value="Payment_Type").font = Font(bold=True)
    ws.cell(row=1, column=2, value="(Todas)")
    ws.cell(row=2, column=1, value="DIA").font = Font(bold=True)
    ws.cell(row=2, column=2, value="(Todas)")

    headers = [
        "Etiquetas de fila", "Suma de Recaudos", "Suma de Servicios", "DIF",
        "Comisión Recaudo", "Retenciones", "TOTAL", "Facturas pendientes", "Porcentaje comisión",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(bold=True); c.fill = PIVOT_HDR_FILL; c.border = BORDER_ALL

    # Filas data
    for i, (emp, info) in enumerate(sorted_emp, start=5):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(info["recaudos"], 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        c = ws.cell(row=i, column=3, value=round(info["servicios"], 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        # DIF = Servicios + Recaudos
        c = ws.cell(row=i, column=4, value=f"=C{i}+B{i}")
        c.number_format = INT_FMT; c.border = BORDER_ALL
        # Comisión Recaudo = SI.ERROR(Recaudos * % * -1, 0)
        c = ws.cell(row=i, column=5, value=f"=IFERROR(B{i}*I{i}*-1,0)")
        c.number_format = INT_FMT; c.border = BORDER_ALL
        # Retenciones = SI(Servicios<>0, -(Servicios*9.66)/1000, 0)
        c = ws.cell(row=i, column=6, value=f"=IF(C{i}<>0,(-(C{i}*9.66)/1000),0)")
        c.number_format = INT_FMT; c.border = BORDER_ALL
        # TOTAL = SUMA(DIF:Retenciones)  → D:F
        c = ws.cell(row=i, column=7, value=f"=SUM(D{i}:F{i})")
        c.number_format = INT_FMT; c.border = BORDER_ALL; c.font = Font(bold=True)
        # Facturas pendientes: vacía
        ws.cell(row=i, column=8, value=None).border = BORDER_ALL
        # Porcentaje comisión: lookup desde companies
        pct = comision_por_company.get(info["company_id"], 0.0)
        c = ws.cell(row=i, column=9, value=pct)
        c.number_format = "0.00%"; c.border = BORDER_ALL

    # Total general
    r_total = 5 + len(sorted_emp)
    c = ws.cell(row=r_total, column=1, value="Total general")
    c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL; c.border = BORDER_ALL
    for col, letra in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G")]:
        c = ws.cell(row=r_total, column=col, value=f"=SUM({letra}5:{letra}{r_total - 1})")
        c.number_format = INT_FMT; c.font = Font(bold=True); c.fill = PIVOT_TOTAL_FILL; c.border = BORDER_ALL

    ws.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A5"


def build_control(wb, td_recaudos_data: list[tuple[str, float]]) -> None:
    """
    Hoja 5 - Control: User Company, Suma de $ Transaction, Pendiente al 31, TOTAL, Comentario.
    Reglas de comentario:
      - PIBOX ADMIN          → 'Pruebas'
      - Surtitodo express    → 'Cero, más servicios que recaudo'
      - Cualquier otro       → 'Cliente ida y vuelta'

    Tabla resumen Surtitodo (filas 12-17 del Control):
      - Recaudos:  ref a fila de Surtitodo en la tabla principal
      - Servicios: VLOOKUP a hoja Mensual columna C
      - Comisión:  = -Recaudos * 1%
      - ICA:       = (-Servicios * 9.66) / 1000
      - Total:     = SUM(B13:B16)
    """
    ws = wb.create_sheet("Control")
    headers = ["User Company", "Suma de $ Transaction", "Pendiente al 31", "TOTAL", "Comentario"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN

    # Mapeo de clientes de prueba (NO son comerciales reales, son cuentas internas).
    # Cualquiera de estos -> comentario "Pruebas".
    CLIENTES_PRUEBA = {"PIBOX ADMIN", "TESTEO 2"}

    def comentario_para(emp: str) -> str:
        if emp.upper().strip() in CLIENTES_PRUEBA:
            return "Pruebas"
        if emp.strip().lower() == "surtitodo express":
            return "Cero, más servicios que recaudo"
        return "Cliente ida y vuelta"

    surtitodo_row = None  # captura la fila donde está Surtitodo para la tabla resumen
    for i, (emp, monto) in enumerate(td_recaudos_data, start=2):
        ws.cell(row=i, column=1, value=emp).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=round(monto, 2))
        c.number_format = INT_FMT; c.border = BORDER_ALL
        ws.cell(row=i, column=3, value=0).border = BORDER_ALL
        ws.cell(row=i, column=4, value=0).border = BORDER_ALL
        ws.cell(row=i, column=5, value=comentario_para(emp)).border = BORDER_ALL
        if emp.strip().lower() == "surtitodo express":
            surtitodo_row = i

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 38
    ws.freeze_panes = "A2"

    # ─── Tabla resumen Surtitodo ───
    if surtitodo_row is None:
        # Si no hay Surtitodo en el mes, no construyo la tabla
        return

    # Calculamos en qué fila empezar (espacio después de la tabla principal)
    table_start = max(12, len(td_recaudos_data) + 4)
    rows = {
        "header":    table_start,
        "recaudos":  table_start + 1,
        "servicios": table_start + 2,
        "comision":  table_start + 3,
        "ica":       table_start + 4,
        "total":     table_start + 5,
    }

    PURPLE_DARK  = PatternFill("solid", start_color="5B2169")
    PURPLE_LIGHT = PatternFill("solid", start_color="E8D7F0")
    YELLOW_FILL  = PatternFill("solid", start_color="FFFF00")
    GREEN_MINT   = PatternFill("solid", start_color="D4EDDA")
    WHITE_FONT   = Font(color="FFFFFF", bold=True, size=12)
    BOLD         = Font(bold=True)
    RED_BOLD     = Font(color="C00000", bold=True)
    RED          = Font(color="C00000")

    # Header (merged A:B)
    ws.merge_cells(f'A{rows["header"]}:B{rows["header"]}')
    c = ws.cell(row=rows["header"], column=1, value="Surtitodo abril")
    c.fill = PURPLE_DARK; c.font = WHITE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[rows["header"]].height = 22

    # Recaudos
    ws.cell(row=rows["recaudos"], column=1, value="Recaudos").fill = PURPLE_LIGHT
    ws.cell(row=rows["recaudos"], column=1).font = BOLD
    c = ws.cell(row=rows["recaudos"], column=2, value=f"=B{surtitodo_row}")
    c.fill = YELLOW_FILL; c.number_format = '#,##0.00'

    # Servicios (VLOOKUP a Mensual)
    ws.cell(row=rows["servicios"], column=1, value="Servicios").fill = PURPLE_LIGHT
    ws.cell(row=rows["servicios"], column=1).font = BOLD
    c = ws.cell(
        row=rows["servicios"], column=2,
        value='=VLOOKUP("Surtitodo express",Mensual!A:C,3,FALSE)',
    )
    c.fill = GREEN_MINT; c.number_format = '#,##0.00'

    # Comisión = -Recaudos * 1%
    ws.cell(row=rows["comision"], column=1, value="Comisión").fill = PURPLE_LIGHT
    ws.cell(row=rows["comision"], column=1).font = BOLD
    c = ws.cell(row=rows["comision"], column=2,
                value=f'=-B{rows["recaudos"]}*1%')
    c.number_format = '#,##0.00'; c.font = RED

    # ICA = (-Servicios * 9.66) / 1000
    ws.cell(row=rows["ica"], column=1, value="ICA").fill = PURPLE_LIGHT
    ws.cell(row=rows["ica"], column=1).font = BOLD
    c = ws.cell(row=rows["ica"], column=2,
                value=f'=(-B{rows["servicios"]}*9.66)/1000')
    c.number_format = '#,##0.00'

    # Total = SUM(Recaudos..ICA)
    ws.cell(row=rows["total"], column=1, value="Total").fill = PURPLE_DARK
    ws.cell(row=rows["total"], column=1).font = WHITE_FONT
    c = ws.cell(row=rows["total"], column=2,
                value=f'=SUM(B{rows["recaudos"]}:B{rows["ica"]})')
    c.number_format = '#,##0.00'; c.font = RED_BOLD

    # Bordes alrededor de la tabla
    for r in range(rows["header"], rows["total"] + 1):
        for c_idx in (1, 2):
            ws.cell(row=r, column=c_idx).border = BORDER_ALL


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Genera el reporte Excel 'Saldo Recaudos' desde ClickHouse",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--desde", help="Fecha desde (YYYY-MM-DD) — aplica a las 2 queries por default")
    parser.add_argument("--hasta", help="Fecha hasta (YYYY-MM-DD)")
    parser.add_argument("--recaudos-desde", help="Override para Query Recaudos (default = --desde)")
    parser.add_argument("--recaudos-hasta", help="Override para Query Recaudos (default = --hasta)")
    parser.add_argument("--tx-desde",       help="Override para Query Transacciones (default = --desde)")
    parser.add_argument("--tx-hasta",       help="Override para Query Transacciones (default = --hasta)")
    parser.add_argument("--output",         help="Nombre del archivo de salida (.xlsx)")
    parser.add_argument("--ch-host", default=os.environ.get("MINTIC_CH_HOST", ""))
    parser.add_argument("--ch-user", default=os.environ.get("MINTIC_CH_USER", ""))
    parser.add_argument("--ch-pass", default=os.environ.get("MINTIC_CH_PASS", ""))
    args = parser.parse_args()

    if not args.ch_host or not args.ch_user or not args.ch_pass:
        sys.exit(
            "❌ Credenciales ClickHouse no configuradas.\n"
            "   Setealas como env vars (MINTIC_CH_HOST/USER/PASS) o pasalas con\n"
            "   --ch-host --ch-user --ch-pass."
        )

    rec_d = args.recaudos_desde or args.desde
    rec_h = args.recaudos_hasta or args.hasta
    tx_d  = args.tx_desde       or args.desde
    tx_h  = args.tx_hasta       or args.hasta
    if not (rec_d and rec_h and tx_d and tx_h):
        sys.exit("❌ Faltan fechas. Usá --desde/--hasta o --recaudos-* + --tx-*.")

    output = args.output or f"Saldo_Recaudos_{rec_d}_a_{rec_h}.xlsx"
    output_path = Path(output)

    print(f"🚀 Generando reporte Recaudos BI")
    print(f"   • Query Recaudos      : {rec_d} → {rec_h}")
    print(f"   • Query Transacciones : {tx_d} → {tx_h}")
    print(f"   • Output              : {output_path}")
    print()

    # ── Query 1: Recaudos
    print("📊 Ejecutando Query Recaudos…")
    t0 = time.time()
    sql_rec = QUERY_RECAUDOS.format(desde=rec_d, hasta=rec_h)
    text_rec = ch_request(sql_rec, args.ch_host, args.ch_user, args.ch_pass)
    headers_rec, rows_rec = parse_tsv(text_rec)
    print(f"   ✓ {len(rows_rec)} filas en {time.time() - t0:.1f}s")

    # ── Query 2: Transacciones
    print("📊 Ejecutando Query Transacciones…")
    t0 = time.time()
    sql_tx = QUERY_TRANSACCIONES.format(desde=tx_d, hasta=tx_h)
    text_tx = ch_request(sql_tx, args.ch_host, args.ch_user, args.ch_pass)
    headers_tx, rows_tx = parse_tsv(text_tx)
    print(f"   ✓ {len(rows_tx)} filas en {time.time() - t0:.1f}s")

    # ── Query 3: Comisiones (collection_fee) para todas las companies de la Q2
    print("📊 Buscando % comisión por empresa…")
    company_idx = headers_tx.index("company_id") if "company_id" in headers_tx else None
    company_ids = list({r[company_idx] for r in rows_tx if company_idx is not None and r[company_idx]})
    comision_por_company = query_company_collection_fee(
        company_ids, args.ch_host, args.ch_user, args.ch_pass
    )
    print(f"   ✓ {len(comision_por_company)} empresas con collection_fee")

    # ── Pre-procesar Query Recaudos como list[dict] para los pivots
    query_recaudos_dicts = [dict(zip(headers_rec, r)) for r in rows_rec]
    query_tx_dicts = [dict(zip(headers_tx, r)) for r in rows_tx]

    # ── Construir Excel
    print()
    print("📝 Generando Excel…")
    wb = Workbook()
    # Hoja 1: Query Recaudos
    ws1 = wb.active
    ws1.title = "Query Recaudos"
    write_query_sheet(ws1, headers_rec, rows_rec)
    print(f"   ✓ Hoja 1: Query Recaudos ({len(rows_rec)} filas × {len(headers_rec)} cols)")

    # Hoja 2: TD Recaudos
    build_td_recaudos(wb, query_recaudos_dicts)
    # Sortear para usar en Control
    por_emp_td = defaultdict(float)
    for r in query_recaudos_dicts:
        emp = r.get("User_Company") or ""
        if emp:
            try:
                por_emp_td[emp] += float(r.get("Transaction_amount") or 0)
            except (ValueError, TypeError):
                pass
    td_sorted = sorted(por_emp_td.items(), key=lambda kv: -kv[1])
    print(f"   ✓ Hoja 2: TD Recaudos ({len(td_sorted)} empresas)")

    # Hoja 3: Query Transacciones
    ws3 = wb.create_sheet("Query Transacciones")
    write_query_transacciones(ws3, headers_tx, rows_tx)
    print(f"   ✓ Hoja 3: Query Transacciones ({len(rows_tx)} filas × {len(headers_tx) + 3} cols)")

    # Hoja 4: Mensual
    build_mensual(wb, query_tx_dicts, comision_por_company)
    print(f"   ✓ Hoja 4: Mensual (pivote por Company_name)")

    # Hoja 5: Control
    build_control(wb, td_sorted)
    print(f"   ✓ Hoja 5: Control ({len(td_sorted)} clientes)")

    # Guardar
    wb.save(output_path)
    print()
    print(f"✅ Listo: {output_path}")
    print(f"   Tamaño: {output_path.stat().st_size / 1024 / 1024:.1f} MB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
